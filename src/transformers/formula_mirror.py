"""
Helpers for mirroring quarter-column formulas to the LTM column and for
evaluating Management Cijfers column formulas against BDO column data.

Both utilities are pure-Python, side-effect-free (the mirror returns a
formula string; the evaluator returns numeric maps). They are used by the
Management Accounts builder, the AI verification orchestrator, and the
post-AI orchestrator refresh path so the LTM column is treated identically
to the quarter column with the only difference being column G vs H in BDO
references.
"""

from __future__ import annotations

import re
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.utils import (
    column_index_from_string,
    get_column_letter,
)


# A1-style reference matcher: optional sheet prefix, optional $-anchors,
# 1-3 letter column, 1+ digit row.
_REF_RE = re.compile(
    r"(?:'(?P<sheet>[^']+)'!)?(?P<col_anchor>\$?)(?P<col>[A-Z]{1,3})"
    r"(?P<row_anchor>\$?)(?P<row>\d+)"
)


# ---------------------------------------------------------------------------
# Mirror utility (Phase 3)
# ---------------------------------------------------------------------------

def mirror_quarter_formula_to_ltm(
    formula: str,
    *,
    bdo_sheet_name: str,
    quarter_col_letter: str,
    ltm_col_letter: str,
    bdo_quarter_letter: str = 'G',
    bdo_data_letter: str = 'H',
) -> str:
    """
    Derive the LTM-column equivalent of a quarter-column formula.

    Two transformations are applied:

    1. ``'<any sheet>'!<bdo_quarter_letter><row>`` → ``'<bdo_sheet_name>'!
       <bdo_data_letter><row>`` (so BDO references move from G to H and
       any stale BDO sheet name is rewritten to the current one).
    2. In-sheet refs to the quarter column letter are rewritten to the
       LTM column letter (e.g. ``SUM(AC23:AC24)`` → ``SUM(AD23:AD24)``)
       so calc subtotals self-reference the LTM column.

    The function is conservative — it only rewrites tokens that are
    unambiguous A1 references and never touches operators, numbers, or
    function names.
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula

    quarter_u = quarter_col_letter.upper()
    ltm_u = ltm_col_letter.upper()
    bdo_qtr_u = bdo_quarter_letter.upper()
    bdo_h_u = bdo_data_letter.upper()

    def _repl(match: re.Match) -> str:
        sheet = match.group('sheet')
        col = match.group('col').upper()
        row = match.group('row')
        col_anchor = match.group('col_anchor')
        row_anchor = match.group('row_anchor')

        if sheet is not None:
            # External (BDO) reference. Swap column letter G → H and
            # rewrite the sheet name to the current BDO sheet.
            new_col = bdo_h_u if col == bdo_qtr_u else col
            return (
                f"'{bdo_sheet_name}'!"
                f"{col_anchor}{new_col}{row_anchor}{row}"
            )

        # In-sheet reference — only rewrite the quarter column letter.
        if col == quarter_u:
            return f"{col_anchor}{ltm_u}{row_anchor}{row}"
        return match.group(0)

    return _REF_RE.sub(_repl, formula)


def mirror_quarter_column_to_ltm(
    sheet,
    *,
    quarter_col: int,
    ltm_col: int,
    bdo_sheet_name: str,
    bdo_quarter_letter: str = 'G',
    bdo_data_letter: str = 'H',
    rows: List[int],
    skip_rows: Optional[Set[int]] = None,
) -> List[Tuple[int, str, str]]:
    """
    Walk ``rows`` and overwrite the LTM-column cell of each row with the
    deterministically mirrored quarter formula (G → H, AC → AD), unless
    the row is in ``skip_rows`` or the quarter cell is not a formula.

    Returns the list of (row, before_formula, after_formula) tuples for
    rows that were actually rewritten so callers can log them.
    """
    skip_rows = skip_rows or set()
    quarter_letter = get_column_letter(quarter_col)
    ltm_letter = get_column_letter(ltm_col)
    rewrites: List[Tuple[int, str, str]] = []

    for row_idx in rows:
        if row_idx in skip_rows:
            continue
        q_val = sheet.cell(row=row_idx, column=quarter_col).value
        if not (isinstance(q_val, str) and q_val.startswith('=')):
            continue

        new_formula = mirror_quarter_formula_to_ltm(
            q_val,
            bdo_sheet_name=bdo_sheet_name,
            quarter_col_letter=quarter_letter,
            ltm_col_letter=ltm_letter,
            bdo_quarter_letter=bdo_quarter_letter,
            bdo_data_letter=bdo_data_letter,
        )
        ltm_cell = sheet.cell(row=row_idx, column=ltm_col)
        before = ltm_cell.value
        if new_formula != before:
            ltm_cell.value = new_formula
            rewrites.append((row_idx, str(before) if before is not None else '', new_formula))

    return rewrites


# ---------------------------------------------------------------------------
# Column evaluator (Phase 2)
# ---------------------------------------------------------------------------

# Patterns used while evaluating individual cell formulas.
_BDO_REF_RE = re.compile(
    r"'(?P<sheet>[^']+)'!\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)"
)
_RANGE_RE = re.compile(
    r"\$?(?P<col1>[A-Z]{1,3})\$?(?P<row1>\d+)"
    r":\$?(?P<col2>[A-Z]{1,3})\$?(?P<row2>\d+)"
)
_INSHEET_REF_RE = re.compile(
    r"\$?(?P<col>[A-Z]{1,3})\$?(?P<row>\d+)"
)


def _read_numeric(value: Any) -> Optional[float]:
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped or stripped.startswith('='):
            return None
        try:
            return float(stripped.replace(',', '.').replace(' ', ''))
        except ValueError:
            return None
    return None


def _resolve_bdo_value(bdo_sheet, bdo_data_only_sheet, row: int, col: int) -> float:
    """Read a BDO cell, preferring the data_only cache for formula cells."""
    if bdo_data_only_sheet is not None:
        cached = bdo_data_only_sheet.cell(row=row, column=col).value
        num = _read_numeric(cached)
        if num is not None:
            return num
    raw = bdo_sheet.cell(row=row, column=col).value if bdo_sheet else None
    num = _read_numeric(raw)
    if num is not None:
        return num
    if isinstance(raw, str) and raw.startswith('=SUM('):
        # Fallback for the common Resultaat na belasting H = SUM(C:G)
        # case so we don't return 0 just because openpyxl skipped caching.
        m = re.match(
            r"^=SUM\(\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)\)$",
            raw, re.IGNORECASE,
        )
        if m:
            c1 = column_index_from_string(m.group(1).upper())
            r1 = int(m.group(2))
            c2 = column_index_from_string(m.group(3).upper())
            r2 = int(m.group(4))
            total = 0.0
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    v = _read_numeric(bdo_sheet.cell(row=rr, column=cc).value)
                    if v is not None:
                        total += v
            return total
    return 0.0


def _evaluate_formula_string(
    formula: str,
    *,
    target_col: int,
    row_values: Dict[int, float],
    target_formula_rows: Set[int],
    bdo_sheet,
    bdo_data_only_sheet,
    sheet,
    bdo_sheet_name: str,
) -> Optional[float]:
    """
    Evaluate a single MA cell formula and return its numeric value, or
    None when the formula references a row that has not been computed yet
    (so the caller can re-queue it for a later pass).

    The evaluator handles:
    - BDO refs: ``'<bdo>'!H123`` → numeric BDO cell value.
    - In-sheet refs in the target column: ``AD23`` → ``row_values[23]``.
    - Ranges in the target column: ``SUM(AD23:AD24)`` and similar.
    - Standard arithmetic operators and SUM/MAX/MIN/IF (best-effort —
      anything beyond simple SUM falls back to Python ``eval``).
    """
    if not isinstance(formula, str) or not formula.startswith('='):
        return _read_numeric(formula)

    expr = formula[1:]

    def _range_repl(match: re.Match) -> str:
        c1_idx = column_index_from_string(match.group('col1').upper())
        c2_idx = column_index_from_string(match.group('col2').upper())
        r1 = int(match.group('row1'))
        r2 = int(match.group('row2'))
        if c1_idx > c2_idx:
            c1_idx, c2_idx = c2_idx, c1_idx
        if r1 > r2:
            r1, r2 = r2, r1
        total = 0.0
        for rr in range(r1, r2 + 1):
            for cc in range(c1_idx, c2_idx + 1):
                if cc == target_col:
                    if rr in row_values:
                        total += row_values[rr]
                    elif rr in target_formula_rows:
                        # Row has a formula that hasn't been evaluated
                        # yet on this pass — re-queue.
                        return 'NaN'
                    # Else: empty cell or static value already excluded;
                    # contributes 0 (Excel SUM ignores blanks).
                else:
                    v = _read_numeric(sheet.cell(row=rr, column=cc).value)
                    if v is not None:
                        total += v
        return repr(total)

    def _bdo_repl(match: re.Match) -> str:
        col_letter = match.group('col').upper()
        row = int(match.group('row'))
        col_idx = column_index_from_string(col_letter)
        return repr(_resolve_bdo_value(
            bdo_sheet, bdo_data_only_sheet, row, col_idx,
        ))

    def _inhseet_repl(match: re.Match) -> str:
        col_letter = match.group('col').upper()
        row = int(match.group('row'))
        col_idx = column_index_from_string(col_letter)
        if col_idx == target_col:
            if row in row_values:
                return repr(row_values[row])
            if row in target_formula_rows:
                return 'NaN'
            return '0'
        v = _read_numeric(sheet.cell(row=row, column=col_idx).value)
        return repr(v if v is not None else 0.0)

    expr = _BDO_REF_RE.sub(_bdo_repl, expr)
    expr = _RANGE_RE.sub(_range_repl, expr)
    expr = _INSHEET_REF_RE.sub(_inhseet_repl, expr)

    expr = re.sub(r'\bSUM\(([^()]*)\)', lambda m: f"({m.group(1).replace(',', '+') or '0'})", expr, flags=re.IGNORECASE)

    if 'NaN' in expr:
        return None

    safe = re.sub(r'[^0-9eE.+\-*/()\s]', '', expr)
    if not safe:
        return None
    try:
        return float(eval(safe, {'__builtins__': {}}, {}))
    except Exception:
        return None


def _evaluate_one_column(
    *,
    sheet,
    bdo_sheet,
    bdo_data_only_sheet,
    bdo_sheet_name: str,
    target_col: int,
    rows_of_interest: List[int],
    max_passes: int,
) -> Dict[str, Any]:
    """
    Inner evaluator that operates on already-opened openpyxl sheets.
    Shared by ``evaluate_ma_column`` and ``evaluate_ma_columns`` so a
    single workbook open can serve multiple target columns.
    """
    formulas: Dict[int, str] = {}
    static_values: Dict[int, float] = {}
    for row_idx in rows_of_interest:
        cell_val = sheet.cell(row=row_idx, column=target_col).value
        if isinstance(cell_val, str) and cell_val.startswith('='):
            formulas[row_idx] = cell_val
        else:
            num = _read_numeric(cell_val)
            if num is not None:
                static_values[row_idx] = num

    row_values: Dict[int, float] = dict(static_values)
    pending = set(formulas.keys())
    target_formula_rows = set(formulas.keys())
    for _ in range(max_passes):
        if not pending:
            break
        progressed = False
        for row_idx in list(pending):
            value = _evaluate_formula_string(
                formulas[row_idx],
                target_col=target_col,
                row_values=row_values,
                target_formula_rows=target_formula_rows,
                bdo_sheet=bdo_sheet,
                bdo_data_only_sheet=bdo_data_only_sheet,
                sheet=sheet,
                bdo_sheet_name=bdo_sheet_name,
            )
            if value is not None:
                row_values[row_idx] = value
                pending.discard(row_idx)
                progressed = True
        if not progressed:
            break

    return {
        'values': row_values,
        'unresolved': sorted(pending),
    }


def _resolve_ma_sheets(
    *,
    wb,
    wb_data,
    summary_sheet_name: str,
    bdo_sheet_name: str,
):
    """Resolve the MA summary and BDO sheet handles from open workbooks."""
    sheet = (
        wb[summary_sheet_name]
        if summary_sheet_name in wb.sheetnames else None
    )
    if sheet is None:
        for sn in wb.sheetnames:
            if 'Management Cijfers' in sn:
                sheet = wb[sn]
                break
    bdo_sheet = (
        wb[bdo_sheet_name] if bdo_sheet_name in wb.sheetnames else None
    )
    bdo_data_only_sheet = (
        wb_data[bdo_sheet_name]
        if bdo_sheet_name in wb_data.sheetnames else None
    )
    return sheet, bdo_sheet, bdo_data_only_sheet


def evaluate_ma_column(
    workbook_path: str,
    *,
    summary_sheet_name: str,
    bdo_sheet_name: str,
    target_col: int,
    rows_of_interest: List[int],
    max_passes: int = 6,
    workbooks: Optional[Tuple[Any, Any]] = None,
) -> Dict[str, Any]:
    """
    Evaluate every formula in ``target_col`` of the Management Cijfers
    sheet for the given rows, resolving BDO refs against the BDO sheet
    and in-sheet refs against already-computed values.

    Returns ``{"values": {row: float}, "unresolved": [row, ...]}``.

    Multiple passes are performed (default 6) so calc subtotals like
    ``SUM(AD3:AD18)`` resolve once their dependencies have a value.

    Pass ``workbooks=(wb, wb_data)`` (already-opened ``openpyxl`` workbooks
    in formulas mode and data-only mode respectively) to share the open
    handles with another evaluator call. When provided, this function
    will not close them.
    """
    if workbooks is not None:
        wb, wb_data = workbooks
        owns_handles = False
    else:
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
        wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
        owns_handles = True

    try:
        sheet, bdo_sheet, bdo_data_only_sheet = _resolve_ma_sheets(
            wb=wb,
            wb_data=wb_data,
            summary_sheet_name=summary_sheet_name,
            bdo_sheet_name=bdo_sheet_name,
        )
        if sheet is None:
            return {'values': {}, 'unresolved': list(rows_of_interest)}

        return _evaluate_one_column(
            sheet=sheet,
            bdo_sheet=bdo_sheet,
            bdo_data_only_sheet=bdo_data_only_sheet,
            bdo_sheet_name=bdo_sheet_name,
            target_col=target_col,
            rows_of_interest=rows_of_interest,
            max_passes=max_passes,
        )
    finally:
        if owns_handles:
            wb.close()
            wb_data.close()


def evaluate_ma_columns(
    workbook_path: str,
    *,
    summary_sheet_name: str,
    bdo_sheet_name: str,
    columns: List[Tuple[str, int]],
    rows_of_interest: List[int],
    max_passes: int = 6,
) -> Dict[str, Dict[str, Any]]:
    """
    Evaluate multiple MA columns in a single workbook session.

    ``columns`` is a list of ``(label, target_col)`` tuples; each label is
    used as the key in the returned dict (e.g. ``[('quarter', 29),
    ('ltm', 30)]``). Returns
    ``{label: {"values": {row: float}, "unresolved": [...]}}``.

    Equivalent to calling :func:`evaluate_ma_column` once per ``(label,
    target_col)`` pair, but opens the workbook only once (saves a few
    seconds per AI verification round when the file is large).
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
    try:
        sheet, bdo_sheet, bdo_data_only_sheet = _resolve_ma_sheets(
            wb=wb,
            wb_data=wb_data,
            summary_sheet_name=summary_sheet_name,
            bdo_sheet_name=bdo_sheet_name,
        )
        results: Dict[str, Dict[str, Any]] = {}
        if sheet is None:
            for label, _ in columns:
                results[label] = {
                    'values': {},
                    'unresolved': list(rows_of_interest),
                }
            return results

        for label, target_col in columns:
            results[label] = _evaluate_one_column(
                sheet=sheet,
                bdo_sheet=bdo_sheet,
                bdo_data_only_sheet=bdo_data_only_sheet,
                bdo_sheet_name=bdo_sheet_name,
                target_col=target_col,
                rows_of_interest=rows_of_interest,
                max_passes=max_passes,
            )
        return results
    finally:
        wb.close()
        wb_data.close()
