"""
Management Accounts Transformer

Builds the Management Accounts Excel file by:
1. Copying ALL data from the BDO file (Cijfers_QSP_) into a new sheet
2. Inserting a new column in Management Cijfers after the latest Q column (before FY/LTM)
3. Building formulas for the new column using formula templates (account code mapping)
4. Updating formulas to reference the new BDO sheet with correct row numbers
5. Ensuring Management Cijfers sheet is always at the end

WORKFLOW:
- Copy all data from BDO file → Create new sheet "BDO - Q{quarter}-{YY}"
- Build account code to row mapping from new BDO sheet
- In Management Cijfers sheet, insert new column after latest quarter (before FY/LTM)
- Use formula templates to build formulas for new column (looking up account codes)
- Update LTM column to reference new BDO sheet with column H
- Move Management Cijfers sheet to be the last sheet

KEY DESIGN:
- Formula templates define which account codes each Management Cijfers row references
- When building new column, look up account codes in the new BDO sheet to get correct row numbers
- Only the NEW column gets updated formulas; existing columns remain unchanged
"""

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.styles.fills import FILL_SOLID
from copy import copy
import yaml
import re
from datetime import datetime
from loguru import logger

from ..parsers.bdo_parser import BDOParseResult, AccountEntry


def _shift_formula_column_letter(formula: str, from_letter: str, to_letter: str) -> str:
    """
    In an Excel formula, replace A1-style references that use from_letter as the
    column with to_letter (exact column match only, so A does not match AA).
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula
    from_u = from_letter.upper()

    def repl(match) -> str:
        d1, col, d2, row_digits = match.group(1), match.group(2), match.group(3), match.group(4)
        if col.upper() == from_u:
            return f'{d1}{to_letter}{d2}{row_digits}'
        return match.group(0)

    return re.sub(
        r'(\$?)([A-Z]{1,3})(\$?)(\d+)',
        repl,
        formula,
        flags=re.IGNORECASE,
    )


_QUARTER_HEADER_RE = re.compile(r'^Q\s*([1-4])\s+(\d{4})\s*$', re.IGNORECASE)
_FY_HEADER_RE = re.compile(r'^FY\s*(\d{4})\s*$', re.IGNORECASE)
_LTM_IN_HEADER_RE = re.compile(r'\bLTM\b', re.IGNORECASE)


@dataclass(frozen=True)
class SummaryColumnLayout:
    """Parsed Management Cijfers row-22 column structure before inserting a new quarter."""
    quarter_columns: Tuple[int, ...]
    fy_columns: Tuple[int, ...]
    ltm_column: Optional[int]
    last_quarter_column: Optional[int]
    insert_column: int
    previous_quarter_column: Optional[int]


def scan_summary_column_layout(
    sheet,
    header_row: int = 22,
) -> SummaryColumnLayout:
    """
    Classify row-22 headers so we insert the new quarter after the latest Q column,
    not after an FY column or before a misplaced LTM anchor.

    The workbook can have many years of historical columns, including old "LTM …"
    headers far to the left of the current quarterly block (e.g. column E in a
    workbook that now reaches column AD+).  Only an LTM header that sits *to the
    right* of the last quarter column is treated as the active LTM anchor; any LTM
    header to the left of the quarterly block is purely historical and is ignored
    for insert-position and post-insert rebuild purposes.
    """
    quarter_columns: List[int] = []
    fy_columns: List[int] = []
    ltm_column_raw: Optional[int] = None  # rightmost LTM found anywhere

    for col_idx in range(1, sheet.max_column + 1):
        raw = sheet.cell(row=header_row, column=col_idx).value
        if raw is None:
            continue
        text = str(raw).strip()
        if not text:
            continue
        if _LTM_IN_HEADER_RE.search(text):
            ltm_column_raw = col_idx   # keep updating so rightmost wins
            continue
        if _FY_HEADER_RE.match(text):
            fy_columns.append(col_idx)
            continue
        if _QUARTER_HEADER_RE.match(text):
            quarter_columns.append(col_idx)

    last_quarter_column = max(quarter_columns) if quarter_columns else None

    # Only count the LTM column as "active" when it is to the right of all
    # quarterly columns.  If it is to the left (historical section), treat it
    # as absent so ltm_column_after_insert falls through to the FY-based logic.
    if ltm_column_raw is not None and last_quarter_column is not None:
        ltm_column: Optional[int] = (
            ltm_column_raw if ltm_column_raw > last_quarter_column else None
        )
    else:
        ltm_column = ltm_column_raw

    if last_quarter_column is not None:
        insert_column = last_quarter_column + 1
        previous_quarter_column = last_quarter_column
    elif ltm_column is not None:
        insert_column = ltm_column
        previous_quarter_column = max(ltm_column - 1, 1)
    else:
        last_date_col = 1
        for col_idx in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(row=2, column=col_idx).value
            if isinstance(cell_val, datetime):
                last_date_col = col_idx
            elif isinstance(cell_val, (int, float)) and 30000 < cell_val < 60000:
                # Excel serial date in balance-sheet row 2 (legacy templates)
                last_date_col = col_idx
        insert_column = last_date_col + 1
        previous_quarter_column = max(insert_column - 1, 1)

    return SummaryColumnLayout(
        quarter_columns=tuple(quarter_columns),
        fy_columns=tuple(fy_columns),
        ltm_column=ltm_column,
        last_quarter_column=last_quarter_column,
        insert_column=insert_column,
        previous_quarter_column=previous_quarter_column,
    )


def ltm_column_after_insert(
    pre_ltm_col: Optional[int],
    insert_column: int,
    fy_columns: Tuple[int, ...],
) -> int:
    """Column index for LTM after insert_cols(insert_column).

    Only treat ``pre_ltm_col`` as the active LTM if it is at or to the right
    of the insert position — meaning it will shift right along with the other
    trailing columns.  If it is to the left it is a historical column that does
    not represent the current-quarter LTM slot; fall through to the FY-based
    (or insert+1) logic in that case.
    """
    if pre_ltm_col is not None and pre_ltm_col >= insert_column:
        return pre_ltm_col + 1
    if fy_columns:
        shifted = [c + 1 if c >= insert_column else c for c in fy_columns]
        return max(shifted) + 1
    return insert_column + 1


def find_column_by_header_label(
    sheet,
    label: str,
    header_row: int = 22,
) -> Optional[int]:
    """Find a column whose row-22 header exactly matches label (case-insensitive)."""
    target = label.strip().upper()
    for col_idx in range(1, sheet.max_column + 1):
        raw = sheet.cell(row=header_row, column=col_idx).value
        if raw is not None and str(raw).strip().upper() == target:
            return col_idx
    return None


def find_ltm_column_near_quarter(
    sheet,
    quarter_col: int,
    header_row: int = 22,
) -> Optional[int]:
    """LTM column for a quarter: layout LTM if to the right, else next LTM header."""
    layout = scan_summary_column_layout(sheet, header_row)
    if layout.ltm_column is not None and layout.ltm_column > quarter_col:
        return layout.ltm_column
    for col_idx in range(quarter_col + 1, sheet.max_column + 1):
        raw = sheet.cell(row=header_row, column=col_idx).value
        if raw and _LTM_IN_HEADER_RE.search(str(raw)):
            return col_idx
    return None


def resolve_quarter_column(
    sheet,
    quarter_label: str,
    header_row: int = 22,
    built_quarter_col: Optional[int] = None,
) -> Optional[int]:
    """
    Resolve the Management Cijfers column for the active quarter.
    Prefer the column index assigned during build, then an exact row-22 header match.
    """
    if built_quarter_col is not None:
        return built_quarter_col
    found = find_column_by_header_label(sheet, quarter_label, header_row)
    if found is not None:
        return found
    layout = scan_summary_column_layout(sheet, header_row)
    return layout.last_quarter_column


def resolve_built_ltm_column(
    sheet,
    header_row: int,
    expected_col: Optional[int],
    quarter_col: Optional[int],
) -> int:
    """
    Resolve the LTM column index after a new quarter column has been inserted.

    Priority:
    1. ``expected_col`` if its row-22 header contains "LTM" (case-insensitive).
    2. The first LTM-labelled header to the right of ``quarter_col``.
    3. ``ValueError`` with a layout dump (so the build fails loudly instead of
       silently targeting the wrong column).
    """
    if expected_col is not None:
        raw = sheet.cell(row=header_row, column=expected_col).value
        if raw is not None and _LTM_IN_HEADER_RE.search(str(raw)):
            return expected_col

    if quarter_col is not None:
        found = find_ltm_column_near_quarter(sheet, quarter_col, header_row)
        if found is not None:
            return found

    layout = scan_summary_column_layout(sheet, header_row)
    if layout.ltm_column is not None:
        return layout.ltm_column

    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col_idx).value
        if val is not None:
            headers.append(f"{get_column_letter(col_idx)}={val!r}")
    raise ValueError(
        "Could not resolve LTM column after column insert. "
        f"expected={expected_col}, quarter_col={quarter_col}, "
        f"row-{header_row} headers: {', '.join(headers)}"
    )


def resolve_built_quarter_column(
    sheet,
    quarter_label: str,
    header_row: int = 22,
    built_quarter_col: Optional[int] = None,
) -> Optional[int]:
    """Compatibility shim — prefer the column index assigned during build."""
    return resolve_quarter_column(
        sheet, quarter_label, header_row, built_quarter_col
    )


@dataclass
class ManagementAccountsConfig:
    """Configuration for Management Accounts generation."""
    quarter: str  # e.g., "Q3 2025"
    period_end: datetime
    bdo_sheet_name: str  # e.g., "BDO - Q3-25"
    summary_sheet_name: str  # e.g., "Management Cijfers - Q3 2025"
    cash_proceeds_sale: float = 0.0  # Row 50: Cash proceeds from sales tracker
    quarter_num: Optional[int] = None  # 1-4; controls FY column retention


class ManagementAccountsBuilder:
    """
    Builds updated Management Accounts workbook.
    
    The Management Accounts file contains:
    - Historical BDO sheets (one per quarter from Q1 2019 onwards)
    - Summary "Management Cijfers" sheet with calculated columns
    
    WORKFLOW:
    1. Copy ALL data from BDO file into new sheet named "BDO - Q{quarter}-{YY}"
    2. Build account code to row mapping from the new BDO sheet
    3. In Management Cijfers, insert new column before LTM column
    4. Build formulas for new quarter column using formula templates
    5. Build formulas for LTM column using same templates but with column H
    6. Move Management Cijfers sheet to be last
    
    KEY DESIGN:
    - Formula templates (config/formula_templates.yaml) define the account codes 
      each Management Cijfers row should reference
    - When adding a new column, we look up those account codes in the NEW BDO sheet
      to get the correct row numbers
    - Only the NEW column gets new formulas; existing columns remain unchanged
    """
    
    def __init__(self, previous_file_path: str, output_path: str, config: ManagementAccountsConfig,
                 bdo_source_path: str = None, mappings_path: str = "config/line_item_mappings.yaml"):
        self.previous_path = Path(previous_file_path)
        self.output_path = Path(output_path)
        self.config = config
        self.bdo_source_path = Path(bdo_source_path) if bdo_source_path else None
        self.workbook = None
        self.mappings_path = Path(mappings_path)
        self.line_item_mappings = self._load_mappings()
        
        # Load formula templates
        self.formula_templates_path = Path("config/formula_templates.yaml")
        self.formula_templates = self._load_formula_templates()
        self.balance_sheet_templates = self._load_balance_sheet_templates()
        
        # Track the previous BDO sheet name for formula updates
        self._prev_bdo_sheet_name = None
        
        # Account code to row mapping in the NEW BDO sheet
        self._new_bdo_row_map = {}
        
        # Label to row mapping in the NEW BDO sheet (for non-numeric account codes)
        self._new_bdo_label_map = {}
        
        # Account code to row mapping in the PREVIOUS BDO sheet (for row offset calculation)
        self._prev_bdo_row_map = {}
        
        # Column indices set during _update_summary_sheet
        self._new_quarter_col: Optional[int] = None
        self._new_ltm_col: Optional[int] = None
        
        # Computed numeric values for every formula cell.
        # Populated at end of build() so downstream consumers (CC builder)
        # can read values without evaluating Excel formulas.
        # Structure: {(row, 'quarter'): float, (row, 'ltm'): float}
        self.computed_values: Dict[Tuple[int, str], float] = {}

        # BDO Resultaat na belasting ground-truth values (sign-adjusted).
        # Populated by _enforce_bdo_alignment for use by validators and the
        # AI verification revalidator. NEVER substituted into computed_values
        # (rows 19/68) — those must stay formula-derived so the
        # downstream pass/fail check actually validates the workbook
        # formulas rather than tautologically comparing the cache to itself.
        # Structure: {'ltm': float | None, 'quarter': float | None}
        self.bdo_ground_truth: Dict[str, Optional[float]] = {
            'ltm': None,
            'quarter': None,
        }
        
        # Sheets that must never be modified (existing BDO/kwartaal sheets)
        self._protected_sheets: set = set()

        # Accounting rules (rounding, alignment, bank map, protected rows)
        self._rules = self._load_accounting_rules()

    # ------------------------------------------------------------------
    # Config loader
    # ------------------------------------------------------------------

    @staticmethod
    def _load_accounting_rules() -> dict:
        """
        Load config/accounting_rules.yaml with safe defaults so the builder
        works even if the file is missing.
        """
        defaults = {
            'identity_alignment': {
                'authoritative_row': 19,
                'dependent_row': 68,
                'scope': 'ltm_only',
                'max_snap_units': 2,
                'enforce_in_workbook': True,
            },
            'rounding': {
                'round_to_integer': True,
                'decimal_rows': [26],
            },
            'bank_bdo_row_map': {
                107: 35, 108: 34, 109: 36,
                110: 32, 111: 33, 112: 37, 113: 31,
            },
            'bank_total_row': 114,
            'protected_total_rows': [19, 68],
            'layout': {
                'balance_sheet_rows': [3, 19],
                'profit_loss_rows': [23, 68],
                'bank_rows': [107, 114],
                'bank_section_rows': [104, 120],
                'header_row': 22,
                'date_row': 2,
                'bs_total_row': 19,
                'bs_border_row': 18,
                'pl_total_row': 68,
                'cash_proceeds_row': 50,
                'warning_start_row': 70,
                'bold_total_rows': [19, 25, 30, 44, 45, 48, 55, 57, 66, 68],
            },
            'styling': {
                'ltm_background_color': 'FFDDEAF6',
                'text_color_on_blue': 'FF000000',
                'text_color_header': 'FF000000',
                'error_font_color': 'FFFF0000',
                'error_fill_color': 'FFFFCCCC',
            },
            'bdo': {
                'data_column': 8,
                'quarter_columns': [4, 5, 6, 7],
                'data_rows': [6, 124],
                'special_rows': [129, 136],
                'label_scan_columns': [1, 2],
            },
        }

        path = Path("config/accounting_rules.yaml")
        if not path.exists():
            logger.warning("accounting_rules.yaml not found — using built-in defaults")
            return defaults

        try:
            with open(path, 'r', encoding='utf-8') as f:
                loaded = yaml.safe_load(f) or {}

            for section, fallback in defaults.items():
                if section not in loaded:
                    loaded[section] = fallback
                elif isinstance(fallback, dict) and isinstance(loaded[section], dict):
                    for key, val in fallback.items():
                        loaded[section].setdefault(key, val)

            # Ensure bank_bdo_row_map keys are ints
            if 'bank_bdo_row_map' in loaded:
                loaded['bank_bdo_row_map'] = {
                    int(k): int(v)
                    for k, v in loaded['bank_bdo_row_map'].items()
                }

            logger.info(f"Loaded accounting rules from {path}")
            return loaded
        except Exception as e:
            logger.warning(f"Error loading accounting_rules.yaml: {e} — using defaults")
            return defaults

    @classmethod
    def reapply_interest_overrides_on_saved_workbook(
        cls, file_path: str, summary_sheet_name: str
    ) -> None:
        """
        Deterministic post-AI safety net for interest rows:
        - Row 60 LTM: copy quarter row 60 formula, swap BDO !G→!H
        - Rows 61/64 LTM: =SUM(Y:quarter_col)
        - Row 60 quarter: extend from previous column
        """
        p = Path(file_path)
        if not p.exists():
            logger.warning(f"reapply_interest_overrides: file not found {file_path}")
            return

        rules = cls._load_accounting_rules()
        shell = cls.__new__(cls)
        shell._rules = rules

        wb = openpyxl.load_workbook(p)
        try:
            if summary_sheet_name not in wb.sheetnames:
                logger.warning(
                    f"reapply_interest_overrides: sheet {summary_sheet_name!r} not in workbook"
                )
                return
            sheet = wb[summary_sheet_name]
            header_row = rules.get('layout', {}).get('header_row', 22)
            layout = scan_summary_column_layout(sheet, header_row)
            ltm_col = layout.ltm_column
            if not ltm_col or ltm_col < 3:
                logger.warning("reapply_interest_overrides: LTM column not found")
                return

            quarter_match = re.search(r'(Q[1-4]\s+\d{4})', summary_sheet_name)
            new_quarter_col = None
            if quarter_match:
                new_quarter_col = find_column_by_header_label(
                    sheet, quarter_match.group(1), header_row
                )
            if not new_quarter_col:
                new_quarter_col = layout.last_quarter_column
            if not new_quarter_col or new_quarter_col < 2:
                logger.warning("reapply_interest_overrides: quarter column not found")
                return

            shell._apply_interest_section_overrides(
                sheet, new_quarter_col, ltm_col
            )
            wb.save(p)
            logger.info(
                f"Re-applied interest overrides: {p.name} "
                f"LTM col {get_column_letter(ltm_col)} (rows 60/61/64)"
            )
        finally:
            wb.close()

    def _load_formula_templates(self) -> Dict[int, Dict[str, Any]]:
        """Load formula templates that define the account code mapping for each row."""
        templates = {}
        
        if not self.formula_templates_path.exists():
            logger.warning(f"Formula templates file not found: {self.formula_templates_path}")
            return templates
        
        try:
            with open(self.formula_templates_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
            
            # Load P&L section templates
            if 'profit_loss' in config:
                for row_str, template in config['profit_loss'].items():
                    if template:
                        templates[int(row_str)] = template
            
            logger.info(f"Loaded {len(templates)} P&L formula templates from {self.formula_templates_path}")
        except Exception as e:
            logger.warning(f"Error loading formula templates: {e}")
        
        return templates
    
    def _load_balance_sheet_templates(self) -> Dict[int, Dict[str, Any]]:
        """Load Balance Sheet formula templates."""
        templates = {}
        
        if not self.formula_templates_path.exists():
            return templates
        
        try:
            with open(self.formula_templates_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
            
            if 'balance_sheet' in config:
                for row_str, template in config['balance_sheet'].items():
                    if template:
                        templates[int(row_str)] = template
            
            logger.info(f"Loaded {len(templates)} Balance Sheet formula templates")
        except Exception as e:
            logger.warning(f"Error loading Balance Sheet templates: {e}")
        
        return templates
    
    def _load_mappings(self) -> Dict[str, Tuple[List[str], str]]:
        """Load line item mappings from YAML config file."""
        mappings = {}
        
        if not self.mappings_path.exists():
            logger.warning(f"Mappings file not found: {self.mappings_path}, using defaults")
            return self._get_default_mappings()
        
        with open(self.mappings_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        
        # Process balance sheet mappings
        if 'balance_sheet' in config:
            for label, item_config in config['balance_sheet'].items():
                if item_config:
                    accounts = item_config.get('accounts', [])
                    calc_type = item_config.get('calc_type', 'sum')
                    mappings[label] = (accounts, calc_type)
        
        # Process P&L mappings
        if 'profit_loss' in config:
            for label, item_config in config['profit_loss'].items():
                if item_config:
                    accounts = item_config.get('accounts', [])
                    calc_type = item_config.get('calc_type', 'sum')
                    mappings[label] = (accounts, calc_type)
        
        logger.info(f"Loaded {len(mappings)} line item mappings from {self.mappings_path}")
        return mappings
    
    def _get_default_mappings(self) -> Dict[str, Tuple[List[str], str]]:
        """Return default mappings if config file not found."""
        return {
            "Deferred Tax Asset": (["1790002"], "direct"),
            "Real estate": (["1600000", "1600200", "1601000", "1610803", "1611003"], "sum"),
            "Financial fixed assets": (["1760*", "1790000"], "sum"),
            "Accounts receivable": (["2000000", "2000100", "2000300"], "sum"),
            "Cash": (["2400*"], "sum"),
            "Equity": (["1000*", "1100000", "1160000"], "sum"),
            "Bank loan": (["1930001", "1930101"], "sum"),
        }
        
    def build(self, bdo_result: BDOParseResult) -> Path:
        """
        Build new Management Accounts file.
        
        Args:
            bdo_result: Parsed BDO data for current quarter
            
        Returns:
            Path to generated file
        """
        logger.info(f"Building Management Accounts for {self.config.quarter}")
        
        # Load previous quarter file
        self.workbook = openpyxl.load_workbook(self.previous_path)
        logger.info(f"Loaded previous file with {len(self.workbook.sheetnames)} sheets")
        
        # Mark all existing BDO/kwartaal sheets as protected - they must never be modified
        self._protected_sheets = {
            name for name in self.workbook.sheetnames
            if name.startswith('BDO') or name != self.config.summary_sheet_name
        }
        # The summary sheet (Management Cijfers) is the only pre-existing sheet we modify
        self._protected_sheets.discard(self.config.summary_sheet_name)
        logger.info(f"Protected sheets (will not be modified): {sorted(self._protected_sheets)}")
        
        # Get previous BDO sheet name and build row map
        bdo_sheets = [name for name in self.workbook.sheetnames if name.startswith('BDO')]
        if bdo_sheets:
            self._prev_bdo_sheet_name = bdo_sheets[-1]  # Most recent
            logger.info(f"Previous BDO sheet: {self._prev_bdo_sheet_name}")
            # Build row map from previous BDO sheet
            self._build_row_map(self.workbook[self._prev_bdo_sheet_name], self._prev_bdo_row_map)
        
        # Step 1: Copy ALL data from BDO file into new sheet
        self._copy_bdo_data_to_new_sheet(bdo_result)
        
        # Step 2: Build row map from the NEW BDO sheet (including label map for special rows)
        new_bdo_sheet = self.workbook[self.config.bdo_sheet_name]
        self._build_row_map(new_bdo_sheet, self._new_bdo_row_map, self._new_bdo_label_map)
        logger.info(f"Built row map for new BDO sheet: {len(self._new_bdo_row_map)} accounts, {len(self._new_bdo_label_map)} labels")
        
        # Step 3: Update summary sheet - insert new column, copy style/formulas
        self._update_summary_sheet()
        
        # Step 4: Update all date references throughout the workbook
        self._update_date_references()
        
        # Step 5: Move BDO sheet to second-to-last position
        self._move_bdo_sheet_to_second_last()
        
        # Step 6: Move Management Cijfers sheet to the end (after BDO)
        self._move_summary_sheet_to_end()
        
        # Step 7: Compute expected numeric values for all formula cells
        self._populate_computed_values(bdo_result)
        
        # Step 8: Enforce BDO alignment (deterministic reconciliation)
        self._enforce_bdo_alignment(bdo_result)
        
        # Re-apply interest LTM formulas (rows 61/64 SUM) and row 60 quarter
        # extension after alignment — coverage logic may overwrite these cells.
        if self._new_quarter_col and self._new_ltm_col:
            sn = self.config.summary_sheet_name
            if sn in self.workbook.sheetnames:
                self._apply_interest_section_overrides(
                    self.workbook[sn], self._new_quarter_col, self._new_ltm_col
                )
        
        # Step 9: Validate calculations
        self.validation_result = self._validate_calculations(bdo_result)

        # Step 9b: Validate LTM column formulas (header, BDO sheet name,
        # calc rows reference LTM letter not quarter letter). Errors are
        # surfaced via validation_result but do not abort the save — the
        # client still needs to see the workbook to triage.
        summary_name = self.config.summary_sheet_name
        if summary_name in self.workbook.sheetnames:
            ltm_validation = self._validate_ltm_column_formulas(
                self.workbook[summary_name]
            )
            self.validation_result['ltm_column'] = ltm_validation
            if ltm_validation['errors']:
                self.validation_result.setdefault('messages', []).extend(
                    f"LTM column: {m}" for m in ltm_validation['errors']
                )
                self.validation_result['is_aligned'] = False

        # Save output
        self.workbook.save(self.output_path)
        logger.info(f"Saved to {self.output_path}")
        
        return self.output_path
    
    def _build_row_map(self, sheet, row_map: dict, label_map: dict = None):
        """
        Build mapping of account codes to row numbers in a BDO sheet.
        Keeps FIRST occurrence of each code to avoid duplicates overwriting.
        Also builds a label map for non-numeric identifiers (like "SWAP").
        """
        row_map.clear()
        if label_map is not None:
            label_map.clear()

        for row_idx in range(1, sheet.max_row + 1):
            code = sheet.cell(row=row_idx, column=1).value
            label = sheet.cell(row=row_idx, column=2).value

            if code and isinstance(code, (str, int, float)):
                code_str = str(code).strip()
                if code_str and code_str[0].isdigit():
                    if code_str not in row_map:
                        row_map[code_str] = row_idx
                elif code_str and label_map is not None:
                    label_map[code_str.lower()] = row_idx

            if label and isinstance(label, str) and label_map is not None:
                label_map[label.strip().lower()] = row_idx

    def _discover_bdo_structure(self, bdo_sheet) -> Dict[str, Any]:
        """
        Scan the entire BDO sheet and build a complete structural map.
        Returns dict with section boundaries, all account rows (including
        duplicates), and unlabeled value rows. Nothing is hardcoded.
        """
        data_col = self._rules.get('bdo', {}).get('data_column', 8)
        structure = {
            'account_to_rows': {},
            'account_to_first': {},
            'unlabeled_value_rows': [],
            'pl_start': None,
            'resultaat_row': None,
            'duplicate_codes': {},
        }

        seen_codes: Dict[str, int] = {}

        for row_idx in range(1, bdo_sheet.max_row + 1):
            code = bdo_sheet.cell(row=row_idx, column=1).value
            h_val = bdo_sheet.cell(row=row_idx, column=data_col).value

            cell_a_str = str(code).strip() if code else ''
            if isinstance(code, str):
                lower = code.lower()
                if 'winst' in lower and 'verlies' in lower and structure['pl_start'] is None:
                    structure['pl_start'] = row_idx + 1
                if 'resultaat na belasting' in lower:
                    structure['resultaat_row'] = row_idx

            is_account = bool(cell_a_str and cell_a_str[0].isdigit())
            if is_account:
                if cell_a_str not in structure['account_to_first']:
                    structure['account_to_first'][cell_a_str] = row_idx
                structure['account_to_rows'].setdefault(cell_a_str, []).append(row_idx)

                if cell_a_str in seen_codes:
                    if cell_a_str not in structure['duplicate_codes']:
                        structure['duplicate_codes'][cell_a_str] = [seen_codes[cell_a_str]]
                    structure['duplicate_codes'][cell_a_str].append(row_idx)
                else:
                    seen_codes[cell_a_str] = row_idx
            elif h_val is not None and h_val != 0 and not is_account:
                has_value = isinstance(h_val, (int, float)) or (
                    isinstance(h_val, str) and h_val.startswith('='))
                if has_value:
                    structure['unlabeled_value_rows'].append(row_idx)

        logger.info(
            f"[BDO Structure] {len(structure['account_to_first'])} unique codes, "
            f"{len(structure['duplicate_codes'])} duplicates, "
            f"{len(structure['unlabeled_value_rows'])} unlabeled value rows, "
            f"PL start={structure['pl_start']}, resultaat={structure['resultaat_row']}"
        )
        return structure

    def _find_sum_range_end(self, bdo_sheet, start_row: int,
                            expected_count: int, prefix: Optional[str]) -> int:
        """
        Extend SUM range ONLY for contiguous rows with matching prefix.
        Stops at the first gap (empty row, no account code) or non-matching
        code. NEVER extends across section boundaries.
        """
        data_col = self._rules.get('bdo', {}).get('data_column', 8)
        end = start_row + expected_count - 1

        if not prefix:
            return end
        prefix = str(prefix)

        while end + 1 <= bdo_sheet.max_row:
            next_code = bdo_sheet.cell(row=end + 1, column=1).value
            next_h = bdo_sheet.cell(row=end + 1, column=data_col).value
            code_str = str(next_code).strip() if next_code else ''
            if not code_str or not code_str[0].isdigit():
                break
            if code_str.startswith(prefix) and next_h is not None:
                end += 1
                logger.info(
                    f"[SUM Range] Extended same-section range: "
                    f"row {end} code {code_str}"
                )
            else:
                break
        return end

    def _build_ltm_interest(self, sheet, ltm_col: int, bdo_sheet,
                            bdo_name: str, structure: Dict[str, Any]):
        """
        Build DMRRP row (65) LTM formula from BDO structure.
        Row 60 LTM is handled separately by _apply_interest_section_overrides
        (copy quarter row 60 formula, swap BDO G→H).
        """
        interest_cfg = self._rules.get('interest_section', {})
        dmrrp_code = interest_cfg.get('dmrrp_code', '')
        dmrrp_ma_row = interest_cfg.get('dmrrp_row')
        data_col_letter = get_column_letter(
            self._rules.get('bdo', {}).get('data_column', 8))

        if dmrrp_code and dmrrp_ma_row:
            dmrrp_rows = structure.get('account_to_rows', {}).get(dmrrp_code, [])
            if dmrrp_rows:
                sheet.cell(row=dmrrp_ma_row, column=ltm_col).value = (
                    f"=-'{bdo_name}'!{data_col_letter}{dmrrp_rows[0]}"
                )
                logger.info(
                    f"[LTM Interest] Row {dmrrp_ma_row}: "
                    f"=-'{bdo_name}'!{data_col_letter}{dmrrp_rows[0]}"
                )

    def _handle_duplicate_accounts(self, sheet, bdo_sheet, ltm_col: int,
                                   q_col: int, bdo_name: str,
                                   structure: Dict[str, Any]):
        """
        For each duplicate account code, check if all occurrences are referenced.
        Add missing occurrences as EXPLICIT +'BDO'!H<row> terms — NEVER extend
        SUM ranges. This prevents bugs 8-10.
        """
        import re as _re
        bdo_cfg = self._rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)
        data_letter = get_column_letter(data_col)
        q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
        q_letter = get_column_letter(q_cols[-1]) if q_cols else 'G'

        duplicates = structure.get('duplicate_codes', {})
        if not duplicates:
            return

        pl_rows = self._rules.get('layout', {}).get('profit_loss_rows', [23, 68])
        bs_rows = self._rules.get('layout', {}).get('balance_sheet_rows', [3, 19])

        for code, info in duplicates.items():
            all_bdo_rows = info.get('rows', info) if isinstance(info, dict) else info

            for ma_row in range(bs_rows[0], pl_rows[1] + 1):
                for col, ref_letter in [(ltm_col, data_letter), (q_col, q_letter)]:
                    formula = sheet.cell(row=ma_row, column=col).value
                    if not formula or not isinstance(formula, str) or not formula.startswith('='):
                        continue

                    refs_in_formula = set()
                    for bdo_row in all_bdo_rows:
                        if f'{ref_letter}{bdo_row}' in formula:
                            refs_in_formula.add(bdo_row)

                    if not refs_in_formula:
                        continue

                    missing = [r for r in all_bdo_rows if r not in refs_in_formula]
                    if missing:
                        for mr in missing:
                            ref = f"'{bdo_name}'!{ref_letter}{mr}"
                            if ref not in formula:
                                formula += f"+{ref}"
                        sheet.cell(row=ma_row, column=col).value = formula
                        logger.info(
                            f"[Duplicates] Added code {code} rows {missing} "
                            f"to MA row {ma_row} col {get_column_letter(col)}"
                        )

    def _post_build_coverage_audit(self, sheet, bdo_sheet, ltm_col: int,
                                   q_col: int, bdo_name: str,
                                   structure: Dict[str, Any]):
        """
        After all formulas are in place, verify that every non-zero BDO P&L
        row is referenced somewhere. If not, add it to the catch-all row.
        """
        import re as _re
        rules = self._rules
        catch_all = rules.get('layout', {}).get('other_costs_row')
        pl_start = structure.get('pl_start')
        resultaat_row = structure.get('resultaat_row')

        if not catch_all or not pl_start or not resultaat_row:
            return

        data_col = rules.get('bdo', {}).get('data_column', 8)
        data_col_letter = get_column_letter(data_col)
        q_cols = rules.get('bdo', {}).get('quarter_columns', [4, 5, 6, 7])
        q_bdo_col_letter = get_column_letter(q_cols[-1]) if q_cols else 'G'

        referenced = set()
        pl_rows = rules.get('layout', {}).get('profit_loss_rows', [23, 68])
        for ma_row in range(pl_rows[0], pl_rows[1] + 1):
            for col in [ltm_col, q_col]:
                f = sheet.cell(row=ma_row, column=col).value
                if f and isinstance(f, str) and f.startswith('='):
                    for m in _re.findall(r'[HG](\d+)', f):
                        referenced.add(int(m))

        q_cols = rules.get('bdo', {}).get('quarter_columns', [4, 5, 6, 7])
        q_bdo_col = q_cols[-1] if q_cols else 7

        missing_count = 0
        for bdo_row in range(pl_start, resultaat_row):
            h = bdo_sheet.cell(row=bdo_row, column=data_col).value
            g = bdo_sheet.cell(row=bdo_row, column=q_bdo_col).value

            h_has_value = (isinstance(h, (int, float)) and h != 0) or \
                          (isinstance(h, str) and h.startswith('='))
            g_has_value = (isinstance(g, (int, float)) and g != 0) or \
                          (isinstance(g, str) and g.startswith('='))

            if (h_has_value or g_has_value) and bdo_row not in referenced:
                for col, ref_letter in [(ltm_col, data_col_letter),
                                        (q_col, q_bdo_col_letter)]:
                    cell = sheet.cell(row=catch_all, column=col)
                    current = str(cell.value) if cell.value else ''
                    ref = f"'{bdo_name}'!{ref_letter}{bdo_row}"
                    if ref not in current and current.startswith('='):
                        cell.value = current + f"-{ref}"
                missing_count += 1
                logger.warning(
                    f"[Coverage Audit] Added unreferenced BDO row {bdo_row} "
                    f"to catch-all row {catch_all}"
                )

        if missing_count:
            logger.info(
                f"[Coverage Audit] Added {missing_count} missing BDO refs "
                f"to row {catch_all}"
            )
        else:
            logger.info("[Coverage Audit] All BDO P&L rows are referenced")

    def _copy_bdo_data_to_new_sheet(self, bdo_result: BDOParseResult):
        """
        Create new BDO sheet as a verbatim copy of the uploaded Cijfers file.

        Only the sheet title is changed to match the automation naming convention.
        All cells (values + formulas), formatting, merged cells, column widths,
        and row heights are preserved exactly as they appear in the source file.
        """
        new_sheet_name = self.config.bdo_sheet_name

        if new_sheet_name in self.workbook.sheetnames:
            logger.warning(f"Sheet {new_sheet_name} already exists, will remove and recreate")
            del self.workbook[new_sheet_name]

        if not self.bdo_source_path or not self.bdo_source_path.exists():
            logger.warning("BDO source file not provided, will clone previous BDO sheet")
            self._clone_and_update_bdo_sheet(bdo_result)
            return

        bdo_wb = openpyxl.load_workbook(self.bdo_source_path, data_only=False)
        source_sheet = bdo_wb.active
        logger.info(f"Loading BDO source (verbatim copy) from: {self.bdo_source_path}")

        insert_index = len(self.workbook.sheetnames)
        for i, name in enumerate(self.workbook.sheetnames):
            if 'Management Cijfers' in name or 'Cijfers' in name:
                insert_index = i
                break

        new_sheet = self.workbook.create_sheet(title=new_sheet_name, index=insert_index)

        max_row = source_sheet.max_row
        max_col = source_sheet.max_column

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                src_cell = source_sheet.cell(row=row, column=col)
                tgt_cell = new_sheet.cell(row=row, column=col)
                tgt_cell.value = src_cell.value
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.alignment = copy(src_cell.alignment)
                tgt_cell.number_format = src_cell.number_format
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)

        for col_letter, dim in source_sheet.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = dim.width

        for row_idx, dim in source_sheet.row_dimensions.items():
            new_sheet.row_dimensions[row_idx].height = dim.height

        for merged_range in source_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))

        logger.info(f"Created verbatim BDO sheet: {new_sheet_name} ({max_row} rows x {max_col} cols)")

        bdo_wb.close()
    
    def _format_quarter_date(self, quarter: str, offset: int = 0) -> str:
        """Format a quarter string into a date string for column headers."""
        import re
        match = re.match(r'Q(\d)\s*(\d{4})', quarter)
        if not match:
            return quarter
        
        q_num, year = int(match.group(1)), int(match.group(2))
        
        # Apply offset (in quarters)
        total_quarters = (year * 4 + q_num) + offset
        new_year = (total_quarters - 1) // 4
        new_q = ((total_quarters - 1) % 4) + 1
        
        # Quarter end dates
        end_dates = {1: "31-03", 2: "30-06", 3: "30-09", 4: "31-12"}
        return f"{end_dates[new_q]}-{new_year}"
    
    def _set_column_h_formulas(self, sheet, max_row: int):
        """
        Set formulas in column H for the BDO sheet.
        
        Column H contains the LTM (Last Twelve Months) sums.
        
        Structure (based on reference file):
        - Rows 6-124: =SUM(C{row}:G{row}) - Balance sheet items with opening balance
        - Row 125 (Resultaat na belasting): =SUM(C125:G125)
        - Row 126: Empty
        - Row 127 (Verschil balans): =SUM(H5:H123) - Note: Sum of H column, not C-G row
        - Row 128: Empty
        - Row 129 (afschrijving SWAP): No H formula
        - Row 130 (Inkomsten SWAP): No H formula
        - Row 131: Empty
        - Row 132 (Rente): =H117+H114+H116+H115
        - Row 133 (SWAP): =SUM(D133:G133)
        - Row 134 (Afschrijving prepaid): =SUM(D134:G134)
        - Row 135 (IC interest): =SUM(D135:G135)
        - Row 136 (Total): =SUM(D136:G136)
        
        This ensures the formulas remain as formulas, not just values.
        """
        formulas_set = 0
        
        # Find the first special row (where labels like 'afschrijving SWAP' start)
        special_start_row = 129  # Default based on reference
        for row in range(120, min(max_row + 1, 140)):
            label = sheet.cell(row=row, column=1).value
            if label and 'afschrijving swap' in str(label).lower():
                special_start_row = row
                break
        
        bdo_cfg = self._rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)
        data_range = bdo_cfg.get('data_rows', [6, 124])
        col_letter = get_column_letter(data_col)
        q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
        first_q_letter = get_column_letter(q_cols[0] - 1) if q_cols else 'C'
        last_q_letter = get_column_letter(q_cols[-1]) if q_cols else 'G'

        for row in range(data_range[0], data_range[1] + 1):
            cell = sheet.cell(row=row, column=data_col)
            if sheet.cell(row=row, column=1).value is not None:
                cell.value = f"=SUM({first_q_letter}{row}:{last_q_letter}{row})"
                formulas_set += 1
        
        sum_total_row = bdo_cfg.get('sum_total_row', data_range[1] + 1)
        sheet.cell(row=sum_total_row, column=data_col).value = (
            f"=SUM({first_q_letter}{sum_total_row}:{last_q_letter}{sum_total_row})"
        )
        formulas_set += 1
        
        diff_row = sum_total_row + 2
        sheet.cell(row=diff_row, column=data_col).value = (
            f"=SUM({col_letter}{data_range[0] - 1}:{col_letter}{data_range[1] - 1})"
        )
        formulas_set += 1
        
        # Rows 128-136 have their H formulas set in _add_special_rows_with_formulas()
        
        logger.info(f"Set {formulas_set} formulas in column H (special rows handled separately)")
    
    def _extract_swap_income_from_sheet(self, sheet) -> float:
        """
        Extract SWAP income value for the new quarter from the BDO sheet.
        Looks up account 4643000 in column A and reads its column G value.
        """
        for row_idx in range(1, sheet.max_row + 1):
            code = sheet.cell(row=row_idx, column=1).value
            if code and str(code).strip() == "4643000":
                val = sheet.cell(row=row_idx, column=7).value
                if val is not None:
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        pass
                logger.warning("Account 4643000 found but column G has no numeric value")
                return 0.0
        logger.warning("Account 4643000 not found in BDO sheet for SWAP income extraction")
        return 0.0

    def _add_special_rows_with_formulas(self, sheet, start_row: int):
        """
        Add special rows with formulas matching the reference file structure.
        
        Reference structure (from correctFiles.xlsx):
        - Row 126: Empty
        - Row 127: "Verschil balans en winst-en-verlies" with SUM(C5:C123) etc.
        - Row 128: Empty
        - Row 129: "afschrijving SWAP" with =36883.33*3 in D-G
        - Row 130: "Inkomsten SWAP" with shifted values
        - Row 131: Empty
        - Row 132: "Rente" with =D117, =E117, etc.
        - Row 133: "SWAP" with =D130, =E130, etc. and G value
        - Row 134: "Afschrijving prepaid" with =D129, =E129, etc.
        - Row 135: "IC interest" with =D115+D120, etc.
        - Row 136: Total row with SUM formulas
        """
        bdo_cfg = self._rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)
        q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
        h = get_column_letter(data_col)
        q_letters = [get_column_letter(c) for c in q_cols]

        prev_inkomsten_values = self._get_shifted_inkomsten_swap_values()
        
        swap_income = self._extract_swap_income_from_sheet(sheet)
        logger.info(f"Extracted SWAP income for new quarter: {swap_income:,.2f}")
        
        row_129 = start_row + 2
        sheet.cell(row=row_129, column=1).value = "afschrijving SWAP"
        for qc in q_cols:
            sheet.cell(row=row_129, column=qc).value = "=36883.33*3"
        
        row_130 = start_row + 3
        sheet.cell(row=row_130, column=1).value = "Inkomsten SWAP"
        prev_letters = ['E', 'F', 'G']
        for i, qc in enumerate(q_cols[:-1]):
            sheet.cell(row=row_130, column=qc).value = prev_inkomsten_values.get(prev_letters[i] if i < len(prev_letters) else 'G', 0)
        sheet.cell(row=row_130, column=q_cols[-1]).value = swap_income
        
        row_132 = start_row + 5
        sheet.cell(row=row_132, column=1).value = "Rente"
        for qc in q_cols:
            ql = get_column_letter(qc)
            sheet.cell(row=row_132, column=qc).value = f"={ql}117"
        sheet.cell(row=row_132, column=data_col).value = f"={h}117+{h}114+{h}116+{h}115"
        
        row_133 = start_row + 6
        sheet.cell(row=row_133, column=1).value = "SWAP"
        for qc in q_cols:
            ql = get_column_letter(qc)
            sheet.cell(row=row_133, column=qc).value = f"={ql}{row_130}"
        sheet.cell(row=row_133, column=data_col).value = (
            f"=SUM({q_letters[0]}{row_133}:{q_letters[-1]}{row_133})"
        )
        
        row_134 = start_row + 7
        sheet.cell(row=row_134, column=1).value = "Afschrijving prepaid derivatives transaction"
        for qc in q_cols:
            ql = get_column_letter(qc)
            sheet.cell(row=row_134, column=qc).value = f"={ql}{row_129}"
        sheet.cell(row=row_134, column=data_col).value = (
            f"=SUM({q_letters[0]}{row_134}:{q_letters[-1]}{row_134})"
        )
        
        row_135 = start_row + 8
        sheet.cell(row=row_135, column=1).value = "IC interest"
        for qc in q_cols:
            ql = get_column_letter(qc)
            sheet.cell(row=row_135, column=qc).value = f"={ql}115+{ql}120"
        sheet.cell(row=row_135, column=data_col).value = (
            f"=SUM({q_letters[0]}{row_135}:{q_letters[-1]}{row_135})"
        )
        
        row_136 = start_row + 9
        for qc in q_cols:
            ql = get_column_letter(qc)
            sheet.cell(row=row_136, column=qc).value = f"=SUM({ql}{row_132}:{ql}{row_135})"
        sheet.cell(row=row_136, column=data_col).value = (
            f"=SUM({q_letters[0]}{row_136}:{q_letters[-1]}{row_136})"
        )
        
        logger.info(f"Added special rows with formulas (rows {row_129}-{row_136})")
    
    def _get_shifted_inkomsten_swap_values(self) -> dict:
        """Get Inkomsten SWAP values from previous BDO sheet for column shifting."""
        values = {'D': 0, 'E': 0, 'F': 0, 'G': 0}
        
        if not self._prev_bdo_sheet_name:
            return values
        
        prev_wb = None
        try:
            prev_wb = openpyxl.load_workbook(str(self.previous_path), data_only=True)
            if self._prev_bdo_sheet_name in prev_wb.sheetnames:
                prev_sheet = prev_wb[self._prev_bdo_sheet_name]
                for row_idx in range(1, prev_sheet.max_row + 1):
                    label = prev_sheet.cell(row=row_idx, column=1).value
                    if label and 'inkomsten swap' in str(label).lower():
                        values['D'] = prev_sheet.cell(row=row_idx, column=4).value or 0
                        values['E'] = prev_sheet.cell(row=row_idx, column=5).value or 0
                        values['F'] = prev_sheet.cell(row=row_idx, column=6).value or 0
                        values['G'] = prev_sheet.cell(row=row_idx, column=7).value or 0
                        logger.debug(f"Found Inkomsten SWAP at row {row_idx}")
                        break
        except Exception as e:
            logger.warning(f"Error getting Inkomsten SWAP values: {e}")
        finally:
            if prev_wb:
                prev_wb.close()
        
        return values
    
    
    def _copy_bdo_data_direct(self, bdo_result: BDOParseResult):
        """
        Fallback: Copy directly from BDO source file when no previous sheet exists.
        """
        new_sheet_name = self.config.bdo_sheet_name
        
        bdo_wb_values = openpyxl.load_workbook(self.bdo_source_path, data_only=True)
        bdo_wb_formatting = openpyxl.load_workbook(self.bdo_source_path, data_only=False)
        source_sheet_values = bdo_wb_values.active
        source_sheet_formatting = bdo_wb_formatting.active
        
        insert_index = len(self.workbook.sheetnames)
        for i, name in enumerate(self.workbook.sheetnames):
            if 'Management Cijfers' in name or 'Cijfers' in name:
                insert_index = i
                break
        
        new_sheet = self.workbook.create_sheet(title=new_sheet_name, index=insert_index)
        
        max_row = source_sheet_values.max_row
        max_col = source_sheet_values.max_column
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                value_cell = source_sheet_values.cell(row=row, column=col)
                format_cell = source_sheet_formatting.cell(row=row, column=col)
                target_cell = new_sheet.cell(row=row, column=col)
                
                target_cell.value = value_cell.value
                target_cell.font = copy(format_cell.font)
                target_cell.alignment = copy(format_cell.alignment)
                target_cell.number_format = format_cell.number_format
                target_cell.border = copy(format_cell.border)
                target_cell.fill = copy(format_cell.fill)
        
        for col_letter, dim in source_sheet_formatting.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = dim.width
        
        for row_idx, dim in source_sheet_formatting.row_dimensions.items():
            new_sheet.row_dimensions[row_idx].height = dim.height
        
        for merged_range in source_sheet_formatting.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
        
        logger.info(f"Copied BDO data directly to new sheet: {new_sheet_name}")
        
        bdo_wb_values.close()
        bdo_wb_formatting.close()
    
    def _clone_and_update_bdo_sheet(self, bdo_result: BDOParseResult):
        """Fallback: Clone the previous BDO sheet and update with new data."""
        sheet_name = self.config.bdo_sheet_name
        
        bdo_sheets = [name for name in self.workbook.sheetnames if name.startswith('BDO')]
        if not bdo_sheets:
            raise ValueError("No BDO sheet found in previous Management Accounts file")
        
        prev_bdo_name = bdo_sheets[-1]
        prev_bdo = self.workbook[prev_bdo_name]
        logger.info(f"Cloning BDO sheet from: {prev_bdo_name}")
        
        if sheet_name in self.workbook.sheetnames:
            del self.workbook[sheet_name]
        
        # Find insert position (before Management Cijfers)
        insert_index = len(self.workbook.sheetnames)
        for i, name in enumerate(self.workbook.sheetnames):
            if 'Management Cijfers' in name:
                insert_index = i
                break
        
        new_sheet = self.workbook.copy_worksheet(prev_bdo)
        new_sheet.title = sheet_name
        
        # Move to correct position
        self.workbook.move_sheet(new_sheet, offset=insert_index - self.workbook.sheetnames.index(sheet_name))
        
        # Find Saldi column and insert new mutations column
        saldi_col = self._find_saldi_column(new_sheet)
        new_sheet.insert_cols(saldi_col)
        new_mutations_col = saldi_col
        new_saldi_col = saldi_col + 1
        
        # Update headers
        period_str = self.config.period_end.strftime('%d-%m-%Y')
        new_sheet.cell(row=1, column=new_mutations_col).value = f"Mutaties {self.config.quarter}"
        new_sheet.cell(row=1, column=new_mutations_col).font = Font(bold=True)
        new_sheet.cell(row=1, column=new_saldi_col).value = f"Saldi per {period_str}"
        new_sheet.cell(row=1, column=new_saldi_col).font = Font(bold=True)
        
        # Populate values
        for row_idx in range(2, new_sheet.max_row + 1):
            code = str(new_sheet.cell(row=row_idx, column=1).value or '').strip()
            if not code or not code[0].isdigit():
                continue
            
            if code in bdo_result.accounts:
                entry = bdo_result.accounts[code]
                mutations = entry.closing_balance - entry.opening_balance
                new_sheet.cell(row=row_idx, column=new_mutations_col).value = mutations
                new_sheet.cell(row=row_idx, column=new_mutations_col).number_format = '#,##0.00'
                new_sheet.cell(row=row_idx, column=new_saldi_col).value = entry.closing_balance
                new_sheet.cell(row=row_idx, column=new_saldi_col).number_format = '#,##0.00'
            else:
                new_sheet.cell(row=row_idx, column=new_mutations_col).value = 0
    
    def _find_saldi_column(self, sheet) -> int:
        """Find the Saldi (closing balance) column in BDO sheet."""
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col_idx).value
            if cell_value and 'Saldi' in str(cell_value):
                return col_idx
        return sheet.max_column
    
    def _move_bdo_sheet_to_second_last(self):
        """Move the newly created BDO sheet to be the second-to-last sheet."""
        bdo_name = self.config.bdo_sheet_name
        if bdo_name in self.workbook.sheetnames:
            current_index = self.workbook.sheetnames.index(bdo_name)
            # Target is second-to-last (before Management Cijfers which will be moved to last)
            target_index = len(self.workbook.sheetnames) - 1
            if current_index != target_index:
                offset = target_index - current_index
                self.workbook.move_sheet(bdo_name, offset=offset)
                logger.info(f"Moved {bdo_name} to second-to-last (position {target_index + 1})")
    
    def _move_summary_sheet_to_end(self):
        """Move the Management Cijfers sheet to be the last sheet in the workbook."""
        summary_name = self.config.summary_sheet_name
        if summary_name in self.workbook.sheetnames:
            current_index = self.workbook.sheetnames.index(summary_name)
            target_index = len(self.workbook.sheetnames) - 1
            if current_index != target_index:
                self.workbook.move_sheet(summary_name, offset=target_index - current_index)
                logger.info(f"Moved {summary_name} to end (position {target_index + 1})")
    
    def _update_summary_sheet(self):
        """
        Update Management Cijfers summary sheet.
        
        IMPORTANT: Balance Sheet (rows 1-20) and P&L (rows 22+) have different structures:
        - Balance Sheet: Data only in sparse columns (C, E, ..., AA for LTM)
        - P&L: Data in consecutive quarterly columns (T, U, V, W, X, Y, Z for quarters, AA for LTM)
        
        Steps:
        1. Insert new column at LTM position (this shifts LTM to the right)
        2. For P&L section (row 22+): Build formulas from templates using account code lookup
        3. For Balance Sheet section (rows 1-20): Copy from LTM column and update BDO refs
        4. Build LTM formulas (same as quarter but with column H instead of G)
        5. Update column headers
        """
        # Find the summary sheet
        summary_sheets = [name for name in self.workbook.sheetnames 
                         if 'Management Cijfers' in name]
        
        if not summary_sheets:
            summary_sheets = [name for name in self.workbook.sheetnames 
                             if 'Cijfers' in name and 'QSP' in name]
        
        if not summary_sheets:
            raise ValueError("Management Cijfers sheet not found")
        
        old_summary_name = summary_sheets[-1]
        summary_sheet = self.workbook[old_summary_name]
        
        # Rename sheet to new quarter
        new_summary_name = self.config.summary_sheet_name
        if old_summary_name != new_summary_name:
            summary_sheet.title = new_summary_name
        
        logger.info(f"Updating summary sheet: {new_summary_name}")
        
        header_row = self._rules.get('layout', {}).get('header_row', 22)
        layout = scan_summary_column_layout(summary_sheet, header_row)

        # Step 0: Remove FY column(s) when processing a non-Q4 quarter.
        # FY 20XX represents the full prior year and only belongs in the
        # workbook after a Q4 build. For Q1/Q2/Q3 it carries stale data
        # (year-end balances + dates) and shifts the LTM slot to the wrong
        # position. Deleting it before insert keeps the layout clean:
        # `... | Q4 YYYY | LTM Q4 YYYY` becomes `... | Q4 | Qn | LTM Qn`.
        if (
            self.config.quarter_num is not None
            and self.config.quarter_num != 4
            and layout.fy_columns
        ):
            for fy_col in sorted(layout.fy_columns, reverse=True):
                fy_letter = get_column_letter(fy_col)
                summary_sheet.delete_cols(fy_col)
                logger.info(
                    f"Removed FY column at {fy_letter} "
                    f"(quarter_num={self.config.quarter_num}, not Q4)"
                )
            # Re-scan so insert_position, prev_quarter_col, and pre_ltm_col
            # all reflect the post-deletion column indices.
            layout = scan_summary_column_layout(summary_sheet, header_row)

        insert_position = layout.insert_column
        prev_quarter_col = layout.previous_quarter_column
        pre_ltm_col = layout.ltm_column

        if not prev_quarter_col:
            raise ValueError(
                "Could not determine previous quarter column from Management Cijfers headers"
            )

        logger.info(
            f"Summary layout: insert={get_column_letter(insert_position)}, "
            f"prev_quarter={get_column_letter(prev_quarter_col)}, "
            f"ltm={get_column_letter(pre_ltm_col) if pre_ltm_col else 'none'}, "
            f"fy={[get_column_letter(c) for c in layout.fy_columns]}"
        )
        
        # Step 1: Insert new column immediately after the latest Q{n} YYYY header
        summary_sheet.insert_cols(insert_position)

        new_quarter_col = insert_position
        provisional_ltm_col = ltm_column_after_insert(
            pre_ltm_col, insert_position, layout.fy_columns
        )

        # Step 1a: Write row-22 headers BEFORE any formula build so the new
        # LTM column is unambiguously identifiable by header_value. This also
        # overwrites any stale label (e.g. "FY 2025") that slid into the LTM
        # slot during insert_cols().
        summary_sheet.cell(row=22, column=new_quarter_col).value = self.config.quarter
        summary_sheet.cell(row=22, column=provisional_ltm_col).value = (
            f"LTM {self.config.quarter}"
        )

        # Step 1b: Re-resolve the LTM column via the header we just wrote.
        # This guards against future layout changes where the provisional index
        # diverges from the actual LTM column (raises with a layout dump if so).
        new_ltm_col = resolve_built_ltm_column(
            summary_sheet,
            header_row=header_row,
            expected_col=provisional_ltm_col,
            quarter_col=new_quarter_col,
        )
        self._new_quarter_col = new_quarter_col
        self._new_ltm_col = new_ltm_col

        new_quarter_letter = get_column_letter(new_quarter_col)
        new_ltm_letter = get_column_letter(new_ltm_col)

        logger.info(f"After insert: New quarter col={new_quarter_letter}, LTM col={new_ltm_letter}")

        # Step 2: Balance Sheet section (rows 1-20)
        # In the Balance Sheet, only the LTM column has formulas - quarterly columns are empty
        # The new quarter column should remain empty, we only update the LTM column
        # Clear any data that might have been shifted into the new column
        for row_idx in range(1, 21):
            summary_sheet.cell(row=row_idx, column=new_quarter_col).value = None

        # Step 3: Build P&L formulas from templates for the new quarter column
        self._build_pl_formulas_from_templates(summary_sheet, new_quarter_col, bdo_column='G',
                                               prev_col_idx=prev_quarter_col,
                                               quarter_col_idx=new_quarter_col)

        # Step 4: Set manual values (Row 50 - Cash proceeds sale) in quarter column
        if self.config.cash_proceeds_sale:
            cash_cell = summary_sheet.cell(row=50, column=new_quarter_col)
            cash_cell.value = self.config.cash_proceeds_sale
            # Copy number format from previous column
            prev_cash_cell = summary_sheet.cell(row=50, column=prev_quarter_col)
            if prev_cash_cell.number_format:
                cash_cell.number_format = prev_cash_cell.number_format
            logger.info(f"Set Row 50 (Cash proceeds sale) to {self.config.cash_proceeds_sale:,.0f}")

        # Step 5: Finalize LTM column — single deterministic rebuild.
        # All LTM cells (P&L, Balance Sheet, bank, identity, interest) are
        # cleared and rewritten from templates + rules. This makes the LTM
        # column independent of whatever shifted in via insert_cols().
        self._finalize_ltm_column(summary_sheet)

        # Step 6: Update headers and copy styling from previous columns
        # Copy header styling from previous quarter column for new quarter (Row 22)
        prev_header_cell = summary_sheet.cell(row=22, column=prev_quarter_col)
        header_cell = summary_sheet.cell(row=22, column=new_quarter_col)
        # Copy ALL styling including fill (background color) AND the number
        # format. Earlier versions left `number_format` untouched which is
        # why the new quarter header carried "General" instead of the
        # accounting format from the previous header.
        if prev_header_cell.has_style:
            header_cell.font = copy(prev_header_cell.font)
            header_cell.alignment = copy(prev_header_cell.alignment)
            header_cell.border = copy(prev_header_cell.border)
            header_cell.fill = copy(prev_header_cell.fill)  # Copy background color
            if prev_header_cell.number_format and prev_header_cell.number_format != 'General':
                header_cell.number_format = prev_header_cell.number_format

        # Row 2: Balance sheet date header - only in LTM column, quarter column stays empty
        # The date cell in the new quarter column should remain None (cleared earlier)
        # Update the LTM column with the current period end date
        ltm_date_cell = summary_sheet.cell(row=2, column=new_ltm_col)
        ltm_date_cell.value = self.config.period_end
        # Copy styling from previous LTM date cell, force black text
        prev_ltm_date_cell = summary_sheet.cell(row=2, column=new_quarter_col + 1) if new_quarter_col + 1 <= summary_sheet.max_column else None
        if prev_ltm_date_cell and prev_ltm_date_cell.has_style:
            ltm_date_cell.number_format = prev_ltm_date_cell.number_format or 'YYYY-MM-DD'
            base = copy(prev_ltm_date_cell.font)
            hdr_font_color = self._rules.get('styling', {}).get('text_color_header', 'FFFFFFFF')
            ltm_date_cell.font = Font(
                name=base.name, size=base.size, bold=base.bold,
                italic=base.italic, underline=base.underline,
                color=hdr_font_color,
            )
            ltm_date_cell.alignment = copy(prev_ltm_date_cell.alignment)
            ltm_date_cell.border = copy(prev_ltm_date_cell.border)
            ltm_date_cell.fill = copy(prev_ltm_date_cell.fill)
        
        client_number_format = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'

        # Rows 2-21 styling for the new quarter column.
        # The previous fix only handled row 22 (header) and rows 23+ (data),
        # which left rows 1-21 with whatever insert_cols() copied — usually
        # `Calibri 11 / General`. That diverged from the client reference
        # (Avenir / accounting format). Copy from the previous quarter column
        # for the date row + the empty Balance Sheet rows so the column
        # carries consistent styling top to bottom.
        for row_idx in range(2, 22):
            prev_cell = summary_sheet.cell(row=row_idx, column=prev_quarter_col)
            cell = summary_sheet.cell(row=row_idx, column=new_quarter_col)
            if not prev_cell.has_style:
                continue
            cell.font = copy(prev_cell.font)
            cell.alignment = copy(prev_cell.alignment)
            cell.border = copy(prev_cell.border)
            cell.fill = copy(prev_cell.fill)
            if prev_cell.number_format and prev_cell.number_format != 'General':
                cell.number_format = prev_cell.number_format

        # New quarter column: copy ALL styling from the previous quarter column.
        # The previous column already carries the correct Avenir book/10/bold font
        # and per-row colours; we must not replace it with a hardcoded font.
        for row_idx in range(23, summary_sheet.max_row + 1):
            prev_cell = summary_sheet.cell(row=row_idx, column=prev_quarter_col)
            cell = summary_sheet.cell(row=row_idx, column=new_quarter_col)
            if prev_cell.has_style:
                cell.font = copy(prev_cell.font)
                cell.alignment = copy(prev_cell.alignment)
                cell.border = copy(prev_cell.border)
                cell.fill = copy(prev_cell.fill)
                cell.number_format = prev_cell.number_format if prev_cell.number_format != 'General' else client_number_format
            elif cell.number_format == 'General':
                cell.number_format = client_number_format
        
        # Read styling and layout from config (no hard-coded values)
        style_cfg = self._rules.get('styling', {})
        layout_cfg = self._rules.get('layout', {})
        ltm_blue = style_cfg.get('ltm_background_color', 'FFDDEAF6')
        header_blue = style_cfg.get('header_background_color', 'FF5B9BD5')
        text_color = style_cfg.get('text_color_on_blue', 'FF000000')
        header_text_color = style_cfg.get('text_color_header', 'FFFFFFFF')
        blue_fill = PatternFill(
            start_color=ltm_blue, end_color=ltm_blue, fill_type='solid')
        header_fill = PatternFill(
            start_color=header_blue, end_color=header_blue, fill_type='solid')

        bs_range = layout_cfg.get('balance_sheet_rows', [3, 19])
        pl_range = layout_cfg.get('profit_loss_rows', [23, 68])
        bank_range = layout_cfg.get('bank_rows', [107, 114])
        date_row = layout_cfg.get('date_row', 2)
        header_row = layout_cfg.get('header_row', 22)
        border_row = layout_cfg.get('bs_border_row', 18)

        header_rows = {date_row, header_row}
        ltm_data_rows = (
            set(range(date_row, bs_range[1] + 1))
            | {header_row}
            | set(range(pl_range[0], pl_range[1] + 1))
            | set(range(bank_range[0], bank_range[1] + 1))
        )

        # Border row: detect from an existing column so it matches the template
        border_row_border = None
        for probe_col in range(2, prev_quarter_col + 1):
            probe_border = summary_sheet.cell(row=border_row, column=probe_col).border
            if probe_border and probe_border.bottom and probe_border.bottom.style:
                border_row_border = copy(probe_border)
                break
        if not border_row_border:
            thin = Side(style='thin', color='FF000000')
            medium = Side(style='medium', color='FF000000')
            border_row_border = Border(bottom=medium, top=thin)

        def _fill_is_theme_based(fill) -> bool:
            """True if a PatternFill carries a theme-color reference.

            The client reference workbook uses theme:4 (+ tint) for the
            header band and the LTM data band. When `insert_cols()` shifts
            the previous LTM column right, those theme fills are preserved
            on the cell. Overwriting them with the hardcoded RGB
            (FF5B9BD5 / FFDDEAF6) loses the theme link and can render
            slightly off in different Excel themes, which is what produced
            the `theme:4` vs `FF5B9BD5` divergence the client reported.
            """
            if fill is None or fill.fill_type != 'solid':
                return False
            for color_attr in (fill.fgColor, fill.bgColor, fill.start_color, fill.end_color):
                if color_attr is None:
                    continue
                if getattr(color_attr, 'type', None) == 'theme':
                    return True
            return False

        # The LTM column source is the cell at `new_ltm_col` itself —
        # `insert_cols` shifted the prior LTM column right and its cells
        # now sit at this index, carrying the original Avenir/themed
        # styling that the client reference uses for rows 2–20. Copying
        # styling from `prev_quarter_col` (the prior Q4 quarter column,
        # whose BS rows 3–18 are typically empty Calibri placeholders)
        # would *overwrite* that correct styling with the wrong source.
        # We therefore source styling from the cell itself wherever
        # possible and only enforce the explicit color rules below.
        for row_idx in range(date_row, summary_sheet.max_row + 1):
            cell = summary_sheet.cell(row=row_idx, column=new_ltm_col)

            base_font = copy(cell.font) if cell.has_style and cell.font else None

            if cell.number_format == 'General':
                cell.number_format = client_number_format

            # Fill restoration runs unconditionally so the LTM column
            # always carries the right header/data fill even when the
            # post-insert cell did not retain a styled fill of its own.
            existing_fill = cell.fill if cell.has_style else None
            if _fill_is_theme_based(existing_fill):
                pass  # keep theme-based fill as-is
            elif row_idx in header_rows:
                cell.fill = header_fill
            elif row_idx in ltm_data_rows:
                cell.fill = blue_fill

            if row_idx == border_row:
                cell.border = border_row_border

            row_text_color = header_text_color if row_idx in header_rows else text_color
            if base_font is not None:
                cell.font = Font(
                    name=base_font.name,
                    size=base_font.size,
                    bold=base_font.bold,
                    italic=base_font.italic,
                    underline=base_font.underline,
                    color=row_text_color,
                )
            else:
                cell.font = Font(color=row_text_color)

        # Header rows: apply header blue fill across ALL columns (B through LTM)
        for hdr_row in header_rows:
            for col_idx in range(2, new_ltm_col + 1):
                hdr_cell = summary_sheet.cell(row=hdr_row, column=col_idx)
                hdr_cell.fill = header_fill

        # Deterministic floor for the LTM column's Balance Sheet section
        # (rows date_row+1 .. bs_range[1]). The client reference uses
        # Avenir Book 10 with black text; if `insert_cols` left a cell
        # with default Calibri or no font, force the canonical font here
        # so rows 2–20 always render identically to the client template.
        font_cfg = style_cfg.get('font', {}) or {}
        canonical_font_name = font_cfg.get('name', 'Avenir Book')
        canonical_font_size = font_cfg.get('size', 10)
        bs_total_row = bs_range[1]
        bs_data_start = bs_range[0]
        for row_idx in range(bs_data_start, bs_total_row + 1):
            cell = summary_sheet.cell(row=row_idx, column=new_ltm_col)
            base = copy(cell.font) if cell.font else Font()
            current_name = (base.name or '').strip().lower()
            if not current_name or current_name in ('calibri', 'general'):
                cell.font = Font(
                    name=canonical_font_name,
                    size=canonical_font_size,
                    bold=base.bold or (row_idx == bs_total_row),
                    italic=base.italic,
                    underline=base.underline,
                    color=text_color,
                )

        identity_cfg = self._rules.get('identity_alignment', {})
        bold_rows = set(identity_cfg.get(
            'formula_total_rows',
            [identity_cfg.get('authoritative_row', 19),
             identity_cfg.get('dependent_row', 68)]
        ))
        for br in bold_rows:
            for col_idx in range(1, new_ltm_col + 1):
                cell = summary_sheet.cell(row=br, column=col_idx)
                base = copy(cell.font) if cell.font else Font()
                cell.font = Font(
                    name=base.name, size=base.size, bold=True,
                    italic=base.italic, underline=base.underline,
                    color=base.color,
                )

        # Enforce border on row 18 (bs_border_row) across all columns to LTM
        for col_idx in range(2, new_ltm_col + 1):
            cell = summary_sheet.cell(row=border_row, column=col_idx)
            cell.border = border_row_border
        
        # LTM column: blue background rows 106-114, green text rows 107-113
        green_text = style_cfg.get('bank_text_color', 'FF00B050')
        for row_idx in range(bank_range[0] - 1, bank_range[1] + 1):
            ltm_cell = summary_sheet.cell(row=row_idx, column=new_ltm_col)
            ltm_cell.fill = blue_fill
            if bank_range[0] <= row_idx <= bank_range[1] - 1:
                base = copy(ltm_cell.font) if ltm_cell.font else Font()
                ltm_cell.font = Font(
                    name=base.name, size=base.size, bold=base.bold,
                    italic=base.italic, underline=base.underline,
                    color=green_text,
                )

        # Copy column width from previous quarter column
        prev_col_letter = get_column_letter(prev_quarter_col)
        new_q_letter = get_column_letter(new_quarter_col)
        new_ltm_letter_dim = get_column_letter(new_ltm_col)
        if prev_col_letter in summary_sheet.column_dimensions:
            prev_width = summary_sheet.column_dimensions[prev_col_letter].width
            if prev_width:
                summary_sheet.column_dimensions[new_q_letter].width = prev_width
                summary_sheet.column_dimensions[new_ltm_letter_dim].width = prev_width
        
        # Step 7: Add new quarter column to column group (outline)
        new_col_letter = get_column_letter(new_quarter_col)
        summary_sheet.column_dimensions[new_col_letter].outlineLevel = 1
        summary_sheet.column_dimensions[new_col_letter].hidden = False
        logger.info(f"Added column {new_col_letter} to group (outline_level=1, hidden=False)")

        # Step 8: Update title
        summary_sheet.cell(row=1, column=1).value = f"Management Accounts QSP ESS B.V. - {self.config.quarter}"

        logger.info(f"Updated summary sheet columns: {new_quarter_letter} and {new_ltm_letter}")
    
    def _copy_number_formatting(self, sheet, source_col: int, target_col: int, 
                                 start_row: int, end_row: int):
        """Copy number formatting from source column to target column for specified rows."""
        for row_idx in range(start_row, end_row + 1):
            source_cell = sheet.cell(row=row_idx, column=source_col)
            target_cell = sheet.cell(row=row_idx, column=target_col)
            
            if source_cell.number_format and source_cell.number_format != 'General':
                target_cell.number_format = source_cell.number_format
    
    def _finalize_ltm_column(self, sheet) -> None:
        """
        Deterministic LTM column rebuild.

        After ``insert_cols()`` shifts the old LTM column right, the cells in
        ``self._new_ltm_col`` still carry stale formulas that point at the
        previous BDO sheet (e.g. ``'BDO - Q4-25'!H78``) and stale subtotal
        ranges (e.g. ``SUM(AC23:AC24)`` instead of ``SUM(AD23:AD24)``).

        Rather than patch those cells in place, this method **clears** the
        formula cells in the layout ranges and rebuilds the LTM column from
        the same sources of truth used for the quarter column:

        1. Clear stale formula cells in P&L / Balance Sheet / Bank ranges.
        2. Re-apply P&L formulas from ``formula_templates.yaml`` (column H).
        3. Re-apply Balance Sheet formulas from ``balance_sheet`` templates.
        4. Apply every calc/ltm-override row from ``accounting_rules.yaml``
           as a safety net so subtotal rows always reference the LTM column.
        5. Sweep any surviving stale BDO sheet name and any self-reference
           that still points at the new quarter column letter.
        6. Refresh rolling 4-quarter SUM ranges, bank section, identity row,
           and interest overrides.
        """
        ltm_col = self._new_ltm_col
        new_quarter_col = self._new_quarter_col
        if not ltm_col or not new_quarter_col:
            raise ValueError("_finalize_ltm_column called before column indices were set")

        ltm_letter = get_column_letter(ltm_col)
        logger.info(
            f"[LTM rebuild] Finalizing LTM column {ltm_letter} (col {ltm_col})"
        )

        # 1) Clear stale formula cells in the layout ranges so partial template
        #    coverage cannot leave shifted formulas behind.
        self._clear_ltm_formula_cells(sheet, ltm_col)

        # 2) Re-apply P&L formulas from templates (column H = closing balance).
        self._build_pl_formulas_from_templates(
            sheet, ltm_col, bdo_column='H',
            prev_col_idx=new_quarter_col,
            quarter_col_idx=new_quarter_col,
        )

        # 3) Re-apply Balance Sheet formulas from templates.
        self._build_balance_sheet_formulas(sheet, ltm_col, bdo_column='H')

        # 4) Apply calc / ltm_calc_overrides from accounting_rules.yaml as a
        #    safety net (covers rows 25, 30, 44, 45, 48, 51, 55, 57, 66, etc.).
        self._apply_ltm_calc_formulas_from_rules(sheet, ltm_col)

        # 5) Rewrite any residual references — stale BDO sheet name and any
        #    self-reference still pointing at the quarter column letter.
        self._rewrite_ltm_column_references(sheet, ltm_col, new_quarter_col)

        # 6) Refresh rolling sums, bank section, identity formula, and the
        #    interest overrides. These are idempotent.
        self._update_ltm_sum_ranges(sheet, ltm_col, new_quarter_col)
        self._update_bank_account_overview_section(sheet, ltm_col)
        self._enforce_identity_formula(sheet, ltm_col)
        self._apply_interest_section_overrides(sheet, new_quarter_col, ltm_col)

        # 7) Deterministic AC -> AD mirror. After step 2 wrote LTM
        #    formulas from templates, this re-syncs every BS / P&L / bank
        #    upstream row to whatever is in the freshly-built quarter
        #    column, with G -> H swapped. Rows owned by other writers
        #    (totals, identity row, interest section, cash proceeds) are
        #    skipped. This guarantees that any row Claude later patches
        #    in AC will also be reflected in AD even if Claude forgets
        #    to emit the paired patch — see formula_mirror.py for the
        #    shared transformation logic.
        self._mirror_quarter_to_ltm(sheet, new_quarter_col, ltm_col)

    def _mirror_quarter_to_ltm(self, sheet, quarter_col: int, ltm_col: int) -> None:
        """
        Deterministically derive every LTM-column upstream formula from
        the quarter-column formula by swapping BDO column G -> H and
        in-sheet column AC -> AD (or whatever the actual letters are).

        Skipped rows are owned by other writers and must keep their own
        formula:
        - identity totals (rows 19, 68)
        - interest consolidation row + absorbed sub-rows (60, 61, 64)
        - cash proceeds (row 50) — manual_with_ltm has asymmetric LTM
          structure (rolling 4-quarter SUM)
        - bank total row (114) — LTM is a SUM of bank rows, not a BDO ref
        """
        from .formula_mirror import mirror_quarter_column_to_ltm

        layout = self._rules.get('layout', {})
        bs_range = layout.get('balance_sheet_rows', [3, 19])
        pl_range = layout.get('profit_loss_rows', [23, 68])
        bank_range = layout.get('bank_rows', [107, 114])
        identity = self._rules.get('identity_alignment', {})
        auth_row = identity.get('authoritative_row', 19)
        dep_row = identity.get('dependent_row', 68)
        interest = self._rules.get('interest_section', {})
        skip_rows: set = {
            auth_row,
            dep_row,
            layout.get('cash_proceeds_row', 50),
            self._rules.get('bank_total_row', bank_range[1]),
            interest.get('management_row', 60),
        }
        skip_rows.update(interest.get('absorbed_rows', []) or [])

        rows = (
            list(range(bs_range[0], bs_range[1] + 1))
            + list(range(pl_range[0], pl_range[1] + 1))
            + list(range(bank_range[0], bank_range[1] + 1))
        )

        bdo_cfg = self._rules.get('bdo', {})
        q_cols_cfg = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
        last_q_col = q_cols_cfg[-1] if q_cols_cfg else 7
        bdo_qtr_letter = get_column_letter(last_q_col)
        bdo_h_letter = get_column_letter(bdo_cfg.get('data_column', 8))

        rewrites = mirror_quarter_column_to_ltm(
            sheet,
            quarter_col=quarter_col,
            ltm_col=ltm_col,
            bdo_sheet_name=self.config.bdo_sheet_name,
            bdo_quarter_letter=bdo_qtr_letter,
            bdo_data_letter=bdo_h_letter,
            rows=rows,
            skip_rows=skip_rows,
        )
        if rewrites:
            logger.info(
                f"[LTM mirror] Re-mirrored {len(rewrites)} row(s) from "
                f"quarter column to LTM (G->H, AC->AD swap)"
            )

    def _validate_ltm_column_formulas(self, sheet) -> Dict[str, Any]:
        """
        Sanity-check the rebuilt LTM column before save.

        Hard failures (returned in ``errors``):
        - Row-22 header is missing the literal "LTM".
        - Any ``='BDO - …'!…`` reference uses a sheet name other than
          ``self.config.bdo_sheet_name``.
        - Any ``calc`` row from the rules has a formula whose only column
          letter is the new quarter column (the ``SUM(AC…)`` in ``AD`` bug).

        Soft warnings (returned in ``warnings``):
        - ``bdo_ref`` rows that ended up with no formula (account not found).

        Returns ``{"errors": [...], "warnings": [...], "ltm_col": int}``.
        The result is attached to ``self.validation_result['ltm_column']`` so
        callers can see exactly which rows failed which rule.
        """
        ltm_col = self._new_ltm_col
        new_quarter_col = self._new_quarter_col
        result: Dict[str, Any] = {
            'errors': [],
            'warnings': [],
            'ltm_col': ltm_col,
        }
        if not ltm_col or not new_quarter_col:
            result['errors'].append("LTM column index was not set during build")
            return result

        layout_cfg = self._rules.get('layout', {})
        header_row = layout_cfg.get('header_row', 22)
        bs_range = layout_cfg.get('balance_sheet_rows', [3, 19])
        pl_range = layout_cfg.get('profit_loss_rows', [23, 68])
        bank_range = layout_cfg.get('bank_rows', [107, 114])

        ltm_letter = get_column_letter(ltm_col)
        quarter_letter = get_column_letter(new_quarter_col)
        expected_bdo = self.config.bdo_sheet_name

        header_val = sheet.cell(row=header_row, column=ltm_col).value
        if not header_val or 'LTM' not in str(header_val).upper():
            result['errors'].append(
                f"Row-{header_row} header at column {ltm_letter} is "
                f"{header_val!r}, expected to contain 'LTM'"
            )

        bdo_ref_re = re.compile(r"'(BDO[^']*)'")
        check_rows = (
            list(range(bs_range[0], bs_range[1] + 1))
            + list(range(pl_range[0], pl_range[1] + 1))
            + list(range(bank_range[0], bank_range[1] + 1))
        )

        # Rows that are calc subtotals — these must reference the LTM column,
        # not the quarter column, in their in-sheet references.
        calc_rows = set()
        for source in (
            layout_cfg.get('calc_formulas', {}) or {},
            layout_cfg.get('ltm_calc_overrides', {}) or {},
        ):
            for r in source.keys():
                try:
                    calc_rows.add(int(r))
                except (TypeError, ValueError):
                    pass
        for row_idx, tpl in self.formula_templates.items():
            if tpl.get('type') == 'calc':
                calc_rows.add(row_idx)

        # bdo_ref rows from templates — used for the soft warning.
        bdo_ref_rows = {
            row_idx for row_idx, tpl in self.formula_templates.items()
            if tpl.get('type') in ('bdo_ref', 'bdo_ref_label')
        }
        bs_bdo_ref_rows = {
            row_idx for row_idx, tpl in self.balance_sheet_templates.items()
            if tpl.get('type') in ('bdo_ref', 'bdo_sum_range')
        }
        bdo_ref_rows = bdo_ref_rows | bs_bdo_ref_rows

        # Pattern for in-sheet column refs (mirrors _rewrite_ltm_column_references).
        ref_pattern = re.compile(r'(?<![!A-Z])\$?([A-Z]+)\$?\d+')

        for row_idx in check_rows:
            cell = sheet.cell(row=row_idx, column=ltm_col)
            val = cell.value

            if not isinstance(val, str) or not val.startswith('='):
                if row_idx in bdo_ref_rows and val is None:
                    result['warnings'].append(
                        f"Row {row_idx} (bdo_ref) LTM cell is empty — "
                        f"account code likely not found in BDO sheet"
                    )
                continue

            # Hard check: BDO sheet name normalised.
            for match in bdo_ref_re.finditer(val):
                if match.group(1) != expected_bdo:
                    result['errors'].append(
                        f"{ltm_letter}{row_idx} references stale BDO sheet "
                        f"{match.group(1)!r}, expected {expected_bdo!r}: {val}"
                    )
                    break

            # Hard check: calc rows should reference the LTM column letter, not
            # the quarter column letter, in any in-sheet reference.
            if row_idx in calc_rows:
                # Strip sheet-qualified parts so 'BDO - Q1-26'!H80 doesn't trip
                # the in-sheet ref check.
                stripped = re.sub(r"'[^']*'![A-Z]+\$?\d+", '', val)
                in_sheet_cols = {m.group(1) for m in ref_pattern.finditer(stripped)}
                if in_sheet_cols and quarter_letter in in_sheet_cols and ltm_letter not in in_sheet_cols:
                    result['errors'].append(
                        f"{ltm_letter}{row_idx} (calc) still references the "
                        f"quarter column {quarter_letter} instead of {ltm_letter}: {val}"
                    )

        if result['errors']:
            for msg in result['errors']:
                logger.error(f"[LTM validate] {msg}")
        if result['warnings']:
            for msg in result['warnings']:
                logger.warning(f"[LTM validate] {msg}")
        if not result['errors'] and not result['warnings']:
            logger.info(
                f"[LTM validate] Column {ltm_letter} passed all checks"
            )

        return result

    def _clear_ltm_formula_cells(self, sheet, ltm_col: int) -> None:
        """
        Reset every formula cell in the LTM column inside the BS/P&L/Bank
        layout ranges to ``None`` so the rebuild starts from a clean slate.

        Non-formula values (manual entries, dates, header strings) are left
        untouched so styling and the row-22 ``LTM {quarter}`` header survive.
        """
        layout_cfg = self._rules.get('layout', {})
        bs_range = layout_cfg.get('balance_sheet_rows', [3, 19])
        pl_range = layout_cfg.get('profit_loss_rows', [23, 68])
        bank_section = layout_cfg.get('bank_section_rows', [104, 120])

        rows_to_clear = (
            set(range(bs_range[0], bs_range[1] + 1))
            | set(range(pl_range[0], pl_range[1] + 1))
            | set(range(bank_section[0], bank_section[1] + 1))
        )

        cleared = 0
        for row_idx in rows_to_clear:
            cell = sheet.cell(row=row_idx, column=ltm_col)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                cell.value = None
                cleared += 1
        if cleared:
            logger.info(
                f"[LTM rebuild] Cleared {cleared} stale formula cells in "
                f"column {get_column_letter(ltm_col)}"
            )

    def _apply_ltm_calc_formulas_from_rules(self, sheet, ltm_col: int) -> None:
        """
        Apply every entry in ``layout.calc_formulas`` and
        ``layout.ltm_calc_overrides`` as a safety net for subtotal rows.

        Templates in ``formula_templates.yaml`` already cover most ``calc``
        rows, but a missing template entry or a partial build would leave the
        cell empty. The accounting_rules definitions are authoritative for
        subtotal rows (19, 25, 26, 30, 44, 45, 48, 51, 55, 57, 66, 68), so we
        re-apply them here using ``{col}`` → LTM letter.

        ``ltm_calc_overrides`` wins over ``calc_formulas`` when a row appears
        in both (e.g. row 19 uses ``=SUM({col}3:{col}18)`` in LTM but
        ``={col}68`` in the quarter column).
        """
        layout_cfg = self._rules.get('layout', {})
        calc_formulas = layout_cfg.get('calc_formulas', {}) or {}
        ltm_overrides = layout_cfg.get('ltm_calc_overrides', {}) or {}

        # ltm_overrides wins for the LTM column.
        merged = {**calc_formulas, **ltm_overrides}
        if not merged:
            return

        ltm_letter = get_column_letter(ltm_col)
        bdo_name = self.config.bdo_sheet_name

        # Resolve the BDO "Verschil balans en winst-en-verlies" row dynamically.
        # The override formula for row 19 includes a {verschil_row} placeholder
        # because the position of that label varies between BDO files.
        bdo_sheet = (
            self.workbook[bdo_name]
            if bdo_name in self.workbook.sheetnames else None
        )
        verschil_row = self._find_verschil_row(bdo_sheet)

        applied = 0
        for row_key, pattern in merged.items():
            try:
                row_idx = int(row_key)
            except (TypeError, ValueError):
                continue
            if not isinstance(pattern, str) or not pattern.startswith('='):
                continue
            # Substitutions:
            #   {col} / {COL}     → LTM column letter
            #   {bdo}             → current BDO sheet name (no quotes; caller
            #                       includes single quotes in the template).
            #   {verschil_row}    → dynamically-discovered row of the
            #                       "Verschil balans en winst-en-verlies"
            #                       label (never hardcoded).
            formula = (
                pattern.replace('{col}', ltm_letter)
                       .replace('{COL}', ltm_letter)
                       .replace('{bdo}', bdo_name)
                       .replace('{verschil_row}', str(verschil_row))
            )
            sheet.cell(row=row_idx, column=ltm_col).value = formula
            applied += 1
        if applied:
            logger.info(
                f"[LTM rebuild] Applied {applied} calc formulas from "
                f"accounting_rules.yaml in column {ltm_letter} "
                f"(verschil_row={verschil_row})"
            )

    def _rewrite_ltm_column_references(
        self, sheet, ltm_col: int, new_quarter_col: int
    ) -> None:
        """
        Final sweep over LTM column formulas:

        - Replace any ``'BDO - …'`` sheet name that is not the current
          ``config.bdo_sheet_name`` (catches references that survived the
          rebuild for rows not covered by templates).
        - Rewrite any in-sheet column reference whose letter still matches the
          new quarter column letter to the LTM column letter. This catches the
          ``SUM(AC23:AC24)`` family of bugs without touching cross-column
          arithmetic (e.g. ``AB66+AB67`` from older columns is left alone).
        """
        ltm_letter = get_column_letter(ltm_col)
        quarter_letter = get_column_letter(new_quarter_col)
        new_bdo_name = self.config.bdo_sheet_name
        bdo_pattern = re.compile(r"'(BDO[^']*)'")

        # Match in-sheet refs: letter(s) + optional $ + row number, NOT
        # preceded by '!' (which would make it a sheet-qualified reference).
        # We rewrite only when the column letter exactly equals the new
        # quarter column letter and the cell sits in the LTM column.
        ref_pattern = re.compile(
            r'(?<![!A-Z])(\$?)([A-Z]+)(\$?\d+)'
        )

        rewritten = 0
        renamed = 0
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=ltm_col)
            val = cell.value
            if not isinstance(val, str) or not val.startswith('='):
                continue

            new_val = val

            # Step 1: Normalise BDO sheet name to the current quarter's sheet.
            if 'BDO' in new_val:
                before = new_val
                new_val = bdo_pattern.sub(
                    lambda m: f"'{new_bdo_name}'" if m.group(1) != new_bdo_name else m.group(0),
                    new_val,
                )
                if new_val != before:
                    renamed += 1

            # Step 2: Rewrite in-sheet references that still point at the new
            # quarter column letter (the classic SUM(AC...) in AD bug).
            def _swap_col(match):
                dollar1, col_letters, row_part = match.groups()
                if col_letters == quarter_letter:
                    return f"{dollar1}{ltm_letter}{row_part}"
                return match.group(0)
            swapped = ref_pattern.sub(_swap_col, new_val)
            if swapped != new_val:
                new_val = swapped

            if new_val != val:
                cell.value = new_val
                rewritten += 1

        if renamed or rewritten:
            logger.info(
                f"[LTM rebuild] Reference sweep in column {ltm_letter}: "
                f"{renamed} BDO-name fixes, {rewritten} cells rewritten"
            )

    def _update_ltm_sum_ranges(self, sheet, ltm_col: int, new_quarter_col: int):
        """
        Update SUM formulas in LTM column to include the new quarter column.
        
        After inserting a column, SUM ranges shift but don't extend to include
        the new column. This method fixes that by updating ranges like:
        =SUM(W50:Z50) -> =SUM(X50:AA50)
        """
        ltm_letter = get_column_letter(ltm_col)
        new_q_letter = get_column_letter(new_quarter_col)
        
        updated_count = 0
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=ltm_col)
            
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                
                # Pattern to match SUM formulas like =SUM(W50:Z50)
                # After insert, old range became shifted but excluded new column
                # Need to update end column to include new quarter
                
                # Match SUM formulas that reference same row
                pattern = rf'=SUM\(([A-Z]+){row_idx}:([A-Z]+){row_idx}\)'
                match = re.match(pattern, formula)
                
                if match:
                    start_col_letter = match.group(1)
                    end_col_letter = match.group(2)
                    
                    start_col_idx = column_index_from_string(start_col_letter)
                    end_col_idx = column_index_from_string(end_col_letter)
                    
                    # LTM should always be sum of 4 quarters
                    # After insert: start should shift +1, end should be new quarter column
                    # Example: =SUM(W50:Z50) should become =SUM(X50:AA50)
                    
                    # Only update if the range doesn't already include new quarter
                    if end_col_idx < new_quarter_col:
                        # Shift start by 1 to drop oldest quarter
                        new_start = get_column_letter(start_col_idx + 1)
                        new_formula = f"=SUM({new_start}{row_idx}:{new_q_letter}{row_idx})"
                        cell.value = new_formula
                        updated_count += 1
        
        if updated_count > 0:
            logger.info(f"Updated {updated_count} LTM SUM formulas to include column {new_q_letter}")
    
    def _enforce_identity_formula(self, sheet, ltm_col: int):
        """
        If accounting_rules.yaml has enforce_in_workbook: true, replace the
        dependent row's LTM formula with a direct reference to the
        authoritative row (e.g. row 68 LTM = ={COL}19).
        """
        align = self._rules.get('identity_alignment', {})
        if not align.get('enforce_in_workbook', False):
            return

        auth_row = align.get('authoritative_row', 19)
        dep_row = align.get('dependent_row', 68)
        col_letter = get_column_letter(ltm_col)

        formula = f"={col_letter}{auth_row}"
        sheet.cell(row=dep_row, column=ltm_col).value = formula
        logger.info(
            f"[Identity] Set row {dep_row} LTM (col {col_letter}) "
            f"to '{formula}' (referencing row {auth_row})")

    def _update_bank_account_overview_section(self, sheet, ltm_col: int):
        """
        Update the Bank Account Overview section (rows 104-120) in the LTM column.
        
        This section contains:
        - Row 106: Date header ("Per DD-MM-YYYY") - needs date update
        - Rows 107-113: BDO sheet references for each bank account - set correct formulas
        - Row 114: SUM formula - needs column reference update
        
        IMPORTANT: We set the CORRECT formulas directly rather than just updating 
        the BDO sheet name, because the input file may have incorrect formulas.
        
        The correct bank account formulas reference column H (closing balance) in BDO:
        - Row 107 (RENT): H35
        - Row 108 (MAINT): H34
        - Row 109 (EXP): H36
        - Row 110 (DEP): H33
        - Row 111 (GEN): H32
        - Row 112 (CAPEX): H37
        - Row 113 (DISPOSAL): H31
        """
        ltm_letter = get_column_letter(ltm_col)
        new_bdo_name = self.config.bdo_sheet_name
        period_end = self.config.period_end
        updated_count = 0

        # Define the correct BDO row references for each bank account
        # These are the closing balance (column H) rows in the BDO sheet.
        # Rows 110 (DEP) and 111 (GEN) ordering matches the client-reference
        # workbook; the previous mapping had these two swapped which produced
        # off-by-one totals in the bank section.
        bank_account_bdo_rows = {
            107: 'H35',  # ABN AMRO RENT account
            108: 'H34',  # ABN AMRO MAINT account
            109: 'H36',  # ABN AMRO EXP account
            110: 'H33',  # ABN AMRO DEP account
            111: 'H32',  # ABN AMRO GEN account
            112: 'H37',  # ABN AMRO CAPEX account
            113: 'H31',  # ABN AMRO DISPOSAL account
        }

        # Row 106: "Per DD-MM-YYYY" label is mandatory in the LTM column.
        # The earlier `_clear_ltm_formula_cells` step can leave this cell as
        # None when the previous workbook contained a date or formula here,
        # so we write the label unconditionally before iterating the bank
        # rows. Without this, `correct.xlsx` shows "Per 31-3-2026" while the
        # automation leaves the cell blank or with a stale FY date.
        label_row = 106
        date_label = f"Per {period_end.day}-{period_end.month}-{period_end.year}"
        sheet.cell(row=label_row, column=ltm_col).value = date_label
        updated_count += 1
        logger.debug(f"Row {label_row}: Set date label to '{date_label}'")

        # Define the Bank Account Overview section range
        start_row = 104
        end_row = min(sheet.max_row, 125)  # Safety limit

        for row_idx in range(start_row, end_row + 1):
            # Row 106 already handled unconditionally above; skip here to
            # avoid double-counting and to prevent stale "Per …" detection
            # from interfering when the cell was cleared.
            if row_idx == label_row:
                continue

            cell = sheet.cell(row=row_idx, column=ltm_col)
            cell_value = cell.value

            # Defensive: legacy templates can still carry "Per …" strings
            # in unexpected rows; re-stamp them with the current period.
            if isinstance(cell_value, str) and cell_value.startswith('Per '):
                cell.value = date_label
                updated_count += 1
                logger.debug(f"Row {row_idx}: Updated date header to '{date_label}'")
                continue

            # Set correct bank account formulas (rows 107-113)
            if row_idx in bank_account_bdo_rows:
                bdo_cell_ref = bank_account_bdo_rows[row_idx]
                correct_formula = f"='{new_bdo_name}'!{bdo_cell_ref}"
                old_value = cell.value
                cell.value = correct_formula
                updated_count += 1
                logger.debug(f"Row {row_idx}: Set formula to '{correct_formula}' (was: {old_value})")
                continue
            
            # Handle SUM formula for Total row (row 114). Unconditionally
            # write the canonical formula so the LTM rebuild produces a
            # correct total even when the cleared cell was previously None.
            bank_total_row = self._rules.get('bank_total_row', 114)
            if row_idx == bank_total_row:
                first_bank = min(bank_account_bdo_rows.keys())
                last_bank = max(bank_account_bdo_rows.keys())
                new_formula = f"=SUM({ltm_letter}{first_bank}:{ltm_letter}{last_bank})"
                cell.value = new_formula
                updated_count += 1
                logger.debug(f"Row {row_idx}: Set SUM total to '{new_formula}'")
        
        if updated_count > 0:
            logger.info(f"Updated {updated_count} cells in Bank Account Overview section (column {ltm_letter})")
    
    def _detect_first_quarter_column_from_header_row(
        self, sheet, header_row: int, up_to_col: int
    ) -> Optional[int]:
        """
        Leftmost column in [2, up_to_col] on header_row whose value looks like a
        quarter label (Q1–Q4). Used so horizontal SUMs include all quarter columns
        when the template starts before column Y.
        """
        if up_to_col < 2:
            return None
        qpat = re.compile(r'\bQ\s*[1-4]\b', re.IGNORECASE)
        for c in range(2, up_to_col + 1):
            v = sheet.cell(row=header_row, column=c).value
            if v is None:
                continue
            if qpat.search(str(v).strip()):
                return c
        return None

    def _apply_interest_section_overrides(
        self, sheet, new_quarter_col: int, ltm_col: int
    ) -> None:
        """
        Deterministic interest-row overrides:
        - Row 60 quarter col: extend previous quarter's formula (shift MA column refs).
        - Row 60 LTM col: copy quarter row 60 formula, swap BDO !G → !H.
        - Rows 61 & 64 LTM col: =SUM(Y{row}:{quarter_letter}{row}).
        """
        bdo_cfg = self._rules.get('bdo', {})
        q_bdo_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
        q_bdo_letter = get_column_letter(q_bdo_cols[-1]) if q_bdo_cols else 'G'
        ltm_bdo_letter = get_column_letter(bdo_cfg.get('data_column', 8))

        layout = self._rules.get('layout', {})
        sum_start_col = layout.get('interest_quarter_horizontal_sum_start_column', 25)

        prev_quarter_col = new_quarter_col - 1
        prev_letter = get_column_letter(prev_quarter_col)
        new_letter = get_column_letter(new_quarter_col)

        # Row 60 quarter: extend from previous quarter column
        prev_60 = sheet.cell(row=60, column=prev_quarter_col).value
        if isinstance(prev_60, str) and prev_60.startswith('='):
            shifted = _shift_formula_column_letter(prev_60, prev_letter, new_letter)
            if shifted != prev_60:
                sheet.cell(row=60, column=new_quarter_col).value = shifted
                logger.info(
                    f"Interest row 60: quarter formula derived from col {prev_letter} → {new_letter}"
                )

        # Row 60 LTM: take quarter row 60 and swap BDO column G→H
        quarter_60 = sheet.cell(row=60, column=new_quarter_col).value
        if isinstance(quarter_60, str) and quarter_60.startswith('='):
            ltm_60 = quarter_60.replace(f'!{q_bdo_letter}', f'!{ltm_bdo_letter}')
            sheet.cell(row=60, column=ltm_col).value = ltm_60
            logger.info(
                f"Interest row 60: LTM formula derived from quarter col "
                f"{new_letter} (BDO !{q_bdo_letter}→!{ltm_bdo_letter})"
            )

        # Rows 61 & 64 LTM: SUM of last 4 quarter columns (skip FY / LTM columns)
        header_row = self._rules.get('layout', {}).get('header_row', 22)
        layout = scan_summary_column_layout(sheet, header_row)
        prior_quarters = [c for c in layout.quarter_columns if c < new_quarter_col]
        if len(prior_quarters) >= 4:
            sum_start = prior_quarters[-4]
        elif prior_quarters:
            sum_start = prior_quarters[0]
        else:
            sum_start = new_quarter_col - 3
        configured_start = self._rules.get('layout', {}).get(
            'interest_quarter_horizontal_sum_start_column', 25
        )
        first_q = self._detect_first_quarter_column_from_header_row(
            sheet, header_row, new_quarter_col
        )
        if first_q is not None:
            sum_start = max(sum_start, first_q)
        sum_start = max(sum_start, configured_start, 2)
        start_letter = get_column_letter(sum_start)
        quarter_letter = get_column_letter(new_quarter_col)
        for row_num in (61, 64):
            formula = f'=SUM({start_letter}{row_num}:{quarter_letter}{row_num})'
            sheet.cell(row=row_num, column=ltm_col).value = formula
        logger.info(
            f"Interest rows 61 & 64: LTM col {get_column_letter(ltm_col)} "
            f"=SUM({start_letter}*:{quarter_letter}*)"
        )

    def _build_pl_formulas_from_templates(self, sheet, col_idx: int, bdo_column: str = 'G',
                                          prev_col_idx: int = None,
                                          quarter_col_idx: int = None):
        """
        Build P&L formulas for a column using formula templates.
        
        For each row defined in the templates:
        - bdo_ref type: Look up account codes in BDO sheet, build formula with correct row numbers
        - calc type: Replace {COL} placeholder with actual column letter
        - manual type: Copy value from previous column (for Cash proceeds sale, etc.)
        - ltm_sum_quarters: SUM of the last N quarterly columns (for LTM)
        - manual_with_ltm: Manual for quarter columns, ltm_sum_quarters for LTM
        
        Args:
            sheet: The summary sheet
            col_idx: Column index to write formulas to
            bdo_column: BDO column to reference (G for quarter, H for LTM)
            prev_col_idx: Previous column index (for copying manual values)
            quarter_col_idx: New quarter column index (needed for ltm_sum_quarters)
        """
        col_letter = get_column_letter(col_idx)
        bdo_sheet_name = self.config.bdo_sheet_name
        formulas_built = 0
        
        for row_idx, template in self.formula_templates.items():
            formula_type = template.get('type', 'calc')
            target_cell = sheet.cell(row=row_idx, column=col_idx)
            
            if formula_type == 'bdo_ref':
                formula = self._build_bdo_ref_formula(template, bdo_sheet_name, bdo_column)
                if formula:
                    target_cell.value = formula
                    formulas_built += 1
                    
            elif formula_type == 'bdo_ref_label':
                formula = self._build_bdo_ref_label_formula(template, bdo_sheet_name, bdo_column)
                if formula:
                    target_cell.value = formula
                    formulas_built += 1
            
            elif formula_type == 'bdo_ref_conditional':
                if bdo_column == 'G':
                    sub_template = template.get('q_config', {})
                else:
                    sub_template = template.get('ltm_config', {})
                
                sub_type = sub_template.get('type', 'bdo_ref')
                if sub_type == 'bdo_ref':
                    formula = self._build_bdo_ref_formula(sub_template, bdo_sheet_name, bdo_column)
                elif sub_type == 'bdo_ref_label':
                    formula = self._build_bdo_ref_label_formula(sub_template, bdo_sheet_name, bdo_column)
                elif sub_type == 'ltm_sum_quarters':
                    formula = self._build_ltm_sum_quarters_formula(
                        row_idx, sub_template.get('num_quarters', 4), quarter_col_idx
                    )
                elif sub_type == 'horizontal_sum_quarters':
                    formula = self._build_horizontal_sum_quarters_formula(
                        sheet, row_idx, col_idx, prev_col_idx
                    )
                elif sub_type == 'static':
                    target_cell.value = sub_template.get('value', 0)
                    formulas_built += 1
                    formula = None
                else:
                    formula = None
                    
                if formula:
                    target_cell.value = formula
                    formulas_built += 1
                    
            elif formula_type == 'calc':
                pattern = template.get('pattern', '')
                formula = pattern.replace('{COL}', col_letter)
                target_cell.value = formula
                formulas_built += 1
                
            elif formula_type == 'manual':
                pass
            
            elif formula_type == 'manual_with_ltm':
                if bdo_column == 'H' and quarter_col_idx:
                    ltm_type = template.get('ltm_type', '')
                    if ltm_type == 'ltm_sum_quarters':
                        formula = self._build_ltm_sum_quarters_formula(
                            row_idx, template.get('num_quarters', 4), quarter_col_idx
                        )
                        if formula:
                            target_cell.value = formula
                            formulas_built += 1
            
            elif formula_type == 'constant':
                target_cell.value = template.get('value', 0)
                formulas_built += 1
            
            if prev_col_idx:
                prev_cell = sheet.cell(row=row_idx, column=prev_col_idx)
                if prev_cell.number_format:
                    target_cell.number_format = prev_cell.number_format
        
        logger.info(f"Built {formulas_built} P&L formulas for column {col_letter} (BDO col {bdo_column})")
    
    def _build_ltm_sum_quarters_formula(self, row_idx: int, num_quarters: int,
                                         quarter_col_idx: int) -> Optional[str]:
        """
        Build =SUM(start:end) formula covering the last N quarterly columns.
        
        The quarter_col_idx is the newest quarter column. The formula sums from
        (quarter_col_idx - num_quarters + 1) to quarter_col_idx.
        """
        if not quarter_col_idx:
            logger.warning(f"Row {row_idx}: Cannot build ltm_sum_quarters without quarter_col_idx")
            return None
        start_col = quarter_col_idx - num_quarters + 1
        if start_col < 1:
            start_col = 1
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(quarter_col_idx)
        formula = f"=SUM({start_letter}{row_idx}:{end_letter}{row_idx})"
        logger.debug(f"Row {row_idx}: LTM SUM formula: {formula}")
        return formula
    
    def _build_horizontal_sum_quarters_formula(
        self, sheet, row_idx: int, new_quarter_col: int,
        prev_col_idx: Optional[int]
    ) -> Optional[str]:
        """
        Build =SUM of last 4 quarter columns for the LTM column.

        Used for rows 61/64 in the LTM column.
        prev_col_idx is the newest quarter column (the column just before LTM).
        """
        prev_quarter_col = prev_col_idx if prev_col_idx else (new_quarter_col - 1)
        sum_start = prev_quarter_col - 3
        if sum_start < 2:
            sum_start = 2

        start_letter = get_column_letter(sum_start)
        prev_letter = get_column_letter(prev_quarter_col)
        formula = f"=SUM({start_letter}{row_idx}:{prev_letter}{row_idx})"
        logger.info(f"Row {row_idx}: horizontal SUM formula: {formula}")
        return formula

    def _build_bdo_ref_formula(self, template: Dict[str, Any], bdo_sheet_name: str, 
                                bdo_column: str) -> Optional[str]:
        """
        Build a formula that references BDO sheet cells by account codes.
        
        Example template:
            account_codes: ["4100410", "4101010", "4101020"]
            sign: "-"
            operator: "-"
        
        Result: =-'BDO - Q3-25'!G91-'BDO - Q3-25'!G92-'BDO - Q3-25'!G93
        """
        account_codes = template.get('account_codes', [])
        sign = template.get('sign', '-')
        operator = template.get('operator', '-')
        
        if not account_codes:
            return None
        
        refs = []
        for code in account_codes:
            row_num = self._new_bdo_row_map.get(code)
            if row_num:
                refs.append(f"'{bdo_sheet_name}'!{bdo_column}{row_num}")
            else:
                # Try to find by partial match or similar code
                row_num = self._find_account_row(code)
                if row_num:
                    refs.append(f"'{bdo_sheet_name}'!{bdo_column}{row_num}")
                else:
                    logger.warning(f"Account code {code} not found in BDO sheet")
        
        if not refs:
            return None
        
        # Build formula with sign and operator
        if len(refs) == 1:
            return f"={sign}{refs[0]}"
        else:
            return f"={sign}{operator.join(refs)}"
    
    def _build_bdo_ref_label_formula(self, template: Dict[str, Any], bdo_sheet_name: str,
                                      bdo_column: str) -> Optional[str]:
        """
        Build a formula that references a BDO sheet cell by label search.
        
        Example template:
            label_search: "SWAP"
            sign: "-"
        
        Result: =-'BDO - Q3-25'!G133
        
        Searches in column A of the BDO sheet for exact or partial matches.
        """
        label_search = template.get('label_search', '')
        sign = template.get('sign', '-')
        
        if not label_search:
            return None
        
        label_search_lower = label_search.lower().strip()
        
        # First try exact match in label map
        row_num = self._new_bdo_label_map.get(label_search_lower)
        
        if not row_num:
            # Try exact match with the full text (case-insensitive)
            for label, row in self._new_bdo_label_map.items():
                if label == label_search_lower:
                    row_num = row
                    break
        
        if not row_num:
            # Try partial match only if exact match not found
            # Be careful with partial matches - "SWAP" shouldn't match "Swap fee"
            # Only match if the label is at the start or is the complete text
            for label, row in self._new_bdo_label_map.items():
                # Match if label is exactly the search term (already checked above)
                # or if search term is at the beginning followed by space or end
                if label.startswith(label_search_lower + ' ') or label.startswith(label_search_lower + '\t'):
                    row_num = row
                    break
        
        if row_num:
            return f"={sign}'{bdo_sheet_name}'!{bdo_column}{row_num}"
        else:
            logger.warning(f"Label '{label_search}' not found in BDO sheet")
            return None
    
    def _find_account_row(self, account_code: str) -> Optional[int]:
        """Find a row for an account code, trying partial matches if exact not found."""
        # First try exact match
        if account_code in self._new_bdo_row_map:
            return self._new_bdo_row_map[account_code]
        
        # Try prefix match (e.g., "4100400" might be stored as "4100400.0")
        for code, row in self._new_bdo_row_map.items():
            if code.startswith(account_code) or account_code.startswith(code.split('.')[0]):
                return row
        
        return None
    
    def _build_balance_sheet_formulas(self, sheet, col_idx: int, bdo_column: str = 'H'):
        """
        Build Balance Sheet formulas for the LTM column using templates.
        
        Balance Sheet section (rows 3-19) only has formulas in the LTM column.
        Uses balance_sheet_templates to build formulas with correct BDO row references.
        """
        col_letter = get_column_letter(col_idx)
        bdo_sheet_name = self.config.bdo_sheet_name
        formulas_built = 0
        
        for row_idx, template in self.balance_sheet_templates.items():
            formula_type = template.get('type', 'bdo_ref')
            target_cell = sheet.cell(row=row_idx, column=col_idx)

            if formula_type == 'bdo_ref':
                formula = self._build_balance_sheet_ref_formula(template, bdo_sheet_name, bdo_column)
                if formula:
                    target_cell.value = formula
                    formulas_built += 1
                    
            elif formula_type == 'bdo_sum_range':
                formula = self._build_balance_sheet_sum_formula(template, bdo_sheet_name, bdo_column)
                if formula:
                    target_cell.value = formula
                    formulas_built += 1
                    
            elif formula_type == 'calc':
                pattern = template.get('pattern', '')
                formula = pattern.replace('{COL}', col_letter)
                target_cell.value = formula
                formulas_built += 1
        
        logger.info(f"Built {formulas_built} Balance Sheet formulas in column {col_letter}")
    
    def _build_balance_sheet_ref_formula(self, template: Dict[str, Any], bdo_sheet_name: str,
                                          bdo_column: str) -> Optional[str]:
        """Build a Balance Sheet formula that references BDO cells by account codes."""
        account_codes = template.get('account_codes', [])
        sign = template.get('sign', '')
        operator = template.get('operator', '+')
        
        if not account_codes:
            return None
        
        refs = []
        for code in account_codes:
            row_num = self._find_account_row(code)
            if row_num:
                refs.append(f"'{bdo_sheet_name}'!{bdo_column}{row_num}")
            else:
                logger.debug(f"Balance Sheet: Account code {code} not found in BDO sheet")
        
        if not refs:
            return None
        
        if len(refs) == 1:
            return f"={sign}{refs[0]}"
        else:
            return f"={sign}{operator.join(refs)}"
    
    def _get_all_rows_for_account(self, code: str) -> List[int]:
        """
        Get ALL BDO rows for an account code, including duplicate occurrences.
        Returns list of row numbers sorted ascending.
        """
        rows = []
        primary = self._find_account_row(code)
        if primary:
            rows.append(primary)

        bdo_sheet_name = self.config.bdo_sheet_name
        if bdo_sheet_name in self.workbook.sheetnames:
            bdo_sheet = self.workbook[bdo_sheet_name]
            for row_idx in range(1, bdo_sheet.max_row + 1):
                cell_code = bdo_sheet.cell(row=row_idx, column=1).value
                if cell_code and str(cell_code).strip() == code:
                    if row_idx not in rows:
                        rows.append(row_idx)

        rows.sort()
        return rows

    def _build_balance_sheet_sum_formula(self, template: Dict[str, Any], bdo_sheet_name: str,
                                          bdo_column: str) -> Optional[str]:
        """Build a Balance Sheet SUM formula for a range of accounts.
        Uses _find_sum_range_end() for same-section prefix extension only.
        Skips additional refs already inside the SUM range.
        Handles duplicate account codes by adding all occurrences."""
        start_account = template.get('start_account')
        end_account = template.get('end_account')
        count = template.get('count')
        prefix = template.get('prefix')
        additional = template.get('additional', [])

        start_row = self._find_account_row(start_account) if start_account else None

        if count and start_row:
            bdo_sheet = None
            if self.config.bdo_sheet_name in self.workbook.sheetnames:
                bdo_sheet = self.workbook[self.config.bdo_sheet_name]
            if prefix is not None:
                prefix = str(prefix)
            if bdo_sheet and prefix:
                end_row = self._find_sum_range_end(
                    bdo_sheet, start_row, count, prefix)
            else:
                end_row = start_row + count - 1
                if bdo_sheet and not prefix and start_account:
                    logger.warning(
                        f"[BS SUM] No prefix for start_account {start_account} "
                        f"(count={count}). SUM range will NOT auto-extend. "
                        f"Add prefix to formula_templates.yaml if extension is needed."
                    )
        elif end_account:
            end_row = self._find_account_row(end_account)
        else:
            end_row = None

        parts = []
        sum_range = set()

        if start_row and end_row and end_row > start_row:
            parts.append(
                f"SUM('{bdo_sheet_name}'!{bdo_column}{start_row}"
                f":{bdo_column}{end_row})")
            sum_range = set(range(start_row, end_row + 1))
        elif start_row and count == 1:
            parts.append(f"SUM('{bdo_sheet_name}'!{bdo_column}{start_row})")
            sum_range = {start_row}
        elif start_row:
            parts.append(f"'{bdo_sheet_name}'!{bdo_column}{start_row}")
            sum_range = {start_row}

        for code in additional:
            all_rows = self._get_all_rows_for_account(code)
            for row_num in all_rows:
                if row_num not in sum_range:
                    parts.append(f"'{bdo_sheet_name}'!{bdo_column}{row_num}")

        if not parts:
            return None

        return "=" + "+".join(parts)
    
    def _update_balance_sheet_bdo_refs(self, sheet, col_idx: int, bdo_column: str = 'G'):
        """
        Update Balance Sheet section (rows 1-20) BDO references to point to new BDO sheet.
        This preserves the existing formula structure but updates the sheet name.
        """
        bdo_sheet_name = self.config.bdo_sheet_name
        updated_count = 0
        
        for row_idx in range(1, 21):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                old_formula = cell.value
                if 'BDO' in old_formula:
                    # Update BDO sheet name and column
                    new_formula = re.sub(
                        r"'BDO[^']+'\!([GH])(\d+)",
                        lambda m: f"'{bdo_sheet_name}'!{bdo_column}{m.group(2)}",
                        old_formula
                    )
                    if new_formula != old_formula:
                        cell.value = new_formula
                        updated_count += 1
        
        if updated_count > 0:
            logger.info(f"Updated {updated_count} Balance Sheet BDO refs in column {get_column_letter(col_idx)}")
    
    def _copy_section_formulas(self, sheet, source_col: int, target_col: int, 
                                start_row: int, end_row: int, section_name: str):
        """
        Copy formulas and formatting from source column to target column for a specific row range.
        """
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)
        copied_count = 0
        
        for row_idx in range(start_row, end_row + 1):
            source_cell = sheet.cell(row=row_idx, column=source_col)
            target_cell = sheet.cell(row=row_idx, column=target_col)
            
            if source_cell.value is not None:
                value = source_cell.value
                
                # If it's a formula, update internal column references
                if isinstance(value, str) and value.startswith('='):
                    value = self._update_internal_column_refs(value, source_letter, target_letter)
                    copied_count += 1
                
                target_cell.value = value
            
            # Copy formatting
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
        
        logger.info(f"Copied {section_name} section (rows {start_row}-{end_row}): {copied_count} formulas from {source_letter} to {target_letter}")
    
    def _find_ltm_column(self, sheet) -> Optional[int]:
        """Find the LTM column from row-22 headers."""
        header_row = self._rules.get('layout', {}).get('header_row', 22)
        return scan_summary_column_layout(sheet, header_row).ltm_column
    
    def _find_last_data_column(self, sheet) -> int:
        """Find the last date column in row 2."""
        last_date_col = 1
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=2, column=col_idx)
            if isinstance(cell.value, datetime):
                last_date_col = col_idx
        return last_date_col
    
    def _copy_column_with_formulas(self, sheet, source_col: int, target_col: int):
        """
        Copy column including formulas and all formatting.
        Also updates internal column references from source to target column.
        """
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)
        
        if source_letter in sheet.column_dimensions:
            sheet.column_dimensions[target_letter].width = sheet.column_dimensions[source_letter].width
        
        for row_idx in range(1, sheet.max_row + 1):
            source_cell = sheet.cell(row=row_idx, column=source_col)
            target_cell = sheet.cell(row=row_idx, column=target_col)
            
            if source_cell.value is not None:
                value = source_cell.value
                
                # If it's a formula, update internal column references
                if isinstance(value, str) and value.startswith('='):
                    value = self._update_internal_column_refs(value, source_letter, target_letter)
                
                target_cell.value = value
            
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
        
        logger.info(f"Copied column {source_letter} to {target_letter} (with column reference updates)")
    
    def _update_internal_column_refs(self, formula: str, source_col: str, target_col: str) -> str:
        """
        Update internal column references in a formula.
        Changes references from source column to target column.
        
        Examples:
            =SUM(Z23:Z24) with source=Z, target=AA -> =SUM(AA23:AA24)
            =Z25/Z23-1 with source=Z, target=AA -> =AA25/AA23-1
        
        Does NOT update references to other sheets (like BDO sheets).
        """
        # Pattern to match column references that are NOT sheet references
        # Match: letter+number but not preceded by '!' (sheet reference)
        # and not preceded by another letter (multi-letter column)
        
        # First, protect sheet references by temporarily replacing them
        sheet_refs = []
        def protect_sheet_ref(match):
            idx = len(sheet_refs)
            sheet_refs.append(match.group(0))
            return f"__SHEET_REF_{idx}__"
        
        # Protect sheet references like 'Sheet Name'!A1 or Sheet!A1
        protected_formula = re.sub(r"'[^']+'![A-Z]+\$?\d+", protect_sheet_ref, formula)
        protected_formula = re.sub(r"[A-Za-z_][A-Za-z0-9_]*![A-Z]+\$?\d+", protect_sheet_ref, protected_formula)
        
        # Now update column references in the main sheet
        # Pattern: column letter followed by optional $ and row number
        # Only match the source column letter when it appears as a column reference
        pattern = rf'(\$?){source_col}(\$?\d+)'
        replacement = rf'\1{target_col}\2'
        updated_formula = re.sub(pattern, replacement, protected_formula)
        
        # Also handle column references in range expressions like SUM(Z3:Z18)
        # The above pattern should handle this, but let's also handle bare column refs
        
        # Restore sheet references
        for idx, ref in enumerate(sheet_refs):
            updated_formula = updated_formula.replace(f"__SHEET_REF_{idx}__", ref)
        
        return updated_formula
    
    def _update_formulas_with_row_validation(self, sheet, col_idx: int):
        """
        Update formulas to reference the new BDO sheet.
        Also validate that row references point to rows with actual data.
        If a row has no data, search nearby rows (+1, +2, +3) to find valid data.
        """
        new_bdo_name = self.config.bdo_sheet_name
        new_bdo_sheet = self.workbook[new_bdo_name]
        updated_count = 0
        adjusted_rows = 0
        
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            if not cell.value or not isinstance(cell.value, str):
                continue
            
            if not cell.value.startswith('='):
                continue
            
            old_formula = cell.value
            new_formula = old_formula
            
            # Find all BDO references in the formula
            # Pattern: 'BDO - Q2-25'!G73 or 'BDO - Q2-25'!$G$73
            pattern = r"'(BDO[^']+)'!(\$?)([A-Z]+)(\$?)(\d+)"
            
            def replace_bdo_ref(match):
                nonlocal adjusted_rows
                old_sheet = match.group(1)
                dollar1 = match.group(2)
                col_letter = match.group(3)
                dollar2 = match.group(4)
                row_num = int(match.group(5))
                
                # Validate that the row has data in the new BDO sheet
                validated_row = self._find_valid_row(new_bdo_sheet, col_letter, row_num)
                
                if validated_row != row_num:
                    adjusted_rows += 1
                    logger.debug(f"Adjusted row reference: {col_letter}{row_num} -> {col_letter}{validated_row}")
                
                return f"'{new_bdo_name}'!{dollar1}{col_letter}{dollar2}{validated_row}"
            
            new_formula = re.sub(pattern, replace_bdo_ref, new_formula)
            
            if new_formula != old_formula:
                cell.value = new_formula
                updated_count += 1
        
        logger.info(f"Updated {updated_count} formulas in column {get_column_letter(col_idx)}, adjusted {adjusted_rows} row references")
    
    def _find_valid_row(self, bdo_sheet, col_letter: str, original_row: int, search_range: int = 5) -> int:
        """
        Find a valid row that has data in the specified column.
        
        Logic:
        1. If original row has ANY value (including 0), keep it - don't adjust
        2. If original row is empty/None, search nearby rows for data
        
        This ensures we only adjust row references when truly necessary.
        """
        col_idx = column_index_from_string(col_letter)
        
        # Check original row first - if it has ANY value (including 0), keep it
        original_value = bdo_sheet.cell(row=original_row, column=col_idx).value
        if original_value is not None and original_value != '':
            # Original row has data (could be 0, which is valid) - keep it
            return original_row
        
        # Original row is empty/None - search for valid data nearby
        # Search forward first (original+1, original+2, etc.)
        for offset in range(1, search_range + 1):
            new_row = original_row + offset
            if new_row <= bdo_sheet.max_row:
                value = bdo_sheet.cell(row=new_row, column=col_idx).value
                if value is not None and value != '':
                    logger.debug(f"Row {original_row} was empty, found data at row {new_row}")
                    return new_row
        
        # Search backward (original-1, original-2, etc.)
        for offset in range(1, search_range + 1):
            new_row = original_row - offset
            if new_row >= 1:
                value = bdo_sheet.cell(row=new_row, column=col_idx).value
                if value is not None and value != '':
                    logger.debug(f"Row {original_row} was empty, found data at row {new_row}")
                    return new_row
        
        # No valid row found - keep original (formula will show 0 or error)
        logger.warning(f"No valid data found near row {original_row} in column {col_letter}")
        return original_row
    
    def _update_date_references(self):
        """Update date references only in the Management Cijfers summary sheet.
        
        IMPORTANT: Only the summary sheet is modified. All other sheets
        (especially BDO/kwartaal sheets) are protected and never touched.
        """
        old_date_patterns = self._generate_date_patterns(self._get_previous_period_end())
        new_date_patterns = self._generate_date_patterns(self.config.period_end)
        
        target_sheet_name = self.config.summary_sheet_name
        if target_sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Summary sheet '{target_sheet_name}' not found for date reference update")
            return
        
        sheet = self.workbook[target_sheet_name]
        
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, min(sheet.max_column + 1, 10)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    original = cell.value
                    new_value = original
                    
                    for old_pattern, new_pattern in zip(old_date_patterns, new_date_patterns):
                        if old_pattern in new_value:
                            new_value = new_value.replace(old_pattern, new_pattern)
                    
                    if new_value != original:
                        cell.value = new_value
    
    def _get_previous_period_end(self) -> datetime:
        """Get the previous quarter's period end date."""
        month = self.config.period_end.month
        year = self.config.period_end.year
        
        if month <= 3:
            prev_month = 12
            prev_year = year - 1
        else:
            prev_month = month - 3
            prev_year = year
        
        if prev_month in [1, 3, 5, 7, 8, 10, 12]:
            day = 31
        elif prev_month == 2:
            day = 28
        else:
            day = 30
        
        return datetime(prev_year, prev_month, day)
    
    def _generate_date_patterns(self, date: datetime) -> List[str]:
        """Generate various date patterns for replacement."""
        return [
            date.strftime('%d-%m-%Y'),
            date.strftime('%d-%m-%y'),
            date.strftime('%d/%m/%Y'),
            date.strftime('%-d-%-m-%Y'),
            f"Per {date.strftime('%d-%m-%Y')}",
            f"Per {date.strftime('%-d-%-m-%Y')}",
            date.strftime('%d %B %Y'),
            date.strftime('%B %d, %Y'),
        ]
    
    def _compute_bdo_ref_value(self, template: dict, bdo_result: BDOParseResult,
                               use_closing: bool = False) -> Tuple[float, List[str]]:
        """
        Compute the numeric value a bdo_ref formula would produce from raw BDO data.

        Args:
            use_closing: When True, use closing_balance directly (mirrors BDO
                         column H references used by BS and P&L LTM formulas).
                         When False, use closing-opening (mutation).
        """
        account_codes = template.get('account_codes', [])
        sign = template.get('sign', '-')
        operator = template.get('operator', '-')
        
        if not account_codes:
            return 0.0, []
        
        vals = []
        missing = []
        for code in account_codes:
            entry = bdo_result.accounts.get(code)
            if entry is None:
                entry = self._fuzzy_account_lookup(code, bdo_result)
            if entry:
                vals.append(entry.closing_balance if use_closing else
                            entry.closing_balance - entry.opening_balance)
            else:
                vals.append(0.0)
                missing.append(code)
        
        result = -vals[0] if sign == '-' else vals[0]
        for v in vals[1:]:
            if operator == '-':
                result -= v
            else:
                result += v
        
        return result, missing

    @staticmethod
    def _fuzzy_account_lookup(code: str, bdo_result: BDOParseResult) -> Optional[AccountEntry]:
        """Prefix/partial match mirroring _find_account_row logic."""
        for key, entry in bdo_result.accounts.items():
            if key.startswith(code) or code.startswith(key.split('.')[0]):
                return entry
        return None

    @staticmethod
    def _get_quarter_mutation(entry: AccountEntry, bdo_result: BDOParseResult) -> float:
        """
        Extract the column-G (last mutation, i.e. current quarter) value
        from an AccountEntry.  Falls back to closing-opening when the
        mutations dict is empty.
        """
        if entry.mutations:
            last_key = max(entry.mutations.keys(), key=lambda k: int(k.split('_')[1]))
            return entry.mutations[last_key]
        return entry.closing_balance - entry.opening_balance

    def _compute_bdo_ref_quarter_value(self, template: dict,
                                       bdo_result: BDOParseResult) -> Tuple[float, List[str]]:
        """
        Like _compute_bdo_ref_value but returns the column-G (quarter
        mutation) instead of closing-opening (LTM).
        """
        account_codes = template.get('account_codes', [])
        sign = template.get('sign', '-')
        operator = template.get('operator', '-')

        if not account_codes:
            return 0.0, []

        vals: List[float] = []
        missing: List[str] = []
        for code in account_codes:
            entry = bdo_result.accounts.get(code)
            if entry is None:
                entry = self._fuzzy_account_lookup(code, bdo_result)
            if entry:
                vals.append(self._get_quarter_mutation(entry, bdo_result))
            else:
                vals.append(0.0)
                missing.append(code)

        result = -vals[0] if sign == '-' else vals[0]
        for v in vals[1:]:
            if operator == '-':
                result -= v
            else:
                result += v

        return result, missing

    def _compute_bdo_sum_range_quarter_value(self, template: dict,
                                              bdo_result: BDOParseResult) -> Tuple[float, List[str]]:
        """
        Like _compute_bdo_sum_range_value but uses column-G (quarter
        mutation) instead of closing-opening.
        """
        start_account = template.get('start_account')
        end_account = template.get('end_account')
        count = template.get('count')
        additional = template.get('additional', [])

        total = 0.0
        missing: List[str] = []

        if start_account:
            start_entry = (bdo_result.accounts.get(start_account)
                           or self._fuzzy_account_lookup(start_account, bdo_result))
            if start_entry:
                row_to_entry = {e.raw_row: e for e in bdo_result.accounts.values()}
                start_raw = start_entry.raw_row
                if count:
                    for r in range(start_raw, start_raw + count):
                        if r in row_to_entry:
                            total += self._get_quarter_mutation(row_to_entry[r], bdo_result)
                elif end_account:
                    end_entry = (bdo_result.accounts.get(end_account)
                                 or self._fuzzy_account_lookup(end_account, bdo_result))
                    if end_entry:
                        for r in range(start_raw, end_entry.raw_row + 1):
                            if r in row_to_entry:
                                total += self._get_quarter_mutation(row_to_entry[r], bdo_result)
                    else:
                        missing.append(end_account)
                else:
                    total += self._get_quarter_mutation(start_entry, bdo_result)
            else:
                missing.append(start_account)

        for code in additional:
            entry = (bdo_result.accounts.get(code)
                     or self._fuzzy_account_lookup(code, bdo_result))
            if entry:
                total += self._get_quarter_mutation(entry, bdo_result)
            else:
                missing.append(code)

        return total, missing
    
    def _compute_bdo_sum_range_value(self, template: dict, bdo_result: BDOParseResult,
                                       use_mutation: bool = True) -> Tuple[float, List[str]]:
        """
        Compute the numeric value a bdo_sum_range formula would produce from raw BDO data.

        Uses raw_row from AccountEntry to replicate the Excel SUM range logic.
        When use_mutation=True, computes closing_balance - opening_balance (period change).
        """
        start_account = template.get('start_account')
        end_account = template.get('end_account')
        count = template.get('count')
        additional = template.get('additional', [])

        total = 0.0
        missing = []

        if start_account:
            start_entry = bdo_result.accounts.get(start_account) or self._fuzzy_account_lookup(start_account, bdo_result)
            if start_entry:
                row_to_entry = {}
                for entry in bdo_result.accounts.values():
                    row_to_entry[entry.raw_row] = entry

                start_raw = start_entry.raw_row
                if count:
                    for r in range(start_raw, start_raw + count):
                        if r in row_to_entry:
                            e = row_to_entry[r]
                            total += (e.closing_balance - e.opening_balance) if use_mutation else e.closing_balance
                elif end_account:
                    end_entry = bdo_result.accounts.get(end_account) or self._fuzzy_account_lookup(end_account, bdo_result)
                    if end_entry:
                        for r in range(start_raw, end_entry.raw_row + 1):
                            if r in row_to_entry:
                                e = row_to_entry[r]
                                total += (e.closing_balance - e.opening_balance) if use_mutation else e.closing_balance
                    else:
                        missing.append(end_account)
                else:
                    total += (start_entry.closing_balance - start_entry.opening_balance) if use_mutation else start_entry.closing_balance
            else:
                missing.append(start_account)

        for code in additional:
            entry = bdo_result.accounts.get(code) or self._fuzzy_account_lookup(code, bdo_result)
            if entry:
                total += (entry.closing_balance - entry.opening_balance) if use_mutation else entry.closing_balance
            else:
                missing.append(code)

        return total, missing

    def _evaluate_calc_pattern(self, pattern: str, shadow: dict) -> Optional[float]:
        """
        Evaluate a calc-type formula pattern using pre-computed shadow values.

        Handles every pattern present in formula_templates.yaml:
          =SUM({COL}a:{COL}b)       — range sum
          =-({COL}a-{COL}b)         — negated difference
          ={COL}a/{COL}b-1          — ratio minus 1
          ={COL}a+{COL}b+{COL}c    — addition chain (any number of terms)
          ={COL}a+{COL}b            — two-term addition

        Falls back to a general expression evaluator that respects + and -
        operators between {COL}N references.
        """
        p = pattern.strip()

        # =SUM({COL}a:{COL}b)
        sum_match = re.match(r'^=SUM\(\{COL\}(\d+):\{COL\}(\d+)\)$', p)
        if sum_match:
            start_row = int(sum_match.group(1))
            end_row = int(sum_match.group(2))
            total = 0.0
            for r in range(start_row, end_row + 1):
                if r in shadow:
                    total += shadow[r]['value']
            return total

        # =-({COL}a-{COL}b)
        neg_sub_match = re.match(r'^=-\(\{COL\}(\d+)-\{COL\}(\d+)\)$', p)
        if neg_sub_match:
            r1 = int(neg_sub_match.group(1))
            r2 = int(neg_sub_match.group(2))
            v1 = shadow.get(r1, {}).get('value', 0.0)
            v2 = shadow.get(r2, {}).get('value', 0.0)
            return -(v1 - v2)

        # ={COL}a/{COL}b-1
        div_match = re.match(r'^=\{COL\}(\d+)/\{COL\}(\d+)-1$', p)
        if div_match:
            r1 = int(div_match.group(1))
            r2 = int(div_match.group(2))
            v1 = shadow.get(r1, {}).get('value', 0.0)
            v2 = shadow.get(r2, {}).get('value', 0.0)
            if v2 != 0:
                return v1 / v2 - 1
            return 0.0

        # General expression: parse each +/- {COL}N token with correct sign.
        # Handles: ={COL}67+{COL}66, ={COL}53+{COL}48, ={COL}25+{COL}30+{COL}44, etc.
        tokens = re.findall(r'([+\-]?)\{COL\}(\d+)', p.replace('=', '', 1))
        if tokens:
            total = 0.0
            for sign_str, row_str in tokens:
                val = shadow.get(int(row_str), {}).get('value', 0.0)
                if sign_str == '-':
                    total -= val
                else:
                    total += val
            return total

        logger.warning(f"Could not parse calc pattern: {p}")
        return None
    
    def _compute_shadow_pl(self, bdo_result: BDOParseResult,
                           use_closing: bool = False) -> dict:
        """
        Compute a 'shadow P&L' from BDO raw data, mirroring the formula chain.

        Args:
            use_closing: True  → use closing_balance (column H, for LTM)
                         False → use closing-opening (mutation, legacy behaviour)
        """
        shadow = {}
        
        for row_idx in sorted(self.formula_templates.keys()):
            template = self.formula_templates[row_idx]
            formula_type = template.get('type', 'calc')
            label = template.get('label', f'Row {row_idx}')
            entry = {'label': label, 'value': 0.0, 'type': formula_type,
                     'account_codes': [], 'missing_codes': []}
            
            if formula_type == 'bdo_ref':
                val, missing = self._compute_bdo_ref_value(
                    template, bdo_result, use_closing=use_closing)
                entry['value'] = val
                entry['account_codes'] = template.get('account_codes', [])
                entry['missing_codes'] = missing

            elif formula_type == 'bdo_sum_range':
                val, missing = self._compute_bdo_sum_range_value(
                    template, bdo_result, use_mutation=not use_closing)
                entry['value'] = val
                codes = []
                if template.get('start_account'):
                    codes.append(template['start_account'])
                if template.get('end_account'):
                    codes.append(f"..{template['end_account']}")
                codes.extend(template.get('additional', []))
                entry['account_codes'] = codes
                entry['missing_codes'] = missing

            elif formula_type == 'bdo_ref_conditional':
                sub = template.get('q_config', {})
                sub_type = sub.get('type', 'bdo_ref')
                if sub_type == 'bdo_ref':
                    val, missing = self._compute_bdo_ref_value(
                        sub, bdo_result, use_closing=use_closing)
                    entry['value'] = val
                    entry['account_codes'] = sub.get('account_codes', [])
                    entry['missing_codes'] = missing
                elif sub_type == 'bdo_sum_range':
                    val, missing = self._compute_bdo_sum_range_value(
                        sub, bdo_result, use_mutation=not use_closing)
                    entry['value'] = val
                    entry['missing_codes'] = missing
                else:
                    entry['value'] = 0.0
            
            elif formula_type == 'calc':
                pattern = template.get('pattern', '')
                val = self._evaluate_calc_pattern(pattern, shadow)
                entry['value'] = val if val is not None else 0.0
            
            elif formula_type == 'constant':
                entry['value'] = float(template.get('value', 0))
            
            elif formula_type in ('manual', 'manual_with_ltm'):
                if row_idx == 50:
                    entry['value'] = self.config.cash_proceeds_sale or 0.0
                else:
                    entry['value'] = 0.0
            
            shadow[row_idx] = entry
        
        logger.info(f"Shadow P&L computed for {len(shadow)} rows "
                     f"(use_closing={use_closing})")
        return shadow
    
    def _compute_shadow_pl_quarter(self, bdo_result: BDOParseResult) -> dict:
        """
        Compute shadow P&L using column-G (quarter mutation) values only.

        _compute_shadow_pl uses closing-opening (= column H = LTM total).
        This method uses the last-mutation value (= column G = quarter only).
        """
        shadow = {}

        for row_idx in sorted(self.formula_templates.keys()):
            template = self.formula_templates[row_idx]
            formula_type = template.get('type', 'calc')
            label = template.get('label', f'Row {row_idx}')
            entry = {'label': label, 'value': 0.0, 'type': formula_type,
                     'account_codes': [], 'missing_codes': []}

            if formula_type == 'bdo_ref':
                val, missing = self._compute_bdo_ref_quarter_value(template, bdo_result)
                entry['value'] = val
                entry['account_codes'] = template.get('account_codes', [])
                entry['missing_codes'] = missing

            elif formula_type == 'bdo_sum_range':
                val, missing = self._compute_bdo_sum_range_quarter_value(template, bdo_result)
                entry['value'] = val
                codes = []
                if template.get('start_account'):
                    codes.append(template['start_account'])
                if template.get('end_account'):
                    codes.append(f"..{template['end_account']}")
                codes.extend(template.get('additional', []))
                entry['account_codes'] = codes
                entry['missing_codes'] = missing

            elif formula_type == 'bdo_ref_conditional':
                sub = template.get('q_config', {})
                sub_type = sub.get('type', 'bdo_ref')
                if sub_type == 'bdo_ref':
                    val, missing = self._compute_bdo_ref_quarter_value(sub, bdo_result)
                    entry['value'] = val
                    entry['account_codes'] = sub.get('account_codes', [])
                    entry['missing_codes'] = missing
                elif sub_type == 'bdo_sum_range':
                    val, missing = self._compute_bdo_sum_range_quarter_value(sub, bdo_result)
                    entry['value'] = val
                    entry['missing_codes'] = missing
                else:
                    entry['value'] = 0.0

            elif formula_type == 'calc':
                pattern = template.get('pattern', '')
                val = self._evaluate_calc_pattern(pattern, shadow)
                entry['value'] = val if val is not None else 0.0

            elif formula_type == 'constant':
                entry['value'] = float(template.get('value', 0))

            elif formula_type in ('manual', 'manual_with_ltm'):
                if row_idx == 50:
                    entry['value'] = self.config.cash_proceeds_sale or 0.0
                else:
                    entry['value'] = 0.0

            shadow[row_idx] = entry

        logger.info(f"Shadow P&L (quarter) computed for {len(shadow)} rows, "
                     f"row 68 = {shadow.get(68, {}).get('value', 'N/A')}")
        return shadow

    def _compute_shadow_bs(self, bdo_result: BDOParseResult) -> dict:
        """
        Compute a 'shadow Balance Sheet' from BDO raw data using balance_sheet_templates.

        BS formulas reference BDO column H (closing balance), so we use
        closing_balance directly — NOT closing-opening (mutation).
        """
        shadow = {}

        for row_idx in sorted(self.balance_sheet_templates.keys()):
            template = self.balance_sheet_templates[row_idx]
            formula_type = template.get('type', 'bdo_ref')
            label = template.get('label', f'Row {row_idx}')
            entry = {'label': label, 'value': 0.0, 'type': formula_type,
                     'account_codes': [], 'missing_codes': []}

            if formula_type == 'bdo_ref':
                val, missing = self._compute_bdo_ref_value(
                    template, bdo_result, use_closing=True)
                entry['value'] = val
                entry['account_codes'] = template.get('account_codes', [])
                entry['missing_codes'] = missing

            elif formula_type == 'bdo_sum_range':
                val, missing = self._compute_bdo_sum_range_value(
                    template, bdo_result, use_mutation=False)
                entry['value'] = val
                codes = []
                if template.get('start_account'):
                    codes.append(template['start_account'])
                if template.get('end_account'):
                    codes.append(f"..{template['end_account']}")
                codes.extend(template.get('additional', []))
                entry['account_codes'] = codes
                entry['missing_codes'] = missing

            elif formula_type == 'calc':
                pattern = template.get('pattern', '')
                val = self._evaluate_calc_pattern(pattern, shadow)
                entry['value'] = val if val is not None else 0.0

            shadow[row_idx] = entry

        logger.info(f"Shadow Balance Sheet computed for {len(shadow)} rows, "
                     f"row 19 = {shadow.get(19, {}).get('value', 'N/A')}")
        return shadow
    
    # ------------------------------------------------------------------
    # Value extractors — use parsed BDO data + _new_bdo_row_map
    # ------------------------------------------------------------------

    def _account_value(self, code: str, bdo_result: BDOParseResult,
                       col: str) -> float:
        """
        Get a single account's value from parsed BDO data.
        col='H' → closing_balance, col='G' → quarter mutation.
        """
        entry = bdo_result.accounts.get(code)
        if entry is None:
            entry = self._fuzzy_account_lookup(code, bdo_result)
        if entry is None:
            return 0.0
        if col.upper() == 'G':
            return self._get_quarter_mutation(entry, bdo_result)
        return entry.closing_balance

    def _eval_bdo_ref(self, template: dict, bdo_result: BDOParseResult,
                      col: str) -> float:
        """
        Evaluate a bdo_ref template by reading cells from the BDO sheet.
        Falls back to parsed data when the sheet is unavailable.
        """
        account_codes = template.get('account_codes', [])
        sign = template.get('sign', '-')
        operator = template.get('operator', '-')

        if not account_codes:
            return 0.0

        col_idx = ord(col.upper()) - ord('A') + 1
        bdo_sheet_name = self.config.bdo_sheet_name
        bdo_sheet = (self.workbook[bdo_sheet_name]
                     if bdo_sheet_name in self.workbook.sheetnames else None)

        vals = []
        for code in account_codes:
            row_num = self._find_account_row(code)
            if row_num and bdo_sheet:
                vals.append(self._read_bdo_cell_value(bdo_sheet, row_num, col_idx))
            else:
                vals.append(self._account_value(code, bdo_result, col))

        result = -vals[0] if sign == '-' else vals[0]
        for v in vals[1:]:
            result = (result - v) if operator == '-' else (result + v)
        return result

    def _read_bdo_cell_value(self, bdo_sheet, row: int, col_idx: int) -> float:
        """
        Read a numeric value from a BDO sheet cell, evaluating simple
        formulas in-place when the cell contains a formula string.
        """
        val = bdo_sheet.cell(row=row, column=col_idx).value
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            val = val.strip()
            if val.startswith('='):
                return self._eval_cell_formula(bdo_sheet, val, row, col_idx)
            try:
                return float(val.replace(',', '.').replace(' ', ''))
            except (ValueError, AttributeError):
                pass
        return 0.0

    def _eval_cell_formula(self, sheet, formula: str, row: int,
                           col_idx: int) -> float:
        """
        Evaluate a formula cell from the BDO sheet.
        Handles: =SUM(Cn:Gn), =expr+expr, =-expr-expr, simple arithmetic.
        """
        import re
        f = formula.strip()

        # =SUM(C71:G71) — sum a row range
        m = re.match(r'^=SUM\(([A-Z])(\d+):([A-Z])(\d+)\)$', f, re.IGNORECASE)
        if m:
            c1 = ord(m.group(1).upper()) - ord('A') + 1
            r1 = int(m.group(2))
            c2 = ord(m.group(3).upper()) - ord('A') + 1
            r2 = int(m.group(4))
            total = 0.0
            if r1 == r2:
                for c in range(c1, c2 + 1):
                    total += self._read_bdo_cell_value(sheet, r1, c)
            else:
                for r in range(r1, r2 + 1):
                    total += self._read_bdo_cell_value(sheet, r, c1)
            return total

        # Simple arithmetic: =-187446.25-1  or  =23666+2
        try:
            return float(eval(f[1:]))  # strip '=' and eval arithmetic
        except Exception:
            pass

        return 0.0

    def _eval_bdo_sum_range(self, template: dict, bdo_result: BDOParseResult,
                            col: str) -> float:
        """
        Evaluate a bdo_sum_range template by reading directly from the
        in-memory BDO sheet cells.

        This handles duplicate account codes and formula cells correctly
        by summing every cell in the Excel row range (same as the SUM
        formula the builder writes).
        """
        start_account = template.get('start_account')
        end_account = template.get('end_account')
        count = template.get('count')
        additional = template.get('additional', [])

        total = 0.0
        col_idx = ord(col.upper()) - ord('A') + 1

        bdo_sheet_name = self.config.bdo_sheet_name
        if bdo_sheet_name not in self.workbook.sheetnames:
            return total
        bdo_sheet = self.workbook[bdo_sheet_name]

        start_row = self._find_account_row(start_account) if start_account else None

        if start_row is not None:
            if count:
                end_row = start_row + count - 1
            elif end_account:
                end_row = self._find_account_row(end_account) or start_row
            else:
                end_row = start_row

            for r in range(start_row, end_row + 1):
                total += self._read_bdo_cell_value(bdo_sheet, r, col_idx)

        for code in additional:
            row_num = self._find_account_row(code)
            if row_num:
                total += self._read_bdo_cell_value(bdo_sheet, row_num, col_idx)

        return total

    def _eval_template(self, template: dict, bdo_result: BDOParseResult,
                       col: str, computed: dict, key_suffix: str) -> float:
        """Evaluate any template type and return its numeric value."""
        formula_type = template.get('type', 'calc')

        if formula_type == 'bdo_ref':
            return self._eval_bdo_ref(template, bdo_result, col)

        if formula_type == 'bdo_sum_range':
            return self._eval_bdo_sum_range(template, bdo_result, col)

        if formula_type == 'bdo_ref_conditional':
            sub = template.get('q_config', {})
            sub_type = sub.get('type', 'bdo_ref')
            if sub_type == 'bdo_ref':
                return self._eval_bdo_ref(sub, bdo_result, col)
            if sub_type == 'bdo_sum_range':
                return self._eval_bdo_sum_range(sub, bdo_result, col)
            return 0.0

        if formula_type == 'calc':
            shadow = {r: {'value': v} for (r, s), v in computed.items()
                      if s == key_suffix}
            pattern = template.get('pattern', '')
            val = self._evaluate_calc_pattern(pattern, shadow)
            return val if val is not None else 0.0

        if formula_type == 'constant':
            return float(template.get('value', 0))

        return 0.0

    # ------------------------------------------------------------------

    def _populate_computed_values(self, bdo_result: BDOParseResult):
        """
        Build self.computed_values using parsed BDO data combined with
        _new_bdo_row_map for correct row mapping.

        This uses the same account-code → Excel-row mapping that the formula
        builder uses, but reads actual numeric values from the parsed BDO
        result (not from worksheet cells that may contain formulas).

        Keys: (row_number, 'quarter' | 'ltm')  ->  float
        """

        # ── BS rows (LTM only, column H = closing_balance) ──
        for row_idx in sorted(self.balance_sheet_templates.keys()):
            tpl = self.balance_sheet_templates[row_idx]
            val = self._eval_template(tpl, bdo_result, 'H',
                                      self.computed_values, 'ltm')
            self.computed_values[(row_idx, 'ltm')] = val

        # ── P&L rows ──
        for row_idx in sorted(self.formula_templates.keys()):
            tpl = self.formula_templates[row_idx]

            # Quarter → column G (quarter mutation)
            q_val = self._eval_template(tpl, bdo_result, 'G',
                                        self.computed_values, 'quarter')
            self.computed_values[(row_idx, 'quarter')] = q_val

            # LTM → column H (closing_balance)
            ltm_val = self._eval_template(tpl, bdo_result, 'H',
                                          self.computed_values, 'ltm')
            self.computed_values[(row_idx, 'ltm')] = ltm_val

        # ── Manual overrides ──
        self.computed_values[(50, 'quarter')] = self.config.cash_proceeds_sale or 0.0

        # ── Bank rows (LTM only, column H) ──
        self._compute_bank_account_values()

        # ── Rounding (from config) ──
        rounding_cfg = self._rules.get('rounding', {})
        if rounding_cfg.get('round_to_integer', True):
            decimal_rows = set(rounding_cfg.get('decimal_rows', [26]))
            for key in list(self.computed_values.keys()):
                if key[0] not in decimal_rows:
                    self.computed_values[key] = round(self.computed_values[key])

        # ── Identity alignment (from config) ──
        align = self._rules.get('identity_alignment', {})
        auth_row = align.get('authoritative_row', 19)
        dep_row = align.get('dependent_row', 68)
        max_snap = align.get('max_snap_units', 2)
        scope = align.get('scope', 'ltm_only')
        enforce_wb = align.get('enforce_in_workbook', False)

        scopes = ['ltm'] if scope == 'ltm_only' else ['quarter', 'ltm']
        for s in scopes:
            auth_val = self.computed_values.get((auth_row, s))
            dep_val = self.computed_values.get((dep_row, s))
            if auth_val is not None and dep_val is not None:
                diff = abs(auth_val - dep_val)
                if enforce_wb and diff > 0:
                    # Workbook formula guarantees identity — mirror in cache
                    logger.info(
                        f"[Alignment] enforce_in_workbook: setting row "
                        f"{dep_row} {s} {dep_val} → row {auth_row} {s} "
                        f"{auth_val} (diff={diff})")
                    self.computed_values[(dep_row, s)] = auth_val
                elif 0 < diff <= max_snap:
                    logger.info(
                        f"[Alignment] Snapping row {dep_row} {s} "
                        f"{dep_val} → row {auth_row} {s} {auth_val}")
                    self.computed_values[(dep_row, s)] = auth_val
                elif diff > max_snap:
                    logger.warning(
                        f"[Alignment] Row {auth_row} {s}={auth_val} vs "
                        f"row {dep_row} {s}={dep_val} differ by {diff} "
                        f"(> max_snap {max_snap})")

        logger.info(
            f"Populated computed_values: {len(self.computed_values)} entries, "
            f"BS19(ltm)={self.computed_values.get((19, 'ltm'), 'N/A')}, "
            f"PL68(quarter)={self.computed_values.get((68, 'quarter'), 'N/A')}, "
            f"PL68(ltm)={self.computed_values.get((68, 'ltm'), 'N/A')}"
        )

    def _enforce_bdo_alignment(self, bdo_result: BDOParseResult):
        """
        Write formulas (not hardcoded values) for rows 19 and 68 in the LTM
        column, consolidate interest accounts dynamically, and run a post-build
        coverage audit. All BDO row positions are discovered at runtime.
        """
        align = self._rules.get('identity_alignment', {})
        auth_row = align.get('authoritative_row', 19)
        dep_row = align.get('dependent_row', 68)

        bdo_qtr_raw = self._read_bdo_resultaat_na_belasting(use_ltm=False)
        bdo_ltm_raw = self._read_bdo_resultaat_na_belasting_cached()

        if bdo_ltm_raw is not None and bdo_ltm_raw != 0.0:
            effective_raw = bdo_ltm_raw
            source = "column H (data_only cache)"
        elif bdo_qtr_raw is not None and bdo_qtr_raw != 0.0:
            effective_raw = bdo_qtr_raw
            source = "quarter sum D-G"
        else:
            effective_raw = None
            source = "none"

        # Persist sign-adjusted BDO totals separately for revalidation /
        # logging. Earlier versions wrote these straight into
        # `computed_values[(19,'ltm')]` and `[(68,'ltm')]`, which silently
        # turned the downstream pass/fail check into a tautology: the AI
        # revalidator was comparing those cached numbers against the same
        # BDO H value they had been derived from, so the loop could PASS
        # even when the actual Excel formulas in column AD did not
        # evaluate to the BDO ground truth. Keep the cache derived from
        # the formula templates (set by _populate_computed_values) so the
        # cross-check stays meaningful.
        if effective_raw is not None:
            bdo_ltm = round(-effective_raw, 2)
            logger.info(
                f"[BDO Align] LTM ground truth={bdo_ltm:,.2f} (source: {source})"
            )
            self.bdo_ground_truth['ltm'] = bdo_ltm
        else:
            logger.warning("[BDO Align] Cannot determine LTM ground truth")

        if bdo_qtr_raw is not None:
            self.bdo_ground_truth['quarter'] = round(-bdo_qtr_raw, 2)

        summary_name = self.config.summary_sheet_name
        if summary_name not in self.workbook.sheetnames:
            return
        ws = self.workbook[summary_name]
        ltm_col = self._new_ltm_col
        ltm_letter = get_column_letter(ltm_col)
        bdo_name = self.config.bdo_sheet_name

        # Write calc formulas from config for total rows
        calc_formulas = self._rules.get('layout', {}).get('calc_formulas', {})
        ltm_overrides = self._rules.get('layout', {}).get('ltm_calc_overrides', {})
        all_calcs = {**calc_formulas, **ltm_overrides}

        # Resolve the BDO "Verschil balans en winst-en-verlies" row dynamically
        # for templates that reference {verschil_row} (currently the row 19
        # override). The position varies between BDO files.
        bdo_sheet_for_verschil = (
            self.workbook[bdo_name]
            if bdo_name in self.workbook.sheetnames else None
        )
        verschil_row = self._find_verschil_row(bdo_sheet_for_verschil)

        for row_key in [dep_row, auth_row]:
            tmpl = all_calcs.get(row_key) or all_calcs.get(str(row_key))
            if tmpl:
                # Same substitutions as _apply_ltm_calc_formulas_from_rules
                # so templates using {bdo} and {verschil_row} (e.g. row 19's
                # -H{verschil_row} reconciliation correction) are written
                # with dynamically-resolved values.
                formula = (
                    tmpl.replace('{col}', ltm_letter)
                        .replace('{COL}', ltm_letter)
                        .replace('{bdo}', bdo_name)
                        .replace('{verschil_row}', str(verschil_row))
                )
                ws.cell(row=row_key, column=ltm_col).value = formula
                logger.info(f"[BDO Align] Row {row_key}: {formula}")
            else:
                logger.warning(f"[BDO Align] No calc_formula for row {row_key}")

        # Discover BDO structure dynamically for interest, duplicates, coverage
        bdo_sheet = (self.workbook[bdo_name]
                     if bdo_name in self.workbook.sheetnames else None)
        if bdo_sheet:
            structure = self._discover_bdo_structure(bdo_sheet)
            header_row = self._rules.get('layout', {}).get('header_row', 22)
            q_col = resolve_quarter_column(
                ws,
                self.config.quarter,
                header_row,
                built_quarter_col=self._new_quarter_col,
            )
            if not q_col and ltm_col:
                q_col = ltm_col - 1

            self._build_ltm_interest(
                ws, ltm_col, bdo_sheet, bdo_name, structure)

            self._handle_duplicate_accounts(
                ws, bdo_sheet, ltm_col, q_col, bdo_name, structure)

            self._post_build_coverage_audit(
                ws, bdo_sheet, ltm_col, q_col, bdo_name, structure)

        # Quarter alignment (computed_values only, for context)
        if bdo_qtr_raw is not None and bdo_qtr_raw != 0.0:
            bdo_qtr = round(-bdo_qtr_raw, 2)
            pl68_q = self.computed_values.get((dep_row, 'quarter'), 0.0)
            if pl68_q != bdo_qtr:
                logger.info(
                    f"[BDO Align] Quarter: PL{dep_row}={pl68_q:,.2f} → "
                    f"BDO ground truth={bdo_qtr:,.2f} (delta for AI)"
                )
            else:
                logger.info("[BDO Align] Quarter PL68 matches ground truth")

    def _find_verschil_row(self, bdo_sheet) -> int:
        """
        Locate the row whose column A label is "Verschil balans en
        winst-en-verlies" in a BDO sheet. The position differs between
        quarterly BDO files, so the row number must always be discovered
        dynamically — never hardcoded.

        Returns the discovered 1-based row number, or 127 as a last-resort
        default that mirrors the historical layout for older BDO files.
        """
        if bdo_sheet is None:
            return 127

        bdo_cfg = self._rules.get('bdo', {})
        label_cols = bdo_cfg.get('label_scan_columns', [1, 2])
        scan_limit = min(bdo_sheet.max_row + 1, 200)

        for row_idx in range(1, scan_limit):
            for col in label_cols:
                cell_val = bdo_sheet.cell(row=row_idx, column=col).value
                if isinstance(cell_val, str) and 'verschil balans' in cell_val.lower():
                    return row_idx

        logger.warning(
            "[BDO Align] 'Verschil balans en winst-en-verlies' row not "
            "found in BDO sheet; defaulting to row 127"
        )
        return 127

    def _read_bdo_resultaat_na_belasting_cached(self) -> Optional[float]:
        """
        Re-open the BDO source xlsx with data_only=True so that column-H
        formulas resolve to their cached numeric values (last-calculated by
        Excel).  Returns None if the file is unavailable or the row isn't found.
        """
        if not self.bdo_source_path or not self.bdo_source_path.exists():
            return None

        try:
            wb = openpyxl.load_workbook(self.bdo_source_path, data_only=True)
            sheet = wb.active
            bdo_cfg = self._rules.get('bdo', {})
            data_col = bdo_cfg.get('data_column', 8)
            label_cols = bdo_cfg.get('label_scan_columns', [1, 2])

            target_row = None
            for row_idx in range(1, min(sheet.max_row + 1, 200)):
                for col in label_cols:
                    cell_val = sheet.cell(row=row_idx, column=col).value
                    if cell_val and isinstance(cell_val, str):
                        if 'resultaat na belasting' in cell_val.lower():
                            target_row = row_idx
                            break
                if target_row:
                    break

            if target_row is None:
                wb.close()
                return None

            val = sheet.cell(row=target_row, column=data_col).value
            wb.close()

            if isinstance(val, (int, float)):
                logger.info(
                    f"[BDO Align] Cached column H value at row {target_row}: {float(val):,.2f}"
                )
                return float(val)
            return None
        except Exception as e:
            logger.warning(f"[BDO Align] data_only re-read failed: {e}")
            return None

    def _compute_bank_account_values(self):
        """
        Compute expected numeric values for bank account rows
        by reading BDO column H closing balance cells.
        Row mapping and total-row number come from accounting_rules.yaml.
        """
        bank_bdo_rows = self._rules.get('bank_bdo_row_map', {})
        bank_total_row = self._rules.get('bank_total_row', 114)

        bdo_sheet_name = self.config.bdo_sheet_name
        if bdo_sheet_name not in self.workbook.sheetnames:
            logger.warning("BDO sheet not found for bank account values")
            return

        bdo_sheet = self.workbook[bdo_sheet_name]
        data_col = self._rules.get('bdo', {}).get('data_column', 8)
        bank_total = 0.0
        for mc_row, bdo_row in bank_bdo_rows.items():
            cell = bdo_sheet.cell(row=bdo_row, column=data_col)
            val = cell.value
            if isinstance(val, (int, float)):
                self.computed_values[(mc_row, 'ltm')] = float(val)
                bank_total += float(val)
            elif isinstance(val, str) and val.startswith('='):
                evaluated = self._eval_simple_sum_formula(bdo_sheet, val, data_col)
                self.computed_values[(mc_row, 'ltm')] = evaluated
                bank_total += evaluated
            else:
                self.computed_values[(mc_row, 'ltm')] = 0.0

        self.computed_values[(bank_total_row, 'ltm')] = bank_total

    def _reconcile_pl_chain(self, shadow: dict, equity_movement: float) -> dict:
        """
        Compare shadow P&L chain against equity movement to find divergence.
        
        Walks backwards from row 68 (Direct Result) through the formula chain
        to identify which specific row(s) contribute to any mismatch.
        """
        shadow_direct = shadow.get(68, {}).get('value', 0.0)
        tolerance = 1.0
        difference = abs(shadow_direct - equity_movement)
        is_reconciled = difference <= tolerance
        
        key_rows = [68, 66, 57, 55, 48, 45, 25, 67, 60, 61, 64, 65, 56, 53, 50,
                    44, 30, 23, 24, 27, 28, 29, 46, 47]
        
        breakdown = []
        for r in key_rows:
            if r in shadow:
                s = shadow[r]
                breakdown.append({
                    'row': r,
                    'label': s['label'],
                    'value': s['value'],
                    'type': s['type'],
                    'missing_codes': s.get('missing_codes', []),
                })
        
        divergence_point = None
        if not is_reconciled:
            chain_pairs = [
                (68, [67, 66], 'Direct result = Delta DTA + EBT'),
                (66, list(range(57, 66)), 'EBT = SUM(57:65)'),
                (57, [56, 55], 'EBIT = Depreciation + Total EBITDA'),
                (55, [53, 48], 'Total EBITDA = Sales + Rental'),
                (48, [47, 46, 45], 'EBITDA = Mgmt fees + Net rental'),
                (45, [25, 30, 44], 'Net rental = Gross + Service + Property'),
            ]
            
            for parent_row, child_rows, desc in chain_pairs:
                parent_val = shadow.get(parent_row, {}).get('value', 0.0)
                children_sum = sum(
                    shadow.get(r, {}).get('value', 0.0) for r in child_rows
                    if r in shadow
                )
                if abs(parent_val - children_sum) > tolerance:
                    divergence_point = {
                        'row': parent_row,
                        'label': shadow.get(parent_row, {}).get('label', ''),
                        'expected': children_sum,
                        'actual': parent_val,
                        'detail': f"{desc}: children sum={children_sum:,.2f}, parent={parent_val:,.2f}"
                    }
                    break
            
            if not divergence_point:
                bdo_ref_rows = [r for r in shadow if shadow[r]['type'] in ('bdo_ref', 'bdo_ref_conditional')]
                for r in bdo_ref_rows:
                    if shadow[r].get('missing_codes'):
                        codes_str = ', '.join(shadow[r]['missing_codes'])
                        divergence_point = {
                            'row': r,
                            'label': shadow[r]['label'],
                            'expected': None,
                            'actual': shadow[r]['value'],
                            'detail': f"Missing account codes in BDO data: {codes_str}"
                        }
                        break
        
        result = {
            'shadow_direct_result': shadow_direct,
            'equity_movement': equity_movement,
            'is_reconciled': is_reconciled,
            'difference': difference,
            'row_breakdown': breakdown,
            'divergence_point': divergence_point,
        }
        
        if is_reconciled:
            logger.info(f"Shadow P&L reconciliation passed: shadow={shadow_direct:,.2f}, equity={equity_movement:,.2f}")
        else:
            logger.warning(
                f"Shadow P&L reconciliation FAILED: shadow={shadow_direct:,.2f}, "
                f"equity={equity_movement:,.2f}, diff={difference:,.2f}"
            )
            if divergence_point:
                logger.warning(f"  Divergence at row {divergence_point['row']}: {divergence_point['detail']}")
        
        return result
    
    def _validate_calculations(self, bdo_result: BDOParseResult) -> dict:
        """
        Validate that calculations are correct and return structured results.
        
        Three-pass validation:
        1. Template-based check: shadow BS row 19 vs shadow P&L row 68
           (uses the exact same account codes as the Excel formulas)
        2. Structural formula check: verify key Excel formulas are present and correct
        3. Shadow reconciliation: row-by-row check to pinpoint divergence
        
        If any check fails, writes warning rows into the Management Cijfers sheet.
        """
        # --- Compute shadow models from BDO data using template account codes ---
        shadow_pl = self._compute_shadow_pl(bdo_result)
        shadow_bs = self._compute_shadow_bs(bdo_result)
        
        direct_result = shadow_pl.get(68, {}).get('value', 0.0)
        equity_movement = shadow_bs.get(19, {}).get('value', 0.0)
        
        tolerance = 1.0
        difference = abs(equity_movement - direct_result)
        is_aligned = difference <= tolerance
        
        validation = {
            'equity_movement': equity_movement,
            'direct_result': direct_result,
            'is_aligned': is_aligned,
            'difference': difference,
            'formula_checks': {},
            'shadow_bs': shadow_bs,
            'shadow_pl': shadow_pl,
            'messages': []
        }
        
        if not is_aligned:
            msg = (
                f"Template check: Total Equity Movement (BS row 19 = {equity_movement:,.2f}) != "
                f"Direct Result (P&L row 68 = {direct_result:,.2f}), difference: {difference:,.2f}"
            )
            validation['messages'].append(msg)
            logger.warning(f"VALIDATION WARNING: {msg}")
            
            bs_missing = []
            for r in sorted(shadow_bs.keys()):
                if r == 19:
                    continue
                codes = shadow_bs[r].get('missing_codes', [])
                if codes:
                    bs_missing.append(f"BS row {r} ({shadow_bs[r]['label']}): {', '.join(codes)}")
            if bs_missing:
                validation['messages'].append(
                    f"Missing BDO accounts in Balance Sheet: {'; '.join(bs_missing)}"
                )
        else:
            logger.info(
                f"Template check passed: BS row 19 = {equity_movement:,.2f} "
                f"≈ P&L row 68 = {direct_result:,.2f}"
            )
        
        # --- Pass 1b: BDO ground truth check ---
        # Profits are stored as NEGATIVE in BDO (Dutch trial balance convention).
        # Management Cijfers uses positive-for-profit (formula templates negate via sign: "-").
        # The ground truth comparison must account for this sign difference.
        bdo_resultaat_raw = self._read_bdo_resultaat_na_belasting()
        validation['bdo_resultaat_raw'] = bdo_resultaat_raw
        
        if bdo_resultaat_raw is not None:
            # Negate BDO value: BDO stores profit as negative, MC stores as positive
            bdo_resultaat = -bdo_resultaat_raw
            validation['bdo_resultaat'] = bdo_resultaat
            
            diff_dr = abs(bdo_resultaat - direct_result)
            diff_eq = abs(bdo_resultaat - equity_movement)
            dr_ok = diff_dr <= tolerance
            eq_ok = diff_eq <= tolerance
            
            validation['bdo_vs_direct_result'] = diff_dr
            validation['bdo_vs_equity_movement'] = diff_eq
            
            if not dr_ok or not eq_ok:
                validation['is_aligned'] = False
                validation['messages'].append(
                    f"BDO ground truth: Resultaat na belasting (kwartaal D-G) = {bdo_resultaat_raw:,.2f} "
                    f"(sign-adjusted: {bdo_resultaat:,.2f})"
                )
                if not dr_ok:
                    validation['messages'].append(
                        f"  Expected (BDO sign-adjusted): {bdo_resultaat:,.2f}"
                    )
                    validation['messages'].append(
                        f"  Current  (Direct Result P&L row 68): {direct_result:,.2f}"
                    )
                    validation['messages'].append(
                        f"  Difference: {diff_dr:,.2f}"
                    )
                if not eq_ok:
                    validation['messages'].append(
                        f"  Expected (BDO sign-adjusted): {bdo_resultaat:,.2f}"
                    )
                    validation['messages'].append(
                        f"  Current  (Equity Movement BS row 19): {equity_movement:,.2f}"
                    )
                    validation['messages'].append(
                        f"  Difference: {diff_eq:,.2f}"
                    )
                logger.warning(
                    f"VALIDATION WARNING: BDO Resultaat na belasting "
                    f"(raw={bdo_resultaat_raw:,.2f}, adjusted={bdo_resultaat:,.2f}) "
                    f"differs from Direct Result ({direct_result:,.2f}, Δ={diff_dr:,.2f}) "
                    f"and/or Equity Movement ({equity_movement:,.2f}, Δ={diff_eq:,.2f})"
                )
            else:
                logger.info(
                    f"BDO ground truth check passed: Resultaat(raw={bdo_resultaat_raw:,.2f}, "
                    f"adjusted={bdo_resultaat:,.2f}) ≈ DR={direct_result:,.2f} ≈ EM={equity_movement:,.2f}"
                )
        else:
            validation['bdo_resultaat'] = None
            validation['messages'].append(
                "BDO ground truth: Could not locate 'Resultaat na belasting' row in BDO sheet"
            )
        
        # --- Pass 1c: LTM shadow vs BDO column H ---
        # The quarter-style shadow_pl above uses use_closing=False (D-G
        # mutation values). Run a second shadow pass with use_closing=True
        # so we can compare the LTM-side P&L total against BDO column H
        # on Resultaat na belasting at build time. Without this any LTM
        # row-mapping error stays invisible until the AI verifier runs —
        # surfacing it here gives an early signal for the deterministic
        # LTM mirror to repair.
        try:
            shadow_pl_ltm = self._compute_shadow_pl(bdo_result, use_closing=True)
            shadow_bs_ltm = self._compute_shadow_bs(bdo_result)
            ltm_dr = shadow_pl_ltm.get(68, {}).get('value', 0.0)
            ltm_em = shadow_bs_ltm.get(19, {}).get('value', 0.0)
            validation['shadow_pl_ltm'] = shadow_pl_ltm
            validation['ltm_direct_result_shadow'] = ltm_dr
            validation['ltm_equity_movement_shadow'] = ltm_em

            bdo_resultaat_h_raw = self._read_bdo_resultaat_na_belasting(use_ltm=True)
            if bdo_resultaat_h_raw is not None:
                bdo_h = -bdo_resultaat_h_raw
                validation['bdo_resultaat_h'] = bdo_h
                ltm_diff_dr = abs(bdo_h - ltm_dr)
                ltm_diff_em = abs(bdo_h - ltm_em)
                if ltm_diff_dr > tolerance or ltm_diff_em > tolerance:
                    validation['is_aligned'] = False
                    msg = (
                        f"LTM shadow check: BDO column H (Resultaat na "
                        f"belasting) sign-adjusted = {bdo_h:,.2f}; "
                        f"shadow PL68 (use_closing=True) = {ltm_dr:,.2f} "
                        f"(Δ={ltm_diff_dr:,.2f}); shadow BS19 = "
                        f"{ltm_em:,.2f} (Δ={ltm_diff_em:,.2f})"
                    )
                    validation['messages'].append(msg)
                    logger.warning(f"VALIDATION WARNING: {msg}")
                    for r in sorted(shadow_pl_ltm.keys()):
                        if r in (66, 68):
                            continue
                        codes = shadow_pl_ltm[r].get('missing_codes') or []
                        if codes:
                            validation['messages'].append(
                                f"LTM shadow row {r} "
                                f"({shadow_pl_ltm[r].get('label')}): "
                                f"missing accounts {', '.join(codes)}"
                            )
                else:
                    logger.info(
                        f"LTM shadow check passed: BDO H={bdo_h:,.2f} ≈ "
                        f"DR_LTM={ltm_dr:,.2f} ≈ EM_LTM={ltm_em:,.2f}"
                    )
        except Exception as exc:
            logger.warning(f"LTM shadow check failed (non-fatal): {exc}")

        # --- Pass 2: Structural formula check ---
        formula_ok = self._validate_structural_formulas(validation)
        
        # --- Pass 3: Shadow reconciliation (detailed row-by-row) ---
        reconciliation = self._reconcile_pl_chain(shadow_pl, equity_movement)
        validation['reconciliation'] = reconciliation
        
        if not reconciliation['is_reconciled']:
            validation['messages'].append(
                f"Row-by-row: P&L chain result = {reconciliation['shadow_direct_result']:,.2f}, "
                f"BS equity movement = {reconciliation['equity_movement']:,.2f}, "
                f"difference = {reconciliation['difference']:,.2f}"
            )
            
            if reconciliation.get('divergence_point'):
                dp = reconciliation['divergence_point']
                validation['messages'].append(
                    f"Divergence point: Row {dp['row']} ({dp['label']}) - {dp['detail']}"
                )
            
            for row_info in reconciliation['row_breakdown']:
                if row_info.get('missing_codes'):
                    codes = ', '.join(row_info['missing_codes'])
                    validation['messages'].append(
                        f"Row {row_info['row']} ({row_info['label']}): "
                        f"missing BDO accounts: {codes}"
                    )
            
            validation['is_aligned'] = False
        
        if not validation['is_aligned'] or not formula_ok or not reconciliation['is_reconciled']:
            self._add_validation_warning_to_sheet(validation['messages'])
        
        return validation
    
    def _read_bdo_resultaat_na_belasting(self, use_ltm: bool = False) -> Optional[float]:
        """
        Read the 'Resultaat na belasting' ground truth value from the BDO sheet.

        Args:
            use_ltm: If True, read column H (LTM total).
                     If False, sum columns D-G (quarter mutations).

        Returns None if the row or sheet cannot be found.
        """
        bdo_sheet_name = self.config.bdo_sheet_name
        if bdo_sheet_name not in self.workbook.sheetnames:
            logger.warning("BDO ground truth check: sheet not found")
            return None

        sheet = self.workbook[bdo_sheet_name]

        target_row = None

        label_key = 'resultaat na belasting'
        if label_key in self._new_bdo_label_map:
            target_row = self._new_bdo_label_map[label_key]
        else:
            label_cols = self._rules.get('bdo', {}).get('label_scan_columns', [1, 2])
            for row_idx in range(1, min(sheet.max_row + 1, 200)):
                for col in label_cols:
                    cell_val = sheet.cell(row=row_idx, column=col).value
                    if cell_val and isinstance(cell_val, str):
                        if 'resultaat na belasting' in cell_val.lower():
                            target_row = row_idx
                            break
                if target_row:
                    break

        if target_row is None:
            logger.warning("BDO ground truth check: 'Resultaat na belasting' row not found")
            return None

        def _read_cell(col_idx):
            val = sheet.cell(row=target_row, column=col_idx).value
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str) and val.startswith('='):
                return self._eval_simple_sum_formula(sheet, val, col_idx)
            if isinstance(val, str):
                try:
                    return float(val.replace(',', '.').replace(' ', ''))
                except (ValueError, AttributeError):
                    pass
            return 0.0

        bdo_cfg = self._rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)
        q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])

        if use_ltm:
            total = _read_cell(data_col)
            logger.info(
                f"BDO ground truth (LTM): Resultaat na belasting "
                f"(row {target_row}, col {get_column_letter(data_col)}) = {total:,.2f}")
        else:
            total = sum(_read_cell(c) for c in q_cols)
            q_letters = [get_column_letter(c) for c in q_cols]
            logger.info(
                f"BDO ground truth (quarter): Resultaat na belasting "
                f"(row {target_row}, {q_letters[0]}-{q_letters[-1]}) = {total:,.2f}")
        return total

    def _eval_simple_sum_formula(self, sheet, formula: str, default_col: int) -> float:
        """
        Evaluate a simple SUM formula by reading cell values.

        Handles both column ranges (=SUM(D6:D123)) and row ranges (=SUM(D135:G135)).
        Returns 0.0 for unparseable formulas.
        """
        m = re.match(r'^=SUM\(([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)\)$', formula, re.IGNORECASE)
        if not m:
            logger.debug(f"Cannot evaluate formula: {formula}")
            return 0.0

        col_start = column_index_from_string(m.group(1).upper())
        row_start = int(m.group(2))
        col_end = column_index_from_string(m.group(3).upper())
        row_end = int(m.group(4))

        total = 0.0
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                cell_val = sheet.cell(row=r, column=c).value
                if isinstance(cell_val, (int, float)):
                    total += cell_val
        return total

    def _validate_structural_formulas(self, validation: dict) -> bool:
        """
        Verify that key Excel formulas in Management Cijfers are structurally correct.
        
        Checks row 19 (Total Equity Movement), row 66 (EBT), row 68 (Direct Result),
        and intermediate P&L rows (60, 61, 64, 65, 67) for the new quarter and LTM columns.
        
        Returns True if all formula checks pass.
        """
        summary_sheets = [name for name in self.workbook.sheetnames
                          if 'Management Cijfers' in name]
        if not summary_sheets:
            logger.warning("Structural validation skipped: Management Cijfers sheet not found")
            return True
        
        sheet = self.workbook[summary_sheets[-1]]
        
        header_row = self._rules.get('layout', {}).get('header_row', 22)
        layout = scan_summary_column_layout(sheet, header_row)
        ltm_col = layout.ltm_column
        if not ltm_col:
            logger.warning("Structural validation skipped: LTM column not found")
            return True

        quarter_col = find_column_by_header_label(
            sheet, self.config.quarter, header_row
        )
        if not quarter_col:
            quarter_col = layout.last_quarter_column
        if not quarter_col:
            logger.warning("Structural validation skipped: quarter column not found")
            return True
        q_letter = get_column_letter(quarter_col)
        ltm_letter = get_column_letter(ltm_col)
        
        all_ok = True
        checks = {}
        
        expected_formulas = {
            66: {'col': quarter_col, 'letter': q_letter,
                 'expected': f'=SUM({q_letter}57:{q_letter}65)',
                 'label': 'EBT (quarter)'},
            68: {'col': quarter_col, 'letter': q_letter,
                 'expected': f'={q_letter}67+{q_letter}66',
                 'label': 'Direct Result (quarter)'},
        }
        
        for row_num, spec in expected_formulas.items():
            cell = sheet.cell(row=row_num, column=spec['col'])
            actual = str(cell.value) if cell.value is not None else '(empty)'
            ok = actual.upper() == spec['expected'].upper()
            checks[f'row_{row_num}'] = {
                'label': spec['label'],
                'expected': spec['expected'],
                'actual': actual,
                'ok': ok
            }
            if not ok:
                all_ok = False
                msg = (
                    f"Formula check failed: {spec['label']} row {row_num} col {spec['letter']} - "
                    f"expected '{spec['expected']}', got '{actual}'"
                )
                validation['messages'].append(msg)
                logger.warning(f"VALIDATION WARNING: {msg}")

        # Row 19 LTM: must contain a formula (e.g. ={ltm}68), not a hardcoded value
        row19_cell = sheet.cell(row=19, column=ltm_col)
        row19_val = row19_cell.value
        if not isinstance(row19_val, str) or not row19_val.startswith('='):
            all_ok = False
            msg = (
                f"Formula check failed: Total Equity Movement (LTM) row 19 col {ltm_letter} - "
                f"expected a formula, got '{row19_val}'"
            )
            validation['messages'].append(msg)
            logger.warning(f"VALIDATION WARNING: {msg}")
        else:
            logger.info(f"Row 19 LTM: formula present — {row19_val}")
        
        intermediate_rows = {
            60: 'Net interest expenses - SFA',
            61: 'Prepaid derivatives - SFA',
            64: 'Net interest expenses - Hedge',
            65: 'Net interest income - DMRRP',
            67: 'Delta DTA & CIT',
        }
        for row_num, label in intermediate_rows.items():
            q_cell = sheet.cell(row=row_num, column=quarter_col)
            ltm_cell = sheet.cell(row=row_num, column=ltm_col)
            q_val = q_cell.value
            ltm_val = ltm_cell.value
            
            q_ok = q_val is not None
            ltm_ok = ltm_val is not None
            checks[f'row_{row_num}_q'] = {
                'label': f'{label} (quarter)',
                'has_formula': q_ok,
                'actual': str(q_val) if q_val is not None else '(empty)',
                'ok': q_ok
            }
            checks[f'row_{row_num}_ltm'] = {
                'label': f'{label} (LTM)',
                'has_formula': ltm_ok,
                'actual': str(ltm_val) if ltm_val is not None else '(empty)',
                'ok': ltm_ok
            }
            if not q_ok:
                all_ok = False
                msg = f"Formula missing: {label} (quarter) row {row_num} col {q_letter} is empty"
                validation['messages'].append(msg)
                logger.warning(f"VALIDATION WARNING: {msg}")
            if not ltm_ok:
                all_ok = False
                msg = f"Formula missing: {label} (LTM) row {row_num} col {ltm_letter} is empty"
                validation['messages'].append(msg)
                logger.warning(f"VALIDATION WARNING: {msg}")
        
        validation['formula_checks'] = checks
        
        if all_ok:
            logger.info("Structural formula validation passed: all key formulas present and correct")
        else:
            validation['is_aligned'] = False
        
        return all_ok
    
    @staticmethod
    def _read_bdo_resultaat_h(
        workbook_path: str,
        bdo_sheet_name: str,
    ) -> Optional[float]:
        """
        Read the sign-adjusted "Resultaat na belasting" total from BDO
        column H in the saved workbook. Returns None when the row or H
        value cannot be resolved.

        Used by the post-AI orchestrator refresh to compare the
        evaluator's LTM totals against the authoritative BDO ground
        truth before snapping ``computed_values``.
        """
        wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
        try:
            if bdo_sheet_name not in wb_data.sheetnames:
                return None
            sheet = wb_data[bdo_sheet_name]
            target_row = None
            for row_idx in range(1, min(sheet.max_row + 1, 200)):
                for col in (1, 2):
                    label = sheet.cell(row=row_idx, column=col).value
                    if (
                        isinstance(label, str)
                        and 'resultaat na belasting' in label.lower()
                    ):
                        target_row = row_idx
                        break
                if target_row is not None:
                    break
            if target_row is None:
                return None
            h_val = sheet.cell(row=target_row, column=8).value
            if isinstance(h_val, (int, float)):
                return -float(h_val)
            running = 0.0
            saw = False
            for col_idx in range(3, 8):
                v = sheet.cell(row=target_row, column=col_idx).value
                if isinstance(v, (int, float)):
                    running += float(v)
                    saw = True
            return -running if saw else None
        finally:
            wb_data.close()

    @staticmethod
    def evaluate_ma_column_from_formulas(
        workbook_path: str,
        summary_sheet_name: str,
        bdo_sheet_name: str,
        column_scope: str = 'ltm',
        rules: Optional[dict] = None,
    ) -> Optional[Dict[str, Any]]:
        """
        Evaluate the actual formulas in the Management Cijfers quarter or
        LTM column against BDO column G/H respectively.

        This is the authoritative pass/fail signal for the AI verification
        loop and the post-AI orchestrator refresh — earlier versions
        compared a pre-computed cache that had been overwritten with the
        BDO ground truth itself, which made the check tautological.

        Args:
            workbook_path: path to the saved .xlsx file.
            summary_sheet_name: e.g. ``"Management Cijfers - Q1 2026"``.
            bdo_sheet_name: current BDO sheet (e.g. ``"BDO - Q1-26"``).
            column_scope: ``"ltm"`` or ``"quarter"``.
            rules: pre-loaded accounting rules (optional; loaded from
                ``config/accounting_rules.yaml`` when omitted).

        Returns ``{"values": {row: float}, "unresolved": [...],
        "column": int, "scope": "ltm"|"quarter"}`` or ``None`` when the
        target column cannot be resolved.
        """
        from .formula_mirror import evaluate_ma_column

        if rules is None:
            rules_path = Path("config/accounting_rules.yaml")
            if rules_path.exists():
                try:
                    with open(rules_path, 'r', encoding='utf-8') as f:
                        rules = yaml.safe_load(f) or {}
                except Exception:
                    rules = {}
            else:
                rules = {}

        layout_cfg = rules.get('layout', {})
        bs_range = layout_cfg.get('balance_sheet_rows', [3, 19])
        pl_range = layout_cfg.get('profit_loss_rows', [23, 68])
        bank_range = layout_cfg.get('bank_rows', [107, 114])
        header_row = layout_cfg.get('header_row', 22)

        rows_of_interest = (
            list(range(bs_range[0], bs_range[1] + 1))
            + list(range(pl_range[0], pl_range[1] + 1))
            + list(range(bank_range[0], bank_range[1] + 1))
        )

        wb = openpyxl.load_workbook(workbook_path, data_only=False)
        try:
            sheet = (
                wb[summary_sheet_name]
                if summary_sheet_name in wb.sheetnames else None
            )
            if sheet is None:
                for sn in wb.sheetnames:
                    if 'Management Cijfers' in sn:
                        sheet = wb[sn]
                        break
            if sheet is None:
                return None

            layout = scan_summary_column_layout(sheet, header_row)
            ltm_col = layout.ltm_column
            quarter_col = layout.last_quarter_column

            # Try to align with the active quarter from the sheet name
            # (``Management Cijfers - Q1 2026`` -> ``Q1 2026``) when the
            # sheet has multiple quarter headers.
            if sheet.title and ' - ' in sheet.title:
                quarter_label = sheet.title.split(' - ', 1)[1].strip()
                aligned = find_column_by_header_label(
                    sheet, quarter_label, header_row
                )
                if aligned:
                    quarter_col = aligned
                    ltm_col = find_ltm_column_near_quarter(
                        sheet, quarter_col, header_row
                    ) or ltm_col

            target_col = ltm_col if column_scope == 'ltm' else quarter_col
            if not target_col:
                return None
        finally:
            wb.close()

        result = evaluate_ma_column(
            workbook_path,
            summary_sheet_name=summary_sheet_name,
            bdo_sheet_name=bdo_sheet_name,
            target_col=target_col,
            rows_of_interest=rows_of_interest,
        )
        result['column'] = target_col
        result['scope'] = column_scope
        return result

    @staticmethod
    def evaluate_ma_columns_from_formulas(
        workbook_path: str,
        summary_sheet_name: str,
        bdo_sheet_name: str,
        rules: Optional[dict] = None,
    ) -> Dict[str, Optional[Dict[str, Any]]]:
        """
        Dual-column convenience wrapper around
        :func:`formula_mirror.evaluate_ma_columns`. Evaluates BOTH the
        quarter and LTM columns in a single workbook session and returns
        ``{"quarter": <result>, "ltm": <result>}`` (each value is shaped
        like :func:`evaluate_ma_column_from_formulas` returns, or ``None``
        if the column cannot be resolved).

        This is the I/O-consolidated path used by the AI verification
        pre-flight + post-flight checks: opening the workbook once and
        evaluating both columns shaves the duplicate ``load_workbook``
        cost off every verification round.
        """
        from .formula_mirror import evaluate_ma_columns

        if rules is None:
            rules_path = Path("config/accounting_rules.yaml")
            if rules_path.exists():
                try:
                    with open(rules_path, 'r', encoding='utf-8') as f:
                        rules = yaml.safe_load(f) or {}
                except Exception:
                    rules = {}
            else:
                rules = {}

        layout_cfg = rules.get('layout', {})
        bs_range = layout_cfg.get('balance_sheet_rows', [3, 19])
        pl_range = layout_cfg.get('profit_loss_rows', [23, 68])
        bank_range = layout_cfg.get('bank_rows', [107, 114])
        header_row = layout_cfg.get('header_row', 22)

        rows_of_interest = (
            list(range(bs_range[0], bs_range[1] + 1))
            + list(range(pl_range[0], pl_range[1] + 1))
            + list(range(bank_range[0], bank_range[1] + 1))
        )

        wb = openpyxl.load_workbook(workbook_path, data_only=False)
        try:
            sheet = (
                wb[summary_sheet_name]
                if summary_sheet_name in wb.sheetnames else None
            )
            if sheet is None:
                for sn in wb.sheetnames:
                    if 'Management Cijfers' in sn:
                        sheet = wb[sn]
                        break
            if sheet is None:
                return {'quarter': None, 'ltm': None}

            layout = scan_summary_column_layout(sheet, header_row)
            ltm_col = layout.ltm_column
            quarter_col = layout.last_quarter_column

            if sheet.title and ' - ' in sheet.title:
                quarter_label = sheet.title.split(' - ', 1)[1].strip()
                aligned = find_column_by_header_label(
                    sheet, quarter_label, header_row
                )
                if aligned:
                    quarter_col = aligned
                    ltm_col = find_ltm_column_near_quarter(
                        sheet, quarter_col, header_row
                    ) or ltm_col
        finally:
            wb.close()

        columns: List[Tuple[str, int]] = []
        if quarter_col:
            columns.append(('quarter', quarter_col))
        if ltm_col:
            columns.append(('ltm', ltm_col))

        if not columns:
            return {'quarter': None, 'ltm': None}

        evaluated = evaluate_ma_columns(
            workbook_path,
            summary_sheet_name=summary_sheet_name,
            bdo_sheet_name=bdo_sheet_name,
            columns=columns,
            rows_of_interest=rows_of_interest,
        )

        out: Dict[str, Optional[Dict[str, Any]]] = {
            'quarter': None,
            'ltm': None,
        }
        if quarter_col and 'quarter' in evaluated:
            out['quarter'] = {
                **evaluated['quarter'],
                'column': quarter_col,
                'scope': 'quarter',
            }
        if ltm_col and 'ltm' in evaluated:
            out['ltm'] = {
                **evaluated['ltm'],
                'column': ltm_col,
                'scope': 'ltm',
            }
        return out

    def _add_validation_warning_to_sheet(self, messages: list):
        """Write visible red warning rows in Management Cijfers if validation fails."""
        summary_sheets = [name for name in self.workbook.sheetnames
                          if 'Management Cijfers' in name]
        if not summary_sheets:
            return
        
        sheet = self.workbook[summary_sheets[-1]]
        layout_cfg = self._rules.get('layout', {})
        style_cfg = self._rules.get('styling', {})
        warning_row = layout_cfg.get('warning_start_row', 70)
        
        err_font_color = style_cfg.get('error_font_color', 'FFFF0000')
        err_fill_color = style_cfg.get('error_fill_color', 'FFFFCCCC')
        red_font = Font(name='Calibri', size=11, bold=True, color=err_font_color)
        red_fill = PatternFill(start_color=err_fill_color, end_color=err_fill_color, fill_type='solid')
        
        header_cell = sheet.cell(row=warning_row, column=2)
        header_cell.value = "VALIDATION WARNING: Reconciliation check failed — see details below"
        header_cell.font = red_font
        header_cell.fill = red_fill
        
        for i, msg in enumerate(messages):
            detail_cell = sheet.cell(row=warning_row + 1 + i, column=2)
            detail_cell.value = f"  - {msg}"
            detail_cell.font = Font(name='Calibri', size=10, color=err_font_color)
        
        logger.info(f"Added validation warnings in Management Cijfers rows {warning_row}-{warning_row + len(messages)}")

