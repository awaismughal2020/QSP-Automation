"""
Compliance Certificate Builder

Builds the Compliance Certificate Excel file by:
1. Copying previous quarter structure
2. Updating Q Management Accounts sheet with new data
3. Adding new quarterly column to Suppl. Calc and Impact Unit Sales
4. Updating formulas in SFA CC sheet
"""

from dataclasses import dataclass
from typing import Dict, Optional
from pathlib import Path
import re
import yaml
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, PatternFill
from copy import copy
from datetime import datetime
from loguru import logger

from ..parsers.bdo_parser import BDOParseResult


@dataclass
class ComplianceConfig:
    """Configuration for Compliance Certificate generation."""
    year: int
    quarter: int
    
    @property
    def quarter_str(self) -> str:
        return f"Q{self.quarter} {self.year}"
    
    @property
    def short_quarter(self) -> str:
        """Short format: 25Q3"""
        return f"{str(self.year)[-2:]}Q{self.quarter}"
    
    @property
    def prev_quarter_str(self) -> str:
        prev_q = self.quarter - 1
        prev_y = self.year
        if prev_q == 0:
            prev_q = 4
            prev_y -= 1
        return f"Q{prev_q} {prev_y}"
    
    @property
    def period_end(self):
        """Calculate period end date for the quarter."""
        from datetime import datetime
        # Q1: March 31, Q2: June 30, Q3: September 30, Q4: December 31
        month = self.quarter * 3
        # March=31, June=30, September=30, December=31
        if month in [3, 12]:
            day = 31
        else:  # 6, 9
            day = 30
        return datetime(self.year, month, day)


class ComplianceBuilder:
    """
    Builds updated Compliance Certificate workbook.
    
    The Compliance Certificate file contains:
    - SFA CC: Main compliance certificate with covenant calculations
    - Suppl. Calc: Supplementary calculations by quarter
    - Impact Unit Sales: Unit sales impact calculations
    - Qx Management Accounts: Management accounts data
    
    This builder:
    1. Copies the previous quarter file
    2. Updates Management Accounts sheet with new data
    3. Adds new column to Suppl. Calc and Impact Unit Sales
    4. Updates formulas in SFA CC
    """
    
    def __init__(self, previous_file_path: str, output_path: str, config: ComplianceConfig):
        self.previous_path = Path(previous_file_path)
        self.output_path = Path(output_path)
        self.config = config
        self.workbook = None
        self._computed_values: Optional[dict] = None
        
    def _get_previous_period_end(self):
        """
        Calculate the PREVIOUS quarter's period end date.
        
        For a Q3 2025 report, returns June 30, 2025 (end of Q2).
        For a Q4 2025 report, returns September 30, 2025 (end of Q3).
        """
        prev_q = self.config.quarter - 1
        prev_y = self.config.year
        if prev_q == 0:
            prev_q = 4
            prev_y -= 1
        
        # Calculate period end for previous quarter
        month = prev_q * 3
        # March=31, June=30, September=30, December=31
        if month in [3, 12]:
            day = 31
        else:  # 6, 9
            day = 30
        
        return datetime(prev_y, month, day)
        
    def build(self, bdo_result: BDOParseResult, management_accounts_path: Optional[str] = None,
              computed_values: Optional[dict] = None) -> Path:
        """
        Build new Compliance Certificate file.
        
        Args:
            bdo_result: Parsed BDO data for current quarter
            management_accounts_path: Path to Management Accounts file for data copy
            computed_values: Pre-computed numeric values from MA builder
                             {(row,'quarter'|'ltm'): float}. When provided,
                             columns B and C are populated directly from this dict
                             instead of parsing MA file formulas.
            
        Returns:
            Path to generated file
        """
        logger.info(f"Building Compliance Certificate for {self.config.quarter_str}")
        self._computed_values = computed_values
        
        # Load previous quarter file
        self.workbook = openpyxl.load_workbook(self.previous_path)
        logger.info(f"Loaded previous file with sheets: {self.workbook.sheetnames}")
        
        # Step 1: Update Management Accounts sheet
        self._update_management_accounts_sheet(bdo_result, management_accounts_path)
        
        # Step 2: Add new column to Suppl. Calc
        self._update_suppl_calc_sheet()
        
        # Step 3: Add new column to Impact Unit Sales
        self._update_impact_unit_sales_sheet()
        
        # Step 4: Update formulas in Suppl. Calc to reference Q3 Management Accounts
        self._update_suppl_calc_formulas()
        
        # Step 5: Update SFA CC formulas (if needed)
        self._update_sfa_cc_sheet()
        
        # Step 6: Copy actual values from Management Cijfers to Suppl. Calc
        if management_accounts_path:
            self._copy_actual_values_to_suppl_calc(management_accounts_path)
        
        # Step 7: Update all dates and text references throughout workbook
        self._update_all_dates_and_text()
        
        # Step 8: Clear manual input cells (must be last, after all other operations)
        self._clear_manual_input_cells()
        
        # Save output
        self.workbook.save(self.output_path)
        logger.info(f"Saved to {self.output_path}")
        
        return self.output_path
    
    def _update_management_accounts_sheet(self, bdo_result: BDOParseResult, 
                                          management_accounts_path: Optional[str]):
        """Update the Management Accounts sheet with new quarter data."""
        # The new sheet name format should match what the formulas expect: "Q3 Management Accounts"
        # NOT "Q3 2025 Management Accounts" as the formulas don't include the year
        new_ma_sheet_name = f"Q{self.config.quarter} Management Accounts"
        
        # Handle different naming patterns
        ma_sheets = [name for name in self.workbook.sheetnames if 'Management Accounts' in name]
        
        if ma_sheets:
            old_sheet_name = ma_sheets[0]  # Use the first one found
            logger.info(f"Found Management Accounts sheet: {old_sheet_name}")
            
            # Get the old sheet structure BEFORE renaming (to preserve it)
            old_sheet = self.workbook[old_sheet_name]
            
            # Create a new sheet with the new name, copying structure from old sheet
            new_sheet = self.workbook.copy_worksheet(old_sheet)
            new_sheet.title = new_ma_sheet_name
            
            # Remove the old sheet
            self.workbook.remove(old_sheet)
            
            logger.info(f"Created new sheet: {new_ma_sheet_name} with structure from {old_sheet_name}")
            
            # Copy data: prefer computed_values cache, fall back to reading the file
            if self._computed_values or management_accounts_path:
                self._copy_ma_data(new_ma_sheet_name, management_accounts_path or "")
        else:
            logger.warning("No Management Accounts sheet found to update")
    
    def _copy_ma_data(self, target_sheet_name: str, source_path: str):
        """
        Copy Management Accounts data from the CURRENT quarter's file.
        
        The Q{n} Management Accounts sheet in the Compliance Certificate
        should contain BOTH the kwartaal (quarterly) AND LTM data:
        - Column B: Kwartaal (quarterly) values
        - Column C: LTM values
        
        When computed_values is available (passed via build()), values are
        copied directly — no formula parsing needed, guaranteeing an exact match.
        """
        if self._computed_values:
            self._copy_ma_data_from_cache(target_sheet_name)
            return

        try:
            source_wb_data = openpyxl.load_workbook(source_path, data_only=True)
            source_wb_formulas = openpyxl.load_workbook(source_path, data_only=False)
            
            target_sheet = self.workbook[target_sheet_name]
            
            cijfers_sheet_data = None
            cijfers_sheet_formulas = None
            bdo_sheet = None
            bdo_sheet_name = None
            
            for name in source_wb_data.sheetnames:
                if 'Management Cijfers' in name:
                    cijfers_sheet_data = source_wb_data[name]
                    cijfers_sheet_formulas = source_wb_formulas[name]
                    logger.info(f"Using Management Cijfers sheet: {name}")
                    break
            
            expected_bdo_name = f"BDO - Q{self.config.quarter}-{str(self.config.year)[-2:]}"
            for name in source_wb_formulas.sheetnames:
                if name == expected_bdo_name:
                    bdo_sheet = source_wb_formulas[name]
                    bdo_sheet_name = name
                    logger.info(f"Found BDO sheet for formula evaluation: {name}")
                    break
            
            if bdo_sheet is None:
                for name in source_wb_formulas.sheetnames:
                    if 'BDO -' in name and f"Q{self.config.quarter}-" in name:
                        bdo_sheet = source_wb_formulas[name]
                        bdo_sheet_name = name
                        logger.info(f"Found BDO sheet (fallback): {name}")
                        break
            
            if cijfers_sheet_data is None:
                logger.warning("Management Cijfers sheet not found in source file")
                source_wb_data.close()
                source_wb_formulas.close()
                return
            
            cijfers_sheet = cijfers_sheet_data
            
            from .management_accounts import (
                find_ltm_column_near_quarter,
                resolve_quarter_column,
                scan_summary_column_layout,
            )

            header_row = 22
            layout = scan_summary_column_layout(cijfers_sheet, header_row)
            ltm_column = layout.ltm_column
            if ltm_column is not None:
                logger.info(
                    f"Found LTM column at {get_column_letter(ltm_column)} "
                    f"({ltm_column}) via row {header_row} layout scan"
                )

            if ltm_column is None:
                for col in range(cijfers_sheet.max_column, 0, -1):
                    cell_val = cijfers_sheet.cell(row=3, column=col).value
                    if cell_val is not None and isinstance(cell_val, (int, float)):
                        ltm_column = col
                        logger.info(
                            f"Found LTM column at {get_column_letter(col)} ({col}) via row 3 value"
                        )
                        break

            if ltm_column is None:
                logger.warning(
                    "Could not find LTM column dynamically, falling back to column 27 (AA)"
                )
                ltm_column = 27

            quarter_header = f"Q{self.config.quarter} {self.config.year}"
            quarter_column = resolve_quarter_column(
                cijfers_sheet, quarter_header, header_row
            )
            if quarter_column is not None:
                logger.info(
                    f"Found quarterly column at {get_column_letter(quarter_column)} "
                    f"({quarter_column}) via header {quarter_header!r}"
                )
            elif ltm_column:
                quarter_column = ltm_column - 1
                logger.info(
                    f"Quarterly column defaulted to {get_column_letter(quarter_column)} "
                    f"({quarter_column})"
                )

            if quarter_column and not layout.ltm_column:
                alt_ltm = find_ltm_column_near_quarter(
                    cijfers_sheet, quarter_column, header_row
                )
                if alt_ltm:
                    ltm_column = alt_ltm
            
            # Row mapping: target row in CC -> source row in Management Cijfers
            target_to_source_mapping = {}
            
            # Balance Sheet section (target rows 2-18 -> source rows 3-19)
            # Row 18 = Total Equity Movement (MA row 19)
            for target_row in range(2, 19):
                target_to_source_mapping[target_row] = target_row + 1
            
            # P&L section (target rows 22-67 -> source rows 23-68)
            for target_row in range(22, 68):
                target_to_source_mapping[target_row] = target_row + 1
            
            # Bank account overview section (target rows 72-78 -> source rows 107-113)
            for target_row in range(72, 79):
                target_to_source_mapping[target_row] = target_row + 35
            
            # Update headers
            period_end = self.config.period_end
            target_sheet.cell(row=1, column=3).value = period_end
            target_sheet.cell(row=1, column=3).number_format = 'YYYY-MM-DD'
            logger.info(f"Set header date to current quarter end: {period_end}")
            
            # Column B header: kwartaal title
            target_sheet.cell(row=21, column=2).value = self.config.quarter_str
            # Column C header: LTM title
            ltm_title = f"LTM {self.config.quarter_str}"
            target_sheet.cell(row=21, column=3).value = ltm_title
            logger.info(f"Set P&L headers: col B = '{self.config.quarter_str}', col C = '{ltm_title}'")
            
            bank_header = f"Per {period_end.day}-{period_end.month}-{period_end.year}"
            target_sheet.cell(row=71, column=3).value = bank_header
            logger.info(f"Set bank account header (row 71) to: {bank_header}")
            
            # --- Evaluate LTM values ---
            evaluated_ltm = {}
            
            def evaluate_ltm_cell(source_row: int) -> float:
                if source_row in evaluated_ltm:
                    return evaluated_ltm[source_row]
                
                value = cijfers_sheet_data.cell(row=source_row, column=ltm_column).value
                if value is not None and isinstance(value, (int, float)):
                    evaluated_ltm[source_row] = float(value)
                    return float(value)
                
                formula = cijfers_sheet_formulas.cell(row=source_row, column=ltm_column).value
                
                if formula is None:
                    evaluated_ltm[source_row] = None
                    return None
                
                if isinstance(formula, (int, float)):
                    evaluated_ltm[source_row] = float(formula)
                    return float(formula)
                
                if isinstance(formula, str) and not formula.startswith('='):
                    evaluated_ltm[source_row] = formula
                    return None
                
                result = self._evaluate_formula_with_internal_refs(
                    formula, bdo_sheet, bdo_sheet_name, ltm_column,
                    cijfers_sheet_formulas, evaluate_ltm_cell
                )
                evaluated_ltm[source_row] = result
                return result
            
            # --- Evaluate quarterly values ---
            evaluated_quarter = {}
            
            def evaluate_quarter_cell(source_row: int) -> float:
                if source_row in evaluated_quarter:
                    return evaluated_quarter[source_row]
                
                value = cijfers_sheet_data.cell(row=source_row, column=quarter_column).value
                if value is not None and isinstance(value, (int, float)):
                    evaluated_quarter[source_row] = float(value)
                    return float(value)
                
                formula = cijfers_sheet_formulas.cell(row=source_row, column=quarter_column).value
                
                if formula is None:
                    evaluated_quarter[source_row] = None
                    return None
                
                if isinstance(formula, (int, float)):
                    evaluated_quarter[source_row] = float(formula)
                    return float(formula)
                
                if isinstance(formula, str) and not formula.startswith('='):
                    evaluated_quarter[source_row] = formula
                    return None
                
                result = self._evaluate_formula_with_internal_refs(
                    formula, bdo_sheet, bdo_sheet_name, quarter_column,
                    cijfers_sheet_formulas, evaluate_quarter_cell
                )
                evaluated_quarter[source_row] = result
                return result
            
            # Evaluate all source rows for both columns
            for target_row, source_row in target_to_source_mapping.items():
                evaluate_ltm_cell(source_row)
                evaluate_quarter_cell(source_row)
            
            # Copy LTM values to column C (existing behavior)
            ltm_copied = 0
            for target_row, source_row in target_to_source_mapping.items():
                value = evaluated_ltm.get(source_row)
                if value is not None:
                    target_sheet.cell(row=target_row, column=3).value = value
                    ltm_copied += 1
            
            # Copy quarterly values to column B (new)
            quarter_copied = 0
            for target_row, source_row in target_to_source_mapping.items():
                value = evaluated_quarter.get(source_row)
                if value is not None:
                    target_sheet.cell(row=target_row, column=2).value = value
                    quarter_copied += 1
            
            logger.info(f"Copied to {target_sheet_name}: {quarter_copied} kwartaal values (col B), {ltm_copied} LTM values (col C)")
            source_wb_data.close()
            source_wb_formulas.close()
            
        except Exception as e:
            logger.error(f"Error copying MA data: {e}")
            import traceback
            traceback.print_exc()
    
    @staticmethod
    def _load_accounting_rules() -> dict:
        rules_path = Path("config/accounting_rules.yaml")
        if rules_path.exists():
            try:
                with open(rules_path, 'r', encoding='utf-8') as f:
                    return yaml.safe_load(f) or {}
            except Exception:
                pass
        return {}

    def _copy_ma_data_from_cache(self, target_sheet_name: str):
        """
        Fast path: copy values from computed_values dict directly into CC columns B/C.
        No formula parsing — guarantees exact match with the MA builder output.

        Row mapping:
          CC  2-17  → MA  3-18   (BS individual rows, LTM only — no quarter)
          CC 18     → MA 19      (Total Equity Movement, LTM only)
          CC 22-67  → MA 23-68   (P&L rows, quarter col B + LTM col C)
          CC 72-78  → MA 107-113 (Bank rows, LTM only)
          CC 79     → MA 114     (Bank Total, LTM only)
        """
        target_sheet = self.workbook[target_sheet_name]
        cv = self._computed_values
        period_end = self.config.period_end

        # Accounting number format: negatives in parentheses
        acct_fmt = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        # ── BS rows: LTM only (CC 2-18 → MA 3-19) ──
        bs_mapping = {}
        for t in range(2, 19):               # CC rows 2-18 inclusive
            bs_mapping[t] = t + 1            # MA rows 3-19 inclusive

        # ── P&L rows: quarter + LTM (CC 22-67 → MA 23-68) ──
        pl_mapping = {}
        for t in range(22, 68):
            pl_mapping[t] = t + 1

        # ── Bank rows: LTM only (CC 72-78 → MA 107-113) ──
        bank_mapping = {}
        for t in range(72, 79):
            bank_mapping[t] = t + 35

        # ── Headers ──
        target_sheet.cell(row=1, column=3).value = period_end
        target_sheet.cell(row=1, column=3).number_format = 'YYYY-MM-DD'
        target_sheet.cell(row=21, column=2).value = self.config.quarter_str
        target_sheet.cell(row=21, column=3).value = f"LTM {self.config.quarter_str}"
        target_sheet.cell(row=71, column=3).value = (
            f"Per {period_end.day}-{period_end.month}-{period_end.year}"
        )

        q_copied = ltm_copied = 0

        # ── BS: column C only (no quarter column for Balance Sheet) ──
        for cc_row, src_row in bs_mapping.items():
            ltm_val = cv.get((src_row, 'ltm'))
            if ltm_val is not None:
                cell = target_sheet.cell(row=cc_row, column=3)
                cell.value = ltm_val
                cell.number_format = acct_fmt
                cell.font = Font(name='Calibri', size=11, color='FF000000')
                ltm_copied += 1

        # ── P&L: column B (quarter) + column C (LTM) ──
        for cc_row, src_row in pl_mapping.items():
            q_val = cv.get((src_row, 'quarter'))
            ltm_val = cv.get((src_row, 'ltm'))

            if q_val is not None:
                cell = target_sheet.cell(row=cc_row, column=2)
                cell.value = q_val
                cell.number_format = acct_fmt
                cell.font = Font(name='Calibri', size=11, color='FF000000')
                q_copied += 1
            if ltm_val is not None:
                cell = target_sheet.cell(row=cc_row, column=3)
                cell.value = ltm_val
                cell.number_format = acct_fmt
                cell.font = Font(name='Calibri', size=11, color='FF000000')
                ltm_copied += 1

        # ── Bank: column C only ──
        for cc_row, src_row in bank_mapping.items():
            ltm_val = cv.get((src_row, 'ltm'))
            if ltm_val is not None:
                cell = target_sheet.cell(row=cc_row, column=3)
                cell.value = ltm_val
                cell.number_format = acct_fmt
                cell.font = Font(name='Calibri', size=11, color='FF000000')
                ltm_copied += 1

        # Bank Total: CC row 79 → MA row 114
        bank_total = cv.get((114, 'ltm'))
        if bank_total is not None:
            cell = target_sheet.cell(row=79, column=3)
            cell.value = bank_total
            cell.number_format = acct_fmt
            cell.font = Font(name='Calibri', size=11, color='FF000000')
            ltm_copied += 1

        # Clear residual template values in the gap between P&L and Bank
        # (CC rows 68-70 are not mapped and may contain stale data)
        for gap_row in range(68, 71):
            for gap_col in (2, 3):
                target_sheet.cell(row=gap_row, column=gap_col).value = None

        # Bold the key total rows (label in col A + values in B/C).
        # CC rows are offset by -1 from MA rows.
        rules = self._load_accounting_rules()
        identity = rules.get('identity_alignment', {})
        bold_ma_rows = identity.get(
            'formula_total_rows',
            [identity.get('authoritative_row', 19),
             identity.get('dependent_row', 68)]
        )
        for ma_row in bold_ma_rows:
            cc_row = ma_row - 1
            for col in range(1, 4):  # columns A, B, C
                cell = target_sheet.cell(row=cc_row, column=col)
                base = copy(cell.font) if cell.font else Font()
                cell.font = Font(
                    name=base.name or 'Calibri', size=base.size or 11,
                    bold=True, italic=base.italic, underline=base.underline,
                    color=base.color or 'FF000000',
                )

        logger.info(
            f"[CC] Copied from computed_values to {target_sheet_name}: "
            f"{q_copied} quarter (col B), {ltm_copied} LTM (col C)"
        )

    def _evaluate_formula_with_internal_refs(self, formula: str, bdo_sheet, bdo_sheet_name: str,
                                               ltm_column: int, cijfers_sheet, evaluate_cell_func):
        """
        Evaluate a formula that may contain:
        - BDO sheet references (='BDO - Q3-25'!H16)
        - Internal column references (=SUM(AB23:AB24))
        - Negation formulas (=-'BDO - Q3-25'!H76)
        - Division formulas (=AB25/AB23-1)
        - Column range sums (=SUM(X50:AA50))
        - Complex internal formulas (=-(AB50-AB53))
        
        Args:
            formula: The formula string to evaluate
            bdo_sheet: The BDO worksheet
            bdo_sheet_name: Name of the BDO sheet
            ltm_column: The LTM column index in Management Cijfers
            cijfers_sheet: The Management Cijfers worksheet (formulas version)
            evaluate_cell_func: Function to recursively evaluate cells
        """
        import re
        from openpyxl.utils import column_index_from_string, get_column_letter
        
        if not formula or not isinstance(formula, str) or not formula.startswith('='):
            return None
        
        ltm_col_letter = get_column_letter(ltm_column)
        
        try:
            formula_body = formula[1:]  # Remove leading '='
            
            # Handle division formulas like =AB25/AB23-1
            division_pattern = rf"{ltm_col_letter}(\d+)/{ltm_col_letter}(\d+)(-?\d*)"
            division_match = re.match(division_pattern, formula_body)
            if division_match:
                numerator_row = int(division_match.group(1))
                denominator_row = int(division_match.group(2))
                offset_str = division_match.group(3)
                
                num_val = evaluate_cell_func(numerator_row)
                denom_val = evaluate_cell_func(denominator_row)
                
                if num_val is not None and denom_val is not None and denom_val != 0:
                    result = float(num_val) / float(denom_val)
                    if offset_str:
                        result += float(offset_str)
                    return result
                return None
            
            # Handle SUM with different column range: =SUM(X50:AA50)
            # But NOT when it's the same column as LTM (that should use internal_sum_pattern)
            cross_col_sum_pattern = r"SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)"
            cross_col_match = re.match(cross_col_sum_pattern, formula_body)
            if cross_col_match:
                start_col = cross_col_match.group(1)
                start_row = int(cross_col_match.group(2))
                end_col = cross_col_match.group(3)
                end_row = int(cross_col_match.group(4))
                
                # If both columns are the LTM column, skip to internal_sum_pattern handling
                if start_col == ltm_col_letter and end_col == ltm_col_letter:
                    pass  # Skip - will be handled by internal_sum_pattern below
                else:
                    # Sum across the column range (different columns)
                    total = 0.0
                    start_col_idx = column_index_from_string(start_col)
                    end_col_idx = column_index_from_string(end_col)
                    
                    for col in range(start_col_idx, end_col_idx + 1):
                        for row in range(start_row, end_row + 1):
                            cell_val = cijfers_sheet.cell(row=row, column=col).value
                            if isinstance(cell_val, (int, float)):
                                total += float(cell_val)
                            elif isinstance(cell_val, str) and not cell_val.startswith('='):
                                try:
                                    total += float(cell_val)
                                except ValueError:
                                    pass
                    return total  # Return 0 if formula evaluates to 0
            
            # Handle complex parenthesized formulas: =-(AB50-AB53)
            paren_pattern = rf"-\({ltm_col_letter}(\d+)-{ltm_col_letter}(\d+)\)"
            paren_match = re.match(paren_pattern, formula_body)
            if paren_match:
                row1 = int(paren_match.group(1))
                row2 = int(paren_match.group(2))
                val1 = evaluate_cell_func(row1)
                val2 = evaluate_cell_func(row2)
                if val1 is not None and val2 is not None:
                    return -(float(val1) - float(val2))
                return None
            
            # Check if this is primarily a BDO reference formula
            # (contains 'BDO' and not just internal column references)
            if 'BDO' in formula_body:
                # Delegate to BDO reference evaluation
                bdo_value = self._evaluate_bdo_reference(formula, bdo_sheet, bdo_sheet_name)
                if bdo_value is not None:
                    return bdo_value
            
            # Check if it starts with negation (=-...)
            negate_result = False
            if formula_body.startswith('-'):
                negate_result = True
                formula_body = formula_body[1:]
            
            # Pattern for internal SUM references: SUM(AB23:AB24)
            internal_sum_pattern = rf"SUM\({ltm_col_letter}(\d+):{ltm_col_letter}(\d+)\)"
            
            # Check for internal SUM pattern
            internal_sum_match = re.search(internal_sum_pattern, formula_body)
            if internal_sum_match:
                start_row = int(internal_sum_match.group(1))
                end_row = int(internal_sum_match.group(2))
                total = 0.0
                for row in range(start_row, end_row + 1):
                    cell_val = evaluate_cell_func(row)
                    if cell_val is not None and isinstance(cell_val, (int, float)):
                        total += float(cell_val)
                
                # Process remaining terms after the SUM
                remaining = formula_body[internal_sum_match.end():]
                total += self._process_additional_terms(remaining, bdo_sheet, ltm_col_letter, evaluate_cell_func)
                
                return -total if negate_result else total
            
            # Pattern for addition of internal cell references: AB57+AB65 or AB47+AB46+AB45
            internal_add_pattern = rf"^{ltm_col_letter}(\d+)(?:\+{ltm_col_letter}(\d+))+$"
            if re.match(internal_add_pattern, formula_body):
                total = 0.0
                for row_str in re.findall(rf"{ltm_col_letter}(\d+)", formula_body):
                    row_num = int(row_str)
                    cell_val = evaluate_cell_func(row_num)
                    if cell_val is not None and isinstance(cell_val, (int, float)):
                        total += float(cell_val)
                return -total if negate_result else total
            
            # Pattern for single internal cell reference
            internal_ref_pattern = rf"^{ltm_col_letter}(\d+)$"
            single_ref_match = re.match(internal_ref_pattern, formula_body)
            if single_ref_match:
                row_num = int(single_ref_match.group(1))
                cell_val = evaluate_cell_func(row_num)
                if cell_val is not None:
                    return -float(cell_val) if negate_result else float(cell_val)
            
            return None
            
        except Exception as e:
            logger.debug(f"Could not evaluate formula with internal refs '{formula}': {e}")
            return None
    
    def _process_additional_terms(self, remaining: str, bdo_sheet, ltm_col_letter: str, evaluate_cell_func):
        """Process additional terms like +AB45 or -'BDO'!H76."""
        import re
        
        total = 0.0
        
        # Process internal column references: +AB45 or -AB45
        internal_pattern = rf"([+-]){ltm_col_letter}(\d+)"
        for match in re.finditer(internal_pattern, remaining):
            sign = 1 if match.group(1) == '+' else -1
            row_num = int(match.group(2))
            cell_val = evaluate_cell_func(row_num)
            if cell_val is not None and isinstance(cell_val, (int, float)):
                total += sign * float(cell_val)
        
        return total
    
    def _evaluate_bdo_reference(self, formula: str, bdo_sheet, bdo_sheet_name: str):
        """
        Evaluate a formula that references the BDO sheet.
        
        Handles patterns like:
        - ='BDO - Q3-25'!H16 (single cell reference)
        - =SUM('BDO - Q3-25'!H6:H10) (SUM range)
        - ='BDO - Q3-25'!H13+'BDO - Q3-25'!H14+'BDO - Q3-25'!H15 (additions)
        - =SUM('BDO - Q3-25'!H67:H72)+'BDO - Q3-25'!H65 (combined)
        
        Also handles nested formulas where BDO column H contains SUM(C:G) formulas.
        """
        import re
        from openpyxl.utils import column_index_from_string
        
        if not formula or not isinstance(formula, str):
            return None
        
        def evaluate_simple_expr(expr: str) -> float:
            """Evaluate a simple arithmetic expression like =-110650-5."""
            try:
                # Remove the leading = if present
                if expr.startswith('='):
                    expr = expr[1:]
                # Only evaluate if it's a simple arithmetic expression (numbers and +/-)
                if re.match(r'^[\d\.\-\+\s]+$', expr):
                    return float(eval(expr))
            except Exception:
                pass
            return 0.0
        
        def get_raw_cell_value(row: int, col: int, depth: int = 0) -> float:
            """Get numeric value from a cell, evaluating nested formulas if needed."""
            if depth > 10:  # Prevent infinite recursion
                return 0.0
            
            val = bdo_sheet.cell(row=row, column=col).value
            if val is None:
                return 0.0
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str) and val.startswith('='):
                # Handle cell reference formulas like =D130
                cell_ref = re.match(r'^=([A-Z]+)(\d+)$', val)
                if cell_ref:
                    ref_col = column_index_from_string(cell_ref.group(1))
                    ref_row = int(cell_ref.group(2))
                    return get_raw_cell_value(ref_row, ref_col, depth + 1)
                
                # Try to evaluate SUM formulas
                inner_sum = re.match(r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', val)
                if inner_sum:
                    start_col = column_index_from_string(inner_sum.group(1))
                    start_row = int(inner_sum.group(2))
                    end_col = column_index_from_string(inner_sum.group(3))
                    end_row = int(inner_sum.group(4))
                    total = 0.0
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            total += get_raw_cell_value(r, c, depth + 1)
                    return total
                
                # Try to evaluate simple expressions like =-110650-5
                return evaluate_simple_expr(val)
            return 0.0
        
        def evaluate_inner_sum(inner_formula: str) -> float:
            """Evaluate a SUM formula within the BDO sheet like =SUM(C6:G6)."""
            inner_sum = re.match(r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', inner_formula)
            if inner_sum:
                start_col = column_index_from_string(inner_sum.group(1))
                start_row = int(inner_sum.group(2))
                end_col = column_index_from_string(inner_sum.group(3))
                end_row = int(inner_sum.group(4))
                
                total = 0.0
                for r in range(start_row, end_row + 1):
                    for c in range(start_col, end_col + 1):
                        total += get_raw_cell_value(r, c, 0)  # Pass depth parameter
                return total
            return 0.0
        
        def get_cell_value(col_letter: str, row_num: int) -> float:
            """Get numeric value from BDO sheet cell, evaluating nested formulas if needed."""
            try:
                col_idx = column_index_from_string(col_letter)
                value = bdo_sheet.cell(row=row_num, column=col_idx).value
                
                # If None, return 0
                if value is None:
                    return 0.0
                
                # If it's a number, return it directly
                if isinstance(value, (int, float)):
                    return float(value)
                
                # If it's a formula, try to evaluate it
                if isinstance(value, str) and value.startswith('='):
                    # First try as a SUM formula
                    result = evaluate_inner_sum(value)
                    if result != 0.0:
                        return result
                    # Then try as a simple expression
                    return evaluate_simple_expr(value)
                    
            except Exception:
                pass
            return 0.0
        
        def evaluate_sum_range(col_letter: str, start_row: int, end_row: int) -> float:
            """Evaluate a SUM range like H6:H10."""
            total = 0.0
            for row in range(start_row, end_row + 1):
                total += get_cell_value(col_letter, row)
            return total
        
        try:
            total = 0.0
            remaining = formula
            
            # Track if we found any matches (SUM or cell references)
            found_matches = False
            
            # Process all SUM(...) parts first
            sum_pattern = r"SUM\('[^']+'\!([A-Z]+)(\d+):([A-Z]+)(\d+)\)"
            for match in re.finditer(sum_pattern, remaining):
                found_matches = True
                col_letter = match.group(1)
                start_row = int(match.group(2))
                end_col = match.group(3)
                end_row = int(match.group(4))
                
                # Assuming same column for range (H6:H10, not H6:I10)
                if col_letter == end_col:
                    total += evaluate_sum_range(col_letter, start_row, end_row)
            
            # Remove processed SUM parts from remaining
            remaining = re.sub(sum_pattern, '', remaining)
            
            # Process all single cell references with + or - signs
            # For formula =-'BDO'!H76-'BDO'!H77, we parse each term with its sign:
            # - First term after '=' has sign determined by what follows '='
            # - Subsequent terms have explicit + or - signs
            
            # Pattern to match: =-'BDO'!H76 or ='BDO'!H76 or +/-'BDO'!H76
            # We need to handle the = at the start specially
            formula_body = remaining[1:] if remaining.startswith('=') else remaining
            
            # Pattern for all cell references with their signs
            # This matches: -'BDO...'!H76 or +'BDO...'!H76 or 'BDO...'!H76 (implicit +)
            cell_pattern = r"([+-]?)'[^']+'\!([A-Z]+)(\d+)"
            
            for match in re.finditer(cell_pattern, formula_body):
                found_matches = True
                sign_str = match.group(1)
                col_letter = match.group(2)
                row_num = int(match.group(3))
                
                # Determine sign: '-' means negative, '+' or '' means positive
                sign = -1 if sign_str == '-' else 1
                total += sign * get_cell_value(col_letter, row_num)
            
            # Return total if we found matches (even if 0), None if no matches
            return total if found_matches else None
            
        except Exception as e:
            logger.debug(f"Could not evaluate formula '{formula}': {e}")
            return None
    
    def _evaluate_simple_formula(self, formula: str, sheet, current_row: int) -> float:
        """Evaluate a simple SUM formula within the same sheet."""
        if 'SUM' in formula.upper():
            try:
                # Match SUM(Col1Row1:Col2Row2) pattern
                match = re.search(r'SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', formula)
                if match:
                    total = 0
                    start_col = ord(match.group(1)) - ord('A') + 1
                    end_col = ord(match.group(3)) - ord('A') + 1
                    start_row = int(match.group(2))
                    end_row = int(match.group(4))
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            v = sheet.cell(row=r, column=c).value
                            if isinstance(v, (int, float)):
                                total += v
                    return total
            except Exception:
                pass
        return 0
    
    def _evaluate_bdo_formula(self, formula: str, bdo_h_values: dict, bdo_sheet_name: str) -> float:
        """
        Evaluate a formula that references the BDO sheet.
        
        Handles patterns like:
        - ='BDO - Q3-25'!H16
        - =SUM('BDO - Q3-25'!H6:H10)
        - ='BDO - Q3-25'!H13+'BDO - Q3-25'!H14+'BDO - Q3-25'!H15
        """
        total = 0.0
        
        try:
            # Pattern 1: SUM formula =SUM('BDO - Q3-25'!H6:H10)
            sum_match = re.search(r"SUM\('.*?'!H(\d+):H(\d+)\)", formula)
            if sum_match:
                start_row = int(sum_match.group(1))
                end_row = int(sum_match.group(2))
                for r in range(start_row, end_row + 1):
                    total += bdo_h_values.get(r, 0)
                
                # There might be additional terms after the SUM, e.g., +...+...
                remaining = formula[sum_match.end():]
                additional = self._parse_additional_terms(remaining, bdo_h_values)
                total += additional
                return total
            
            # Pattern 2: Multiple cell references added together
            # ='BDO - Q3-25'!H13+'BDO - Q3-25'!H14+...
            cell_refs = re.findall(r"'.*?'!H(\d+)", formula)
            if cell_refs:
                for row_str in cell_refs:
                    row_num = int(row_str)
                    total += bdo_h_values.get(row_num, 0)
                return total
            
            # Pattern 3: Single cell reference without quotes
            simple_match = re.search(r"H(\d+)", formula)
            if simple_match:
                row_num = int(simple_match.group(1))
                return bdo_h_values.get(row_num, 0)
                
        except Exception as e:
            logger.warning(f"Error evaluating formula '{formula}': {e}")
        
        return None
    
    def _parse_additional_terms(self, remaining: str, bdo_h_values: dict) -> float:
        """Parse additional terms like +'BDO - Q3-25'!H65"""
        total = 0.0
        cell_refs = re.findall(r"\+'.*?'!H(\d+)", remaining)
        for row_str in cell_refs:
            row_num = int(row_str)
            total += bdo_h_values.get(row_num, 0)
        return total
    
    def _copy_ma_data_fallback(self, target_sheet, bdo_sheet, bdo_h_values: dict) -> int:
        """Fallback method using direct BDO row mappings."""
        # Direct mapping from target row to BDO rows (based on reference formulas)
        row_to_bdo_rows = {
            2: [16],                    # Deferred Tax Asset: H16
            3: list(range(6, 11)),      # Real estate: H6:H10
            4: [13, 14, 15],            # Financial fixed assets: H13+H14+H15
            5: [19, 21, 22],            # Accounts receivable: H19+H21+H22
            6: list(range(67, 73)) + [65, 61, 62],  # Service costs
            7: list(range(23, 26)) + [30, 26],      # Prepaid expenses
            8: list(range(31, 38)),     # Cash: H31:H37
            9: [40, 41, 42],            # Equity: H40:H42
            10: [20],                   # AC Shareholder: H20
            11: [45, 46],               # Bank loan: H45+H46
            12: [47],                   # Amortised fee: H47
            13: [50, 60, 63, 64, 29],   # Accounts payable
            14: list(range(51, 57)),    # Current account: H51:H56
            15: [57],                   # VAT payable: H57
            16: [59],                   # Deposits: H59
            17: [58],                   # Rent Invoiced in advance: H58
        }
        
        items_copied = 0
        for target_row, bdo_rows in row_to_bdo_rows.items():
            total = sum(bdo_h_values.get(r, 0) for r in bdo_rows)
            target_sheet.cell(row=target_row, column=3).value = total
            items_copied += 1
        
        return items_copied
    
    @staticmethod
    def _parse_quarter_label(label: str):
        """Parse a quarter label like '26Q3' into (year_short, quarter_num)."""
        label = str(label).strip()
        if 'Q' not in label:
            return None
        parts = label.split('Q')
        try:
            return int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            return None

    @staticmethod
    def _next_quarter_label(yr: int, q: int) -> str:
        """Advance one quarter and return label like '26Q4'."""
        q += 1
        if q > 4:
            q, yr = 1, yr + 1
        return f"{yr}Q{q}"

    def _update_suppl_calc_sheet(self):
        """
        Update Suppl. Calc sheet by inserting all missing forecast quarters
        up to and including next_forecast_quarter (current_quarter + 4).

        For Q3 2025 the target label is 26Q3 (one insert needed).
        For Q4 2025 the target label is 26Q4 — if 26Q3 is also missing it
        is inserted first so NTM covers four consecutive forecast quarters.
        """
        if 'Suppl. Calc' not in self.workbook.sheetnames:
            logger.warning("Suppl. Calc sheet not found")
            return

        sheet = self.workbook['Suppl. Calc']

        ntm_col = None
        for col in range(1, sheet.max_column + 5):
            cell = sheet.cell(row=2, column=col)
            if cell.value and str(cell.value).strip().upper() == 'NTM':
                ntm_col = col
                break

        if ntm_col is None:
            logger.warning("NTM column not found in Suppl. Calc")
            return

        logger.info(f"Found NTM column at {get_column_letter(ntm_col)} ({ntm_col})")

        target_label = f"{str(self.config.year + 1)[-2:]}Q{self.config.quarter}"

        # Find the last existing forecast column (the one just before NTM)
        last_fc_hdr = sheet.cell(row=2, column=ntm_col - 1).value
        last_parsed = self._parse_quarter_label(last_fc_hdr) if last_fc_hdr else None
        target_parsed = self._parse_quarter_label(target_label)

        if target_parsed is None:
            logger.warning(f"Cannot parse target quarter label: {target_label}")
            return

        # Build the list of quarters to insert (everything after last existing
        # forecast up to and including the target).
        to_insert: list[str] = []
        if last_parsed is not None:
            yr, q = last_parsed
            while True:
                nxt = self._next_quarter_label(yr, q)
                to_insert.append(nxt)
                if nxt == target_label:
                    break
                p = self._parse_quarter_label(nxt)
                if p is None:
                    break
                yr, q = p
                if len(to_insert) > 8:
                    break
        else:
            to_insert = [target_label]

        # Remove any labels that already exist
        existing = set()
        for col in range(1, sheet.max_column + 1):
            hdr = sheet.cell(row=2, column=col).value
            if hdr:
                existing.add(str(hdr).strip())
        to_insert = [lbl for lbl in to_insert if lbl not in existing]

        if not to_insert:
            logger.info(f"All forecast quarters up to {target_label} already exist")
            return

        # Insert each missing quarter at the current NTM position
        inserted = 0
        for label in to_insert:
            logger.info(f"Inserting column for {label} at position {ntm_col + inserted} (NTM column)")
            ins_pos = ntm_col + inserted
            sheet.insert_cols(ins_pos)
            sheet.cell(row=2, column=ins_pos).value = label
            sheet.cell(row=2, column=ins_pos).font = Font(bold=True)
            sheet.cell(row=2, column=ins_pos).alignment = Alignment(horizontal='center')
            sheet.cell(row=3, column=ins_pos).value = "Forecast"
            sheet.cell(row=3, column=ins_pos).alignment = Alignment(horizontal='center')
            self._copy_column_formatting(sheet, ins_pos - 1, ins_pos)
            inserted += 1

        self._new_ntm_col = ntm_col + inserted
        self._old_ntm_col = ntm_col
        logger.info(f"Inserted {inserted} forecast column(s) ({', '.join(to_insert)}), "
                    f"NTM shifted to column {get_column_letter(self._new_ntm_col)} ({self._new_ntm_col})")
    
    def _update_impact_unit_sales_sheet(self):
        """
        Update Impact Unit Sales sheet.
        
        IMPORTANT: When we add a new quarter column to Suppl. Calc, the Impact Unit Sales
        sheet's references to Suppl. Calc must shift forward by one column.
        
        For a Q3 2025 report:
        - The current quarter is Q3 2025 (25Q3)
        - The forecast should start at the NEXT quarter: Q4 2025 (25Q4)
        - Row 9 references like ='Suppl. Calc'!O2 need to shift to ='Suppl. Calc'!P2
        
        The structure is:
        - Column C: Labels (Average Sale Price LTM, etc.)
        - Column D onwards: Data for different forecasts
        - Row 9: References to Suppl. Calc for quarter labels
        - Row 10: Unit Sales counts
        
        We shift all Suppl. Calc column references forward by one to correctly
        start the forecast at the next quarter (not the current quarter).
        """
        if 'Impact Unit Sales' not in self.workbook.sheetnames:
            logger.warning("Impact Unit Sales sheet not found")
            return
        
        sheet = self.workbook['Impact Unit Sales']
        
        # Shift Suppl. Calc references forward by one column
        # This ensures the forecast horizon starts at the NEXT quarter, not the current quarter
        # E.g., for Q3 2025 report: O->P, P->Q, Q->R, R->S (so forecast starts at 25Q4, not 25Q3)
        
        updated_count = 0
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is None:
                    continue
                
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    original_value = cell_value
                    new_value = original_value
                    
                    # Find and update Suppl. Calc column references
                    # Pattern: 'Suppl. Calc'!O2 -> 'Suppl. Calc'!P2
                    # We need to shift all column letters forward by one
                    suppl_calc_pattern = r"'Suppl\. Calc'!(\$?)([A-Z]+)(\$?)(\d+)"
                    
                    def shift_column(match):
                        dollar1 = match.group(1)
                        col_letter = match.group(2)
                        dollar2 = match.group(3)
                        row_num = match.group(4)
                        
                        # Convert column letter to index, add 1, convert back
                        from openpyxl.utils import column_index_from_string
                        col_idx = column_index_from_string(col_letter)
                        new_col_letter = get_column_letter(col_idx + 1)
                        
                        return f"'Suppl. Calc'!{dollar1}{new_col_letter}{dollar2}{row_num}"
                    
                    new_value = re.sub(suppl_calc_pattern, shift_column, new_value)
                    
                    if new_value != original_value:
                        cell.value = new_value
                        updated_count += 1
                        logger.debug(f"Row {row}, Col {col}: '{original_value}' -> '{new_value}'")
        
        if updated_count > 0:
            logger.info(f"Updated {updated_count} Suppl. Calc references in Impact Unit Sales (shifted forward by 1 column)")
        else:
            logger.info("Impact Unit Sales sheet - no Suppl. Calc references found to update")
    
    def _update_suppl_calc_formulas(self):
        """
        Update formulas in Suppl. Calc sheet to reference Q3 Management Accounts correctly.
        
        The formulas in Suppl. Calc reference the Management Accounts sheet, and we need to ensure
        they reference the correct sheet name and column (column C for Q3 data).
        """
        if 'Suppl. Calc' not in self.workbook.sheetnames:
            return
        
        sheet = self.workbook['Suppl. Calc']
        ma_sheet_name = f"Q{self.config.quarter} Management Accounts"
        
        prev_q = self.config.quarter - 1
        prev_y = self.config.year
        if prev_q == 0:
            prev_q = 4
            prev_y -= 1
        
        old_ma_patterns = [
            f"'Q{prev_q} {prev_y} Management Accounts'",
            f"'Q{prev_q} Management Accounts'",
            f"Q{prev_q} Management Accounts",
        ]
        new_ma_name = f"'{ma_sheet_name}'"
        
        updated_count = 0
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell_value = cell.value
                
                if cell_value is None:
                    continue
                
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    original_value = cell_value
                    new_value = original_value
                    
                    # Update Management Accounts sheet references
                    for old_pattern in old_ma_patterns:
                        if old_pattern in new_value:
                            new_value = new_value.replace(old_pattern, new_ma_name)
                            # Ensure proper exclamation mark
                            if "'" in new_value and "!" not in new_value.split("'")[-1]:
                                # Add ! if missing after sheet name
                                parts = new_value.split("'")
                                for i in range(len(parts) - 1):
                                    if parts[i].endswith(" Management Accounts") and i + 1 < len(parts):
                                        if not parts[i + 1].startswith("!"):
                                            parts[i + 1] = "!" + parts[i + 1]
                                new_value = "'".join(parts)
                    
                    if new_value != original_value:
                        cell.value = new_value
                        updated_count += 1
        
        if updated_count > 0:
            logger.info(f"Updated {updated_count} formulas in Suppl. Calc to reference {ma_sheet_name}")
        
        # Verify and fix column references: Suppl. Calc should reference column B
        # in Q{n} Management Accounts (kwartaal values), not column C (LTM values)
        fix_count = 0
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    val = cell.value
                    if f"'{ma_sheet_name}'!" in val:
                        # Check if it references column C and fix to column B
                        old_ref = f"'{ma_sheet_name}'!C"
                        new_ref = f"'{ma_sheet_name}'!B"
                        if old_ref in val:
                            cell.value = val.replace(old_ref, new_ref)
                            fix_count += 1
        if fix_count > 0:
            logger.info(f"Fixed {fix_count} Suppl. Calc formula(s) to reference column B in {ma_sheet_name}")
    
    def _update_sfa_cc_sheet(self):
        """
        Update SFA CC sheet with current quarter references.
        
        This includes:
        1. Updating sheet name references in formulas (Q2 Management Accounts → Q3 Management Accounts)
        2. Updating Suppl. Calc column references for NTM (column S → T after insert)
        3. Updating text references to quarter
        4. Fixing any #NAME? errors by ensuring correct references
        """
        if 'SFA CC' not in self.workbook.sheetnames:
            logger.warning("SFA CC sheet not found")
            return
        
        sheet = self.workbook['SFA CC']
        
        # Get the NTM column shift info
        old_ntm_col_letter = None
        new_ntm_col_letter = None
        
        if 'Suppl. Calc' in self.workbook.sheetnames:
            suppl = self.workbook['Suppl. Calc']
            for col in range(1, suppl.max_column + 5):
                val = suppl.cell(row=2, column=col).value
                if val and str(val).upper() == 'NTM':
                    new_ntm_col_letter = get_column_letter(col)
                    # The old NTM was one column to the left (before we inserted)
                    if col > 1:
                        old_ntm_col_letter = get_column_letter(col - 1)
                    break
            
            if old_ntm_col_letter and new_ntm_col_letter:
                logger.info(f"NTM column shift: {old_ntm_col_letter} → {new_ntm_col_letter}")
        
        # Build replacement patterns for Management Accounts sheet name
        prev_q = self.config.quarter - 1
        prev_y = self.config.year
        if prev_q == 0:
            prev_q = 4
            prev_y -= 1
        
        old_ma_patterns = [
            f"'Q{prev_q} {prev_y} Management Accounts'",
            f"'Q{prev_q} Management Accounts'",
            f"Q{prev_q} Management Accounts",
            f"'Q{prev_q} {prev_y} Management Accounts'!",
            f"'Q{prev_q} Management Accounts'!",
        ]
        new_ma_name = f"'Q{self.config.quarter} Management Accounts'"
        
        updated_count = 0
        formula_count = 0
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Handle both string formulas and formula objects
                cell_value = cell.value
                if cell_value is None:
                    continue
                
                # Convert to string for processing
                if hasattr(cell_value, 'text'):
                    original_value = cell_value.text
                else:
                    original_value = str(cell_value)
                
                new_value = original_value
                
                # Check if it's a formula
                if isinstance(cell_value, str) and new_value.startswith('='):
                    # Update Management Accounts sheet references
                    for old_pattern in old_ma_patterns:
                        if old_pattern in new_value:
                            new_value = new_value.replace(old_pattern, new_ma_name + "!")
                            # Remove double exclamation if created
                            new_value = new_value.replace("!!", "!")
                            logger.debug(f"Updated MA reference: {old_pattern} → {new_ma_name}")
                    
                    # Update ALL Suppl. Calc column references at or beyond the
                    # insertion point (old NTM position). When we insert a column
                    # at the NTM position, all columns at or to the right shift +1.
                    # This handles both the NTM shift and any columns beyond it
                    # (e.g., growth rate columns), and works regardless of whether
                    # old quarters were hidden or removed.
                    if old_ntm_col_letter and "'Suppl. Calc'" in new_value:
                        from openpyxl.utils import column_index_from_string
                        old_ntm_idx = column_index_from_string(old_ntm_col_letter)
                        
                        def _shift_suppl_calc_col(match):
                            dollar1 = match.group(1)
                            col_letter = match.group(2)
                            dollar2 = match.group(3)
                            row_num = match.group(4)
                            col_idx = column_index_from_string(col_letter)
                            if col_idx >= old_ntm_idx:
                                shifted = get_column_letter(col_idx + 1)
                                return f"'Suppl. Calc'!{dollar1}{shifted}{dollar2}{row_num}"
                            return match.group(0)
                        
                        suppl_pattern = r"'Suppl\. Calc'!(\$?)([A-Z]+)(\$?)(\d+)"
                        new_value = re.sub(suppl_pattern, _shift_suppl_calc_col, new_value)
                    
                    # Update Impact Unit Sales column references similarly
                    if old_ntm_col_letter and "'Impact Unit Sales'" in new_value:
                        def _shift_ius_col(match):
                            dollar1 = match.group(1)
                            col_letter = match.group(2)
                            dollar2 = match.group(3)
                            row_num = match.group(4)
                            col_idx = column_index_from_string(col_letter)
                            if col_idx >= old_ntm_idx:
                                shifted = get_column_letter(col_idx + 1)
                                return f"'Impact Unit Sales'!{dollar1}{shifted}{dollar2}{row_num}"
                            return match.group(0)
                        
                        ius_pattern = r"'Impact Unit Sales'!(\$?)([A-Z]+)(\$?)(\d+)"
                        new_value = re.sub(ius_pattern, _shift_ius_col, new_value)
                    
                    if new_value != original_value:
                        formula_count += 1
                        cell.value = new_value
                        updated_count += 1
                
                # Update text references to quarter (non-formula)
                elif isinstance(cell_value, str) and not new_value.startswith('='):
                    if self.config.prev_quarter_str in new_value:
                        new_value = new_value.replace(self.config.prev_quarter_str, self.config.quarter_str)
                        if new_value != original_value:
                            cell.value = new_value
                            updated_count += 1
        
        logger.info(f"Updated {updated_count} cells in SFA CC ({formula_count} formulas)")
        
        # Update G44 (Proceeds from disposals) to reference Q3 Management Accounts C49
        ma_sheet_name = f"Q{self.config.quarter} Management Accounts"
        g44_formula = f"='{ma_sheet_name}'!C49"
        sheet.cell(row=44, column=7).value = g44_formula
        logger.info(f"Set G44 to {g44_formula}")
        
        # Update G68 (links to bank loan / senior loan in MA sheet)
        g68_formula = f"=-'{ma_sheet_name}'!C11"
        sheet.cell(row=68, column=7).value = g68_formula
        logger.info(f"Set G68 to {g68_formula}")
        
        # Update signature page dates (typically in rows 82+)
        self._update_signature_page_dates(sheet)
        
        # Ensure signature page content exists
        self._ensure_signature_page(sheet)
    
    def _copy_actual_values_to_suppl_calc(self, management_accounts_path: str):
        """
        Copy actual values from Management Cijfers to Suppl. Calc sheet.
        
        This matches keywords in Suppl. Calc Column B with Management Cijfers Column A,
        and copies the corresponding values from the LTM column to the current quarter column.
        
        Quarter to column mapping:
        - Q3 2025 → Column O (15)
        - Q4 2025 → Column P (16)
        - Q1 2026 → Column Q (17)
        - Q2 2026 → Column R (18)
        """
        if 'Suppl. Calc' not in self.workbook.sheetnames:
            logger.warning("Suppl. Calc sheet not found")
            return
        
        suppl_sheet = self.workbook['Suppl. Calc']
        
        # Load Management Cijfers sheet with both formula and data views
        try:
            ma_wb_formulas = openpyxl.load_workbook(management_accounts_path, data_only=False)
            ma_wb_data = openpyxl.load_workbook(management_accounts_path, data_only=True)
            
            # Find Management Cijfers sheet (formulas)
            cijfers_sheet = None
            cijfers_sheet_data = None
            for name in ma_wb_formulas.sheetnames:
                if 'Management Cijfers' in name:
                    cijfers_sheet = ma_wb_formulas[name]
                    cijfers_sheet_data = ma_wb_data[name]
                    break
            
            if cijfers_sheet is None:
                logger.warning("Management Cijfers sheet not found in Management Accounts file")
                ma_wb_formulas.close()
                ma_wb_data.close()
                return
            
            # Find BDO sheet for formula evaluation
            expected_bdo_name = f"BDO - Q{self.config.quarter}-{str(self.config.year)[-2:]}"
            bdo_sheet = None
            for name in ma_wb_formulas.sheetnames:
                if expected_bdo_name in name:
                    bdo_sheet = ma_wb_formulas[name]
                    logger.info(f"Found BDO sheet for Suppl. Calc: {name}")
                    break
            
            # Find the current quarter column in Management Cijfers (NOT the LTM column)
            # For Q3 2025, this is Column AA (the quarter column, not AB which is LTM)
            # The quarter column is one to the left of the LTM column
            source_column = None
            quarter_header = f"Q{self.config.quarter} {self.config.year}"
            
            for col in range(cijfers_sheet.max_column, 0, -1):
                cell_val = cijfers_sheet.cell(row=22, column=col).value
                if cell_val:
                    cell_str = str(cell_val)
                    # Look for the quarter header (e.g., "Q3 2025") NOT the LTM header
                    if quarter_header in cell_str and 'LTM' not in cell_str:
                        source_column = col
                        logger.info(f"Found quarter column at {get_column_letter(col)} ({col}) with header: '{cell_str}'")
                        break
            
            if source_column is None:
                logger.warning(f"Could not find quarter column '{quarter_header}' in Management Cijfers")
                ma_wb_formulas.close()
                ma_wb_data.close()
                return
            
            source_col_letter = get_column_letter(source_column)
            
            # Determine target column based on quarter
            # Q3=O(15), Q4=P(16), Q1=Q(17), Q2=R(18)
            quarter_offset = (self.config.quarter - 3) % 4
            target_column = 15 + quarter_offset  # O=15 for Q3
            target_col_letter = get_column_letter(target_column)
            
            logger.info(f"Target column for Q{self.config.quarter}: {target_col_letter} ({target_column})")
            
            # Find LTM column for Balance Sheet items (one column to the right of quarter column)
            ltm_column = source_column + 1
            logger.info(f"LTM column for Balance Sheet items: {get_column_letter(ltm_column)} ({ltm_column})")
            
            # Keyword mapping: Suppl. Calc keyword -> (Management Cijfers keyword variations, row, is_balance_sheet)
            # P&L items (rows 23-68) use quarterly column AA
            # Balance Sheet items (rows 3-19) use LTM column AB
            keyword_mapping = {
                'Theoretical rental income after sales': (['Gross Theoretical rental income'], 23, False),
                '(Financial vacancy)': (['(Financial vacancy)'], 24, False),
                '(loss of rent from Unit Sales)': (None, None, None),  # Not in Management Cijfers - skip
                'Gross rental income': (['Gross rental income'], 25, False),
                'Service costs charged (100% occupancy)': (None, None, None),  # Skip - keep empty
                '(Service costs recoverable)': (None, None, None),  # Skip - keep empty
                '(Vacancy costs)': (['(Vacancy costs)'], 28, False),
                'Non recoverable service charges': (['Service charges'], 30, False),
                '(Maintenance & repair costs)': (['(Maintenance & repair costs actual)'], 32, False),
                '(Owners society costs (VVE))': (['(Owners society costs (VVE))'], 33, False),
                '(Insurance costs)': (['(Insurance costs)'], 34, False),
                '(Landlord tax)': (['(Landlord tax costs)'], 35, False),
                '(Property tax)': (['(Property tax)'], 36, False),
                '(Other costs)': (['(Other costs)'], 40, False),
                '(Agent costs) - in financing costs on CC': (['(Agent costs)'], 38, False),
                '(Brokerage costs)': (['(Brokerage costs)'], 39, False),
                '(Audit/fee costs)': (['(Accountant costs)', '(Advisory costs)'], [41, 42], False),  # Sum of both
                'Property related expenses': (['Property related expenses'], 44, False),
                'Net rental income': (['Net rental income'], 45, False),
                '(Management fees)': (['(Management fees)'], 46, False),
                'EBIT Rental operation': (['EBITDA rental operation'], 48, False),
                'Total EBIT': (['Total EBITDA'], 55, False),
                '(Interest expenses Senior)': (['(Net interest expenses - SFA)'], 60, False),
                '(Interest expenses Hedge)': (['(Net interest expenses - Hedge)'], 64, False),
                'EBT': (['EBT'], 66, False),
                'Senior loan - EoP': (['Bank loan'], 12, True),  # Balance Sheet - use LTM column
                'CAPEX loan - EoP': (None, None, None),  # Not in Management Cijfers - skip
            }
            
            # Shift row 7 values to preserve forecast values by moving them forward
            # Row 7 is "(loss of rent from Unit Sales)"
            # The shift starts from the current quarter's target column and moves values right
            # Q3 2025: O→P→Q→R→S, clear O
            # Q4 2025: P→Q→R→S→T, clear P
            # Q1 2026: Q→R→S→T→U, clear Q
            # Q2 2026: R→S→T→U→V, clear R
            shift_start_col = target_column  # Dynamic based on quarter
            shift_cols = [shift_start_col + i for i in range(4)]  # 4 columns to shift
            for i in range(len(shift_cols) - 1, -1, -1):
                src_col = shift_cols[i]
                dst_col = src_col + 1
                val = suppl_sheet.cell(row=7, column=src_col).value
                if val is not None:
                    suppl_sheet.cell(row=7, column=dst_col).value = val
            # Clear the current quarter's column (row 7) after shifting
            suppl_sheet.cell(row=7, column=target_column).value = None
            logger.info(f"Shifted row 7 values from {target_col_letter} onwards and cleared {target_col_letter}7")
            
            cv = getattr(self, '_computed_values', None) or {}

            def get_cell_value(row: int, col: int = source_column):
                """Get a cell's value — prefer computed_values cache, then formula eval."""
                if cv:
                    col_key = 'ltm' if col == ltm_column else 'quarter'
                    cached = cv.get((row, col_key))
                    if cached is not None:
                        return cached

                cached_val = cijfers_sheet_data.cell(row=row, column=col).value
                if cached_val is not None and isinstance(cached_val, (int, float)):
                    return cached_val

                formula = cijfers_sheet.cell(row=row, column=col).value
                if formula and isinstance(formula, str) and formula.startswith('='):
                    if bdo_sheet is not None:
                        result = self._evaluate_bdo_reference(formula, bdo_sheet, expected_bdo_name)
                        if result is not None:
                            return result
                        result = self._evaluate_formula_with_internal_refs(
                            formula, bdo_sheet, expected_bdo_name, source_column, cijfers_sheet, get_cell_value
                        )
                        if result is not None:
                            return result

                return None
            
            # Now process each keyword in Suppl. Calc
            values_copied = 0
            for suppl_row in range(1, suppl_sheet.max_row + 1):
                suppl_keyword = suppl_sheet.cell(row=suppl_row, column=2).value  # Column B
                if suppl_keyword is None:
                    continue
                
                suppl_keyword_str = str(suppl_keyword).strip()
                
                # Check if this keyword is in our mapping
                if suppl_keyword_str in keyword_mapping:
                    keywords, source_rows, is_balance_sheet = keyword_mapping[suppl_keyword_str]
                    if keywords is None:
                        continue  # Skip keywords that don't exist in Management Cijfers
                    
                    # Determine which column to use
                    # P&L items use quarterly column (AA), Balance Sheet uses LTM column (AB)
                    use_column = ltm_column if is_balance_sheet else source_column
                    
                    # Get the value from Management Cijfers
                    found_value = None
                    
                    if isinstance(source_rows, list):
                        # Sum multiple rows (e.g., Accountant + Advisory for Audit/fee costs)
                        total = 0.0
                        all_found = True
                        for src_row in source_rows:
                            val = get_cell_value(src_row, use_column)
                            if val is not None:
                                total += val
                            else:
                                all_found = False
                        if all_found or total != 0:
                            found_value = total
                    else:
                        # Single row
                        found_value = get_cell_value(source_rows, use_column)
                    
                    if found_value is not None:
                        target_cell = suppl_sheet.cell(row=suppl_row, column=target_column)
                        target_cell.value = found_value
                        prev_fmt_cell = suppl_sheet.cell(row=suppl_row, column=target_column - 1)
                        if prev_fmt_cell.number_format and prev_fmt_cell.number_format != 'General':
                            target_cell.number_format = prev_fmt_cell.number_format
                        if prev_fmt_cell.has_style:
                            target_cell.font = copy(prev_fmt_cell.font)
                            target_cell.border = copy(prev_fmt_cell.border)
                            target_cell.fill = copy(prev_fmt_cell.fill)
                            target_cell.alignment = copy(prev_fmt_cell.alignment)
                        values_copied += 1
                        logger.debug(f"Copied {found_value:,.2f} to Suppl. Calc row {suppl_row} ({suppl_keyword_str})")
            
            # Update the column header from "Forecast" to "Actual"
            header_cell = suppl_sheet.cell(row=3, column=target_column)
            if header_cell.value and 'Forecast' in str(header_cell.value):
                header_cell.value = 'Actual'
                logger.info(f"Updated {target_col_letter}3 from 'Forecast' to 'Actual'")
            
            logger.info(f"Copied {values_copied} actual values to Suppl. Calc column {target_col_letter}")
            
            # Set formula rows - these should never be replaced with values
            # These are summary/total formulas that must always reference their column
            self._set_suppl_calc_formula_rows(suppl_sheet, target_column)
            
            # Update formulas and shift columns for next quarter
            self._update_suppl_calc_quarter_formulas(suppl_sheet, target_column)
            
            ma_wb_formulas.close()
            ma_wb_data.close()
            
        except Exception as e:
            logger.error(f"Error copying actual values to Suppl. Calc: {e}")
            import traceback
            traceback.print_exc()
    
    def _set_suppl_calc_formula_rows(self, suppl_sheet, target_column: int):
        """
        Set formula rows in Suppl. Calc that should never be replaced with values.
        
        These are summary/total rows that must always contain formulas:
        - Row 8: =X6+X5+X7
        - Row 13: =SUM(X10:X12)
        - Row 26: =SUM(X17:X25)
        - Row 27: =X26+X13+X8
        - Row 29: =X27+X28
        - Row 31: =X29
        - Row 35: =X33+X32+X31
        - Row 43: =SUM(X41:X42)
        
        Args:
            suppl_sheet: The Suppl. Calc worksheet
            target_column: Column index to set formulas for
        """
        col_letter = get_column_letter(target_column)
        
        # Define formula templates for each row
        formula_rows = {
            8: f"={col_letter}6+{col_letter}5+{col_letter}7",
            13: f"=SUM({col_letter}10:{col_letter}12)",
            26: f"=SUM({col_letter}17:{col_letter}25)",
            27: f"={col_letter}26+{col_letter}13+{col_letter}8",
            29: f"={col_letter}27+{col_letter}28",
            31: f"={col_letter}29",
            35: f"={col_letter}33+{col_letter}32+{col_letter}31",
            43: f"=SUM({col_letter}41:{col_letter}42)",
        }
        
        for row, formula in formula_rows.items():
            cell = suppl_sheet.cell(row=row, column=target_column)
            cell.value = formula
            prev_cell = suppl_sheet.cell(row=row, column=target_column - 1)
            if prev_cell.has_style:
                cell.font = copy(prev_cell.font)
                cell.alignment = copy(prev_cell.alignment)
                cell.border = copy(prev_cell.border)
                cell.fill = copy(prev_cell.fill)
                if prev_cell.number_format and prev_cell.number_format != 'General':
                    cell.number_format = prev_cell.number_format
        
        logger.info(f"Set {len(formula_rows)} formula rows in column {col_letter}")
        
        # Also set formulas for all subsequent quarterly columns (P, Q, R, S)
        # to ensure they maintain formula structure
        for col in range(target_column + 1, 20):  # Up to S(19)
            c_letter = get_column_letter(col)
            for row, _ in formula_rows.items():
                if row == 8:
                    formula = f"={c_letter}6+{c_letter}5+{c_letter}7"
                elif row == 13:
                    formula = f"=SUM({c_letter}10:{c_letter}12)"
                elif row == 26:
                    formula = f"=SUM({c_letter}17:{c_letter}25)"
                elif row == 27:
                    formula = f"={c_letter}26+{c_letter}13+{c_letter}8"
                elif row == 29:
                    formula = f"={c_letter}27+{c_letter}28"
                elif row == 31:
                    formula = f"={c_letter}29"
                elif row == 35:
                    formula = f"={c_letter}33+{c_letter}32+{c_letter}31"
                elif row == 43:
                    formula = f"=SUM({c_letter}41:{c_letter}42)"
                
                cell = suppl_sheet.cell(row=row, column=col)
                cell.value = formula
                prev_cell = suppl_sheet.cell(row=row, column=col - 1)
                if prev_cell.has_style:
                    cell.font = copy(prev_cell.font)
                    cell.alignment = copy(prev_cell.alignment)
                    cell.border = copy(prev_cell.border)
                    cell.fill = copy(prev_cell.fill)
                    if prev_cell.number_format and prev_cell.number_format != 'General':
                        cell.number_format = prev_cell.number_format
    
    def _update_suppl_calc_quarter_formulas(self, suppl_sheet, target_column: int):
        """
        Update Suppl. Calc sheet formulas for quarter progression.

        All column positions (NTM, forecast range, rate) are resolved dynamically
        from row-2 headers rather than hardcoded column indices.
        """
        from openpyxl.utils import get_column_letter

        target_col_letter = get_column_letter(target_column)
        prev_col_letter = get_column_letter(target_column - 1) if target_column > 1 else 'N'

        # --- Resolve NTM column dynamically ---
        ntm_col = None
        for col in range(1, suppl_sheet.max_column + 5):
            hdr = suppl_sheet.cell(row=2, column=col).value
            if hdr and str(hdr).strip().upper() == 'NTM':
                ntm_col = col
                break

        if ntm_col is None:
            logger.warning("NTM column not found in Suppl. Calc – skipping formula updates")
            return

        ntm_letter = get_column_letter(ntm_col)
        last_fc_col = ntm_col - 1
        last_fc_letter = get_column_letter(last_fc_col)
        logger.info(f"Dynamic NTM at {ntm_letter} ({ntm_col}), "
                    f"forecast range {get_column_letter(target_column + 1)}-{last_fc_letter}")

        # === 1. Populate newly-inserted forecast column if mostly empty ===
        new_fc = last_fc_col
        prev_fc = last_fc_col - 1

        if prev_fc > target_column:
            new_cnt = sum(1 for r in range(5, 46)
                         if suppl_sheet.cell(row=r, column=new_fc).value is not None)
            prev_cnt = sum(1 for r in range(5, 46)
                          if suppl_sheet.cell(row=r, column=prev_fc).value is not None)

            if new_cnt < prev_cnt * 0.5:
                p_letter = get_column_letter(prev_fc)
                n_letter = get_column_letter(new_fc)
                p2_letter = get_column_letter(prev_fc - 1) if prev_fc > 1 else p_letter
                logger.info(f"Populating empty forecast col {n_letter} from {p_letter}")

                for row in range(4, suppl_sheet.max_row + 1):
                    src_cell = suppl_sheet.cell(row=row, column=prev_fc)
                    tgt_cell = suppl_sheet.cell(row=row, column=new_fc)
                    src = src_cell.value
                    if src is not None and tgt_cell.value is None:
                        if isinstance(src, str) and src.startswith('='):
                            shifted = re.sub(rf'\b{re.escape(p_letter)}(\d+)',
                                             rf'{n_letter}\1', src)
                            shifted = re.sub(rf'\b{re.escape(p2_letter)}(\d+)',
                                             rf'{p_letter}\1', shifted)
                            tgt_cell.value = shifted
                        else:
                            tgt_cell.value = src
                        if src_cell.has_style:
                            tgt_cell.font = copy(src_cell.font)
                            tgt_cell.alignment = copy(src_cell.alignment)
                            tgt_cell.border = copy(src_cell.border)
                            tgt_cell.fill = copy(src_cell.fill)
                            if src_cell.number_format and src_cell.number_format != 'General':
                                tgt_cell.number_format = src_cell.number_format

                prev_hdr = suppl_sheet.cell(row=2, column=prev_fc).value
                if prev_hdr and 'Q' in str(prev_hdr):
                    parts = str(prev_hdr).split('Q')
                    yr, q = parts[0], int(parts[1])
                    q += 1
                    if q > 4:
                        q, yr = 1, str(int(yr) + 1)
                    suppl_sheet.cell(row=2, column=new_fc).value = f"{yr}Q{q}"
                suppl_sheet.cell(row=3, column=new_fc).value = "Forecast"

        # === 2. Row 6 formula pattern: =col5*$RATE$6 ===
        rate_col = None
        for col in range(ntm_col + 1, ntm_col + 10):
            val = suppl_sheet.cell(row=6, column=col).value
            if val is not None and isinstance(val, (int, float)) and val != 0:
                rate_col = col
                break
        if rate_col is None:
            rate_col = ntm_col + 3
        rate_letter = get_column_letter(rate_col)
        logger.info(f"Row 6 rate column: {rate_letter} "
                    f"(value: {suppl_sheet.cell(row=6, column=rate_col).value})")

        for col in range(target_column + 1, ntm_col):
            cl = get_column_letter(col)
            cell = suppl_sheet.cell(row=6, column=col)
            cell.value = f"={cl}5*${rate_letter}$6"
            prev_cell = suppl_sheet.cell(row=6, column=col - 1)
            if prev_cell.has_style:
                cell.font = copy(prev_cell.font)
                cell.alignment = copy(prev_cell.alignment)
                cell.border = copy(prev_cell.border)
                cell.fill = copy(prev_cell.fill)
                if prev_cell.number_format and prev_cell.number_format != 'General':
                    cell.number_format = prev_cell.number_format

        # === 3. Rows 31-35 formulas for actual + forecast columns ===
        for r, formula in [
            (31, f"={target_col_letter}29"),
            (32, f"=-({prev_col_letter}41*(1.75%)+({prev_col_letter}41*4%*30%))/4"),
            (33, f"=-({prev_col_letter}41*70%*2.6%)/4"),
            (35, f"={target_col_letter}33+{target_col_letter}32+{target_col_letter}31"),
        ]:
            cell = suppl_sheet.cell(row=r, column=target_column)
            cell.value = formula
            prev_cell = suppl_sheet.cell(row=r, column=target_column - 1)
            if prev_cell.has_style:
                cell.font = copy(prev_cell.font)
                cell.alignment = copy(prev_cell.alignment)
                cell.border = copy(prev_cell.border)
                cell.fill = copy(prev_cell.fill)
                if prev_cell.number_format and prev_cell.number_format != 'General':
                    cell.number_format = prev_cell.number_format

        for col in range(target_column + 1, ntm_col):
            cl = get_column_letter(col)
            pl = get_column_letter(col - 1)
            for r, formula in [
                (31, f"={cl}29"),
                (32, f"=-({pl}41*(1.75%)+({pl}41*4%*30%))/4"),
                (33, f"=-({pl}41*70%*2.6%)/4"),
                (35, f"={cl}33+{cl}32+{cl}31"),
            ]:
                cell = suppl_sheet.cell(row=r, column=col)
                cell.value = formula
                prev_cell = suppl_sheet.cell(row=r, column=col - 1)
                if prev_cell.has_style:
                    cell.font = copy(prev_cell.font)
                    cell.alignment = copy(prev_cell.alignment)
                    cell.border = copy(prev_cell.border)
                    cell.fill = copy(prev_cell.fill)
                    if prev_cell.number_format and prev_cell.number_format != 'General':
                        cell.number_format = prev_cell.number_format

        # === 4. NTM SUM formulas: 4 forecast columns (excl. actual) ===
        ntm_sum_start = get_column_letter(target_column + 1)
        ntm_sum_end = last_fc_letter

        for row in range(4, 42):
            cell = suppl_sheet.cell(row=row, column=ntm_col)
            val = cell.value
            if val and isinstance(val, str) and '=SUM(' in val:
                new_formula = re.sub(
                    r'=SUM\([A-Z]+(\d+):[A-Z]+(\d+)\)',
                    f'=SUM({ntm_sum_start}\\1:{ntm_sum_end}\\2)',
                    val
                )
                if new_formula != val:
                    cell.value = new_formula
                    logger.debug(f"Updated {ntm_letter}{row}: {val} -> {new_formula}")
            elif val is None or (isinstance(val, str) and not val.startswith('=')):
                has_data = any(
                    suppl_sheet.cell(row=row, column=c).value is not None
                    for c in range(target_column, ntm_col)
                )
                if has_data:
                    cell.value = f"=SUM({ntm_sum_start}{row}:{ntm_sum_end}{row})"
                    logger.debug(f"Set new NTM formula at {ntm_letter}{row}")
            prev_ntm_cell = suppl_sheet.cell(row=row, column=ntm_col - 1)
            if prev_ntm_cell.has_style:
                if prev_ntm_cell.number_format and prev_ntm_cell.number_format != 'General':
                    cell.number_format = prev_ntm_cell.number_format
                cell.font = copy(prev_ntm_cell.font)
                cell.alignment = copy(prev_ntm_cell.alignment)
                cell.border = copy(prev_ntm_cell.border)
                cell.fill = copy(prev_ntm_cell.fill)

        # === 5. Suppl. Calc header styling: rows 2-3 white text on #43546a ===
        header_fill = PatternFill(start_color='FF43546A', end_color='FF43546A', fill_type='solid')
        for row in (2, 3):
            for col in range(1, suppl_sheet.max_column + 1):
                cell = suppl_sheet.cell(row=row, column=col)
                base = copy(cell.font) if cell.font else Font()
                cell.font = Font(
                    name=base.name, size=base.size, bold=base.bold,
                    italic=base.italic, underline=base.underline,
                    color='FFFFFFFF',
                )
                cell.fill = header_fill

        # === 6. NTM column rows 4-35: light grey background + black text ===
        ntm_data_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
        for row in range(4, 36):
            cell = suppl_sheet.cell(row=row, column=ntm_col)
            base = copy(cell.font) if cell.font else Font()
            make_bold = row >= 5
            cell.font = Font(
                name=base.name, size=base.size, bold=make_bold,
                italic=base.italic, underline=base.underline,
                color='FF000000',
            )
            cell.fill = ntm_data_fill

        logger.info("Completed Suppl. Calc formula updates (dynamic NTM)")
    
    def _update_signature_page_dates(self, sheet):
        """
        Update dates in the signature page section of SFA CC.
        
        The signature page typically contains:
        - Date in format "DD-MM-YYYY" or "DD Month YYYY"
        - Quarter references
        - Interest Period dates
        
        These need to be updated to the current quarter's period end.
        """
        # Calculate dates
        period_end = self.config.period_end
        
        # Report date is typically ~20 days after quarter end
        from datetime import timedelta
        report_date = period_end + timedelta(days=21)
        
        # Previous quarter dates
        prev_q = self.config.quarter - 1
        prev_y = self.config.year
        if prev_q == 0:
            prev_q = 4
            prev_y -= 1
        
        # Calculate previous period end
        prev_month = prev_q * 3
        if prev_month in [3, 12]:
            prev_day = 31
        else:
            prev_day = 30
        
        # New date formats
        new_period_date = f"{period_end.day}-{period_end.month}-{period_end.year}"
        new_report_date = f"{report_date.day}-{report_date.month}-{report_date.year}"
        
        # Date pattern replacements - sorted by specificity/length to avoid partial matches
        date_replacements = [
            # Interest Period formats (various)
            (f"Interest Period: {prev_month}/{prev_day}/{prev_y}", 
             f"Interest Period: {period_end.month}/{period_end.day}/{period_end.year}"),
            (f"Interest Period: {prev_day}-{prev_month}-{prev_y}", 
             f"Interest Period: {new_period_date}"),
            # US format mm/dd/yyyy
            (f"{prev_month}/{prev_day}/{prev_y}", 
             f"{period_end.month}/{period_end.day}/{period_end.year}"),
            # EU format dd-mm-yyyy
            (f"{prev_day}-{prev_month}-{prev_y}", new_period_date),
            # Dated field - report date (21 days after quarter end)
            (f"Dated {self._calc_dated_field(prev_q, prev_y)}",
             f"Dated {new_report_date}"),
            # Just "21-7-2025" or similar
            (f"21-{prev_month + 1 if prev_month < 12 else 1}-{prev_y}", new_report_date),
        ]
        
        updated_dates = 0
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and not cell.value.startswith('='):
                    original = cell.value
                    new_value = original
                    
                    # Apply date replacements
                    for old_pattern, new_pattern in date_replacements:
                        if old_pattern in new_value:
                            new_value = new_value.replace(old_pattern, new_pattern)
                    
                    # Replace quarter text references
                    old_q_str = f"Q{prev_q} {prev_y}"
                    if old_q_str in new_value:
                        new_value = new_value.replace(old_q_str, self.config.quarter_str)
                    
                    if new_value != original:
                        cell.value = new_value
                        updated_dates += 1
        
        if updated_dates > 0:
            logger.info(f"Updated {updated_dates} dates in signature section")
    
    def _calc_dated_field(self, quarter: int, year: int) -> str:
        """Calculate the 'Dated' field date (typically 21 days after quarter end)."""
        from datetime import datetime, timedelta
        month = quarter * 3
        day = 31 if month in [3, 12] else 30
        period_end = datetime(year, month, day)
        dated = period_end + timedelta(days=21)
        return f"{dated.day}-{dated.month}-{dated.year}"
    
    def _ensure_signature_page(self, sheet):
        """
        Ensure the signature page content exists in the SFA CC sheet.
        
        If row 81 says "Signature page follows" but the actual signature section
        (rows 82+) is missing content, add the standard signature fields.
        """
        from datetime import timedelta
        
        # Check if signature page content already exists
        signature_content_exists = False
        for row in range(82, min(sheet.max_row + 1, 100)):
            for col in range(1, 10):
                cell_val = sheet.cell(row=row, column=col).value
                if cell_val and isinstance(cell_val, str) and len(cell_val.strip()) > 0:
                    # Actual content found
                    if 'Name' in str(cell_val) or 'Title' in str(cell_val) or 'Signature' in str(cell_val):
                        signature_content_exists = True
                        break
            if signature_content_exists:
                break
        
        if signature_content_exists:
            logger.debug("Signature page content already exists")
            return
        
        # Add signature page content
        # Calculate dates
        period_end = self.config.period_end
        report_date = period_end + timedelta(days=21)
        report_date_str = report_date.strftime("%d %B %Y")
        
        # Standard signature page layout starting at row 83
        signature_content = [
            # Row 83: Title
            (83, 1, "SIGNATURE PAGE"),
            (83, 2, None),
            # Row 85: QSP ESS B.V.
            (85, 1, "For and on behalf of QSP ESS B.V."),
            # Row 87-88: Name fields
            (87, 1, "Name:"),
            (87, 4, "_________________________________"),
            (88, 1, "Title:"),
            (88, 4, "_________________________________"),
            (89, 1, "Date:"),
            (89, 4, report_date_str),
            # Row 91: Authorised Signatory
            (91, 1, "Authorised Signatory"),
            # Row 93-94: Second signer
            (93, 1, "Name:"),
            (93, 4, "_________________________________"),
            (94, 1, "Title:"),
            (94, 4, "_________________________________"),
            (95, 1, "Date:"),
            (95, 4, report_date_str),
            # Row 97
            (97, 1, "Authorised Signatory"),
        ]
        
        for row, col, value in signature_content:
            cell = sheet.cell(row=row, column=col)
            cell.value = value
            if row == 83:  # Title row
                cell.font = Font(bold=True, size=14)
            elif 'Name' in str(value) or 'Title' in str(value) or 'Date' in str(value):
                cell.font = Font(bold=True)
        
        # Update print area to include signature page
        if sheet.print_area:
            import re
            match = re.match(r".*\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", sheet.print_area)
            if match:
                start_col, start_row, end_col, _ = match.groups()
                new_print_area = f"${start_col}${start_row}:${end_col}$99"
                sheet.print_area = new_print_area
                logger.debug(f"Extended print area to include signature page: {new_print_area}")
        
        logger.info("Added signature page content to SFA CC sheet")
    
    def _update_all_dates_and_text(self):
        """
        Update ALL dates and text references throughout the entire workbook.
        
        This ensures consistency across all sheets:
        - SFA CC signature dates
        - Interest Period references  
        - Quarter text (Q2 2025 → Q3 2025)
        - Date formats (21-7-2025 → 21-10-2025)
        
        IMPORTANT: We MUST NOT replace short quarter patterns (like "25Q2") in 
        column headers of Suppl. Calc and Impact Unit Sales sheets, as these
        represent historical quarters, not the current quarter being updated.
        """
        from datetime import timedelta
        
        period_end = self.config.period_end
        report_date = period_end + timedelta(days=21)
        
        # Previous quarter info
        prev_q = self.config.quarter - 1
        prev_y = self.config.year
        if prev_q == 0:
            prev_q = 4
            prev_y -= 1
        
        prev_month = prev_q * 3
        prev_day = 31 if prev_month in [3, 12] else 30
        
        # Build comprehensive replacement map (sorted by length to prevent partial matches)
        replacements = []
        
        # Quarter text patterns - FULL format only (e.g., "Q2 2025")
        # We DO NOT include the short format (e.g., "25Q2") because it would
        # incorrectly replace historical quarter column headers in Suppl. Calc
        replacements.extend([
            (f"Q{prev_q} {prev_y}", self.config.quarter_str),
            (f"Q{prev_q}-{prev_y}", f"Q{self.config.quarter}-{self.config.year}"),
            # NOTE: We intentionally OMIT the short quarter pattern replacement
            # (f"{str(prev_y)[-2:]}Q{prev_q}", self.config.short_quarter),
            # as it would incorrectly replace column headers in Suppl. Calc and Impact Unit Sales
        ])
        
        # Date patterns - multiple formats
        # Dated field (report date, ~21 days after quarter)
        prev_report_month = prev_month + 1 if prev_month < 12 else 1
        prev_report_year = prev_y if prev_month < 12 else prev_y + 1
        replacements.extend([
            # Full dated field
            (f"Dated 21-{prev_report_month}-{prev_report_year}",
             f"Dated {report_date.day}-{report_date.month}-{report_date.year}"),
            # Just the date
            (f"21-{prev_report_month}-{prev_report_year}",
             f"{report_date.day}-{report_date.month}-{report_date.year}"),
        ])
        
        # Interest Period dates - US format mm/dd/yyyy
        replacements.extend([
            (f"Interest Period: {prev_month}/{prev_day}/{prev_y}",
             f"Interest Period: {period_end.month}/{period_end.day}/{period_end.year}"),
            (f"{prev_month}/{prev_day}/{prev_y}",
             f"{period_end.month}/{period_end.day}/{period_end.year}"),
        ])
        
        # Period end dates - EU format dd-mm-yyyy
        replacements.extend([
            (f"{prev_day}-{prev_month}-{prev_y}",
             f"{period_end.day}-{period_end.month}-{period_end.year}"),
        ])
        
        # Sort by length (longest first) to prevent partial matches
        replacements.sort(key=lambda x: len(x[0]), reverse=True)
        
        total_updated = 0
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_updated = 0
            
            # Skip row 2 in Suppl. Calc and Impact Unit Sales (column headers)
            # These contain historical quarter labels that should NOT be updated
            skip_header_rows = sheet_name in ['Suppl. Calc', 'Impact Unit Sales']
            
            # For Management Accounts sheet, skip Row 21 (P&L title) since we set it explicitly
            # in _copy_ma_data to ensure it shows the correct LTM title
            is_management_accounts = 'Management Accounts' in sheet_name
            
            for row in range(1, min(sheet.max_row + 1, 200)):
                # Skip header row (row 2) for specific sheets with quarter column headers
                if skip_header_rows and row == 2:
                    continue
                # Skip row 21 in Management Accounts (P&L title set explicitly in _copy_ma_data)
                if is_management_accounts and row == 21:
                    continue
                
                for col in range(1, min(sheet.max_column + 1, 50)):
                    cell = sheet.cell(row=row, column=col)
                    
                    if cell.value is None:
                        continue
                    
                    # Only process string values (not formulas)
                    if isinstance(cell.value, str) and not cell.value.startswith('='):
                        original = cell.value
                        new_value = original
                        
                        for old_pattern, new_pattern in replacements:
                            if old_pattern in new_value:
                                new_value = new_value.replace(old_pattern, new_pattern)
                        
                        if new_value != original:
                            cell.value = new_value
                            sheet_updated += 1
            
            if sheet_updated > 0:
                logger.debug(f"Updated {sheet_updated} cells in sheet '{sheet_name}'")
                total_updated += sheet_updated
        
        if total_updated > 0:
            logger.info(f"Updated {total_updated} date/text references across all sheets")
    
    def _clear_manual_input_cells(self):
        """Clear cells that require manual input - they must remain empty.
        
        These cells are populated manually by the client after automation runs.
        Setting them to None ensures no stale or incorrect values remain.
        """
        manual_cells = {
            'SFA CC': [
                (46, 7),   # G46
                (73, 7),   # G73
            ],
            'Suppl. Calc': [
                (41, 15),  # O41
                (42, 15),  # O42
            ],
            'Impact Unit Sales': [
                (3, 4),    # D3
                (4, 4),    # D4
                (5, 4),    # D5
                (10, 4),   # D10
                (10, 5),   # E10
                (10, 6),   # F10
                (10, 7),   # G10
            ],
        }
        
        for sheet_name, cells in manual_cells.items():
            if sheet_name not in self.workbook.sheetnames:
                logger.debug(f"Sheet '{sheet_name}' not found, skipping manual cell clearing")
                continue
            
            sheet = self.workbook[sheet_name]
            cleared = 0
            for row, col in cells:
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell.value = None
                    cleared += 1
            
            if cleared > 0:
                logger.info(f"Cleared {cleared} manual input cell(s) in '{sheet_name}'")

    def _copy_column_formatting(self, sheet, source_col: int, target_col: int):
        """Copy column formatting from source to target."""
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)
        
        # Copy column width
        if source_letter in sheet.column_dimensions:
            sheet.column_dimensions[target_letter].width = sheet.column_dimensions[source_letter].width
        
        # Copy cell formatting for data rows
        for row in range(1, sheet.max_row + 1):
            source_cell = sheet.cell(row=row, column=source_col)
            target_cell = sheet.cell(row=row, column=target_col)
            
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)

