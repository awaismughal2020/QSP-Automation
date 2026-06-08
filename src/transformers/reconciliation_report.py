"""
Post-AI reconciliation reporting for Management Accounts workbooks.

Builds a single authoritative report from the **formula evaluator** on the
saved workbook (not the build-time shadow model). Used by the orchestrator
for user-facing warnings and by the API for per-row mismatch visibility.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import yaml
from loguru import logger
from openpyxl.utils import get_column_letter

from .management_accounts import (
    ManagementAccountsBuilder,
    find_column_by_header_label,
    find_ltm_column_near_quarter,
    scan_summary_column_layout,
)


def _load_rules(rules: Optional[dict]) -> dict:
    if rules is not None:
        return rules
    rules_path = Path('config/accounting_rules.yaml')
    if rules_path.exists():
        try:
            with open(rules_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        except Exception:
            pass
    return {}


def read_bdo_ground_truth_pair(
    workbook_path: str,
    bdo_sheet_name: str,
    rules: Optional[dict] = None,
) -> Dict[str, Optional[float]]:
    """
    Sign-adjusted ``Resultaat na belasting`` totals from the BDO sheet.

    Returns ``{'ltm': float|None, 'quarter': float|None}`` where LTM uses
    column H (with C:G fallback) and quarter uses the sum of mutation columns.
    """
    rules = _load_rules(rules)
    result: Dict[str, Optional[float]] = {'ltm': None, 'quarter': None}

    bdo_cfg = rules.get('bdo', {})
    data_col = bdo_cfg.get('data_column', 8)
    q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
    label_cols = bdo_cfg.get('label_scan_columns', [1, 2])

    try:
        wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
    except Exception as exc:
        logger.warning(f"[Reconcile] BDO ground-truth open failed: {exc}")
        return result

    try:
        if bdo_sheet_name not in wb_data.sheetnames:
            return result
        bdo_sheet = wb_data[bdo_sheet_name]

        target_row = None
        for row_idx in range(1, min(bdo_sheet.max_row + 1, 200)):
            for col in label_cols:
                label = bdo_sheet.cell(row=row_idx, column=col).value
                if (
                    isinstance(label, str)
                    and 'resultaat na belasting' in label.lower()
                ):
                    target_row = row_idx
                    break
            if target_row is not None:
                break
        if target_row is None:
            return result

        def _sum_cols(cols: List[int]) -> Optional[float]:
            running = 0.0
            saw = False
            for col_idx in cols:
                v = bdo_sheet.cell(row=target_row, column=col_idx).value
                if isinstance(v, (int, float)):
                    running += float(v)
                    saw = True
            return -running if saw else None

        h_val = bdo_sheet.cell(row=target_row, column=data_col).value
        if isinstance(h_val, (int, float)) and float(h_val) != 0.0:
            result['ltm'] = -float(h_val)
        else:
            cols_for_sum = (
                list(range(3, q_cols[-1] + 1)) if q_cols else [3, 4, 5, 6, 7]
            )
            result['ltm'] = _sum_cols(cols_for_sum)

        result['quarter'] = _sum_cols(list(q_cols))
        return result
    finally:
        wb_data.close()


def diff_quarter_vs_ltm_rows(
    sheet,
    *,
    ltm_col: int,
    quarter_col: int,
    bdo_sheet_name: str,
    rules: Optional[dict] = None,
) -> List[Dict[str, Any]]:
    """
    Row-by-row AC vs AD BDO reference parity check (G rows vs H rows).
    """
    rules = _load_rules(rules)
    layout_cfg = rules.get('layout', {})
    bs_range = layout_cfg.get('balance_sheet_rows', [3, 19])
    pl_range = layout_cfg.get('profit_loss_rows', [23, 68])
    bank_range = layout_cfg.get('bank_rows', [107, 114])
    bank_total_row = rules.get('bank_total_row', bank_range[1])
    cash_row = layout_cfg.get('cash_proceeds_row', 50)
    identity = rules.get('identity_alignment', {})
    auth_row = identity.get('authoritative_row', 19)
    dep_row = identity.get('dependent_row', 68)

    interest_cfg = rules.get('interest_section', {})
    interest_ma_row = interest_cfg.get('management_row', 60)
    interest_skip_rows = {interest_ma_row, 61, 64}

    bdo_cfg = rules.get('bdo', {})
    q_cols_cfg = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
    last_q_col = q_cols_cfg[-1] if q_cols_cfg else 7
    bdo_qtr_letter = get_column_letter(last_q_col)
    bdo_data_letter = get_column_letter(bdo_cfg.get('data_column', 8))

    rows_of_interest = (
        list(range(bs_range[0], bs_range[1] + 1))
        + list(range(pl_range[0], pl_range[1] + 1))
        + list(range(bank_range[0], bank_range[1] + 1))
    )
    skip = {auth_row, dep_row, cash_row, bank_total_row} | interest_skip_rows

    bdo_qtr_re = re.compile(rf"'[^']*'!{re.escape(bdo_qtr_letter)}(\d+)")
    bdo_ltm_re = re.compile(rf"'[^']*'!{re.escape(bdo_data_letter)}(\d+)")
    bdo_quoted_re = re.compile(r"'([^']*)'!")
    bdo_col_swap_re = re.compile(
        rf"(?P<sheet>'[^']*'!)(?P<col>{re.escape(bdo_qtr_letter)})(?P<row>\d+)"
    )

    mismatches: List[Dict[str, Any]] = []
    for row_idx in rows_of_interest:
        if row_idx in skip:
            continue
        q_val = sheet.cell(row=row_idx, column=quarter_col).value
        l_val = sheet.cell(row=row_idx, column=ltm_col).value
        q_is_formula = isinstance(q_val, str) and q_val.startswith('=')
        l_is_formula = isinstance(l_val, str) and l_val.startswith('=')

        q_rows = (
            {int(m) for m in bdo_qtr_re.findall(q_val)} if q_is_formula else set()
        )
        l_rows = (
            {int(m) for m in bdo_ltm_re.findall(l_val)} if l_is_formula else set()
        )
        if not q_rows and not l_rows:
            continue
        if q_rows == l_rows:
            continue

        expected_ltm = None
        if q_is_formula:
            expected_ltm = bdo_quoted_re.sub(f"'{bdo_sheet_name}'!", q_val)
            expected_ltm = bdo_col_swap_re.sub(
                lambda m: (
                    f"{m.group('sheet')}{bdo_data_letter}{m.group('row')}"
                ),
                expected_ltm,
            )
        mismatches.append({
            'row': row_idx,
            'actual': l_val,
            'expected': expected_ltm,
            'quarter_bdo_rows': sorted(q_rows),
            'ltm_bdo_rows': sorted(l_rows),
        })
    return mismatches


def _resolve_ma_columns(
    workbook_path: str,
    summary_sheet_name: str,
) -> Tuple[Optional[Any], Optional[int], Optional[int], Optional[str]]:
    """Return (sheet, quarter_col, ltm_col, sheet_name) from one workbook open."""
    wb = openpyxl.load_workbook(workbook_path, data_only=False)
    try:
        sheet = (
            wb[summary_sheet_name]
            if summary_sheet_name in wb.sheetnames else None
        )
        if sheet is None:
            for sn in wb.sheetnames:
                if 'Management Cijfers' in sn:
                    sheet = wb[sn]
                    summary_sheet_name = sn
                    break
        if sheet is None:
            return None, None, None, None

        header_row = 22
        quarter_label = None
        if sheet.title and ' - ' in sheet.title:
            quarter_label = sheet.title.split(' - ', 1)[1].strip()

        quarter_col = None
        if quarter_label:
            quarter_col = find_column_by_header_label(
                sheet, quarter_label, header_row
            )
        if not quarter_col:
            layout = scan_summary_column_layout(sheet, header_row)
            quarter_col = layout.last_quarter_column
        ltm_col = None
        if quarter_col:
            ltm_col = find_ltm_column_near_quarter(
                sheet, quarter_col, header_row
            )
        if ltm_col is None:
            layout = scan_summary_column_layout(sheet, header_row)
            ltm_col = layout.ltm_column

        return sheet, quarter_col, ltm_col, summary_sheet_name
    finally:
        wb.close()


def build_reconciliation_report(
    workbook_path: str,
    summary_sheet_name: str,
    bdo_sheet_name: str,
    rules: Optional[dict] = None,
    tolerance: float = 1.0,
) -> Dict[str, Any]:
    """
    Authoritative post-AI reconciliation report from formula evaluation.

    Evaluates both quarter and LTM columns in one session, compares totals
    against BDO ground truth, and lists AC/AD formula parity mismatches.
    """
    rules = _load_rules(rules)
    identity = rules.get('identity_alignment', {})
    auth_row = identity.get('authoritative_row', 19)
    dep_row = identity.get('dependent_row', 68)

    report: Dict[str, Any] = {
        'passed': False,
        'source': 'formula_evaluator_post_ai',
        'messages': [],
        'auth_row': auth_row,
        'dep_row': dep_row,
        'evaluated_ltm': {},
        'evaluated_quarter': {},
        'ltm_row_mismatches': [],
        'bdo_ground_truth_ltm': None,
        'bdo_ground_truth_quarter': None,
        'ltm_col_letter': None,
        'quarter_col_letter': None,
    }

    evaluated = ManagementAccountsBuilder.evaluate_ma_columns_from_formulas(
        workbook_path,
        summary_sheet_name,
        bdo_sheet_name,
        rules=rules,
    )
    eval_ltm = evaluated.get('ltm') or {}
    eval_q = evaluated.get('quarter') or {}
    ltm_values = eval_ltm.get('values') or {}
    q_values = eval_q.get('values') or {}
    report['evaluated_ltm'] = ltm_values
    report['evaluated_quarter'] = q_values

    ltm_col = eval_ltm.get('column')
    quarter_col = eval_q.get('column')
    if ltm_col:
        report['ltm_col_letter'] = get_column_letter(ltm_col)
    if quarter_col:
        report['quarter_col_letter'] = get_column_letter(quarter_col)

    ltm_bs19 = ltm_values.get(auth_row)
    ltm_pl68 = ltm_values.get(dep_row)
    q_pl68 = q_values.get(dep_row)
    report['auth_row_value'] = ltm_bs19
    report['dep_row_value'] = ltm_pl68
    report['quarter_dep_row_value'] = q_pl68

    bdo_pair = read_bdo_ground_truth_pair(
        workbook_path, bdo_sheet_name, rules=rules
    )
    report['bdo_ground_truth_ltm'] = bdo_pair.get('ltm')
    report['bdo_ground_truth_quarter'] = bdo_pair.get('quarter')

    if ltm_col and quarter_col:
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
        try:
            sheet = (
                wb[summary_sheet_name]
                if summary_sheet_name in wb.sheetnames else None
            )
            if sheet is None:
                for sn in wb.sheetnames:
                    if 'Management Cijfers' in sn:
                        sheet = wb[sn]
                        break
            if sheet is not None:
                report['ltm_row_mismatches'] = diff_quarter_vs_ltm_rows(
                    sheet,
                    ltm_col=ltm_col,
                    quarter_col=quarter_col,
                    bdo_sheet_name=bdo_sheet_name,
                    rules=rules,
                )
        finally:
            wb.close()

    internal_passed = True
    internal_diff = None
    if isinstance(ltm_pl68, (int, float)) and isinstance(ltm_bs19, (int, float)):
        internal_diff = abs(float(ltm_pl68) - float(ltm_bs19))
        internal_passed = internal_diff <= tolerance

    ground_passed = True
    ground_diff = None
    bdo_ltm = bdo_pair.get('ltm')
    if (
        bdo_ltm is not None
        and isinstance(ltm_pl68, (int, float))
        and isinstance(ltm_bs19, (int, float))
    ):
        ground_diff = max(
            abs(float(ltm_pl68) - bdo_ltm),
            abs(float(ltm_bs19) - bdo_ltm),
        )
        ground_passed = ground_diff <= tolerance

    quarter_passed = True
    quarter_diff = None
    bdo_q = bdo_pair.get('quarter')
    if bdo_q is not None and isinstance(q_pl68, (int, float)):
        quarter_diff = abs(float(q_pl68) - bdo_q)
        quarter_passed = quarter_diff <= tolerance

    parity_passed = not bool(report['ltm_row_mismatches'])
    report['passed'] = (
        internal_passed
        and ground_passed
        and quarter_passed
        and parity_passed
    )
    report['internal_diff'] = internal_diff
    report['ground_diff'] = ground_diff
    report['quarter_diff'] = quarter_diff

    report['messages'] = format_reconciliation_messages(report, rules=rules)
    return report


def format_reconciliation_messages(
    report: Dict[str, Any],
    rules: Optional[dict] = None,
) -> List[str]:
    """User-facing warning lines from a reconciliation report."""
    rules = _load_rules(rules)
    identity = rules.get('identity_alignment', {})
    auth_row = report.get('auth_row') or identity.get('authoritative_row', 19)
    dep_row = report.get('dep_row') or identity.get('dependent_row', 68)
    ltm = report.get('ltm_col_letter') or 'AD'
    quarter = report.get('quarter_col_letter') or 'AC'
    tolerance = 1.0

    messages: List[str] = []
    ltm_bs19 = report.get('auth_row_value')
    ltm_pl68 = report.get('dep_row_value')
    q_pl68 = report.get('quarter_dep_row_value')
    bdo_ltm = report.get('bdo_ground_truth_ltm')
    bdo_q = report.get('bdo_ground_truth_quarter')

    messages.append(
        'Post-AI reconciliation (formula evaluator on saved workbook):'
    )
    messages.append(
        f"  {ltm}{auth_row} (Total Equity Movement, evaluated) = "
        f"{ltm_bs19:,.2f}"
        if isinstance(ltm_bs19, (int, float))
        else f"  {ltm}{auth_row} (Total Equity Movement) = (unresolved)"
    )
    messages.append(
        f"  {ltm}{dep_row} (Direct Result, evaluated) = "
        f"{ltm_pl68:,.2f}"
        if isinstance(ltm_pl68, (int, float))
        else f"  {ltm}{dep_row} (Direct Result) = (unresolved)"
    )

    if bdo_ltm is not None:
        messages.append(
            f"  BDO ground truth LTM (column H, sign-adjusted) = "
            f"{bdo_ltm:,.2f}"
        )
    if bdo_q is not None:
        messages.append(
            f"  BDO ground truth quarter (columns D–G, sign-adjusted) = "
            f"{bdo_q:,.2f}"
        )
        if isinstance(q_pl68, (int, float)):
            q_diff = report.get('quarter_diff')
            if q_diff is not None and q_diff > tolerance:
                messages.append(
                    f"  {quarter}{dep_row} vs quarter ground truth: "
                    f"evaluated={q_pl68:,.2f}, expected={bdo_q:,.2f}, "
                    f"diff={q_diff:,.2f}"
                )

    internal_diff = report.get('internal_diff')
    if (
        isinstance(ltm_bs19, (int, float))
        and isinstance(ltm_pl68, (int, float))
        and internal_diff is not None
        and internal_diff > tolerance
    ):
        messages.append(
            f"  {ltm}{auth_row} vs {ltm}{dep_row} differ by "
            f"{internal_diff:,.2f} (tolerance {tolerance:.2f})"
        )

    ground_diff = report.get('ground_diff')
    if bdo_ltm is not None and ground_diff is not None and ground_diff > tolerance:
        messages.append(
            f"  LTM column does not reconcile with BDO Resultaat na "
            f"belasting (max divergence {ground_diff:,.2f} > "
            f"tolerance {tolerance:.2f})"
        )

    mismatches = report.get('ltm_row_mismatches') or []
    if mismatches:
        preview = ', '.join(f"row {m['row']}" for m in mismatches[:12])
        extra = (
            f" (+{len(mismatches) - 12} more)"
            if len(mismatches) > 12 else ''
        )
        messages.append(
            f"  AC/AD BDO row parity mismatches ({len(mismatches)}): "
            f"{preview}{extra}"
        )

    if report.get('passed'):
        messages.append('  Reconciliation PASSED.')
    else:
        messages.append(
            '  Reconciliation FAILED — review upstream row BDO mappings '
            f'on column {quarter} (AD mirrors automatically).'
        )

    return messages
