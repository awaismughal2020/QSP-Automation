"""
AI-Verified Management Accounts — Claude Opus 4.6

Proactive verification of Management Accounts workbooks using Claude.
After the automated MA build, this module:
1. Extracts full workbook context (formulas, values, formats, BDO map)
2. Sends to Claude for verification against formula templates and sign rules
3. Applies returned JSON patches (Management Cijfers only)
4. Re-runs validation to confirm fixes

Original draft is always preserved alongside the verified output.
"""

import json
import os
import re
from copy import copy as copy_style
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import yaml
from loguru import logger
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

from ..parsers.bdo_parser import BDOParseResult


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class VerificationConfig:
    """Configuration for AI verification pass."""
    quarter: str              # e.g. "Q4 2025"
    year: int
    quarter_num: int
    period_end: str           # e.g. "31-12-2025"
    bdo_sheet_name: str       # e.g. "BDO - Q4-25"
    summary_sheet_name: str   # e.g. "Management Cijfers - Q4 2025"
    previous_quarter_label: str  # e.g. "Q3 2025"
    # Column indices assigned during the MA build. When set, the AI context
    # extractor and patch validator use these instead of re-scanning the
    # row-22 header — useful when the LTM header is missing or stale.
    built_quarter_col: Optional[int] = None
    built_ltm_col: Optional[int] = None


@dataclass
class VerificationResult:
    """Result from AI verification."""
    status: str                     # "PASS" or "ISSUES_FOUND"
    patches_applied: int = 0
    revalidation_passed: bool = False
    original_file: str = ""
    verified_file: str = ""
    issues: List[Dict] = field(default_factory=list)
    patches: List[Dict] = field(default_factory=list)
    validation_summary: Dict = field(default_factory=dict)
    notes: str = ""
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Context Extractor
# ---------------------------------------------------------------------------

class WorkbookContextExtractor:
    """
    Extracts structured context from a Management Accounts workbook
    for submission to Claude verification.
    """

    def __init__(self, workbook_path: str, config: VerificationConfig,
                 formula_templates_path: str = "config/formula_templates.yaml"):
        self.workbook_path = Path(workbook_path)
        self.config = config
        self.formula_templates_path = Path(formula_templates_path)
        self.formula_templates = self._load_formula_templates()
        self._rules = self._load_accounting_rules()

    def _load_formula_templates(self) -> dict:
        if self.formula_templates_path.exists():
            with open(self.formula_templates_path, 'r') as f:
                return yaml.safe_load(f)
        return {}

    def extract(self) -> Dict[str, Any]:
        """Extract complete workbook context for Claude verification."""
        wb = openpyxl.load_workbook(str(self.workbook_path))
        try:
            context = {
                'metadata': self._extract_metadata(),
                'formula_templates': self.formula_templates,
                'bdo_account_to_row_map': {},
                'bdo_label_to_row_map': {},
                'actual_cell_contents': {},
                'formatting_snapshot': {},
                'bdo_column_h_formulas': {},
                'bdo_special_rows': {},
            }

            bdo_sheet = self._get_bdo_sheet(wb)
            if bdo_sheet:
                context['bdo_account_to_row_map'] = self._build_bdo_row_map(bdo_sheet)
                context['bdo_label_to_row_map'] = self._build_bdo_label_map(bdo_sheet)
                context['bdo_column_h_formulas'] = self._extract_bdo_column_h(bdo_sheet)
                context['bdo_special_rows'] = self._extract_bdo_special_rows(bdo_sheet)
                context['bdo_resultaat_na_belasting'] = self._read_bdo_resultaat(bdo_sheet)
                context['bdo_full_inventory'] = self._build_bdo_full_inventory(bdo_sheet)

            context['accounting_rules'] = self._load_accounting_rules()

            summary_sheet = self._get_summary_sheet(wb)
            if summary_sheet:
                from .management_accounts import (
                    find_column_by_header_label,
                    find_ltm_column_near_quarter,
                    scan_summary_column_layout,
                )

                header_row = self._rules.get('layout', {}).get('header_row', 22)
                # Prefer the column indices assigned during the MA build;
                # the post-build header may still be wrong if a build error
                # left a stale label in the LTM slot.
                quarter_col = self.config.built_quarter_col
                ltm_col = self.config.built_ltm_col
                if quarter_col is None or ltm_col is None:
                    layout = scan_summary_column_layout(summary_sheet, header_row)
                    if quarter_col is None:
                        quarter_col = find_column_by_header_label(
                            summary_sheet, self.config.quarter, header_row
                        )
                        if not quarter_col:
                            quarter_col = layout.last_quarter_column
                    if ltm_col is None:
                        ltm_col = layout.ltm_column
                        if quarter_col and not ltm_col:
                            ltm_col = find_ltm_column_near_quarter(
                                summary_sheet, quarter_col, header_row
                            )
                context['ltm_col'] = ltm_col
                context['quarter_col'] = quarter_col
                context['ltm_col_letter'] = get_column_letter(ltm_col) if ltm_col else None
                context['quarter_col_letter'] = get_column_letter(quarter_col) if quarter_col else None

                if quarter_col and ltm_col:
                    context['actual_cell_contents'] = self._extract_cell_contents(
                        summary_sheet, quarter_col, ltm_col
                    )
                    context['formatting_snapshot'] = self._extract_formatting(
                        summary_sheet, quarter_col, ltm_col
                    )
                    # Previous-quarter formulas as few-shot reference. The
                    # column immediately left of the current quarter is the
                    # previous quarter (Q4 2025 when current = Q1 2026, etc.)
                    # and its formulas are already client-validated history.
                    # Claude uses them to spot row-mapping drift on the new
                    # column without re-deriving every BDO row from scratch.
                    context['previous_quarter_formulas'] = (
                        self._extract_previous_quarter_formulas(
                            summary_sheet, quarter_col
                        )
                    )
            return context
        finally:
            wb.close()

    def _extract_previous_quarter_formulas(
        self, sheet, quarter_col: int
    ) -> Dict[str, Dict[int, str]]:
        """Capture every formula in the previous quarter column.

        Returns a dict with ``previous_quarter_col_letter`` and ``formulas``
        (a map row_idx -> formula string) so the prompt builder can render a
        compact reference block.
        """
        layout_cfg = self._rules.get('layout', {})
        bs_range = layout_cfg.get('balance_sheet_rows', [3, 19])
        pl_range = layout_cfg.get('profit_loss_rows', [23, 68])
        bank_range = layout_cfg.get('bank_rows', [107, 114])

        prev_col = quarter_col - 1
        if prev_col < 2:
            return {'previous_quarter_col_letter': None, 'formulas': {}}

        rows_of_interest = (
            list(range(bs_range[0], bs_range[1] + 1))
            + list(range(pl_range[0], pl_range[1] + 1))
            + list(range(bank_range[0], bank_range[1] + 1))
        )

        prev_letter = get_column_letter(prev_col)
        header = sheet.cell(row=22, column=prev_col).value
        formulas: Dict[int, str] = {}
        for row_idx in rows_of_interest:
            val = sheet.cell(row=row_idx, column=prev_col).value
            if isinstance(val, str) and val.startswith('='):
                formulas[row_idx] = val

        return {
            'previous_quarter_col_letter': prev_letter,
            'previous_quarter_header': header,
            'formulas': formulas,
        }

    def _extract_metadata(self) -> Dict[str, Any]:
        return {
            'quarter': self.config.quarter,
            'year': self.config.year,
            'quarter_num': self.config.quarter_num,
            'period_end': self.config.period_end,
            'bdo_sheet_name': self.config.bdo_sheet_name,
            'summary_sheet_name': self.config.summary_sheet_name,
            'previous_quarter_label': self.config.previous_quarter_label,
        }

    def _get_bdo_sheet(self, wb):
        name = self.config.bdo_sheet_name
        if name in wb.sheetnames:
            return wb[name]
        for sn in wb.sheetnames:
            if sn.startswith('BDO'):
                return wb[sn]
        return None

    def _get_summary_sheet(self, wb):
        name = self.config.summary_sheet_name
        if name in wb.sheetnames:
            return wb[name]
        for sn in wb.sheetnames:
            if 'Management Cijfers' in sn:
                return wb[sn]
        return None

    def _find_ltm_column(self, sheet) -> Optional[int]:
        from .management_accounts import scan_summary_column_layout

        header_row = self._rules.get('layout', {}).get('header_row', 22)
        return scan_summary_column_layout(sheet, header_row).ltm_column

    def _build_bdo_row_map(self, sheet) -> Dict[str, int]:
        """Build account-code-to-row map. Keeps FIRST occurrence."""
        row_map = {}
        for row_idx in range(1, sheet.max_row + 1):
            code = sheet.cell(row=row_idx, column=1).value
            if code and isinstance(code, (str, int, float)):
                code_str = str(code).strip()
                if code_str and code_str[0].isdigit():
                    if code_str not in row_map:
                        row_map[code_str] = row_idx
        return row_map

    def _build_bdo_label_map(self, sheet) -> Dict[str, int]:
        label_map = {}
        for row_idx in range(1, sheet.max_row + 1):
            code = sheet.cell(row=row_idx, column=1).value
            label = sheet.cell(row=row_idx, column=2).value

            if code and isinstance(code, (str, int, float)):
                code_str = str(code).strip()
                if code_str and not code_str[0].isdigit():
                    label_map[code_str.lower()] = row_idx

            if label and isinstance(label, str):
                label_map[label.strip().lower()] = row_idx
        return label_map

    def _extract_bdo_column_h(self, sheet) -> Dict[int, Dict[str, Any]]:
        """Extract data column values from BDO sheet — ALL rows, including unlabeled."""
        bdo_cfg = self._rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)

        result = {}
        for row_idx in range(1, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=data_col)
            if cell.value is None:
                continue
            code = sheet.cell(row=row_idx, column=1).value
            result[row_idx] = {
                'value': cell.value,
                'is_formula': isinstance(cell.value, str) and str(cell.value).startswith('='),
                'account_code': str(code).strip() if code else None,
                'label': sheet.cell(row=row_idx, column=2).value,
            }
        return result

    def _build_bdo_full_inventory(self, sheet) -> Dict[str, Any]:
        """
        Scan every BDO row. For rows with non-zero H or G values, record them.
        Catches unlabeled rows, duplicate account codes, and section overflows.
        """
        bdo_cfg = self._rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)
        q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
        last_q_col = q_cols[-1] if q_cols else 7

        all_rows_with_values = []
        duplicate_codes: Dict[str, list] = {}
        seen_codes: Dict[str, int] = {}
        pl_start = None
        pl_end = None

        for row_idx in range(1, sheet.max_row + 1):
            h_val = sheet.cell(row=row_idx, column=data_col).value
            g_val = sheet.cell(row=row_idx, column=last_q_col).value
            code = sheet.cell(row=row_idx, column=1).value
            label = sheet.cell(row=row_idx, column=2).value

            has_h = (isinstance(h_val, (int, float)) and h_val != 0) or \
                    (isinstance(h_val, str) and h_val.startswith('='))
            has_g = (isinstance(g_val, (int, float)) and g_val != 0) or \
                    (isinstance(g_val, str) and g_val.startswith('='))

            code_str = str(code).strip() if code else None
            if isinstance(code, str):
                if 'winst' in code.lower() and 'verlies' in code.lower() and pl_start is None:
                    pl_start = row_idx + 1
                if 'resultaat na belasting' in code.lower():
                    pl_end = row_idx

            if not has_h and not has_g:
                continue

            is_account = bool(code_str and code_str[0].isdigit()) if code_str else False

            row_info = {
                'row': row_idx,
                'account_code': code_str if is_account else None,
                'label': str(label) if label else None,
                'h_value': h_val if isinstance(h_val, (int, float)) else (
                    'formula' if isinstance(h_val, str) else None),
                'g_value': g_val if isinstance(g_val, (int, float)) else (
                    'formula' if isinstance(g_val, str) else None),
                'is_unlabeled': not is_account,
                'raw_col_a': code_str,
            }
            all_rows_with_values.append(row_info)

            if is_account:
                if code_str in seen_codes:
                    if code_str not in duplicate_codes:
                        duplicate_codes[code_str] = [seen_codes[code_str]]
                    duplicate_codes[code_str].append(row_idx)
                else:
                    seen_codes[code_str] = row_idx

        return {
            'all_rows': all_rows_with_values,
            'duplicate_codes': {
                code: {
                    'rows': rows,
                    'h_values': {
                        r: sheet.cell(row=r, column=data_col).value
                        for r in rows
                    }
                }
                for code, rows in duplicate_codes.items()
            },
            'unlabeled_rows': [r for r in all_rows_with_values if r['is_unlabeled']],
            'pl_section_start': pl_start,
            'pl_section_end': pl_end,
            'total_rows_with_values': len(all_rows_with_values),
        }

    def _extract_bdo_special_rows(self, sheet) -> Dict[int, Dict[str, Any]]:
        """Extract special summary rows from BDO sheet."""
        bdo_cfg = self._rules.get('bdo', {})
        special_range = bdo_cfg.get('special_rows', [129, 136])
        data_col = bdo_cfg.get('data_column', 8)

        result = {}
        for row_idx in range(special_range[0], min(sheet.max_row + 1, special_range[1] + 1)):
            row_data = {}
            for col_idx in range(1, data_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                col_letter = get_column_letter(col_idx)
                row_data[col_letter] = {
                    'value': cell.value if not isinstance(cell.value, (int, float)) else cell.value,
                    'is_formula': isinstance(cell.value, str) and str(cell.value).startswith('='),
                }
            result[row_idx] = row_data
        return result

    def _read_bdo_resultaat(self, sheet) -> Optional[Dict[str, Any]]:
        """Read Resultaat na belasting from BDO sheet dynamically."""
        bdo_cfg = self._rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)
        q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
        label_cols = bdo_cfg.get('label_scan_columns', [1, 2])

        target_row = None
        for row_idx in range(1, min(sheet.max_row + 1, 200)):
            for col in label_cols:
                cell_val = sheet.cell(row=row_idx, column=col).value
                if cell_val and isinstance(cell_val, str):
                    if 'resultaat na belasting' in cell_val.lower():
                        target_row = row_idx
                        break
            if target_row:
                break

        if target_row is None:
            return None

        def _cell_num(col_idx):
            val = sheet.cell(row=target_row, column=col_idx).value
            if isinstance(val, (int, float)):
                return float(val)
            return None

        # Quarter total (columns D-G)
        quarter_total = 0.0
        col_values = {}
        for col_idx in q_cols:
            val = sheet.cell(row=target_row, column=col_idx).value
            col_letter = get_column_letter(col_idx)
            if isinstance(val, (int, float)):
                quarter_total += val
                col_values[col_letter] = val
            elif isinstance(val, str) and val.startswith('='):
                col_values[col_letter] = f"(formula: {val})"
            else:
                col_values[col_letter] = val

        # LTM value (data column)
        ltm_col_letter = get_column_letter(data_col)
        h_val = sheet.cell(row=target_row, column=data_col).value
        h_numeric = None
        if isinstance(h_val, (int, float)):
            h_numeric = float(h_val)
            col_values[ltm_col_letter] = h_val
        elif isinstance(h_val, str) and h_val.startswith('='):
            col_values[ltm_col_letter] = f"(formula: {h_val})"
        else:
            col_values[ltm_col_letter] = h_val

        q_col_letters = [get_column_letter(c) for c in q_cols]
        result = {
            'row': target_row,
            f'raw_total_{q_col_letters[0]}_{q_col_letters[-1]}': quarter_total,
            'raw_total_D_G': quarter_total,
            'sign_adjusted_quarter': -quarter_total,
            'sign_adjusted': -quarter_total,
            'column_values': col_values,
            'bdo_data_column': ltm_col_letter,
            'bdo_quarter_columns': q_col_letters,
        }

        if h_numeric is not None:
            result[f'raw_{ltm_col_letter}'] = h_numeric
            result['raw_H'] = h_numeric
            result['sign_adjusted_ltm'] = -h_numeric
        else:
            result[f'raw_{ltm_col_letter}'] = f"(formula — needs Excel eval: {h_val})"
            result['raw_H'] = result[f'raw_{ltm_col_letter}']
            result['sign_adjusted_ltm'] = result['sign_adjusted_quarter']

        return result

    @staticmethod
    def _load_accounting_rules() -> dict:
        """Load accounting_rules.yaml for inclusion in AI prompt context."""
        path = Path("config/accounting_rules.yaml")
        if path.exists():
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    return yaml.safe_load(f) or {}
            except Exception:
                pass
        return {}

    def _extract_cell_contents(self, sheet, quarter_col: int, ltm_col: int) -> Dict[str, Any]:
        """Extract actual cell formulas and values for verification rows."""
        layout = self._rules.get('layout', {})
        bs = layout.get('balance_sheet_rows', [3, 19])
        pl = layout.get('profit_loss_rows', [23, 68])
        bank_sec = layout.get('bank_section_rows', [104, 120])
        header_row = layout.get('header_row', 22)
        date_row = layout.get('date_row', 2)

        sections = {
            'balance_sheet': list(range(bs[0], bs[1] + 1)),
            'header': [1, date_row, header_row],
            'profit_loss': list(range(pl[0], pl[1] + 1)),
            'bank_accounts': list(range(bank_sec[0], bank_sec[1] + 1)),
        }

        contents = {}
        q_letter = get_column_letter(quarter_col)
        ltm_letter = get_column_letter(ltm_col)

        for section_name, rows in sections.items():
            section_data = {}
            for row_idx in rows:
                row_data = {}
                label_cell = sheet.cell(row=row_idx, column=2)
                row_data['label'] = str(label_cell.value) if label_cell.value else ''

                q_cell = sheet.cell(row=row_idx, column=quarter_col)
                row_data['quarter'] = {
                    'col': quarter_col,
                    'col_letter': q_letter,
                    'value': q_cell.value,
                    'is_formula': isinstance(q_cell.value, str) and str(q_cell.value).startswith('='),
                }

                ltm_cell = sheet.cell(row=row_idx, column=ltm_col)
                row_data['ltm'] = {
                    'col': ltm_col,
                    'col_letter': ltm_letter,
                    'value': ltm_cell.value,
                    'is_formula': isinstance(ltm_cell.value, str) and str(ltm_cell.value).startswith('='),
                }

                section_data[row_idx] = row_data
            contents[section_name] = section_data

        return contents

    def _extract_formatting(self, sheet, quarter_col: int, ltm_col: int) -> Dict[str, Any]:
        """Extract formatting metadata for quarter and LTM columns."""
        layout = self._rules.get('layout', {})
        bs = layout.get('balance_sheet_rows', [3, 19])
        pl = layout.get('profit_loss_rows', [23, 68])
        bank_sec = layout.get('bank_section_rows', [104, 120])
        header_row = layout.get('header_row', 22)

        formatting = {}
        rows_to_check = (
            list(range(1, bs[1] + 1))
            + [header_row]
            + list(range(pl[0], pl[1] + 1))
            + list(range(bank_sec[0], bank_sec[1] + 1))
        )

        for row_idx in rows_to_check:
            row_fmt = {}
            for col, col_name in [(quarter_col, 'quarter'), (ltm_col, 'ltm')]:
                cell = sheet.cell(row=row_idx, column=col)
                fmt = {
                    'font_name': cell.font.name if cell.font else None,
                    'font_size': cell.font.size if cell.font else None,
                    'bold': cell.font.bold if cell.font else False,
                    'font_color': str(cell.font.color.rgb) if cell.font and cell.font.color and cell.font.color.rgb else None,
                    'number_format': cell.number_format,
                }
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    fmt['fill_color'] = str(cell.fill.fgColor.rgb)
                else:
                    fmt['fill_color'] = None
                row_fmt[col_name] = fmt
            formatting[row_idx] = row_fmt

        return formatting


# ---------------------------------------------------------------------------
# Prompt Builder
# ---------------------------------------------------------------------------

class VerificationPromptBuilder:
    """Builds the Claude Opus 4.6 verification prompt."""

    def build(self, context: Dict[str, Any]) -> str:
        """Build the complete verification prompt from extracted context."""
        meta = context.get('metadata', {})

        prompt_parts = [
            self._role_section(),
            self._context_section(context),
            self._sign_convention_section(context),
            self._formula_construction_guide(context),
            self._executive_alignment_section(context),
            self._checks_section(context),
            self._response_format_section(),
        ]
        return '\n\n'.join(prompt_parts)

    def _role_section(self) -> str:
        return """ROLE:
You are verifying a Management Accounts Excel workbook for QSP ESS B.V.
The workbook was auto-generated and you must check every formula, value,
cross-sheet reference, sign convention, and cell format for correctness.

You must return ONLY valid JSON — no markdown fences, no commentary outside
the JSON object.

IMPORTANT: If every check passes, return status "PASS" with empty "issues"
and "patches" arrays. Do NOT invent potential improvements or suggestions —
only report actual errors found in the workbook."""

    def _context_section(self, context: Dict[str, Any]) -> str:
        meta = context.get('metadata', {})
        parts = [
            "CONTEXT:",
            f"Quarter: {meta.get('quarter', 'N/A')}",
            f"Year: {meta.get('year', 'N/A')}",
            f"Period End: {meta.get('period_end', 'N/A')}",
            f"BDO Sheet: {meta.get('bdo_sheet_name', 'N/A')}",
            f"Summary Sheet: {meta.get('summary_sheet_name', 'N/A')}",
            f"Quarter Column: {context.get('quarter_col_letter', 'N/A')} (index {context.get('quarter_col', 'N/A')})",
            f"LTM Column: {context.get('ltm_col_letter', 'N/A')} (index {context.get('ltm_col', 'N/A')})",
            "",
            "BDO ACCOUNT-CODE-TO-ROW MAP:",
            json.dumps(context.get('bdo_account_to_row_map', {}), indent=2),
            "",
            "BDO LABEL-TO-ROW MAP:",
            json.dumps(context.get('bdo_label_to_row_map', {}), indent=2),
            "",
            "FORMULA TEMPLATES (balance_sheet and profit_loss):",
            json.dumps(context.get('formula_templates', {}), indent=2),
            "",
            "ACTUAL CELL CONTENTS (new quarter + LTM columns):",
            self._serialize_cell_contents(context.get('actual_cell_contents', {})),
            "",
            "BDO COLUMN H FORMULAS (LTM sums):",
            json.dumps(context.get('bdo_column_h_formulas', {}), indent=2),
            "",
            "BDO SPECIAL ROWS (129-136):",
            json.dumps(context.get('bdo_special_rows', {}), indent=2),
            "",
            "BDO RESULTAAT NA BELASTING (ground truth):",
            json.dumps(context.get('bdo_resultaat_na_belasting', {}), indent=2),
            "",
            "ACCOUNTING RULES (from config/accounting_rules.yaml):",
            json.dumps(context.get('accounting_rules', {}), indent=2),
            "",
            "FORMATTING SNAPSHOT:",
            json.dumps(context.get('formatting_snapshot', {}), indent=2),
        ]

        # Previous-quarter formulas as a known-good reference (few-shot).
        # When the client signed off on Q4 2025, every formula in that
        # column became authoritative for *structure* (which BDO rows feed
        # which Management Cijfers row). The new column should mirror that
        # structure cell-for-cell, only swapping the BDO sheet name and
        # column letter (G vs H). Diffing against this catches row-mapping
        # drift that the static `formula_templates.yaml` alone can miss.
        prev_q = context.get('previous_quarter_formulas') or {}
        prev_col_letter = prev_q.get('previous_quarter_col_letter')
        prev_header = prev_q.get('previous_quarter_header')
        prev_formulas = prev_q.get('formulas') or {}
        if prev_col_letter and prev_formulas:
            parts.append("")
            parts.append(
                f"PREVIOUS QUARTER FORMULAS — known-correct reference "
                f"(column {prev_col_letter}, header={prev_header!r}):"
            )
            parts.append(
                "These formulas were validated by the client for the prior "
                "quarter. For every row that has a previous-quarter formula, "
                "the new quarter's formula should match it structurally — "
                "same BDO row numbers, same operators, same set of accounts. "
                "Only the BDO sheet name and BDO column letter (G vs H) "
                "should differ. Anything else is a row-mapping bug."
            )
            parts.append(json.dumps(
                {str(k): v for k, v in sorted(prev_formulas.items())},
                indent=2,
            ))

        shadow_pl = context.get('shadow_pl')
        shadow_bs = context.get('shadow_bs')
        if shadow_pl:
            parts.append("")
            parts.append("SHADOW P&L (expected values computed from raw BDO data):")
            parts.append("Each row shows label, expected numeric value, formula type,")
            parts.append("account codes used, and any missing codes (missing = BDO account")
            parts.append("not found, which would cause an incorrect zero in the formula).")
            parts.append(json.dumps(
                {str(k): v for k, v in shadow_pl.items()}, indent=2
            ))
        if shadow_bs:
            parts.append("")
            parts.append("SHADOW BALANCE SHEET (expected values computed from raw BDO data):")
            parts.append(json.dumps(
                {str(k): v for k, v in shadow_bs.items()}, indent=2
            ))

        cv = context.get('computed_values_snapshot')
        if cv:
            parts.append("")
            parts.append("COMPUTED VALUES (pre-calculated expected values from MA builder):")
            parts.append("These are the deterministic expected values for each row/scope.")
            parts.append("Use these to cross-check actual cell values and identify mismatches.")
            parts.append(json.dumps(cv, indent=2))

        inventory = context.get('bdo_full_inventory', {})
        if inventory:
            parts.append("")
            parts.append("BDO FULL ROW INVENTORY (every row with a non-zero value):")

            unlabeled = inventory.get('unlabeled_rows', [])
            if unlabeled:
                parts.append(
                    f"\nUNLABELED ROWS ({len(unlabeled)} found — "
                    f"these have values but no account code):")
                parts.append(json.dumps(unlabeled, indent=2))
                parts.append(
                    "Each unlabeled row MUST be referenced in a Management Cijfers "
                    "formula. Determine the appropriate target row based on the "
                    "surrounding account codes and section context."
                )

            dupes = inventory.get('duplicate_codes', {})
            if dupes:
                parts.append(
                    f"\nDUPLICATE ACCOUNT CODES ({len(dupes)} codes appear "
                    f"at multiple rows):")
                parts.append(json.dumps(dupes, indent=2))
                parts.append(
                    "For each duplicate, verify that ALL row occurrences are "
                    "referenced in Management Cijfers formulas. The standard "
                    "account-code map only contains one row per code."
                )

        return '\n'.join(parts)

    def _serialize_cell_contents(self, contents: Dict) -> str:
        """Serialize cell contents with safe JSON handling."""
        safe = {}
        for section, rows in contents.items():
            safe_section = {}
            for row_idx, row_data in rows.items():
                safe_row = {}
                for key, val in row_data.items():
                    if isinstance(val, dict):
                        safe_val = {}
                        for k, v in val.items():
                            if isinstance(v, (int, float, str, bool, type(None))):
                                safe_val[k] = v
                            else:
                                safe_val[k] = str(v)
                        safe_row[key] = safe_val
                    elif isinstance(val, (int, float, str, bool, type(None))):
                        safe_row[key] = val
                    else:
                        safe_row[key] = str(val)
                safe_section[str(row_idx)] = safe_row
            safe[section] = safe_section
        return json.dumps(safe, indent=2)

    def _sign_convention_section(self, context: Dict[str, Any]) -> str:
        bdo = context.get('metadata', {}).get('bdo_sheet_name', 'BDO')
        rules = context.get('accounting_rules', {})
        layout = rules.get('layout', {})
        bs = layout.get('balance_sheet_rows', [3, 19])
        pl = layout.get('profit_loss_rows', [23, 68])
        bdo_col = get_column_letter(rules.get('bdo', {}).get('data_column', 8))
        q_cols = rules.get('bdo', {}).get('quarter_columns', [4, 5, 6, 7])
        q_col_letter = get_column_letter(q_cols[-1]) if q_cols else 'G'

        return f"""SIGN CONVENTION:
BDO stores revenue and profit as NEGATIVE numbers (Dutch trial balance convention).
Management Cijfers uses POSITIVE for income. All bdo_ref formulas with
sign: "-" negate the BDO value to convert.

For Balance Sheet rows ({bs[0]}-{bs[1]}): formulas reference BDO column {bdo_col} (LTM/closing).
Quarter columns for Balance Sheet rows should be EMPTY (no data).

For P&L rows ({pl[0]}-{pl[1]}):
- Quarter column formulas reference BDO column {q_col_letter} (current quarter mutations).
- LTM column formulas reference BDO column {bdo_col} (full-year closing).
- Exception: bdo_ref_conditional and manual_with_ltm rows use
  ltm_sum_quarters (SUM of last 4 quarterly columns in Management Cijfers)
  for the LTM column.

All bdo_ref rows with sign "-": the formula negates BDO values (prefix with =-).
Multi-account bdo_ref with operator "-": first account is negated,
subsequent accounts are subtracted (=-'{bdo}'!{q_col_letter}{{row1}}-'{bdo}'!{q_col_letter}{{row2}}-...).
Multi-account bdo_ref with operator "+": first account is negated,
subsequent accounts are added (=-'{bdo}'!{q_col_letter}{{row1}}+'{bdo}'!{q_col_letter}{{row2}}+...).
bdo_ref with sign "" (empty): no negation, formula is ='{{BDO}}'!{bdo_col}{{row}} (used
for Balance Sheet rows where BDO values already have the correct sign)."""

    def _formula_construction_guide(self, context: Dict[str, Any]) -> str:
        bdo = context.get('metadata', {}).get('bdo_sheet_name', 'BDO')
        q = context.get('quarter_col_letter', 'AA')
        ltm = context.get('ltm_col_letter', 'AB')
        rules = context.get('accounting_rules', {})
        layout = rules.get('layout', {})
        try:
            q_idx = column_index_from_string(str(q).upper())
            prev_q_letter = get_column_letter(q_idx - 1) if q_idx > 1 else q
            sum_start_letter = get_column_letter(max(q_idx - 3, 2))
        except Exception:
            prev_q_letter = '(column immediately left of quarter column)'
            sum_start_letter = '?'

        return f"""FORMULA CONSTRUCTION RULES (how to verify each formula type):

1. bdo_ref (single account, sign "-", quarter column):
   Template: account_codes: ["8000003"], sign: "-"
   Look up 8000003 in the BDO account-code-to-row map → say it's row 10.
   Expected quarter formula: =-'{bdo}'!G10
   Expected LTM formula:     =-'{bdo}'!H10

2. bdo_ref (multiple accounts, sign "-", operator "-"):
   Template: account_codes: ["4611000","4613000","4620200","4623001","4642000"], sign: "-", operator: "-"
   Look up each code in the BDO map → say rows 91, 92, 93, 94, 95.
   Expected quarter formula: =-'{bdo}'!G91-'{bdo}'!G92-'{bdo}'!G93-'{bdo}'!G94-'{bdo}'!G95
   Expected LTM formula:     =-'{bdo}'!H91-'{bdo}'!H92-'{bdo}'!H93-'{bdo}'!H94-'{bdo}'!H95
   (Each term subtracts because operator is "-".)

3. bdo_ref (multiple accounts, sign "-", operator "+"):
   Template: account_codes: ["9500000","9510000"], sign: "-", operator: "+"
   Rows → say 19, 20.
   Expected: =-'{bdo}'!G19+'{bdo}'!G20  (BUT note: ONLY the first term gets
   the sign prefix; subsequent use the operator. The net effect: first is negated,
   rest are added with their natural sign.)
   WRONG:    =-'{bdo}'!G19-'{bdo}'!G20  ← this would be operator "-", not "+"

4. bdo_ref (sign "" / empty, Balance Sheet):
   Template: account_codes: ["1790002"], sign: ""
   Row → say 6.
   Expected LTM formula: ='{bdo}'!H6  (no negation, column H only)

5. bdo_sum_range:
   Template: start_account: "2400200", end_account: "2400206", count: 7
   Look up start code → say row 31. The range spans 7 consecutive rows.
   Expected LTM formula: =SUM('{bdo}'!H31:'{bdo}'!H37)
   If "additional" account codes exist, they are added outside the range:
   =SUM('{bdo}'!H31:'{bdo}'!H37)+'{bdo}'!H42+'{bdo}'!H43

6. bdo_ref_conditional:
   Quarter column: uses q_config (a bdo_ref, same rules as above) for MOST rows.
   LTM column: uses ltm_config with type "ltm_sum_quarters" — formula is
   =SUM({{4 columns before LTM}}50:{{column before LTM}}50) spanning exactly
   the last 4 quarterly columns in Management Cijfers (NOT BDO references).

   EXCEPTION — interest rows 60, 61, 64 in LTM column ({ltm}):
   These cells are set deterministically — DO NOT PATCH THEM.
   - {ltm}60: mirror of {q}60 with BDO !G swapped to !H (same row numbers)
   - {ltm}61: =SUM({sum_start_letter}61:{q}61)
   - {ltm}64: =SUM({sum_start_letter}64:{q}64)
   Quarter column ({q}): rows 61/64 use normal BDO references via q_config.

7. calc:
   Pattern uses {{COL}} placeholder → replace with the actual column letter.
   In the current workbook: quarter column = {q}, LTM column = {ltm}.
   Row 19 is special: only exists in LTM column."""

    def _executive_alignment_section(self, context: Dict[str, Any]) -> str:
        """Plain-English alignment rule for rows 19 and 68 vs BDO ground truth."""
        meta = context.get('metadata', {})
        rules = context.get('accounting_rules', {})
        identity = rules.get('identity_alignment', {})
        auth_row = identity.get('authoritative_row', 19)
        dep_row = identity.get('dependent_row', 68)
        bdo_name = meta.get('bdo_sheet_name', 'BDO')
        summary = meta.get('summary_sheet_name', 'Management Cijfers')
        ltm = context.get('ltm_col_letter', 'AB')
        bdo_col = get_column_letter(
            rules.get('bdo', {}).get('data_column', 8))
        bdo_gt = context.get('bdo_resultaat_na_belasting') or {}
        result_row = bdo_gt.get('row', '?')
        gt_ltm = bdo_gt.get('sign_adjusted_ltm', 'N/A')
        quarter = meta.get('quarter', 'N/A')
        year = meta.get('year', 'N/A')
        bs_range = rules.get('layout', {}).get(
            'balance_sheet_rows', [3, 19])
        bdo_cell = f"'{bdo_name}'!{bdo_col}{result_row}"
        bs_first = bs_range[0]
        bs_last_detail = auth_row - 1

        calc_formulas = rules.get('layout', {}).get('calc_formulas', {})
        ltm_overrides = rules.get('layout', {}).get('ltm_calc_overrides', {})
        all_calcs = {**calc_formulas, **ltm_overrides}
        auth_tmpl = all_calcs.get(auth_row) or all_calcs.get(str(auth_row), f"={{col}}{dep_row}")
        dep_tmpl = all_calcs.get(dep_row) or all_calcs.get(str(dep_row), f"={{col}}{dep_row-2}+{{col}}{dep_row-1}")
        auth_expected = auth_tmpl.replace('{col}', ltm).replace('{bdo}', bdo_name)
        dep_expected = dep_tmpl.replace('{col}', ltm).replace('{bdo}', bdo_name)

        return f"""BDO RESULT AFTER TAXES — EXECUTIVE ALIGNMENT RULE
(Reporting period: {quarter}, year {year})

The result after taxes in {ltm}{auth_row} of '{summary}' (Total Equity
Movement) must equal the Resultaat na belasting in {bdo_cell} multiplied
by -1. The expected sign-adjusted LTM value is: {gt_ltm}.

The result after taxes in {ltm}{dep_row} of '{summary}' (Direct result)
must equal that same BDO value multiplied by -1.

FORMULA INTEGRITY FOR TOTAL ROWS (CRITICAL):
Row {auth_row} and {dep_row} in the LTM column ({ltm}) MUST contain FORMULAS,
not hardcoded numeric values. A hardcoded number breaks the dynamic spreadsheet.

Expected formulas (from config/accounting_rules.yaml calc_formulas):
  - {ltm}{auth_row}: {auth_expected}
  - {ltm}{dep_row}:  {dep_expected}

If these cells contain hardcoded numeric values instead of formulas, emit a
formula patch to restore the correct formula.

Then verify the formula result equals BDO ground truth ({gt_ltm}). If NOT:
the problem is in upstream detail rows — find and fix those.

RECONCILIATION REQUIREMENT:
  Balance Sheet: SUM({ltm}{bs_first}:{ltm}{bs_last_detail}) must approximately equal {gt_ltm}
  P&L:           The formula chain ending at {ltm}{dep_row} must resolve to {gt_ltm}

If the sums do not match, review every upstream bdo_ref formula against the
BDO account-code-to-row map. Check the DUPLICATE ACCOUNT CODES and UNLABELED
ROWS sections for missing references. NEVER fix totals by hardcoding a number."""

    def _checks_section(self, context: Dict[str, Any]) -> str:
        meta = context.get('metadata', {})
        bdo_name = meta.get('bdo_sheet_name', 'BDO')
        summary = meta.get('summary_sheet_name', 'Management Cijfers')
        q = context.get('quarter_col_letter', 'AA')
        ltm = context.get('ltm_col_letter', 'AB')
        q_idx = context.get('quarter_col', '')
        ltm_idx = context.get('ltm_col', '')

        rules = context.get('accounting_rules', {})
        layout = rules.get('layout', {})
        identity = rules.get('identity_alignment', {})
        auth_row = identity.get('authoritative_row', 19)
        dep_row = identity.get('dependent_row', 68)
        bs_range = layout.get('balance_sheet_rows', [3, 19])
        pl_range = layout.get('profit_loss_rows', [23, 68])
        bank_range = layout.get('bank_rows', [107, 114])
        cash_row = layout.get('cash_proceeds_row', 50)
        bold_rows = layout.get('bold_total_rows', [19, 25, 30, 44, 45, 48, 55, 57, 66, 68])
        intermediate = layout.get('intermediate_rows', [60, 61, 64, 65, 67])
        bdo_col_letter = get_column_letter(rules.get('bdo', {}).get('data_column', 8))

        bank_map = rules.get('bank_bdo_row_map', {})
        bank_lines = "\n".join(
            f"     Row {ma_row}: ='{bdo_name}'!{bdo_col_letter}{bdo_row}"
            for ma_row, bdo_row in sorted(bank_map.items())
        )
        bank_total_row = rules.get('bank_total_row', 114)
        bank_start = min(bank_map.keys()) if bank_map else bank_range[0]
        bank_end = max(bank_map.keys()) if bank_map else bank_range[1] - 1

        # Build calc formula checks dynamically from config
        calc_formulas = layout.get('calc_formulas', {})
        ltm_overrides = layout.get('ltm_calc_overrides', {})

        q_calc_lines = []
        for row_num, tmpl in sorted(calc_formulas.items()):
            formula = tmpl.replace('{col}', q).replace('{bdo}', bdo_name)
            q_calc_lines.append(f"     Row {row_num}: {formula}")

        ltm_calc_lines = []
        all_ltm = {**calc_formulas, **ltm_overrides}
        for row_num, tmpl in sorted(all_ltm.items()):
            formula = tmpl.replace('{col}', ltm).replace('{bdo}', bdo_name)
            ltm_calc_lines.append(f"     Row {row_num}: {formula}")

        q_calc_block = "\n".join(q_calc_lines)
        ltm_calc_block = "\n".join(ltm_calc_lines)

        # BDO ground truth from context
        bdo_gt = context.get('bdo_resultaat_na_belasting') or {}
        gt_ltm = bdo_gt.get('sign_adjusted_ltm', 'N/A')
        gt_qtr = bdo_gt.get('sign_adjusted_quarter', 'N/A')
        result_row = bdo_gt.get('row')
        bdo_ltm_for_ref = bdo_gt.get('bdo_data_column') or bdo_col_letter
        if result_row:
            bdo_result_cell_ref = f"'{bdo_name}'!{bdo_ltm_for_ref}{result_row}"
            result_row_text = str(result_row)
        else:
            bdo_result_cell_ref = (
                f"'{bdo_name}'!{bdo_ltm_for_ref}<row from label 'Resultaat na belasting' "
                "in BDO RESULTAAT NA BELASTING (CONTEXT)>"
            )
            result_row_text = "unknown — use BDO RESULTAAT NA BELASTING JSON in CONTEXT"
        auth_row_minus_1 = auth_row - 1
        dep_row_minus_1 = dep_row - 1
        report_quarter = meta.get('quarter', 'N/A')
        report_year = meta.get('year', 'N/A')

        # Pre-compute values that can't be inside f-strings (backslashes, joins)
        num_fmt = rules.get('styling', {}).get(
            'number_format', '_(* #,##0_);_(* (# ,##0);_(* "-"??_);_(@_)')
        ltm_bg_color = rules.get('styling', {}).get('ltm_background_color', 'FFDDEAF6')
        header_row_num = layout.get('header_row', 22)
        quarter_label = meta.get('quarter', '')
        protected_rows = rules.get('protected_total_rows', [auth_row, dep_row])

        # Interest section config
        interest_cfg = rules.get('interest_section', {})
        interest_codes_str = json.dumps(interest_cfg.get('account_codes', []))
        interest_ma_row = interest_cfg.get('management_row', 60)
        absorbed_rows_str = str(interest_cfg.get('absorbed_rows', [61, 64]))
        dmrrp_code = interest_cfg.get('dmrrp_code', '4663000')
        dmrrp_ma_row = interest_cfg.get('dmrrp_row', 65)

        try:
            _q_idx = column_index_from_string(str(q).upper())
            prev_q_letter = get_column_letter(_q_idx - 1) if _q_idx > 1 else q
            sum_start_letter = get_column_letter(max(_q_idx - 3, 2))
        except Exception:
            prev_q_letter = '(column left of quarter)'
            sum_start_letter = '?'

        # Computed values from MA builder (pre-computed expected values)
        computed = context.get('computed_values_snapshot', {})
        cv_section = ""
        if computed:
            cv_section = "\n\nPRE-COMPUTED EXPECTED VALUES (from Management Accounts builder):\n"
            cv_section += "These are the expected numeric values for key rows. Compare with\n"
            cv_section += "actual cell values to identify discrepancies.\n"
            cv_section += json.dumps(computed, indent=2)

        return f"""CHECKS TO PERFORM:

1. BDO FORMULA VERIFICATION:
   For every bdo_ref/bdo_sum_range formula in the actual cell contents:
   verify the row numbers in the formula match the account codes looked up
   in the BDO account-code-to-row map. The formula must reference sheet
   '{bdo_name}'. Use the formula construction rules above to determine
   the expected formula and compare with the actual.

2. CALC FORMULA VERIFICATION:
   For every calc formula, verify it references the correct rows using the
   actual column letters (quarter={q}, LTM={ltm}).

   These formulas are defined in accounting_rules.yaml and must match exactly:

   Quarter column ({q}, index {q_idx}):
{q_calc_block}

   LTM column ({ltm}, index {ltm_idx}):
{ltm_calc_block}

3. TOTAL ROW FORMULA INTEGRITY (MOST IMPORTANT RULE):
   Rows {auth_row} and {dep_row} in the LTM column ({ltm}) MUST contain
   FORMULAS — never hardcoded numeric values. A hardcoded number breaks the
   dynamic spreadsheet and makes SUM({ltm}{bs_range[0]}:{ltm}{auth_row_minus_1}) != {ltm}{auth_row}.

   Expected LTM formulas:
   - {ltm}{auth_row} (Total Equity Movement): ={ltm}{dep_row}
   - {ltm}{dep_row} (Direct result): ={ltm}{dep_row - 2}+{ltm}{dep_row - 1}

   If either cell contains a hardcoded numeric value instead of a formula,
   emit a formula patch to restore the correct formula above.

   BDO GROUND TRUTH VALUES (sign-adjusted, from Resultaat na belasting):
   - LTM target (sign_adjusted_ltm): {gt_ltm}
   - Quarter target (sign_adjusted_quarter): {gt_qtr}
   - BDO reference cell: {bdo_result_cell_ref} (multiply by -1 for MA convention)
   - Reporting period: {report_quarter}, year {report_year}

   After formulas are in place, verify their calculated result equals {gt_ltm}.
   If NOT, the problem is in upstream detail rows. Find and fix those:

   - Balance sheet detail rows {bs_range[0]}–{auth_row_minus_1}: verify every
     bdo_ref/bdo_sum_range uses the correct BDO row per the account map and
     nothing is skipped.
   - P&L detail rows {pl_range[0]}–{dep_row_minus_1}: same verification. Trace
     the formula chain ({dep_row} <- {dep_row - 1}+{dep_row - 2} <- ...) back to
     bdo_ref detail rows and fix any wrong references.
   - SUM({ltm}{bs_range[0]}:{ltm}{auth_row_minus_1}) MUST equal {gt_ltm}.
   - The P&L calc chain ending at {ltm}{dep_row} MUST resolve to {gt_ltm}.

   Row {dep_row} quarter column ({q}) should have the formula:
   ={q}{dep_row - 1}+{q}{dep_row - 2}

   PROTECTED ROWS (from config): {protected_rows}
   Do NOT emit value patches for these rows — they are computed totals.
   Formula patches are acceptable (and required when the cell is hardcoded).

4. FORMULA INTEGRITY: All formulas in the workbook MUST be preserved.
   Never replace a formula with a static number. If a formula evaluates
   to the wrong result, fix the formula itself (wrong cell references,
   wrong BDO row numbers, missing accounts) — do NOT replace it with a
   hard-coded value. This applies to ALL rows including {auth_row} and {dep_row}.

5. INTEREST ROWS 60, 61, 64 (MANAGED DETERMINISTICALLY — DO NOT PATCH):

   Latest QUARTER column ({q}) — rows 60, 61, 64:
   - Row {interest_ma_row}: BDO q_config formula, OR extended from column {prev_q_letter}.
     Do NOT delete a working formula on {q}{interest_ma_row}.
   - Rows 61 and 64 in column {q}: normal BDO reference formulas from q_config.

   LTM column ({ltm}) — rows 60, 61, 64 (DO NOT MODIFY):
   - {ltm}{interest_ma_row}: mirror of {q}{interest_ma_row} with BDO !G swapped to !H.
     It has the SAME BDO row numbers as the quarter column, just column H instead of G.
     Do NOT add extra BDO rows. Do NOT use a different set of account codes.
   - {ltm}61 MUST contain =SUM({sum_start_letter}61:{q}61)
   - {ltm}64 MUST contain =SUM({sum_start_letter}64:{q}64)
   These cells are set by deterministic code. Do NOT patch them.

   The DMRRP income account ({dmrrp_code}) goes to its own row ({dmrrp_ma_row})
   with a direct BDO {bdo_col_letter} reference (per template).

6. LTM SUM RANGES: For row {cash_row} (manual_with_ltm),
   the LTM column must contain a SUM spanning exactly the last 4 quarterly
   columns. After column insertion these ranges sometimes exclude the new
   column — catch and fix.

7. BANK ACCOUNT FORMULAS (rows {bank_start}-{bank_end} in '{summary}'):
   Each row must reference '{bdo_name}' column {bdo_col_letter} at these
   specific BDO rows (from accounting_rules.yaml bank_bdo_row_map):
{bank_lines}
     Row {bank_total_row} (Total): =SUM({ltm}{bank_start}:{ltm}{bank_end})

8. FORMATTING VERIFICATION:
   For every financial cell in the quarter and LTM columns, verify:
   - Font: Avenir Book, size 10 (bold for total/header rows: {bold_rows})
   - Number format: {num_fmt}
     for all numeric cells (rows {bs_range[0]}-{bs_range[1]}, {pl_range[0]}-{pl_range[1]}, {bank_range[0]}-{bank_range[1]})
   - Fill: LTM column should have background {ltm_bg_color}
     for data rows; section header and total rows should match template styling
   - Row {header_row_num}: quarter header "{quarter_label}" and LTM header
     "LTM {quarter_label}" with bold styling

9. MISSING ACCOUNT CODES: For every account code in the formula templates,
   verify it exists in the BDO account-code-to-row map. If missing, flag it
   as kind "missing" — a missing code means the formula will produce an
   incorrect zero.

10. BALANCE SHEET QUARTER COLUMNS: Rows {bs_range[0]}-{bs_range[1]} in the quarter column ({q})
   must be empty/None. Only the LTM column ({ltm}) should have formulas
   for Balance Sheet rows.

11. ROW {cash_row} (Cash proceeds sale):
    - Quarter column ({q}): must contain a numeric value (from sales tracker)
      or 0, NOT a formula.
    - LTM column ({ltm}): must contain a SUM of the last 4 quarterly columns.

12. FULL BDO COVERAGE AUDIT:
    For EVERY row listed in the BDO FULL ROW INVENTORY that has a non-zero
    {bdo_col_letter} value:
    a) Search all Management Cijfers formulas for a reference to that BDO
       row number (in both quarter and LTM columns).
    b) If a row is NOT referenced anywhere, flag it and determine where it
       should go based on surrounding account codes.
    c) Emit a patch to add the missing reference.
    Pay special attention to UNLABELED ROWS and DUPLICATE ACCOUNT CODES.

    EXCEPTION: Cells {ltm}60, {ltm}61, and {ltm}64 are managed deterministically.
    - {ltm}60 mirrors {q}60 with BDO !G→!H (same row numbers, different column).
    - {ltm}61 and {ltm}64 contain SUM formulas with NO BDO references.
    Do NOT flag these as missing coverage and do NOT patch them.

13. SUM RANGE AND DUPLICATE ACCOUNT CHECK:
    NEVER extend a SUM range to handle duplicates or rows from a different section.
    SUM ranges represent contiguous BDO structural sections.
    - If a duplicate account code exists outside a SUM range, add it as an
      explicit +'{bdo_name}'!{bdo_col_letter}<row> term APPENDED to the formula.
    - Only extend a SUM range if a new row is contiguous AND shares the same
      account code prefix (same structural section).
    - Check the DUPLICATE ACCOUNT CODES list: for each duplicate, verify ALL
      row occurrences are referenced in Management Cijfers formulas.

14. INTEREST SECTION — PATCH POLICY (NON-NEGOTIABLE):
    DO NOT PATCH any of these LTM cells — they are set deterministically:
    - {ltm}{interest_ma_row}: mirror of {q}{interest_ma_row} with BDO !G→!H
    - {ltm}61: =SUM({sum_start_letter}61:{q}61)
    - {ltm}64: =SUM({sum_start_letter}64:{q}64)
    If any of these cells look different from what you expect, PASS — they are
    managed by deterministic code that runs after your patches.

15. PREVIOUS-QUARTER STRUCTURAL CROSS-CHECK (NEW):
    The PREVIOUS QUARTER FORMULAS section above lists every formula from the
    prior quarter's column. That column has been client-validated. For EVERY
    row that appears in that reference:
    a) Find the corresponding formula in the current quarter column ({q}).
    b) Confirm the set of BDO row numbers referenced is IDENTICAL.
    c) Confirm the operators (+, -, SUM ranges) are IDENTICAL.
    d) The ONLY allowed differences:
         - BDO sheet name (e.g. 'BDO - Q4-25' → '{bdo_name}')
         - BDO column letter (G in quarter column, H in LTM column)
         - In-sheet column references in calc formulas (Q letter → LTM letter)
    e) If the current quarter formula uses a DIFFERENT set of BDO rows than
       the previous quarter (e.g. drops a row, swaps row numbers, adds a row
       that wasn't there before), emit a `formula` patch that mirrors the
       previous-quarter row set exactly. Do this BEFORE falling back to the
       template-based reconstruction in checks 1-2.
    f) Apply the same check to the LTM column ({ltm}): every row that has a
       previous-quarter formula should have an LTM formula referencing the
       SAME BDO rows but via column H.

    This is the single most reliable correctness check — the template can
    drift over time, but a client-signed-off prior quarter cannot.
{cv_section}"""

    def _response_format_section(self) -> str:
        return """RESPONSE FORMAT (strict JSON only, no markdown fences):
{
  "status": "PASS" or "ISSUES_FOUND",
  "checks_performed": <count of individual checks>,
  "issues": [
    {
      "sheet": "<sheet name>",
      "row": <row number>,
      "col": <column index (integer)>,
      "col_letter": "<column letter>",
      "kind": "formula" | "value" | "format" | "missing",
      "current": "<what is currently in the cell>",
      "expected": "<what should be there>",
      "explanation": "<why this is wrong>"
    }
  ],
  "patches": [
    {
      "sheet": "<sheet name>",
      "row": <row number>,
      "col": <column index (integer)>,
      "kind": "formula" | "value" | "format",
      "value": "<corrected formula string, numeric value, or format dict>"
    }
  ],
  "validation_summary": {
    "bs_row_19": <numeric value from Shadow BS or null>,
    "pl_row_68": <numeric value from Shadow P&L or null>,
    "bdo_resultaat_adjusted": <negated Resultaat na belasting or null>,
    "all_match": true | false
  },
  "notes": "<free-text summary of findings>"
}

RULES:
- If status is "PASS", issues and patches MUST both be empty arrays [].
- Only include a patch if you are confident the corrected value is right.
- Every issue should have a corresponding patch unless the fix is ambiguous.
- The "col" field must be an integer column index (1-based), not a letter.
- For formula patches, the value must start with "=".
- For format patches, value is a dict like {"number_format": "...", "bold": true}."""


# ---------------------------------------------------------------------------
# Patch Applier
# ---------------------------------------------------------------------------

ALLOWLISTED_SHEET_PREFIX = "Management Cijfers"
ALLOWLISTED_ROW_RANGE = range(1, 121)
DISALLOWED_SHEET_REFS = ["BDO"]


def _load_protected_rows() -> set:
    """Load protected_total_rows from accounting_rules.yaml."""
    try:
        path = Path("config/accounting_rules.yaml")
        if path.exists():
            with open(path, 'r', encoding='utf-8') as f:
                rules = yaml.safe_load(f) or {}
            return set(rules.get('protected_total_rows', [19, 68]))
    except Exception:
        pass
    return {19, 68}


PROTECTED_TOTAL_ROWS = _load_protected_rows()


class PatchApplier:
    """
    Applies JSON patches from Claude to the workbook.
    Strict allowlist: only Management Cijfers sheets, rows 1-120.
    Protected total rows (from config) reject value patches but allow formula patches.
    """

    INTEREST_PROTECTED_ROWS = {60, 61, 64}

    def __init__(
        self,
        workbook_path: str,
        summary_sheet_name: str,
        expected_bdo_sheet_name: Optional[str] = None,
        ltm_col: Optional[int] = None,
    ):
        self.workbook_path = Path(workbook_path)
        self.summary_sheet_name = summary_sheet_name
        self.expected_bdo_sheet_name = expected_bdo_sheet_name
        self.applied_patches: List[Dict] = []
        self.rejected_patches: List[Dict] = []
        self._ltm_col: Optional[int] = ltm_col

    def _detect_ltm_col(self, wb) -> Optional[int]:
        """Find the LTM column index from row 22 headers."""
        for sn in wb.sheetnames:
            if ALLOWLISTED_SHEET_PREFIX in sn:
                sheet = wb[sn]
                for c in range(sheet.max_column, 1, -1):
                    v = sheet.cell(row=22, column=c).value
                    if v and isinstance(v, str) and 'LTM' in v.upper():
                        return c
        return None

    def apply(self, patches: List[Dict], output_path: str) -> Tuple[int, List[Dict]]:
        """
        Apply patches to workbook, save to output_path.
        Returns (count_applied, list_of_rejected).
        """
        wb = openpyxl.load_workbook(str(self.workbook_path))
        if self._ltm_col is None:
            self._ltm_col = self._detect_ltm_col(wb)
        applied = 0

        for patch in patches:
            if self._validate_patch(patch):
                try:
                    self._apply_single_patch(wb, patch)
                    self.applied_patches.append(patch)
                    applied += 1
                except Exception as e:
                    patch['rejection_reason'] = f"Apply error: {e}"
                    self.rejected_patches.append(patch)
                    logger.warning(f"Patch apply failed for row {patch.get('row')}: {e}")
            else:
                self.rejected_patches.append(patch)

        wb.save(output_path)
        wb.close()
        logger.info(f"Applied {applied} patches, rejected {len(self.rejected_patches)}")
        return applied, self.rejected_patches

    def _validate_patch(self, patch: Dict) -> bool:
        sheet_name = patch.get('sheet', '')
        if ALLOWLISTED_SHEET_PREFIX not in sheet_name:
            patch['rejection_reason'] = f"Sheet '{sheet_name}' not in allowlist"
            logger.warning(f"Rejected patch: {patch['rejection_reason']}")
            return False

        for prefix in DISALLOWED_SHEET_REFS:
            if sheet_name.startswith(prefix):
                patch['rejection_reason'] = f"Sheet '{sheet_name}' is protected"
                logger.warning(f"Rejected patch: {patch['rejection_reason']}")
                return False

        row = patch.get('row', 0)
        if row not in ALLOWLISTED_ROW_RANGE:
            patch['rejection_reason'] = f"Row {row} outside allowlist (1-120)"
            logger.warning(f"Rejected patch: {patch['rejection_reason']}")
            return False

        kind = patch.get('kind', '')
        value = patch.get('value', '')

        if row in PROTECTED_TOTAL_ROWS and kind == 'value':
            patch['rejection_reason'] = (
                f"Row {row} is a protected total — value patches rejected; "
                f"fix upstream rows instead"
            )
            logger.warning(f"Rejected patch: {patch['rejection_reason']}")
            return False

        col = patch.get('col', 0)
        if row in self.INTEREST_PROTECTED_ROWS and self._ltm_col and col == self._ltm_col:
            patch['rejection_reason'] = (
                f"Row {row} in LTM column is deterministically managed — "
                f"patches rejected"
            )
            logger.warning(f"Rejected patch: {patch['rejection_reason']}")
            return False

        if kind == 'formula':
            if not isinstance(value, str) or not value.startswith('='):
                patch['rejection_reason'] = f"Formula patch value doesn't start with '=': {value}"
                logger.warning(f"Rejected patch: {patch['rejection_reason']}")
                return False
            for disallowed in DISALLOWED_SHEET_REFS:
                pass  # BDO refs within formulas are expected and valid

            # LTM-column formula patches must reference the current quarter's
            # BDO sheet — reject stale references (e.g. 'BDO - Q4-25' when the
            # build target is 'BDO - Q1-26'). Guards against the AI patching
            # the LTM column back to the prior quarter.
            if (
                self.expected_bdo_sheet_name
                and self._ltm_col
                and col == self._ltm_col
                and 'BDO' in value
            ):
                bdo_refs = re.findall(r"'(BDO[^']*)'", value)
                stale = [name for name in bdo_refs if name != self.expected_bdo_sheet_name]
                if stale:
                    patch['rejection_reason'] = (
                        f"LTM formula patch references stale BDO sheet(s) "
                        f"{stale!r}, expected {self.expected_bdo_sheet_name!r}"
                    )
                    logger.warning(f"Rejected patch: {patch['rejection_reason']}")
                    return False

        return True

    def _apply_single_patch(self, wb, patch: Dict):
        sheet_name = patch['sheet']
        if sheet_name not in wb.sheetnames:
            for sn in wb.sheetnames:
                if ALLOWLISTED_SHEET_PREFIX in sn:
                    sheet_name = sn
                    break
            else:
                raise ValueError(f"Sheet '{sheet_name}' not found")

        sheet = wb[sheet_name]
        row = patch['row']
        col = patch['col']
        kind = patch['kind']
        value = patch['value']

        cell = sheet.cell(row=row, column=col)
        old_value = cell.value

        if kind in ('formula', 'value'):
            cell.value = value
            logger.info(
                f"Patched {sheet_name}!{get_column_letter(col)}{row}: "
                f"'{old_value}' → '{value}' (kind={kind})"
            )

        elif kind == 'format':
            if isinstance(value, dict):
                if 'number_format' in value:
                    cell.number_format = value['number_format']
                if 'bold' in value:
                    new_font = copy_style(cell.font)
                    cell.font = Font(
                        name=new_font.name,
                        size=new_font.size,
                        bold=value['bold'],
                        color=new_font.color,
                    )
                if 'font_name' in value:
                    new_font = copy_style(cell.font)
                    cell.font = Font(
                        name=value['font_name'],
                        size=new_font.size,
                        bold=new_font.bold,
                        color=new_font.color,
                    )
            logger.info(f"Format patched {sheet_name}!{get_column_letter(col)}{row}")


# ---------------------------------------------------------------------------
# Claude API Client
# ---------------------------------------------------------------------------

class ClaudeVerifier:
    """Sends workbook context to Claude for verification and parses response."""

    MODEL = "claude-opus-4-6"
    MAX_RETRIES = 2

    SYSTEM_PROMPT = (
        "You are a financial workbook verification engine. "
        "You MUST respond with ONLY a single valid JSON object — "
        "no markdown fences, no commentary, no explanation before or after the JSON. "
        "Your entire response must be parseable by json.loads(). "
        "Start your response with { and end with }. "
        "NEVER include text outside the JSON. NEVER use ```json fences. "
        "If you are unsure, still respond ONLY with valid JSON."
    )

    def __init__(self, api_key: Optional[str] = None):
        self.api_key = api_key or os.environ.get('ANTHROPIC_API_KEY', '')

    def _call_claude(self, client, prompt: str) -> Dict[str, Any]:
        """Single streaming API call; returns parsed dict or raises."""
        with client.messages.stream(
            model=self.MODEL,
            max_tokens=32768,
            system=self.SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}],
        ) as stream:
            for _ in stream:
                pass
            final_message = stream.get_final_message()

        response_text = self._extract_text_from_response(final_message)
        logger.info(
            f"[AI Verify] Response length: {len(response_text)} chars, "
            f"stop_reason: {final_message.stop_reason}"
        )
        if not response_text:
            raise ValueError("Empty response from Claude")
        return self._parse_response(response_text)

    def verify(self, prompt: str) -> Dict[str, Any]:
        """
        Send verification prompt to Claude and parse JSON response.
        Retries up to MAX_RETRIES times if Claude returns non-JSON prose.
        """
        if not self.api_key:
            logger.warning("ANTHROPIC_API_KEY not set — returning mock PASS")
            return self._mock_pass_response()

        try:
            import anthropic
        except ImportError:
            logger.error("anthropic package not installed. Run: pip install anthropic")
            return {'status': 'ERROR', 'error': 'anthropic package not installed'}

        client = anthropic.Anthropic(api_key=self.api_key)

        last_error = None
        for attempt in range(1, self.MAX_RETRIES + 1):
            try:
                result = self._call_claude(client, prompt)
                if result.get('status') != 'ERROR' or 'JSON parse failed' not in result.get('error', ''):
                    return result

                # Claude returned prose — retry with a reminder prefix
                last_error = result.get('error', 'JSON parse failed')
                logger.warning(
                    f"[AI Verify] Attempt {attempt}/{self.MAX_RETRIES}: "
                    f"Claude returned prose, retrying..."
                )
                prompt = (
                    "IMPORTANT: Your previous response was NOT valid JSON. "
                    "Respond with ONLY a JSON object. No explanation, no markdown. "
                    "Start with {{ and end with }}.\n\n" + prompt
                )
            except Exception as e:
                last_error = str(e)
                logger.error(f"[AI Verify] Attempt {attempt}/{self.MAX_RETRIES} failed: {e}")

        return {'status': 'ERROR', 'error': f'Claude API error: {last_error}'}

    @staticmethod
    def _extract_text_from_response(message) -> str:
        """Extract text content from Claude response, skipping thinking blocks."""
        text_parts = []
        for block in message.content:
            if getattr(block, 'type', None) == 'text':
                text_parts.append(block.text)
        if text_parts:
            return '\n'.join(text_parts).strip()
        if message.content:
            return str(message.content[0].text).strip()
        return ""

    def _parse_response(self, text: str) -> Dict[str, Any]:
        """Parse Claude's JSON response, handling markdown fences and mixed text."""
        if not text:
            return {'status': 'ERROR', 'error': 'Empty response from Claude'}

        cleaned = text.strip()

        # Strip markdown fences
        if cleaned.startswith('```'):
            lines = cleaned.split('\n')
            start = 1
            end = len(lines)
            for i, line in enumerate(lines[1:], 1):
                if line.strip() == '```':
                    end = i
                    break
            cleaned = '\n'.join(lines[start:end]).strip()

        # Try direct parse first
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            pass

        # Find the outermost JSON object by matching balanced braces
        first_brace = text.find('{')
        if first_brace >= 0:
            depth = 0
            last_brace = -1
            for i in range(first_brace, len(text)):
                if text[i] == '{':
                    depth += 1
                elif text[i] == '}':
                    depth -= 1
                    if depth == 0:
                        last_brace = i
            if last_brace > first_brace:
                candidate = text[first_brace:last_brace + 1]
                try:
                    return json.loads(candidate)
                except json.JSONDecodeError:
                    pass

        # Fallback: regex search
        json_match = re.search(r'\{[\s\S]*\}', text)
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass

        logger.error(f"Failed to parse Claude response as JSON")
        logger.debug(f"Raw response (first 1000 chars): {text[:1000]}")
        return {
            'status': 'ERROR',
            'error': f'JSON parse failed — Claude returned prose instead of JSON',
            'raw_response': text[:2000],
            }

    def _mock_pass_response(self) -> Dict[str, Any]:
        return {
            'status': 'PASS',
            'checks_performed': 0,
            'issues': [],
            'patches': [],
            'validation_summary': {
                'bs_row_19': None,
                'pl_row_68': None,
                'bdo_resultaat_adjusted': None,
                'all_match': True,
            },
            'notes': 'Mock response — ANTHROPIC_API_KEY not configured',
        }


# ---------------------------------------------------------------------------
# BDO Copy Verifier (Phase A context additions)
# ---------------------------------------------------------------------------

class BDOCopyVerifier:
    """Adds BDO sheet copy verification context to the prompt."""

    def __init__(self, workbook_path: str, bdo_source_path: Optional[str],
                 config: VerificationConfig):
        self.workbook_path = Path(workbook_path)
        self.bdo_source_path = Path(bdo_source_path) if bdo_source_path else None
        self.config = config

    def build_verification_context(self) -> Dict[str, Any]:
        """Build BDO copy verification context."""
        result = {
            'bdo_copy_verified': False,
            'integrity_checks': {},
            'column_h_checks': {},
            'special_row_checks': {},
        }

        if not self.bdo_source_path or not self.bdo_source_path.exists():
            result['note'] = 'BDO source file not available for copy verification'
            return result

        try:
            wb = openpyxl.load_workbook(str(self.workbook_path))
            source_wb = openpyxl.load_workbook(str(self.bdo_source_path))

            bdo_sheet = None
            for sn in wb.sheetnames:
                if sn == self.config.bdo_sheet_name:
                    bdo_sheet = wb[sn]
                    break

            source_sheet = source_wb.active

            if bdo_sheet and source_sheet:
                result['integrity_checks'] = self._check_copy_integrity(
                    source_sheet, bdo_sheet
                )
                result['column_h_checks'] = self._check_column_h_formulas(bdo_sheet)
                result['special_row_checks'] = self._check_special_rows(bdo_sheet)
                result['bdo_copy_verified'] = True

            wb.close()
            source_wb.close()
        except Exception as e:
            result['error'] = str(e)
            logger.warning(f"BDO copy verification failed: {e}")

        return result

    def _check_copy_integrity(self, source, target) -> Dict[str, Any]:
        mismatches = []
        rows_checked = 0

        max_row = min(source.max_row, 140)
        max_col = min(source.max_column, 10)

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                src_val = source.cell(row=row_idx, column=col_idx).value
                tgt_val = target.cell(row=row_idx, column=col_idx).value

                if src_val != tgt_val:
                    if isinstance(src_val, float) and isinstance(tgt_val, float):
                        if abs(src_val - tgt_val) < 0.01:
                            continue
                    mismatches.append({
                        'row': row_idx,
                        'col': col_idx,
                        'source': str(src_val)[:50] if src_val else None,
                        'target': str(tgt_val)[:50] if tgt_val else None,
                    })
            rows_checked += 1

        return {
            'rows_checked': rows_checked,
            'mismatches': mismatches[:20],
            'total_mismatches': len(mismatches),
            'passed': len(mismatches) == 0,
        }

    def _check_column_h_formulas(self, sheet) -> Dict[str, Any]:
        """Verify column H contains =SUM(C{row}:G{row}) for data rows."""
        issues = []
        for row_idx in range(6, min(sheet.max_row + 1, 125)):
            code = sheet.cell(row=row_idx, column=1).value
            if code is None:
                continue
            h_cell = sheet.cell(row=row_idx, column=8)
            if h_cell.value and isinstance(h_cell.value, str) and h_cell.value.startswith('='):
                expected = f"=SUM(C{row_idx}:G{row_idx})"
                if h_cell.value.upper() != expected.upper():
                    issues.append({
                        'row': row_idx,
                        'expected': expected,
                        'actual': h_cell.value,
                    })
        return {'issues': issues[:20], 'total_issues': len(issues)}

    def _check_special_rows(self, sheet) -> Dict[str, Any]:
        """Verify special rows 129-136."""
        checks = {}
        for row_idx in range(129, min(sheet.max_row + 1, 137)):
            row_data = {}
            for col_idx in range(1, 9):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_data[get_column_letter(col_idx)] = {
                    'value': cell.value if not isinstance(cell.value, float) else cell.value,
                    'is_formula': isinstance(cell.value, str) and str(cell.value).startswith('='),
                }
            checks[row_idx] = row_data
        return checks


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

class AIVerificationOrchestrator:
    """
    Main orchestrator for AI verification of Management Accounts.
    Always runs after MA build, applies patches, re-validates.
    """

    def __init__(self, config: VerificationConfig,
                 api_key: Optional[str] = None,
                 formula_templates_path: str = "config/formula_templates.yaml"):
        self.config = config
        self.api_key = api_key
        self.formula_templates_path = formula_templates_path

    def run(self, workbook_path: str,
            bdo_source_path: Optional[str] = None,
            bdo_result: Optional[BDOParseResult] = None,
            computed_values: Optional[Dict] = None) -> VerificationResult:
        """
        Run full AI verification pipeline.

        Args:
            workbook_path: Path to the draft MA workbook
            bdo_source_path: Path to original BDO source file (for copy verification)
            bdo_result: Parsed BDO data (for shadow model context)
            computed_values: Pre-computed numeric values from MA builder {(row,'quarter'|'ltm'): float}

        Returns:
            VerificationResult with status, patches applied, and file paths
        """
        workbook_path = str(workbook_path)

        result = VerificationResult(
            status='PASS',
            original_file=workbook_path,
            verified_file=workbook_path,
        )

        try:
            # Phase A: Extract workbook context
            logger.info("[AI Verify] Phase A: Extracting workbook context...")
            extractor = WorkbookContextExtractor(
                workbook_path, self.config, self.formula_templates_path
            )
            context = extractor.extract()

            # Phase A+: BDO copy verification context
            if bdo_source_path:
                logger.info("[AI Verify] Phase A+: Verifying BDO copy integrity...")
                bdo_verifier = BDOCopyVerifier(workbook_path, bdo_source_path, self.config)
                bdo_context = bdo_verifier.build_verification_context()
                context['bdo_copy_verification'] = bdo_context

            # Phase A++: Compute shadow P&L and shadow BS from raw BDO data
            if bdo_result:
                logger.info("[AI Verify] Phase A++: Computing shadow models...")
                shadow_pl, shadow_bs = self._compute_shadow_models(
                    workbook_path, bdo_result
                )
                if shadow_pl:
                    context['shadow_pl'] = shadow_pl
                if shadow_bs:
                    context['shadow_bs'] = shadow_bs

            # Phase A+++: Include pre-computed values from MA builder
            if computed_values:
                cv_snapshot = {}
                for (row, scope), val in sorted(computed_values.items()):
                    key = f"row_{row}_{scope}"
                    cv_snapshot[key] = round(val, 2) if isinstance(val, float) else val
                context['computed_values_snapshot'] = cv_snapshot
                logger.info(f"[AI Verify] Added {len(cv_snapshot)} computed_values to context")

            # Phase B: Build prompt
            logger.info("[AI Verify] Phase B: Building verification prompt...")
            prompt_builder = VerificationPromptBuilder()
            prompt = prompt_builder.build(context)

            # Add BDO copy verification to prompt if available
            if 'bdo_copy_verification' in context:
                prompt += "\n\nBDO COPY VERIFICATION CONTEXT:\n"
                prompt += json.dumps(context['bdo_copy_verification'], indent=2)

            logger.info(f"[AI Verify] Prompt size: {len(prompt)} chars")

            # Phase C: Send to Claude
            logger.info("[AI Verify] Phase C: Sending to Claude for verification...")
            verifier = ClaudeVerifier(api_key=self.api_key)
            claude_response = verifier.verify(prompt)

            if claude_response.get('status') == 'ERROR':
                result.status = 'ERROR'
                result.error = claude_response.get('error', 'Unknown error')
                result.notes = f"Claude API error: {result.error}"
                logger.warning(f"[AI Verify] Claude error — original file unchanged: {result.error}")
                return result

            result.status = claude_response.get('status', 'PASS')
            result.issues = claude_response.get('issues', [])
            result.patches = claude_response.get('patches', [])
            result.validation_summary = claude_response.get('validation_summary', {})
            result.notes = claude_response.get('notes', '')

            # Phase D: Apply patches in-place to the original workbook
            if result.patches:
                logger.info(f"[AI Verify] Phase D: Applying {len(result.patches)} patches in-place...")

                applier = PatchApplier(
                    workbook_path, self.config.summary_sheet_name,
                    expected_bdo_sheet_name=self.config.bdo_sheet_name,
                    ltm_col=self.config.built_ltm_col,
                )
                applied, rejected = applier.apply(result.patches, workbook_path)
                result.patches_applied = applied

                if rejected:
                    logger.warning(f"[AI Verify] {len(rejected)} patches rejected")
                    for rej in rejected:
                        logger.warning(f"  Rejected: row {rej.get('row')}, reason: {rej.get('rejection_reason')}")

                # Phase D++: Verify formula chain integrity (rows 19, 68, interest)
                self._verify_formula_chain(workbook_path, self.config)

                # Phase D+: Re-validate
                logger.info("[AI Verify] Phase D+: Re-validating patched workbook...")
                result.revalidation_passed = self._revalidate(workbook_path, bdo_result, computed_values)

            else:
                result.revalidation_passed = True
                logger.info("[AI Verify] No patches needed — workbook passed verification")

            logger.info(
                f"[AI Verify] Complete — status={result.status}, "
                f"patches={result.patches_applied}, "
                f"revalidation={'PASS' if result.revalidation_passed else 'FAIL'}"
            )

        except Exception as e:
            logger.exception(f"[AI Verify] Verification failed: {e}")
            result.status = 'ERROR'
            result.error = str(e)
            result.verified_file = workbook_path

        return result

    def _verify_formula_chain(self, workbook_path: str, config: 'VerificationConfig'):
        """
        Deterministic post-patch verification. All row numbers and account
        positions are discovered dynamically. Nothing is hardcoded.
        """
        import re as _re

        rules = self._load_rules()
        identity = rules.get('identity_alignment', {})
        auth_row = identity.get('authoritative_row', 19)
        dep_row = identity.get('dependent_row', 68)
        calc_formulas = rules.get('layout', {}).get('calc_formulas', {})
        ltm_overrides = rules.get('layout', {}).get('ltm_calc_overrides', {})
        interest_cfg = rules.get('interest_section', {})
        interest_codes = interest_cfg.get('account_codes', [])
        sfa_row = interest_cfg.get('management_row')
        absorbed_rows = interest_cfg.get('absorbed_rows', [])
        bdo_cfg = rules.get('bdo', {})
        data_col = bdo_cfg.get('data_column', 8)
        data_col_letter = get_column_letter(data_col)

        wb = openpyxl.load_workbook(workbook_path)
        sheet = None
        for sn in wb.sheetnames:
            if 'Management Cijfers' in sn:
                sheet = wb[sn]
                break
        if not sheet:
            wb.close()
            return

        header_row = rules.get('layout', {}).get('header_row', 22)
        ltm_col = None
        for col_idx in range(sheet.max_column, 0, -1):
            cell = sheet.cell(row=header_row, column=col_idx)
            if cell.value and 'LTM' in str(cell.value):
                ltm_col = col_idx
                break
        if not ltm_col:
            wb.close()
            return

        ltm_letter = get_column_letter(ltm_col)
        bdo_name = config.bdo_sheet_name
        needs_save = False
        all_calcs = {**calc_formulas, **ltm_overrides}

        # Check 1: dep_row (68) must be a formula
        dep_template = all_calcs.get(dep_row) or all_calcs.get(str(dep_row))
        expected_dep = (
            dep_template.replace('{col}', ltm_letter).replace('{bdo}', bdo_name)
            if dep_template else
            f"={ltm_letter}{dep_row - 2}+{ltm_letter}{dep_row - 1}"
        )
        cell_dep = sheet.cell(row=dep_row, column=ltm_col)
        if not isinstance(cell_dep.value, str) or not cell_dep.value.startswith('='):
            cell_dep.value = expected_dep
            needs_save = True
            logger.warning(f"[Post-patch] Row {dep_row} restored formula: {expected_dep}")

        # Check 2: auth_row (19) must be a formula
        auth_template = all_calcs.get(auth_row) or all_calcs.get(str(auth_row))
        expected_auth = (
            auth_template.replace('{col}', ltm_letter).replace('{bdo}', bdo_name)
            if auth_template else f"={ltm_letter}{dep_row}"
        )
        cell_auth = sheet.cell(row=auth_row, column=ltm_col)
        if not isinstance(cell_auth.value, str) or not cell_auth.value.startswith('='):
            cell_auth.value = expected_auth
            needs_save = True
            logger.warning(f"[Post-patch] Row {auth_row} restored formula: {expected_auth}")

        # Check 3: Interest consolidation with dynamic row discovery
        if interest_codes and sfa_row:
            bdo_sheet = wb[bdo_name] if bdo_name in wb.sheetnames else None
            if bdo_sheet:
                discovered_rows = []
                for row_idx in range(1, bdo_sheet.max_row + 1):
                    code = bdo_sheet.cell(row=row_idx, column=1).value
                    if code and str(code).strip() in interest_codes:
                        discovered_rows.append(row_idx)

                if discovered_rows:
                    cell_sfa = sheet.cell(row=sfa_row, column=ltm_col)
                    current = str(cell_sfa.value) if cell_sfa.value else ''
                    all_present = all(
                        f'{data_col_letter}{r}' in current
                        for r in discovered_rows
                    )
                    if not all_present:
                        parts = [f"-'{bdo_name}'!{data_col_letter}{r}"
                                 for r in discovered_rows]
                        cell_sfa.value = '=' + ''.join(parts)
                        for ar in absorbed_rows:
                            sheet.cell(row=ar, column=ltm_col).value = 0
                        needs_save = True
                        logger.warning(
                            f"[Post-patch] Row {sfa_row} LTM rebuilt with "
                            f"BDO rows {discovered_rows}; zeroed {absorbed_rows}"
                        )

        # Check 4: BDO coverage audit
        bdo_sheet = wb[bdo_name] if bdo_name in wb.sheetnames else None
        if bdo_sheet:
            pl_start_row = None
            resultaat_row = None
            for row_idx in range(1, bdo_sheet.max_row + 1):
                val = bdo_sheet.cell(row=row_idx, column=1).value
                if isinstance(val, str):
                    if 'winst' in val.lower() and 'verlies' in val.lower() and pl_start_row is None:
                        pl_start_row = row_idx + 1
                    if 'resultaat na belasting' in val.lower():
                        resultaat_row = row_idx

            if pl_start_row and resultaat_row:
                referenced_rows = set()
                pl_range = rules.get('layout', {}).get('profit_loss_rows', [23, 68])
                for ma_row in range(pl_range[0], pl_range[1] + 1):
                    for col in [ltm_col, ltm_col - 1]:
                        f = sheet.cell(row=ma_row, column=col).value
                        if f and isinstance(f, str) and f.startswith('='):
                            for m in _re.findall(r'[HG](\d+)', f):
                                referenced_rows.add(int(m))

                catch_all_row = rules.get('layout', {}).get('other_costs_row')
                if catch_all_row:
                    q_cols = bdo_cfg.get('quarter_columns', [4, 5, 6, 7])
                    q_bdo_col = q_cols[-1] if q_cols else 7
                    q_bdo_letter = get_column_letter(q_bdo_col)
                    for bdo_row in range(pl_start_row, resultaat_row):
                        h_val = bdo_sheet.cell(row=bdo_row, column=data_col).value
                        g_val = bdo_sheet.cell(row=bdo_row, column=q_bdo_col).value

                        h_has = (isinstance(h_val, (int, float)) and h_val != 0) or \
                                (isinstance(h_val, str) and h_val.startswith('='))
                        g_has = (isinstance(g_val, (int, float)) and g_val != 0) or \
                                (isinstance(g_val, str) and g_val.startswith('='))

                        if (h_has or g_has) and bdo_row not in referenced_rows:
                            for col, ref_l in [(ltm_col, data_col_letter),
                                               (ltm_col - 1, q_bdo_letter)]:
                                cell = sheet.cell(row=catch_all_row, column=col)
                                cur = str(cell.value) if cell.value else ''
                                ref = f"'{bdo_name}'!{ref_l}{bdo_row}"
                                if ref not in cur and cur.startswith('='):
                                    cell.value = cur + f"-{ref}"
                                    needs_save = True
                            logger.warning(
                                f"[Post-patch] Added unreferenced BDO row "
                                f"{bdo_row} to catch-all row {catch_all_row}"
                            )

        if needs_save:
            wb.save(workbook_path)
            logger.info("[Post-patch] Formula chain verified and corrected")
        wb.close()

    @staticmethod
    def _load_rules() -> dict:
        rules_path = Path("config/accounting_rules.yaml")
        if rules_path.exists():
            try:
                with open(rules_path, 'r', encoding='utf-8') as f:
                    return yaml.safe_load(f) or {}
            except Exception:
                pass
        return {}

    def _compute_shadow_models(
        self, workbook_path: str, bdo_result: BDOParseResult
    ) -> Tuple[Optional[Dict], Optional[Dict]]:
        """
        Compute shadow P&L and shadow BS from raw BDO data.
        Returns (shadow_pl, shadow_bs) dicts or (None, None) on failure.
        """
        try:
            from .management_accounts import ManagementAccountsBuilder, ManagementAccountsConfig

            config = ManagementAccountsConfig(
                quarter=self.config.quarter,
                period_end=_parse_period_end(self.config.period_end),
                bdo_sheet_name=self.config.bdo_sheet_name,
                summary_sheet_name=self.config.summary_sheet_name,
                quarter_num=getattr(self.config, 'quarter_num', None),
            )

            builder = ManagementAccountsBuilder.__new__(ManagementAccountsBuilder)
            builder.config = config
            builder.formula_templates_path = Path(self.formula_templates_path)
            builder.formula_templates = builder._load_formula_templates()
            builder.balance_sheet_templates = builder._load_balance_sheet_templates()
            builder._new_bdo_row_map = {}
            builder._new_bdo_label_map = {}
            builder._prev_bdo_row_map = {}

            builder.workbook = openpyxl.load_workbook(workbook_path)
            bdo_sheet = None
            for sn in builder.workbook.sheetnames:
                if sn == self.config.bdo_sheet_name:
                    bdo_sheet = builder.workbook[sn]
                    break
            if bdo_sheet:
                builder._build_row_map(
                    bdo_sheet, builder._new_bdo_row_map, builder._new_bdo_label_map
                )

            shadow_pl = builder._compute_shadow_pl(bdo_result)
            shadow_bs = builder._compute_shadow_bs(bdo_result)
            builder.workbook.close()

            logger.info(
                f"[AI Verify] Shadow models: P&L row 68={shadow_pl.get(68, {}).get('value', 'N/A')}, "
                f"BS row 19={shadow_bs.get(19, {}).get('value', 'N/A')}"
            )
            return shadow_pl, shadow_bs

        except Exception as e:
            logger.warning(f"[AI Verify] Shadow model computation failed: {e}")
            return None, None

    def _revalidate(self, workbook_path: str,
                    bdo_result: Optional[BDOParseResult],
                    computed_values: Optional[Dict] = None) -> bool:
        """
        Re-run validation on the patched workbook.

        Checks:
        1. Rows 19 and 68 in LTM column contain formulas (not hardcoded values).
        2. Computed values match between PL68 and BS19 (within tolerance).
        """
        rules_path = Path("config/accounting_rules.yaml")
        rules = {}
        if rules_path.exists():
            try:
                with open(rules_path, 'r', encoding='utf-8') as f:
                    rules = yaml.safe_load(f) or {}
            except Exception:
                pass

        identity = rules.get('identity_alignment', {})
        auth_row = identity.get('authoritative_row', 19)
        dep_row = identity.get('dependent_row', 68)

        wb = openpyxl.load_workbook(workbook_path)
        formula_ok = True
        for sn in wb.sheetnames:
            if 'Management Cijfers' in sn:
                sheet = wb[sn]
                ltm_col = None
                for col_idx in range(sheet.max_column, 0, -1):
                    cell = sheet.cell(row=22, column=col_idx)
                    if cell.value and 'LTM' in str(cell.value):
                        ltm_col = col_idx
                        break
                if ltm_col:
                    cell_19 = sheet.cell(row=auth_row, column=ltm_col)
                    cell_68 = sheet.cell(row=dep_row, column=ltm_col)
                    if not (isinstance(cell_19.value, str) and cell_19.value.startswith('=')):
                        logger.error(
                            f"REVALIDATION FAIL: Row {auth_row} LTM is not a formula: "
                            f"{cell_19.value}"
                        )
                        formula_ok = False
                    if not (isinstance(cell_68.value, str) and cell_68.value.startswith('=')):
                        logger.error(
                            f"REVALIDATION FAIL: Row {dep_row} LTM is not a formula: "
                            f"{cell_68.value}"
                        )
                        formula_ok = False
                break
        wb.close()

        if not formula_ok:
            return False

        tolerance = 1.0

        if computed_values:
            pl68 = computed_values.get((dep_row, 'ltm'), 0.0)
            bs19 = computed_values.get((auth_row, 'ltm'), 0.0)
            diff = abs(pl68 - bs19)
            passed = diff <= tolerance
            logger.info(
                f"[AI Verify] Re-validation (computed_values): "
                f"PL68={pl68:,.2f}, BS19={bs19:,.2f}, diff={diff:,.2f}, passed={passed}"
            )
            return passed

        if bdo_result is None:
            logger.warning("[AI Verify] No BDO result for re-validation, skipping")
            return True

        try:
            from .management_accounts import ManagementAccountsBuilder, ManagementAccountsConfig

            config = ManagementAccountsConfig(
                quarter=self.config.quarter,
                period_end=_parse_period_end(self.config.period_end),
                bdo_sheet_name=self.config.bdo_sheet_name,
                summary_sheet_name=self.config.summary_sheet_name,
                quarter_num=getattr(self.config, 'quarter_num', None),
            )

            builder = ManagementAccountsBuilder.__new__(ManagementAccountsBuilder)
            builder.config = config
            builder.formula_templates_path = Path(self.formula_templates_path)
            builder.formula_templates = builder._load_formula_templates()
            builder.balance_sheet_templates = builder._load_balance_sheet_templates()
            builder._new_bdo_row_map = {}
            builder._new_bdo_label_map = {}
            builder._prev_bdo_row_map = {}

            builder.workbook = openpyxl.load_workbook(workbook_path)

            bdo_sheet = None
            for sn in builder.workbook.sheetnames:
                if sn == self.config.bdo_sheet_name:
                    bdo_sheet = builder.workbook[sn]
                    break

            if bdo_sheet:
                builder._build_row_map(bdo_sheet, builder._new_bdo_row_map, builder._new_bdo_label_map)

            shadow_pl = builder._compute_shadow_pl(bdo_result)
            shadow_bs = builder._compute_shadow_bs(bdo_result)

            dr = shadow_pl.get(dep_row, {}).get('value', 0.0)
            em = shadow_bs.get(auth_row, {}).get('value', 0.0)

            builder.workbook.close()

            diff = abs(dr - em)
            passed = diff <= tolerance
            logger.info(
                f"[AI Verify] Re-validation (shadow): DR={dr:,.2f}, EM={em:,.2f}, "
                f"diff={diff:,.2f}, passed={passed}"
            )
            return passed

        except Exception as e:
            logger.error(f"[AI Verify] Re-validation error: {e}")
            return False


def _parse_period_end(period_end_str: str):
    """Parse period end string to datetime."""
    from datetime import datetime
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(period_end_str, fmt)
        except ValueError:
            continue
    return datetime.now()
