"""
PDF Assembler

Assembles final quarterly report PDF from multiple sources:
1. Main report (Word -> PDF)
2. Rent roll (Excel sheet -> PDF)
3. Management Accounts (Excel sheet -> PDF)
4. Sales tracker (Excel sheet -> PDF)
5. Compliance certificate (Excel sheets -> PDF)
"""

from dataclasses import dataclass
from typing import List, Optional
from pathlib import Path
import subprocess
import tempfile
import os
import shutil
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from loguru import logger

# Headless LibreOffice conversions can hang on formula-heavy workbooks; cap wait time.
LIBREOFFICE_TIMEOUT_SECONDS = 300


@dataclass
class PDFSource:
    """Definition of a PDF source for merging."""
    name: str
    source_path: Path
    source_type: str  # 'pdf', 'docx', 'xlsx'
    sheet_name: Optional[str] = None  # For xlsx sources
    page_range: Optional[tuple] = None  # (start, end) 1-indexed, None = all


class PDFAssembler:
    """
    Assembles final merged PDF for ASX submission.
    
    Merge order (per client specification):
    1. Main report (from Word) - Pages 1-7
    2. Rent roll (from Excel) - Pages 8-19
    3. Management Accounts (from Excel) - Pages 20-21
    4. Sales tracker (from Excel) - Pages 22-23
    5. Compliance certificate (from Excel) - Pages 24-28
    """
    
    def __init__(self, output_path: str, working_dir: str = None):
        self.output_path = Path(output_path)
        if working_dir is None:
            self.working_dir = Path(tempfile.gettempdir()) / "pdf_assembly"
        else:
            self.working_dir = Path(working_dir)
        self.working_dir.mkdir(parents=True, exist_ok=True)
    
    def _find_soffice(self) -> str:
        """
        Find LibreOffice soffice command.
        
        Checks common installation paths on different platforms.
        """
        import shutil
        import platform
        
        # Common paths to check
        paths_to_check = [
            'soffice',  # In PATH
            '/usr/bin/soffice',  # Linux
            '/usr/local/bin/soffice',  # Linux homebrew
        ]
        
        # macOS specific paths
        if platform.system() == 'Darwin':
            paths_to_check.extend([
                '/Applications/LibreOffice.app/Contents/MacOS/soffice',
                '/Applications/LibreOffice.app/Contents/MacOS/libreoffice',
                os.path.expanduser('~/Applications/LibreOffice.app/Contents/MacOS/soffice'),
            ])
        
        # Windows paths
        if platform.system() == 'Windows':
            paths_to_check.extend([
                r'C:\Program Files\LibreOffice\program\soffice.exe',
                r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
            ])
        
        # Check each path
        for path in paths_to_check:
            if path == 'soffice':
                # Check if in PATH
                if shutil.which('soffice'):
                    return 'soffice'
            elif os.path.isfile(path):
                return path
        
        raise FileNotFoundError(
            "LibreOffice not found. Please install LibreOffice:\n"
            "  - macOS: brew install --cask libreoffice\n"
            "  - Linux: sudo apt install libreoffice\n"
            "  - Windows: Download from https://www.libreoffice.org/download/"
        )

    def _run_soffice(self, cmd: List[str], description: str) -> subprocess.CompletedProcess:
        """Run LibreOffice with a timeout so API/n8n requests fail fast instead of hanging."""
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=LIBREOFFICE_TIMEOUT_SECONDS,
            )
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError(
                f"LibreOffice timed out after {LIBREOFFICE_TIMEOUT_SECONDS}s "
                f"while {description}. Command: {' '.join(cmd)}"
            ) from exc
        if result.returncode != 0:
            stderr = (result.stderr or result.stdout or '').strip()
            raise RuntimeError(
                f"LibreOffice conversion failed while {description}: {stderr or 'unknown error'}"
            )
        return result
        
    def assemble(self, sources: List[PDFSource]) -> Path:
        """
        Assemble final PDF from multiple sources with interleaved structure.
        
        The Word document contains section intro pages with "File enclosed" text.
        Each attachment should be inserted immediately after its section intro page.
        
        Args:
            sources: Ordered list of PDF sources. First source should be the Word doc,
                    followed by attachments in order: Rent Roll, MA, Sales, Compliance.
            
        Returns:
            Path to merged PDF
        """
        logger.info(f"Assembling PDF from {len(sources)} sources")
        
        # Identify Word doc (first docx source) and attachment sources
        word_source = None
        attachment_sources = []
        
        for source in sources:
            if source.source_type == 'docx' and word_source is None:
                word_source = source
            else:
                attachment_sources.append(source)
        
        if word_source is None:
            # Fallback to simple linear merge if no Word doc
            return self._assemble_linear(sources)
        
        # Convert Word doc to PDF
        logger.info("Converting Word document to PDF")
        word_pdf = self._convert_docx_to_pdf(word_source.source_path)
        
        # Convert all attachments to PDF
        attachment_pdfs = []
        attachment_markers = [
            ("Rent Roll", "rent roll"),
            ("Management Accounts", "management accounts"),
            ("Sales", "sales report"),
            ("Compliance", "compliance"),
        ]
        
        for source in attachment_sources:
            logger.info(f"Processing attachment: {source.name}")
            if source.source_type == 'xlsx':
                pdf_path = self._convert_xlsx_to_pdf(source.source_path, source.sheet_name)
            elif source.source_type == 'pdf':
                pdf_path = source.source_path
            else:
                pdf_path = self._convert_docx_to_pdf(source.source_path)
            
            if source.page_range:
                pdf_path = self._extract_pages(pdf_path, source.page_range)
            
            attachment_pdfs.append((source.name, pdf_path))
        
        # Interleave Word PDF pages with attachments
        merged_path = self._interleave_attachments(word_pdf, attachment_pdfs, attachment_markers)
        
        # Move to final output location
        if merged_path.exists():
            if self.output_path.exists():
                self.output_path.unlink()
            shutil.copy2(merged_path, self.output_path)
            merged_path.unlink()
        
        logger.info(f"Final PDF saved to: {self.output_path}")
        return self.output_path
    
    def _assemble_linear(self, sources: List[PDFSource]) -> Path:
        """Fallback linear merge without interleaving."""
        pdf_parts = []
        for source in sources:
            if source.source_type == 'pdf':
                pdf_path = source.source_path
            elif source.source_type == 'docx':
                pdf_path = self._convert_docx_to_pdf(source.source_path)
            elif source.source_type == 'xlsx':
                pdf_path = self._convert_xlsx_to_pdf(source.source_path, source.sheet_name)
            else:
                raise ValueError(f"Unknown source type: {source.source_type}")
            
            if source.page_range:
                pdf_path = self._extract_pages(pdf_path, source.page_range)
            
            pdf_parts.append(pdf_path)
        
        merged_path = self._merge_pdfs(pdf_parts)
        
        if merged_path.exists():
            if self.output_path.exists():
                self.output_path.unlink()
            shutil.copy2(merged_path, self.output_path)
            merged_path.unlink()
        
        return self.output_path
    
    def _interleave_attachments(self, word_pdf: Path, attachment_pdfs: List[tuple], 
                                 markers: List[tuple]) -> Path:
        """
        Interleave Word PDF pages with attachments based on section markers.
        
        Scans Word PDF for "File enclosed" pages and inserts corresponding
        attachment immediately after each.
        """
        word_reader = PdfReader(word_pdf)
        writer = PdfWriter()
        
        # Find pages with "File enclosed" and identify which section they belong to
        section_pages = []  # List of (page_num, section_type)
        
        for page_num in range(len(word_reader.pages)):
            text = word_reader.pages[page_num].extract_text().lower()
            if "file enclosed" in text:
                # Identify which section this is
                for marker_name, marker_text in markers:
                    if marker_text in text:
                        section_pages.append((page_num, marker_name))
                        logger.debug(f"Found '{marker_name}' section intro at page {page_num + 1}")
                        break
        
        logger.info(f"Found {len(section_pages)} section intro pages: {[s[1] for s in section_pages]}")
        
        # Build mapping of section name to attachment PDFs
        attachment_map = {}
        for name, pdf_path in attachment_pdfs:
            # Match attachment to section
            name_lower = name.lower()
            matched = False
            for marker_name, _ in markers:
                if marker_name.lower() in name_lower or name_lower in marker_name.lower():
                    if marker_name not in attachment_map:
                        attachment_map[marker_name] = []
                    attachment_map[marker_name].append(pdf_path)
                    matched = True
                    break
            
            # Special handling for Compliance sheets (multiple)
            if not matched and "compliance" in name_lower or "sfa" in name_lower or "suppl" in name_lower or "impact" in name_lower:
                if "Compliance" not in attachment_map:
                    attachment_map["Compliance"] = []
                attachment_map["Compliance"].append(pdf_path)
        
        logger.debug(f"Attachment map: {list(attachment_map.keys())}")
        
        # Build the final PDF with interleaving
        sections_inserted = set()
        current_page = 0
        
        for page_num in range(len(word_reader.pages)):
            # Add the Word page
            writer.add_page(word_reader.pages[page_num])
            
            # Check if this is a section intro page
            for section_page, section_name in section_pages:
                if page_num == section_page and section_name not in sections_inserted:
                    # Insert attachment(s) for this section
                    if section_name in attachment_map:
                        for attachment_pdf in attachment_map[section_name]:
                            logger.info(f"Inserting {section_name} attachment after page {page_num + 1}")
                            att_reader = PdfReader(attachment_pdf)
                            for att_page in att_reader.pages:
                                writer.add_page(att_page)
                    sections_inserted.add(section_name)
        
        # Add any remaining attachments (e.g., Compliance after all Word pages)
        for section_name, pdfs in attachment_map.items():
            if section_name not in sections_inserted:
                for pdf_path in pdfs:
                    logger.info(f"Appending remaining attachment: {section_name}")
                    reader = PdfReader(pdf_path)
                    for page in reader.pages:
                        writer.add_page(page)
        
        # Write the interleaved PDF
        output_path = self.working_dir / "interleaved_output.pdf"
        with open(output_path, 'wb') as f:
            writer.write(f)
        
        logger.info(f"Interleaved PDF created with {len(writer.pages)} pages")
        return output_path
    
    def _convert_docx_to_pdf(self, docx_path: Path) -> Path:
        """Convert Word document to PDF using LibreOffice."""
        output_dir = self.working_dir / "docx_converted"
        output_dir.mkdir(exist_ok=True)
        
        soffice_cmd = self._find_soffice()
        
        cmd = [
            soffice_cmd,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', str(output_dir),
            str(docx_path)
        ]
        
        self._run_soffice(cmd, f"converting Word document {docx_path.name}")
        
        # Find the output file
        pdf_name = docx_path.stem + '.pdf'
        pdf_path = output_dir / pdf_name
        
        if not pdf_path.exists():
            raise FileNotFoundError(f"Expected PDF not found: {pdf_path}")
        
        return pdf_path
    
    def _convert_xlsx_to_pdf(self, xlsx_path: Path, sheet_name: str = None) -> Path:
        """
        Convert Excel sheet to PDF.
        
        Uses LibreOffice for conversion. If sheet_name specified,
        exports only that sheet.
        """
        output_dir = self.working_dir / "xlsx_converted"
        output_dir.mkdir(exist_ok=True)
        
        if sheet_name:
            # For specific sheet, use Python to extract and convert
            pdf_path = self._export_excel_sheet_to_pdf(xlsx_path, sheet_name, output_dir)
        else:
            # Convert entire workbook
            soffice_cmd = self._find_soffice()
            cmd = [
                soffice_cmd,
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(output_dir),
                str(xlsx_path)
            ]
            
            self._run_soffice(cmd, f"converting workbook {xlsx_path.name}")
            
            pdf_name = xlsx_path.stem + '.pdf'
            pdf_path = output_dir / pdf_name
        
        return pdf_path
    
    def _recalculate_workbook(self, xlsx_path: Path) -> Path:
        """
        Use LibreOffice to recalculate all formulas in a workbook.
        This is necessary because openpyxl cannot calculate formulas.
        
        Returns the path to the recalculated workbook.
        """
        recalc_dir = self.working_dir / "recalculated"
        recalc_dir.mkdir(exist_ok=True)
        
        soffice_cmd = self._find_soffice()
        
        # LibreOffice macro to recalculate and save
        # We use --convert-to xlsx which forces recalculation
        cmd = [
            soffice_cmd,
            '--headless',
            '--calc',
            '--convert-to', 'xlsx',
            '--outdir', str(recalc_dir),
            str(xlsx_path)
        ]
        
        try:
            self._run_soffice(cmd, f"recalculating workbook {xlsx_path.name}")
        except RuntimeError as exc:
            logger.warning(f"Recalculation may have failed: {exc}")
            return xlsx_path
        
        recalc_path = recalc_dir / xlsx_path.name
        if recalc_path.exists():
            logger.debug(f"Workbook recalculated: {recalc_path}")
            return recalc_path
        
        return xlsx_path
    
    def _export_excel_sheet_to_pdf(self, xlsx_path: Path, sheet_name: str, 
                                    output_dir: Path) -> Path:
        """
        Export a specific Excel sheet to PDF.
        
        IMPORTANT (Issue 9 fix): #NAME? errors in PDF but not Excel
        This happens because:
        1. LibreOffice may not recalculate formulas properly
        2. Formulas referencing external/other sheets may not resolve
        
        Solution: Prefer the ORIGINAL file's Excel-cached values (what the user
        sees in their spreadsheet). Only fall back to LibreOffice recalculation
        for cells where the original has no cached value (e.g. files generated
        by openpyxl which doesn't cache formula results).
        """
        import openpyxl
        from copy import copy
        
        # Load ORIGINAL file with data_only=True to get Excel's cached values.
        # This preserves exactly what the user sees in their spreadsheet.
        wb_original = openpyxl.load_workbook(xlsx_path, data_only=True)
        
        if sheet_name not in wb_original.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {xlsx_path}")
        
        source_sheet_original = wb_original[sheet_name]
        
        # Check if the original file has any cached values at all.
        # Files generated by openpyxl won't have cached formula results.
        has_cached = any(
            cell.value is not None
            for row in source_sheet_original.iter_rows(max_row=min(source_sheet_original.max_row, 20))
            for cell in row
        )
        
        # Recalculate via LibreOffice as fallback for cells with no cached value
        recalc_xlsx = self._recalculate_workbook(xlsx_path)
        wb_recalc = openpyxl.load_workbook(recalc_xlsx, data_only=True)
        source_sheet_recalc = wb_recalc[sheet_name] if sheet_name in wb_recalc.sheetnames else None
        
        # Load with formulas to preserve formatting
        wb_format = openpyxl.load_workbook(xlsx_path, data_only=False)
        
        temp_xlsx = output_dir / f"temp_{sheet_name.replace(' ', '_')}.xlsx"
        
        source_sheet_format = wb_format[sheet_name]
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        new_sheet.title = sheet_name
        
        # Copy cell VALUES and formatting.
        # Prefer original Excel-cached values; fall back to recalculated for None cells.
        fallback_count = 0
        for row in source_sheet_original.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                
                val = cell.value
                if val is None and source_sheet_recalc is not None:
                    recalc_val = source_sheet_recalc.cell(row=cell.row, column=cell.column).value
                    if recalc_val is not None:
                        val = recalc_val
                        fallback_count += 1
                new_cell.value = val
                
                format_cell = source_sheet_format.cell(row=cell.row, column=cell.column)
                if format_cell.has_style:
                    new_cell.font = copy(format_cell.font)
                    new_cell.alignment = copy(format_cell.alignment)
                    new_cell.number_format = format_cell.number_format
                    new_cell.border = copy(format_cell.border)
                    new_cell.fill = copy(format_cell.fill)
        
        if fallback_count > 0:
            logger.info(f"Used LibreOffice fallback for {fallback_count} cells with no Excel-cached value")
        
        wb_recalc.close()
        wb_original.close()
        
        # Copy column widths
        for col_letter, dim in source_sheet_format.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = dim.width
        
        # Copy row heights
        for row_idx, dim in source_sheet_format.row_dimensions.items():
            new_sheet.row_dimensions[row_idx].height = dim.height
        
        # Copy merged cells
        for merged_range in source_sheet_format.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
        
        # Set print settings
        new_sheet.page_setup.orientation = 'landscape'
        new_sheet.page_setup.fitToPage = True
        new_sheet.page_setup.fitToWidth = 1
        new_sheet.page_setup.fitToHeight = 0
        
        # Handle print area:
        # For SFA CC sheet, extend print area to include signature page (rows 82+)
        # This fixes Issue 10: signature page incomplete
        if source_sheet_format.print_area:
            print_area = source_sheet_format.print_area
            
            # Check if this is SFA CC and signature rows exist
            if 'SFA CC' in sheet_name and new_sheet.max_row > 81:
                # Extend print area to include all rows (for signature section)
                # Original print area might be like $A$2:$J$81
                import re
                match = re.match(r"['\"]?[^'\"]*['\"]?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", print_area)
                if match:
                    start_col, start_row, end_col, end_row = match.groups()
                    # Extend to include all content rows
                    new_print_area = f"${start_col}${start_row}:${end_col}${new_sheet.max_row}"
                    new_sheet.print_area = new_print_area
                    logger.debug(f"Extended print area for {sheet_name}: {print_area} → {new_print_area}")
                else:
                    new_sheet.print_area = print_area
            else:
                new_sheet.print_area = print_area
        
        new_wb.save(temp_xlsx)
        wb_format.close()
        
        # Convert to PDF with LibreOffice
        soffice_cmd = self._find_soffice()
        cmd = [
            soffice_cmd,
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', str(output_dir),
            str(temp_xlsx)
        ]
        
        self._run_soffice(cmd, f"exporting sheet {sheet_name!r} from {xlsx_path.name}")
        
        pdf_path = output_dir / f"temp_{sheet_name.replace(' ', '_')}.pdf"
        
        # Cleanup temp xlsx
        if temp_xlsx.exists():
            temp_xlsx.unlink()
        
        logger.debug(f"Exported sheet '{sheet_name}' to PDF with calculated values")
        return pdf_path
    
    def _extract_pages(self, pdf_path: Path, page_range: tuple) -> Path:
        """Extract specific pages from PDF."""
        start_page, end_page = page_range
        
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        
        # Convert to 0-indexed
        for page_num in range(start_page - 1, min(end_page, len(reader.pages))):
            writer.add_page(reader.pages[page_num])
        
        output_path = self.working_dir / f"extracted_{pdf_path.stem}.pdf"
        with open(output_path, 'wb') as f:
            writer.write(f)
        
        return output_path
    
    def _merge_pdfs(self, pdf_paths: List[Path]) -> Path:
        """Merge multiple PDFs into one."""
        merger = PdfMerger()
        
        for pdf_path in pdf_paths:
            logger.debug(f"Adding to merge: {pdf_path}")
            merger.append(str(pdf_path))
        
        output_path = self.working_dir / "merged_output.pdf"
        merger.write(str(output_path))
        merger.close()
        
        logger.info(f"Merged {len(pdf_paths)} PDFs into {output_path}")
        return output_path
    
    def add_page_numbers(self, pdf_path: Path) -> Path:
        """Add page numbers to merged PDF."""
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from io import BytesIO
        
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        
        total_pages = len(reader.pages)
        
        for page_num, page in enumerate(reader.pages, 1):
            # Create page number overlay
            packet = BytesIO()
            c = canvas.Canvas(packet, pagesize=A4)
            
            # Add page number at bottom center
            c.drawString(A4[0] / 2 - 20, 30, f"Page {page_num} of {total_pages}")
            c.save()
            packet.seek(0)
            
            # Merge overlay with page
            overlay = PdfReader(packet)
            page.merge_page(overlay.pages[0])
            writer.add_page(page)
        
        output_path = self.working_dir / "numbered_output.pdf"
        with open(output_path, 'wb') as f:
            writer.write(f)
        
        return output_path

