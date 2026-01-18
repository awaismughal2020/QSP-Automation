"""
Shared pytest fixtures for QSP Quarterly Reporting tests.
"""

import pytest
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl import Workbook

from src.parsers.bdo_parser import BDOParseResult, AccountEntry
from src.parsers.rent_roll_parser import RentRollResult, RentRollEntry
from src.parsers.sales_tracker_parser import SalesTrackerResult, SaleEntry


@pytest.fixture
def fixtures_dir():
    """Path to test fixtures directory."""
    return Path(__file__).parent / "fixtures"


@pytest.fixture
def sample_bdo_file_8col(fixtures_dir, tmp_path):
    """Create a sample BDO file with 8 columns (old format)."""
    file_path = tmp_path / "bdo_q1_2019_sample.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "BalansenWinstverlies"
    
    # Header rows
    ws['A1'] = "BDO Quarterly Report"
    ws['A2'] = "Period: Q1 2019"
    ws['A3'] = "Date: 31-03-2019"
    
    # Column headers (row 4)
    headers = ["Grootboek", "Omschrijving", "Beginsaldo", "Mutatie 1", "Mutatie 2", 
               "Mutatie 3", "Mutatie 4", "Eindsaldo"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)
    
    # Sample account data (row 5+)
    accounts = [
        ("1600000", "Verkrijgingsprijs", 10000000.00, 0, 0, 0, 0, 10000000.00),
        ("1600200", "Cumulatieve afschrijving", -500000.00, -25000.00, 0, 0, 0, -525000.00),
        ("1760000", "Te vorderen DMRRP", 500000.00, 0, 0, 0, 0, 500000.00),
        ("1100000", "Kas", 100000.00, 50000.00, 0, 0, 0, 150000.00),
        ("1300000", "Debiteuren", 200000.00, 0, 0, 0, 0, 200000.00),
        ("0500000", "Aandelenkapitaal", 5000000.00, 0, 0, 0, 0, 5000000.00),
        ("0600000", "Ingehouden winst", 4500000.00, 25000.00, 0, 0, 0, 4525000.00),
        ("0900000", "Senior lening", 3000000.00, 0, 0, 0, 0, 3000000.00),
        ("8000000", "Huuropbrengsten", 0, 500000.00, 0, 0, 0, 500000.00),
        ("4000000", "Vastgoed kosten", 0, -100000.00, 0, 0, 0, -100000.00),
    ]
    
    for row_idx, account_data in enumerate(accounts, 5):
        for col_idx, value in enumerate(account_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_bdo_file_9col(fixtures_dir, tmp_path):
    """Create a sample BDO file with 9 columns (new format)."""
    file_path = tmp_path / "bdo_q3_2025_sample.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "BalansenWinstverlies"
    
    # Header rows
    ws['A1'] = "BDO Quarterly Report"
    ws['A2'] = "Period: Q3 2025"
    ws['A3'] = "Date: 30-09-2025"
    
    # Column headers (row 4)
    headers = ["Grootboek", "Omschrijving", "Beginsaldo", "Mutatie 1", "Mutatie 2", 
               "Mutatie 3", "Mutatie 4", "Mutatie 5", "Eindsaldo"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)
    
    # Sample account data (row 5+)
    accounts = [
        ("1600000", "Verkrijgingsprijs", 12000000.00, 0, 0, 0, 0, 0, 12000000.00),
        ("1600200", "Cumulatieve afschrijving", -600000.00, -30000.00, 0, 0, 0, 0, -630000.00),
        ("1601000", "Verkoop verkrijgingsprijs", 0, -200000.00, 0, 0, 0, 0, -200000.00),
        ("1760000", "Te vorderen DMRRP", 600000.00, 0, 0, 0, 0, 0, 600000.00),
        ("1790002", "Latente belastingvordering", 100000.00, 0, 0, 0, 0, 0, 100000.00),
        ("1100000", "Kas", 150000.00, 50000.00, 0, 0, 0, 0, 200000.00),
        ("1110000", "Bank", 500000.00, 100000.00, 0, 0, 0, 0, 600000.00),
        ("1300000", "Debiteuren", 250000.00, 0, 0, 0, 0, 0, 250000.00),
        ("0500000", "Aandelenkapitaal", 5000000.00, 0, 0, 0, 0, 0, 5000000.00),
        ("0510000", "Agio", 1000000.00, 0, 0, 0, 0, 0, 1000000.00),
        ("0600000", "Ingehouden winst", 5000000.00, 120000.00, 0, 0, 0, 0, 5120000.00),
        ("0900000", "Senior lening", 3500000.00, 0, 0, 0, 0, 0, 3500000.00),
        ("0920000", "Opgelopen rente", 50000.00, 25000.00, 0, 0, 0, 0, 75000.00),
        ("8000000", "Huuropbrengsten", 0, 600000.00, 0, 0, 0, 0, 600000.00),
        ("8100000", "Servicekosten", 0, 50000.00, 0, 0, 0, 0, 50000.00),
        ("4000000", "Vastgoed kosten", 0, -120000.00, 0, 0, 0, 0, -120000.00),
        ("4200000", "Onderhoud", 0, -50000.00, 0, 0, 0, 0, -50000.00),
        ("4300000", "Afschrijving", 0, -30000.00, 0, 0, 0, 0, -30000.00),
        ("4400000", "Rentelasten", 0, -25000.00, 0, 0, 0, 0, -25000.00),
    ]
    
    for row_idx, account_data in enumerate(accounts, 5):
        for col_idx, value in enumerate(account_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_rent_roll_file(tmp_path):
    """Create a sample rent roll file."""
    file_path = tmp_path / "rent_roll_sample.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "huurlijst"
    
    # Headers
    headers = ["Owner", "Complex", "Subcomplex", "Unit", "Address", "Monthly Rent", "Annual Rent", "Status"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Sample units
    units = [
        ("QSP", "Complex A", "Block 1", "A101", "Street 1, Unit A101", 1200.00, 14400.00, "Occupied"),
        ("QSP", "Complex A", "Block 1", "A102", "Street 1, Unit A102", 1200.00, 14400.00, "Occupied"),
        ("QSP", "Complex A", "Block 1", "A103", "Street 1, Unit A103", 0.00, 0.00, "Vacant"),
        ("QSP", "Complex B", "Block 2", "B201", "Street 2, Unit B201", 1500.00, 18000.00, "Occupied"),
    ]
    
    for row_idx, unit_data in enumerate(units, 2):
        for col_idx, value in enumerate(unit_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_sales_tracker_file(tmp_path):
    """Create a sample sales tracker file."""
    file_path = tmp_path / "sales_tracker_sample.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    
    # Headers
    headers = ["Unit", "Address", "Valuation", "Sale Price", "Delta", "Closing Date", "Status"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Sample sales
    sales = [
        ("A101", "Street 1, Unit A101", 150000.00, 160000.00, 10000.00, "2025-09-15", "Sold"),
        ("B201", "Street 2, Unit B201", 180000.00, 203555.00, 23555.00, "2025-09-20", "Sold"),
    ]
    
    for row_idx, sale_data in enumerate(sales, 2):
        for col_idx, value in enumerate(sale_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    return file_path


@pytest.fixture
def sample_bdo_result():
    """Create a sample BDOParseResult for testing."""
    accounts = {
        "1600000": AccountEntry(
            code="1600000",
            name="Verkrijgingsprijs",
            opening_balance=12000000.00,
            closing_balance=12000000.00,
            mutations={},
            raw_row=5
        ),
        "1600200": AccountEntry(
            code="1600200",
            name="Cumulatieve afschrijving",
            opening_balance=-600000.00,
            closing_balance=-630000.00,
            mutations={"mutation_4": -30000.00},
            raw_row=6
        ),
        "1100000": AccountEntry(
            code="1100000",
            name="Kas",
            opening_balance=150000.00,
            closing_balance=200000.00,
            mutations={"mutation_4": 50000.00},
            raw_row=7
        ),
        "0600000": AccountEntry(
            code="0600000",
            name="Ingehouden winst",
            opening_balance=5000000.00,
            closing_balance=5120000.00,
            mutations={"mutation_4": 120000.00},
            raw_row=8
        ),
        "8000000": AccountEntry(
            code="8000000",
            name="Huuropbrengsten",
            opening_balance=0.00,
            closing_balance=600000.00,
            mutations={"mutation_4": 600000.00},
            raw_row=9
        ),
    }
    
    return BDOParseResult(
        accounts=accounts,
        period_end="2025-09-30",
        quarter="Q3 2025",
        column_mapping={"account_code": 0, "account_name": 1, "opening_balance": 2, "closing_balance": 8},
        schema_version="standard",
        warnings=[]
    )

