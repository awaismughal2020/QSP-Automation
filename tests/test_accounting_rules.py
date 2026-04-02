"""
Tests for config-driven accounting rules:
- YAML loader (defaults, missing file, override)
- Identity enforcement in workbook formula
- PatchApplier protected-row rejection
- Orchestrator post-AI verify & refresh
"""

import os
import pytest
import yaml
from pathlib import Path
from unittest.mock import patch as mock_patch, MagicMock

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.transformers.management_accounts import ManagementAccountsBuilder
from src.transformers.ai_verification import PatchApplier, PROTECTED_TOTAL_ROWS
from src.orchestrator import QuarterlyReportOrchestrator


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_rules(tmp_path: Path, overrides: dict = None) -> Path:
    """Write a temporary accounting_rules.yaml and return its path."""
    defaults = {
        'identity_alignment': {
            'authoritative_row': 19,
            'dependent_row': 68,
            'scope': 'ltm_only',
            'max_snap_units': 2,
            'enforce_in_workbook': True,
        },
        'rounding': {'round_to_integer': True, 'decimal_rows': [26]},
        'bank_bdo_row_map': {107: 35, 108: 34, 109: 36, 110: 32, 111: 33, 112: 37, 113: 31},
        'bank_total_row': 114,
        'protected_total_rows': [19, 68],
    }
    if overrides:
        for k, v in overrides.items():
            if isinstance(v, dict) and isinstance(defaults.get(k), dict):
                defaults[k].update(v)
            else:
                defaults[k] = v
    path = tmp_path / 'accounting_rules.yaml'
    with open(path, 'w') as f:
        yaml.dump(defaults, f)
    return path


def _make_mc_workbook(tmp_path: Path, bs19_val=100, pl68_val=101,
                      ltm_col: int = 3) -> Path:
    """Create a minimal MC workbook with row 19 and 68 in an LTM column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Management Cijfers - Q4 2025"
    ws.cell(row=22, column=ltm_col, value="LTM Q4 2025")
    ws.cell(row=19, column=ltm_col, value=bs19_val)
    ws.cell(row=68, column=ltm_col, value=pl68_val)
    path = tmp_path / "ma_test.xlsx"
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# 1. YAML loader
# ---------------------------------------------------------------------------

class TestAccountingRulesLoader:

    def test_defaults_when_file_missing(self, tmp_path):
        """Loader should return built-in defaults when yaml missing."""
        with mock_patch(
            'src.transformers.management_accounts.Path',
            side_effect=lambda p: tmp_path / "nonexistent.yaml" if 'accounting_rules' in str(p) else Path(p),
        ):
            rules = ManagementAccountsBuilder._load_accounting_rules()

        assert rules['identity_alignment']['authoritative_row'] == 19
        assert rules['rounding']['decimal_rows'] == [26]
        assert 107 in rules['bank_bdo_row_map']

    def test_loads_custom_snap(self, tmp_path, monkeypatch):
        """Custom max_snap_units should be honoured."""
        rules_file = _write_rules(tmp_path, {
            'identity_alignment': {'max_snap_units': 5}
        })
        monkeypatch.chdir(tmp_path)
        os.makedirs(tmp_path / 'config', exist_ok=True)
        os.rename(rules_file, tmp_path / 'config' / 'accounting_rules.yaml')

        rules = ManagementAccountsBuilder._load_accounting_rules()
        assert rules['identity_alignment']['max_snap_units'] == 5

    def test_bank_map_keys_are_ints(self, tmp_path, monkeypatch):
        """YAML loads dict keys as ints even though YAML defaults to strings."""
        rules_file = _write_rules(tmp_path)
        monkeypatch.chdir(tmp_path)
        os.makedirs(tmp_path / 'config', exist_ok=True)
        os.rename(rules_file, tmp_path / 'config' / 'accounting_rules.yaml')

        rules = ManagementAccountsBuilder._load_accounting_rules()
        for k in rules['bank_bdo_row_map']:
            assert isinstance(k, int), f"Key {k} should be int"


# ---------------------------------------------------------------------------
# 2. Workbook identity formula enforcement
# ---------------------------------------------------------------------------

class TestIdentityFormulaEnforcement:

    def test_enforce_sets_formula(self, tmp_path, monkeypatch):
        """When enforce_in_workbook is True the dependent row gets ={COL}{auth}."""
        rules_file = _write_rules(tmp_path)
        monkeypatch.chdir(tmp_path)
        os.makedirs(tmp_path / 'config', exist_ok=True)
        os.rename(rules_file, tmp_path / 'config' / 'accounting_rules.yaml')

        wb = Workbook()
        ws = wb.active
        ws.title = "MC"
        ltm_col = 5
        ws.cell(row=68, column=ltm_col, value="=AA67+AA66")

        from src.transformers.management_accounts import ManagementAccountsConfig
        cfg = ManagementAccountsConfig(
            quarter="Q4 2025", period_end=None,
            bdo_sheet_name="BDO", summary_sheet_name="MC",
        )
        builder = ManagementAccountsBuilder.__new__(ManagementAccountsBuilder)
        builder._rules = ManagementAccountsBuilder._load_accounting_rules()
        builder._enforce_identity_formula(ws, ltm_col)

        expected = f"={get_column_letter(ltm_col)}19"
        assert ws.cell(row=68, column=ltm_col).value == expected

    def test_no_enforcement_when_disabled(self, tmp_path, monkeypatch):
        """When enforce_in_workbook is False the cell should remain unchanged."""
        rules_file = _write_rules(tmp_path, {
            'identity_alignment': {'enforce_in_workbook': False}
        })
        monkeypatch.chdir(tmp_path)
        os.makedirs(tmp_path / 'config', exist_ok=True)
        os.rename(rules_file, tmp_path / 'config' / 'accounting_rules.yaml')

        wb = Workbook()
        ws = wb.active
        ws.title = "MC"
        ltm_col = 5
        original = "=AA67+AA66"
        ws.cell(row=68, column=ltm_col, value=original)

        builder = ManagementAccountsBuilder.__new__(ManagementAccountsBuilder)
        builder._rules = ManagementAccountsBuilder._load_accounting_rules()
        builder._enforce_identity_formula(ws, ltm_col)
        assert ws.cell(row=68, column=ltm_col).value == original


# ---------------------------------------------------------------------------
# 3. PatchApplier protected-row rejection
# ---------------------------------------------------------------------------

class TestProtectedRowPatch:

    @pytest.fixture
    def sample_wb(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Management Cijfers - Q4 2025"
        for r in range(1, 120):
            ws.cell(row=r, column=1, value=f"Row {r}")
        path = tmp_path / "ma.xlsx"
        wb.save(path)
        wb.close()
        return path

    def test_value_patch_on_row19_rejected(self, sample_wb):
        applier = PatchApplier(str(sample_wb), "Management Cijfers - Q4 2025")
        patches = [{
            'sheet': "Management Cijfers - Q4 2025",
            'row': 19,
            'col': 3,
            'kind': 'value',
            'value': 999999,
        }]
        output = str(sample_wb).replace('.xlsx', '_out.xlsx')
        applied, rejected = applier.apply(patches, output)
        assert applied == 0
        assert len(rejected) == 1
        assert 'protected' in rejected[0]['rejection_reason'].lower()

    def test_formula_patch_on_row19_allowed(self, sample_wb):
        """Formula patches on protected rows should still be accepted."""
        applier = PatchApplier(str(sample_wb), "Management Cijfers - Q4 2025")
        patches = [{
            'sheet': "Management Cijfers - Q4 2025",
            'row': 19,
            'col': 3,
            'kind': 'formula',
            'value': "=SUM(C3:C18)",
        }]
        output = str(sample_wb).replace('.xlsx', '_out.xlsx')
        applied, rejected = applier.apply(patches, output)
        assert applied == 1
        assert len(rejected) == 0

    def test_value_patch_on_row68_rejected(self, sample_wb):
        applier = PatchApplier(str(sample_wb), "Management Cijfers - Q4 2025")
        patches = [{
            'sheet': "Management Cijfers - Q4 2025",
            'row': 68,
            'col': 3,
            'kind': 'value',
            'value': 123456,
        }]
        output = str(sample_wb).replace('.xlsx', '_out.xlsx')
        applied, rejected = applier.apply(patches, output)
        assert applied == 0
        assert len(rejected) == 1

    def test_value_patch_on_unprotected_row_ok(self, sample_wb):
        applier = PatchApplier(str(sample_wb), "Management Cijfers - Q4 2025")
        patches = [{
            'sheet': "Management Cijfers - Q4 2025",
            'row': 50,
            'col': 3,
            'kind': 'value',
            'value': 42,
        }]
        output = str(sample_wb).replace('.xlsx', '_out.xlsx')
        applied, rejected = applier.apply(patches, output)
        assert applied == 1
        assert len(rejected) == 0


# ---------------------------------------------------------------------------
# 4. Orchestrator _verify_and_refresh
# ---------------------------------------------------------------------------

class TestVerifyAndRefresh:

    def test_refresh_snaps_values(self, tmp_path):
        """After AI, computed_values should reflect the file's row-19 value."""
        ma_path = _make_mc_workbook(tmp_path, bs19_val=5000, pl68_val=5001)

        cv = {(19, 'ltm'): 4999.0, (68, 'ltm'): 4998.0}
        results = {'warnings': []}
        refreshed = QuarterlyReportOrchestrator._verify_and_refresh(
            ma_path, cv, results)

        assert refreshed[(19, 'ltm')] == 5000.0
        assert refreshed[(68, 'ltm')] == 5000.0

    def test_mismatch_over_tolerance_warns(self, tmp_path):
        """Large diff should generate a warning."""
        ma_path = _make_mc_workbook(tmp_path, bs19_val=5000, pl68_val=6000)

        cv = {(19, 'ltm'): 5000.0, (68, 'ltm'): 6000.0}
        results = {'warnings': []}
        QuarterlyReportOrchestrator._verify_and_refresh(ma_path, cv, results)

        assert any('differ by' in w for w in results['warnings'])

    def test_no_crash_on_missing_sheet(self, tmp_path):
        """Workbook without Management Cijfers should not crash."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Other Sheet"
        path = tmp_path / "no_mc.xlsx"
        wb.save(path)
        wb.close()

        cv = {(19, 'ltm'): 100.0}
        results = {'warnings': []}
        out = QuarterlyReportOrchestrator._verify_and_refresh(path, cv, results)
        assert out[(19, 'ltm')] == 100.0
