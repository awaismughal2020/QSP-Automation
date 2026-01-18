"""
Tests for BDO Parser - Schema Drift Handling
"""

import pytest
from pathlib import Path
from src.parsers.bdo_parser import BDOParser, BDOParseResult, AccountEntry


class TestBDOParser:
    """Test BDO parser with schema drift handling."""
    
    @pytest.fixture
    def parser(self):
        """Create BDO parser instance."""
        return BDOParser("config/account_mappings.yaml")
    
    def test_parse_8_column_format(self, parser, sample_bdo_file_8col):
        """Test parsing BDO file with 8 columns (old format)."""
        result = parser.parse(str(sample_bdo_file_8col))
        
        assert isinstance(result, BDOParseResult)
        assert len(result.accounts) > 0
        assert "1600000" in result.accounts
        assert result.quarter == "Q1 2019"
        assert result.schema_version in ["compact", "standard"]
    
    def test_parse_9_column_format(self, parser, sample_bdo_file_9col):
        """Test parsing BDO file with 9 columns (new format)."""
        result = parser.parse(str(sample_bdo_file_9col))
        
        assert isinstance(result, BDOParseResult)
        assert len(result.accounts) > 0
        assert "1600000" in result.accounts
        assert result.quarter == "Q3 2025"
        assert result.schema_version in ["standard", "extended"]
    
    def test_schema_drift_handling(self, parser, sample_bdo_file_8col, sample_bdo_file_9col):
        """Test that parser handles different column counts."""
        result_8col = parser.parse(str(sample_bdo_file_8col))
        result_9col = parser.parse(str(sample_bdo_file_9col))
        
        # Both should parse successfully
        assert result_8col.accounts
        assert result_9col.accounts
        
        # Critical accounts should be present in both
        assert "1600000" in result_8col.accounts
        assert "1600000" in result_9col.accounts
        assert "1600200" in result_8col.accounts
        assert "1600200" in result_9col.accounts
    
    def test_account_code_matching(self, parser, sample_bdo_file_9col):
        """Test wildcard account code matching."""
        result = parser.parse(str(sample_bdo_file_9col))
        
        # Get all cash accounts (11xxxxx)
        cash_accounts = [code for code in result.accounts.keys() if code.startswith("11")]
        assert len(cash_accounts) > 0
        
        # Get all equity accounts (05xxxxx, 06xxxxx)
        equity_accounts = [code for code in result.accounts.keys() 
                          if code.startswith("05") or code.startswith("06")]
        assert len(equity_accounts) > 0
    
    def test_calculate_totals(self, parser, sample_bdo_file_9col):
        """Test total calculations."""
        result = parser.parse(str(sample_bdo_file_9col))
        totals = parser.calculate_totals(result)
        
        # Balance sheet equation should hold
        total_assets = (totals['real_estate_net'] + 
                       totals['financial_fixed_assets'] + 
                       totals['current_assets'])
        total_liab_equity = totals['total_equity'] + totals['total_liabilities']
        
        # Should be within EUR 1 tolerance
        assert abs(total_assets - total_liab_equity) < 1.00
    
    def test_missing_critical_accounts(self, parser, tmp_path):
        """Test validation for missing critical accounts."""
        # Create a file without critical accounts
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "BalansenWinstverlies"
        
        headers = ["Grootboek", "Omschrijving", "Beginsaldo", "Eindsaldo"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Only non-critical accounts
        ws.cell(row=2, column=1, value="1100000")
        ws.cell(row=2, column=2, value="Kas")
        ws.cell(row=2, column=3, value=100000.00)
        ws.cell(row=2, column=4, value=100000.00)
        
        file_path = tmp_path / "missing_accounts.xlsx"
        wb.save(file_path)
        
        result = parser.parse(str(file_path))
        
        # Should have warnings for missing critical accounts
        assert len(result.warnings) > 0
        assert any("1600000" in warning or "1600200" in warning for warning in result.warnings)
    
    def test_period_extraction_from_filename(self, parser, sample_bdo_file_9col):
        """Test period extraction from filename."""
        result = parser.parse(str(sample_bdo_file_9col))
        
        assert result.period_end == "2025-09-30"
        assert result.quarter == "Q3 2025"
    
    def test_get_account_by_category(self, parser, sample_bdo_file_9col):
        """Test getting accounts by category."""
        result = parser.parse(str(sample_bdo_file_9col))
        
        # Get equity accounts
        equity_accounts = parser.get_account_by_category(result, "equity")
        assert len(equity_accounts) > 0
        
        # Get current assets
        current_assets = parser.get_account_by_category(result, "current_assets")
        assert len(current_assets) > 0
    
    def test_parse_number_formats(self, parser):
        """Test parsing various number formats."""
        # Test comma decimal
        assert parser._parse_number("1234,56") == 1234.56
        # Test dot thousand separator
        assert parser._parse_number("1.234,56") == 1234.56
        # Test negative
        assert parser._parse_number("-500,00") == -500.00
        # Test empty
        assert parser._parse_number("") == 0.0
        # Test dash
        assert parser._parse_number("-") == 0.0

