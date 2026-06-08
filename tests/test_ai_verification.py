"""
Tests for AI Verification of Management Accounts.

Covers:
- WorkbookContextExtractor
- VerificationPromptBuilder
- PatchApplier (allowlist enforcement, formula/value/format patches)
- ClaudeVerifier (response parsing, mock mode)
- BDOCopyVerifier
- AIVerificationOrchestrator (end-to-end with mock Claude)
- Synthetic mismatch scenarios
"""

import json
import pytest
from copy import copy
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from src.parsers.bdo_parser import AccountEntry, BDOParseResult
from src.transformers.ai_verification import (
    AIVerificationOrchestrator,
    BDOCopyVerifier,
    ClaudeVerifier,
    PatchApplier,
    VerificationConfig,
    VerificationPromptBuilder,
    VerificationResult,
    WorkbookContextExtractor,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def verification_config():
    return VerificationConfig(
        quarter="Q4 2025",
        year=2025,
        quarter_num=4,
        period_end="31-12-2025",
        bdo_sheet_name="BDO - Q4-25",
        summary_sheet_name="Management Cijfers - Q4 2025",
        previous_quarter_label="Q3 2025",
    )


@pytest.fixture
def sample_bdo_result():
    """Minimal BDO parse result with key accounts."""
    accounts = {
        "8000003": AccountEntry(code="8000003", name="Huuropbrengsten theoretisch",
                                opening_balance=0, closing_balance=-3200000,
                                mutations={"q4": -800000}, raw_row=10),
        "8000004": AccountEntry(code="8000004", name="Leegstand financieel",
                                opening_balance=0, closing_balance=-100000,
                                mutations={"q4": -25000}, raw_row=11),
        "8400600": AccountEntry(code="8400600", name="Servicekosten doorberekend",
                                opening_balance=0, closing_balance=-50000,
                                mutations={"q4": -12500}, raw_row=12),
        "8400700": AccountEntry(code="8400700", name="Leegstandskosten",
                                opening_balance=0, closing_balance=-5000,
                                mutations={"q4": -1250}, raw_row=13),
        "8400800": AccountEntry(code="8400800", name="Servicekosten",
                                opening_balance=0, closing_balance=-40000,
                                mutations={"q4": -10000}, raw_row=14),
        "4100400": AccountEntry(code="4100400", name="Onderhoud",
                                opening_balance=0, closing_balance=80000,
                                mutations={"q4": 20000}, raw_row=15),
        "4310400": AccountEntry(code="4310400", name="Verzekeringen",
                                opening_balance=0, closing_balance=10000,
                                mutations={"q4": 2500}, raw_row=16),
        "4051400": AccountEntry(code="4051400", name="Beheerkosten",
                                opening_balance=0, closing_balance=60000,
                                mutations={"q4": 15000}, raw_row=17),
        "4310500": AccountEntry(code="4310500", name="Managementfee aandeelhouder",
                                opening_balance=0, closing_balance=20000,
                                mutations={"q4": 5000}, raw_row=18),
        "9500000": AccountEntry(code="9500000", name="Vennootschapsbelasting",
                                opening_balance=0, closing_balance=30000,
                                mutations={"q4": 7500}, raw_row=19),
        "9510000": AccountEntry(code="9510000", name="DTA mutatie",
                                opening_balance=0, closing_balance=5000,
                                mutations={"q4": 1250}, raw_row=20),
        "1790002": AccountEntry(code="1790002", name="DTA",
                                opening_balance=100000, closing_balance=100000,
                                mutations={}, raw_row=6),
    }
    return BDOParseResult(
        accounts=accounts,
        period_end="2025-12-31",
        quarter="Q4 2025",
        column_mapping={},
        schema_version="standard",
        warnings=[],
    )


@pytest.fixture
def sample_workbook(tmp_path, verification_config):
    """Create a minimal MA workbook with BDO + Management Cijfers sheets."""
    path = tmp_path / "Management Accounts Q4 2025 - Draft 1.xlsx"
    wb = Workbook()

    bdo = wb.active
    bdo.title = verification_config.bdo_sheet_name
    bdo.cell(row=1, column=1, value="Grootboek")
    bdo.cell(row=1, column=2, value="Omschrijving")
    test_accounts = [
        (6, "1790002", "DTA"),
        (10, "8000003", "Huuropbrengsten theoretisch"),
        (11, "8000004", "Leegstand financieel"),
        (12, "8400600", "Servicekosten doorberekend"),
        (15, "4100400", "Onderhoud"),
        (17, "4051400", "Beheerkosten"),
        (18, "4310500", "Managementfee"),
        (19, "9500000", "VPB"),
        (20, "9510000", "DTA mutatie"),
    ]
    for row, code, name in test_accounts:
        bdo.cell(row=row, column=1, value=code)
        bdo.cell(row=row, column=2, value=name)
        bdo.cell(row=row, column=7, value=-100000)  # col G
        bdo.cell(row=row, column=8, value=f"=SUM(C{row}:G{row})")

    bdo.cell(row=125, column=2, value="Resultaat na belasting")
    bdo.cell(row=125, column=4, value=-50000)
    bdo.cell(row=125, column=5, value=-50000)
    bdo.cell(row=125, column=6, value=-50000)
    bdo.cell(row=125, column=7, value=-50000)

    mc = wb.create_sheet(verification_config.summary_sheet_name)
    mc.cell(row=1, column=1, value=f"Management Accounts QSP ESS B.V. - Q4 2025")
    mc.cell(row=2, column=27, value=datetime(2025, 12, 31))

    q_col = 27
    ltm_col = 28
    mc.cell(row=22, column=q_col, value="Q4 2025")
    mc.cell(row=22, column=ltm_col, value="LTM Q4 2025")

    q_letter = get_column_letter(q_col)
    ltm_letter = get_column_letter(ltm_col)

    for row in range(3, 20):
        mc.cell(row=row, column=q_col, value=None)
    mc.cell(row=3, column=ltm_col, value=f"='{verification_config.bdo_sheet_name}'!H6")
    mc.cell(row=19, column=ltm_col, value=f"=SUM({ltm_letter}3:{ltm_letter}18)")

    bdo_name = verification_config.bdo_sheet_name
    mc.cell(row=23, column=q_col, value=f"=-'{bdo_name}'!G10")
    mc.cell(row=23, column=ltm_col, value=f"=-'{bdo_name}'!H10")
    mc.cell(row=24, column=q_col, value=f"=-'{bdo_name}'!G11")
    mc.cell(row=24, column=ltm_col, value=f"=-'{bdo_name}'!H11")
    mc.cell(row=25, column=q_col, value=f"=SUM({q_letter}23:{q_letter}24)")
    mc.cell(row=25, column=ltm_col, value=f"=SUM({ltm_letter}23:{ltm_letter}24)")

    mc.cell(row=66, column=q_col, value=f"=SUM({q_letter}57:{q_letter}65)")
    mc.cell(row=66, column=ltm_col, value=f"=SUM({ltm_letter}57:{ltm_letter}65)")
    mc.cell(row=67, column=q_col, value=f"=-'{bdo_name}'!G19-'{bdo_name}'!G20")
    mc.cell(row=68, column=q_col, value=f"={q_letter}67+{q_letter}66")
    mc.cell(row=68, column=ltm_col, value=f"={ltm_letter}67+{ltm_letter}66")

    for row in range(107, 114):
        mc.cell(row=row, column=ltm_col, value=f"='{bdo_name}'!H{row - 75}")

    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def sample_bdo_source(tmp_path):
    """Create a sample BDO source file for copy verification."""
    path = tmp_path / "Cijfers_QSP_Q4_2025.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Grootboek")
    ws.cell(row=6, column=1, value="1790002")
    ws.cell(row=6, column=2, value="DTA")
    ws.cell(row=6, column=7, value=-100000)
    ws.cell(row=6, column=8, value="=SUM(C6:G6)")
    wb.save(str(path))
    wb.close()
    return path


# ---------------------------------------------------------------------------
# WorkbookContextExtractor Tests
# ---------------------------------------------------------------------------

class TestWorkbookContextExtractor:

    def test_extract_returns_all_keys(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        assert 'metadata' in context
        assert 'formula_templates' in context
        assert 'bdo_account_to_row_map' in context
        assert 'bdo_label_to_row_map' in context
        assert 'actual_cell_contents' in context
        assert 'formatting_snapshot' in context
        assert 'ltm_col' in context
        assert 'quarter_col' in context

    def test_bdo_row_map_populated(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        row_map = context['bdo_account_to_row_map']
        assert '1790002' in row_map
        assert row_map['1790002'] == 6
        assert '8000003' in row_map

    def test_cell_contents_extracted(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        contents = context['actual_cell_contents']
        assert 'balance_sheet' in contents
        assert 'profit_loss' in contents
        assert 'bank_accounts' in contents

        row_23 = contents['profit_loss'].get(23)
        assert row_23 is not None
        assert row_23['quarter']['is_formula'] is True

    def test_ltm_column_detected(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        assert context['ltm_col'] == 28
        assert context['quarter_col'] == 27

    def test_bdo_resultaat_extracted(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        resultaat = context.get('bdo_resultaat_na_belasting')
        assert resultaat is not None
        assert resultaat['row'] == 125
        assert resultaat['raw_total_D_G'] == -200000

    def test_metadata_correct(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        meta = context['metadata']
        assert meta['quarter'] == "Q4 2025"
        assert meta['year'] == 2025
        assert meta['bdo_sheet_name'] == "BDO - Q4-25"


# ---------------------------------------------------------------------------
# VerificationPromptBuilder Tests
# ---------------------------------------------------------------------------

class TestVerificationPromptBuilder:

    def test_prompt_contains_key_sections(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert 'ROLE:' in prompt
        assert 'CONTEXT:' in prompt
        assert 'SIGN CONVENTION:' in prompt
        assert 'FORMULA CONSTRUCTION RULES' in prompt
        assert 'CHECKS TO PERFORM:' in prompt
        assert 'RESPONSE FORMAT' in prompt

    def test_prompt_includes_bdo_map(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert '1790002' in prompt
        assert '8000003' in prompt

    def test_prompt_includes_sign_rules(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert 'NEGATIVE' in prompt
        assert 'POSITIVE' in prompt
        assert 'Dutch trial balance' in prompt

    def test_prompt_includes_formula_construction_guide(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert 'bdo_ref (single account' in prompt
        assert 'bdo_ref (multiple accounts' in prompt
        assert 'bdo_sum_range' in prompt
        assert 'bdo_ref_conditional' in prompt
        assert "BDO - Q4-25" in prompt

    def test_prompt_uses_real_column_letters(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        q_letter = context['quarter_col_letter']
        ltm_letter = context['ltm_col_letter']
        assert f'=SUM({q_letter}23:{q_letter}24)' in prompt
        assert f'=SUM({ltm_letter}3:{ltm_letter}18)' in prompt

    def test_prompt_includes_bank_account_mappings(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert "Row 107" in prompt
        assert "!H35" in prompt
        assert "Row 113" in prompt
        assert "!H31" in prompt

    def test_prompt_includes_formatting_spec(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert 'Avenir Book' in prompt
        assert 'size 10' in prompt
        assert '#,##0' in prompt

    def test_prompt_includes_pass_semantics(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert 'PASS' in prompt
        assert 'empty' in prompt.lower()
        assert 'Do NOT invent' in prompt

    def test_prompt_includes_shadow_models_when_present(self, sample_workbook, verification_config):
        extractor = WorkbookContextExtractor(
            str(sample_workbook), verification_config
        )
        context = extractor.extract()

        context['shadow_pl'] = {
            68: {'label': 'Direct result', 'value': 150000.0, 'type': 'calc',
                 'account_codes': [], 'missing_codes': []}
        }
        context['shadow_bs'] = {
            19: {'label': 'Total', 'value': 150000.0, 'type': 'calc',
                 'account_codes': [], 'missing_codes': []}
        }

        builder = VerificationPromptBuilder()
        prompt = builder.build(context)

        assert 'SHADOW P&L' in prompt
        assert 'SHADOW BALANCE SHEET' in prompt
        assert '150000.0' in prompt


# ---------------------------------------------------------------------------
# PatchApplier Tests
# ---------------------------------------------------------------------------

class TestPatchApplier:

    def test_apply_formula_patch(self, sample_workbook, verification_config):
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [{
            'sheet': verification_config.summary_sheet_name,
            'row': 23,
            'col': 27,
            'kind': 'formula',
            'value': "=-'BDO - Q4-25'!G10",
        }]

        output = str(sample_workbook).replace('.xlsx', '_patched.xlsx')
        applied, rejected = applier.apply(patches, output)

        assert applied == 1
        assert len(rejected) == 0

        wb = openpyxl.load_workbook(output)
        sheet = wb[verification_config.summary_sheet_name]
        assert sheet.cell(row=23, column=27).value == "=-'BDO - Q4-25'!G10"
        wb.close()

    def test_reject_patch_on_bdo_sheet(self, sample_workbook, verification_config):
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [{
            'sheet': 'BDO - Q4-25',
            'row': 10,
            'col': 7,
            'kind': 'value',
            'value': 999999,
        }]

        output = str(sample_workbook).replace('.xlsx', '_patched.xlsx')
        applied, rejected = applier.apply(patches, output)

        assert applied == 0
        assert len(rejected) == 1
        assert 'allowlist' in rejected[0].get('rejection_reason', '').lower()

    def test_reject_patch_outside_row_range(self, sample_workbook, verification_config):
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [{
            'sheet': verification_config.summary_sheet_name,
            'row': 200,
            'col': 27,
            'kind': 'value',
            'value': 0,
        }]

        output = str(sample_workbook).replace('.xlsx', '_patched.xlsx')
        applied, rejected = applier.apply(patches, output)

        assert applied == 0
        assert len(rejected) == 1

    def test_reject_invalid_formula(self, sample_workbook, verification_config):
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [{
            'sheet': verification_config.summary_sheet_name,
            'row': 23,
            'col': 27,
            'kind': 'formula',
            'value': "not a formula",
        }]

        output = str(sample_workbook).replace('.xlsx', '_patched.xlsx')
        applied, rejected = applier.apply(patches, output)

        assert applied == 0
        assert len(rejected) == 1

    def test_apply_value_patch(self, sample_workbook, verification_config):
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [{
            'sheet': verification_config.summary_sheet_name,
            'row': 50,
            'col': 27,
            'kind': 'value',
            'value': 762500,
        }]

        output = str(sample_workbook).replace('.xlsx', '_patched.xlsx')
        applied, rejected = applier.apply(patches, output)
        assert applied == 1

        wb = openpyxl.load_workbook(output)
        sheet = wb[verification_config.summary_sheet_name]
        assert sheet.cell(row=50, column=27).value == 762500
        wb.close()

    def test_apply_format_patch(self, sample_workbook, verification_config):
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [{
            'sheet': verification_config.summary_sheet_name,
            'row': 23,
            'col': 27,
            'kind': 'format',
            'value': {'number_format': '#,##0', 'bold': True},
        }]

        output = str(sample_workbook).replace('.xlsx', '_patched.xlsx')
        applied, rejected = applier.apply(patches, output)
        assert applied == 1

        wb = openpyxl.load_workbook(output)
        sheet = wb[verification_config.summary_sheet_name]
        cell = sheet.cell(row=23, column=27)
        assert cell.number_format == '#,##0'
        assert cell.font.bold is True
        wb.close()

    def test_multiple_patches_partial_apply(self, sample_workbook, verification_config):
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [
            {
                'sheet': verification_config.summary_sheet_name,
                'row': 23,
                'col': 27,
                'kind': 'formula',
                'value': "=-'BDO - Q4-25'!G10",
            },
            {
                'sheet': 'BDO - Q4-25',
                'row': 10,
                'col': 7,
                'kind': 'value',
                'value': 999999,
            },
            {
                'sheet': verification_config.summary_sheet_name,
                'row': 25,
                'col': 27,
                'kind': 'formula',
                'value': "=SUM(AA23:AA24)",
            },
        ]

        output = str(sample_workbook).replace('.xlsx', '_patched.xlsx')
        applied, rejected = applier.apply(patches, output)

        assert applied == 2
        assert len(rejected) == 1


# ---------------------------------------------------------------------------
# ClaudeVerifier Tests
# ---------------------------------------------------------------------------

class TestClaudeVerifier:

    def test_mock_pass_when_no_api_key(self):
        verifier = ClaudeVerifier(api_key='')
        result = verifier.verify("test prompt")

        assert result['status'] == 'PASS'
        assert 'Mock response' in result.get('notes', '')

    def test_parse_valid_json(self):
        verifier = ClaudeVerifier(api_key='test')
        response = json.dumps({
            'status': 'PASS',
            'checks_performed': 42,
            'issues': [],
            'patches': [],
            'validation_summary': {'all_match': True},
            'notes': 'All good',
        })

        parsed = verifier._parse_response(response)
        assert parsed['status'] == 'PASS'
        assert parsed['checks_performed'] == 42

    def test_parse_json_with_markdown_fence(self):
        verifier = ClaudeVerifier(api_key='test')
        response = '```json\n{"status": "PASS", "issues": []}\n```'

        parsed = verifier._parse_response(response)
        assert parsed['status'] == 'PASS'

    def test_parse_invalid_json_fallback(self):
        verifier = ClaudeVerifier(api_key='test')
        response = 'Some preamble text {"status": "ISSUES_FOUND", "issues": []} trailing text'

        parsed = verifier._parse_response(response)
        assert parsed['status'] == 'ISSUES_FOUND'

    def test_parse_totally_invalid(self):
        verifier = ClaudeVerifier(api_key='test')
        response = 'This is not JSON at all'

        parsed = verifier._parse_response(response)
        assert parsed['status'] == 'ERROR'


# ---------------------------------------------------------------------------
# BDOCopyVerifier Tests
# ---------------------------------------------------------------------------

class TestBDOCopyVerifier:

    def test_copy_integrity_check(self, sample_workbook, sample_bdo_source, verification_config):
        verifier = BDOCopyVerifier(
            str(sample_workbook), str(sample_bdo_source), verification_config
        )
        context = verifier.build_verification_context()

        assert 'integrity_checks' in context
        assert 'column_h_checks' in context

    def test_no_source_file(self, sample_workbook, verification_config):
        verifier = BDOCopyVerifier(
            str(sample_workbook), None, verification_config
        )
        context = verifier.build_verification_context()

        assert context['bdo_copy_verified'] is False
        assert 'not available' in context.get('note', '')


# ---------------------------------------------------------------------------
# AIVerificationOrchestrator Tests
# ---------------------------------------------------------------------------

class TestAIVerificationOrchestrator:

    def test_end_to_end_mock_pass(self, sample_workbook, verification_config, sample_bdo_result):
        """With no API key, Claude returns mock PASS — no patches applied."""
        orchestrator = AIVerificationOrchestrator(
            config=verification_config,
            api_key='',
        )

        result = orchestrator.run(
            workbook_path=str(sample_workbook),
            bdo_result=sample_bdo_result,
        )

        assert result.status == 'PASS'
        assert result.patches_applied == 0
        assert Path(result.verified_file).exists()
        assert Path(result.original_file).exists()

    def test_end_to_end_with_patches(self, sample_workbook, verification_config, sample_bdo_result):
        """
        Simulate Claude finding issues and returning patches.

        Under the single-Opus pipeline (``AI_VERIFY_MAX_ROUNDS=1``,
        the default) the orchestrator calls Claude exactly once and
        applies the returned patch in-place — no retry. We assert the
        patch landed and the cell value reflects Claude's response.
        """
        mock_response = {
            'status': 'ISSUES_FOUND',
            'checks_performed': 50,
            'issues': [{
                'sheet': verification_config.summary_sheet_name,
                'row': 23,
                'col': 27,
                'col_letter': 'AA',
                'kind': 'formula',
                'current': "=-'BDO - Q4-25'!G99",
                'expected': "=-'BDO - Q4-25'!G10",
                'explanation': 'Wrong BDO row reference',
            }],
            'patches': [{
                'sheet': verification_config.summary_sheet_name,
                'row': 23,
                'col': 27,
                'kind': 'formula',
                'value': "=-'BDO - Q4-25'!G10",
            }],
            'validation_summary': {
                'bs_row_19': None,
                'pl_row_68': None,
                'bdo_resultaat_adjusted': None,
                'all_match': False,
            },
            'notes': 'Fixed wrong BDO row reference in row 23',
        }

        with patch.object(ClaudeVerifier, 'verify', return_value=mock_response) as mock_verify:
            orchestrator = AIVerificationOrchestrator(
                config=verification_config,
                api_key='fake-key',
            )

            result = orchestrator.run(
                workbook_path=str(sample_workbook),
                bdo_result=sample_bdo_result,
            )

        assert result.status == 'ISSUES_FOUND'
        # Single-Opus: exactly one Claude call, exactly one patch applied.
        assert mock_verify.call_count == 1
        assert result.patches_applied == 1
        assert result.claude_rounds == 1
        assert len(result.issues) == 1
        assert Path(result.verified_file).exists()

        wb = openpyxl.load_workbook(result.verified_file)
        sheet = wb[verification_config.summary_sheet_name]
        assert sheet.cell(row=23, column=27).value == "=-'BDO - Q4-25'!G10"
        wb.close()

    def test_claude_error_preserves_original(self, sample_workbook, verification_config, sample_bdo_result):
        """If Claude returns ERROR, verified file is a copy of original."""
        mock_response = {'status': 'ERROR', 'error': 'API timeout'}

        with patch.object(ClaudeVerifier, 'verify', return_value=mock_response):
            orchestrator = AIVerificationOrchestrator(
                config=verification_config,
                api_key='fake-key',
            )

            result = orchestrator.run(
                workbook_path=str(sample_workbook),
                bdo_result=sample_bdo_result,
            )

        assert result.status == 'ERROR'
        assert result.patches_applied == 0
        assert Path(result.verified_file).exists()

    def test_with_bdo_source_verification(self, sample_workbook, sample_bdo_source,
                                          verification_config, sample_bdo_result):
        """BDO copy verification context is included when source is provided."""
        orchestrator = AIVerificationOrchestrator(
            config=verification_config,
            api_key='',
        )

        result = orchestrator.run(
            workbook_path=str(sample_workbook),
            bdo_source_path=str(sample_bdo_source),
            bdo_result=sample_bdo_result,
        )

        assert result.status == 'PASS'
        assert Path(result.verified_file).exists()


# ---------------------------------------------------------------------------
# Single Opus Verification Pipeline
# ---------------------------------------------------------------------------

class TestSingleOpusPipeline:
    """
    Verify the default single-Opus pipeline: exactly one Claude call per
    run, pre-flight evaluator output injected into the prompt, and the
    legacy multi-round retry loop available behind the
    ``AI_VERIFY_MAX_ROUNDS`` env override.
    """

    def test_default_max_rounds_is_one(self, monkeypatch):
        """``AI_VERIFY_MAX_ROUNDS`` unset → resolved value is 1."""
        monkeypatch.delenv('AI_VERIFY_MAX_ROUNDS', raising=False)
        assert AIVerificationOrchestrator._resolve_max_rounds() == 1

    def test_legacy_multi_round_env(self, monkeypatch):
        """``AI_VERIFY_MAX_ROUNDS=3`` re-enables the legacy retry loop."""
        monkeypatch.setenv('AI_VERIFY_MAX_ROUNDS', '3')
        assert AIVerificationOrchestrator._resolve_max_rounds() == 3

    def test_invalid_env_falls_back_to_default(self, monkeypatch):
        monkeypatch.setenv('AI_VERIFY_MAX_ROUNDS', 'not-an-int')
        assert AIVerificationOrchestrator._resolve_max_rounds() == 1

    def test_zero_env_clamped_to_one(self, monkeypatch):
        monkeypatch.setenv('AI_VERIFY_MAX_ROUNDS', '0')
        assert AIVerificationOrchestrator._resolve_max_rounds() == 1

    def test_single_round_only_one_claude_call(
        self, sample_workbook, verification_config, sample_bdo_result, monkeypatch
    ):
        """
        Even when revalidation reports failure, the single-Opus pipeline
        must call Claude exactly once. The legacy 3-round loop would
        retry on failure; that behaviour is now opt-in via the env var.
        """
        monkeypatch.delenv('AI_VERIFY_MAX_ROUNDS', raising=False)

        mock_response = {
            'status': 'ISSUES_FOUND',
            'issues': [],
            'patches': [],
            'validation_summary': {},
            'notes': '',
        }

        with patch.object(ClaudeVerifier, 'verify', return_value=mock_response) as mock_verify, \
             patch.object(
                 AIVerificationOrchestrator,
                 '_revalidate_detailed',
                 return_value={'passed': False, 'failure_reasons': ['mock failure']},
             ):
            orchestrator = AIVerificationOrchestrator(
                config=verification_config,
                api_key='fake-key',
            )
            result = orchestrator.run(
                workbook_path=str(sample_workbook),
                bdo_result=sample_bdo_result,
            )

        assert mock_verify.call_count == 1
        assert result.claude_rounds == 1
        assert result.preflight_passed is False
        assert result.postflight_passed is False

    def test_preflight_report_injected_in_prompt(
        self, sample_workbook, verification_config, sample_bdo_result, monkeypatch
    ):
        """
        The pre-flight re-validator output must reach the prompt under
        the ``PRE-FLIGHT EVALUATOR`` header before Claude is called.
        """
        monkeypatch.delenv('AI_VERIFY_MAX_ROUNDS', raising=False)

        preflight_report = {
            'passed': False,
            'failure_reasons': ['Synthetic preflight failure'],
            'auth_row_value': 1234.5,
            'dep_row_value': 1230.0,
            'bdo_ground_truth': 1234.5,
            'bdo_ground_truth_quarter': 1000.0,
            'evaluated_ltm': {3: 100.0, 19: 1234.5},
            'evaluated_quarter': {3: 80.0, 19: 1000.0},
            'ltm_row_mismatches': [
                {'row': 5, 'actual': "='X'!H10", 'expected': "='X'!H11"},
            ],
        }

        captured_prompts = []

        def fake_verify(self, prompt):
            captured_prompts.append(prompt)
            return {
                'status': 'PASS',
                'issues': [],
                'patches': [],
                'validation_summary': {},
                'notes': '',
            }

        with patch.object(ClaudeVerifier, 'verify', fake_verify), \
             patch.object(
                 AIVerificationOrchestrator,
                 '_revalidate_detailed',
                 return_value=preflight_report,
             ):
            orchestrator = AIVerificationOrchestrator(
                config=verification_config,
                api_key='fake-key',
            )
            orchestrator.run(
                workbook_path=str(sample_workbook),
                bdo_result=sample_bdo_result,
            )

        assert len(captured_prompts) == 1
        prompt = captured_prompts[0]
        assert 'PRE-FLIGHT EVALUATOR' in prompt
        assert 'Synthetic preflight failure' in prompt
        assert 'EVALUATED LTM COLUMN VALUES' in prompt

    def test_legacy_retry_loop_runs_multiple_rounds(
        self, sample_workbook, verification_config, sample_bdo_result, monkeypatch
    ):
        """With ``AI_VERIFY_MAX_ROUNDS=3``, a failing revalidation triggers retries."""
        monkeypatch.setenv('AI_VERIFY_MAX_ROUNDS', '3')

        mock_response = {
            'status': 'ISSUES_FOUND',
            'issues': [],
            'patches': [],
            'validation_summary': {},
            'notes': '',
        }

        with patch.object(ClaudeVerifier, 'verify', return_value=mock_response) as mock_verify, \
             patch.object(
                 AIVerificationOrchestrator,
                 '_revalidate_detailed',
                 return_value={'passed': False, 'failure_reasons': ['mock']},
             ):
            orchestrator = AIVerificationOrchestrator(
                config=verification_config,
                api_key='fake-key',
            )
            result = orchestrator.run(
                workbook_path=str(sample_workbook),
                bdo_result=sample_bdo_result,
            )

        assert mock_verify.call_count == 3
        assert result.claude_rounds == 3

    def test_timing_attached_to_result(
        self, sample_workbook, verification_config, sample_bdo_result
    ):
        """Verification result carries per-phase wall-clock timings."""
        with patch.object(
            ClaudeVerifier, 'verify',
            return_value={
                'status': 'PASS',
                'issues': [],
                'patches': [],
                'validation_summary': {},
                'notes': '',
            },
        ):
            orchestrator = AIVerificationOrchestrator(
                config=verification_config,
                api_key='',
            )
            result = orchestrator.run(
                workbook_path=str(sample_workbook),
                bdo_result=sample_bdo_result,
            )

        assert isinstance(result.timing, dict)
        assert 'context_extract' in result.timing
        assert 'preflight' in result.timing
        assert 'claude' in result.timing
        assert 'total' in result.timing


# ---------------------------------------------------------------------------
# Synthetic Mismatch Scenarios
# ---------------------------------------------------------------------------

class TestSyntheticMismatches:

    def test_wrong_bdo_row_reference(self, tmp_path, verification_config, sample_bdo_result):
        """Simulate a formula referencing wrong BDO row — Claude should catch it."""
        path = tmp_path / "mismatch_test.xlsx"
        wb = Workbook()

        bdo = wb.active
        bdo.title = verification_config.bdo_sheet_name
        bdo.cell(row=10, column=1, value="8000003")
        bdo.cell(row=10, column=2, value="Huuropbrengsten")
        bdo.cell(row=10, column=7, value=-800000)
        bdo.cell(row=99, column=1, value="WRONG")

        mc = wb.create_sheet(verification_config.summary_sheet_name)
        mc.cell(row=22, column=27, value="Q4 2025")
        mc.cell(row=22, column=28, value="LTM Q4 2025")
        mc.cell(row=23, column=27, value=f"=-'{verification_config.bdo_sheet_name}'!G99")

        wb.save(str(path))
        wb.close()

        extractor = WorkbookContextExtractor(str(path), verification_config)
        context = extractor.extract()

        row_23 = context['actual_cell_contents']['profit_loss'][23]
        formula = row_23['quarter']['value']
        assert 'G99' in formula

        bdo_map = context['bdo_account_to_row_map']
        assert bdo_map.get('8000003') == 10

    def test_missing_account_code(self, tmp_path, verification_config):
        """Account code in template but not in BDO sheet."""
        path = tmp_path / "missing_code_test.xlsx"
        wb = Workbook()

        bdo = wb.active
        bdo.title = verification_config.bdo_sheet_name
        bdo.cell(row=10, column=1, value="8000003")

        mc = wb.create_sheet(verification_config.summary_sheet_name)
        mc.cell(row=22, column=27, value="Q4 2025")
        mc.cell(row=22, column=28, value="LTM Q4 2025")

        wb.save(str(path))
        wb.close()

        extractor = WorkbookContextExtractor(str(path), verification_config)
        context = extractor.extract()

        bdo_map = context['bdo_account_to_row_map']
        assert '8000004' not in bdo_map

    def test_sign_flip_detection(self, tmp_path, verification_config, sample_bdo_result):
        """P&L formula without negation when sign: '-' is required."""
        path = tmp_path / "sign_flip_test.xlsx"
        wb = Workbook()

        bdo = wb.active
        bdo.title = verification_config.bdo_sheet_name
        bdo.cell(row=10, column=1, value="8000003")
        bdo.cell(row=10, column=7, value=-800000)

        mc = wb.create_sheet(verification_config.summary_sheet_name)
        mc.cell(row=22, column=27, value="Q4 2025")
        mc.cell(row=22, column=28, value="LTM Q4 2025")
        mc.cell(row=23, column=27, value=f"='{verification_config.bdo_sheet_name}'!G10")

        wb.save(str(path))
        wb.close()

        extractor = WorkbookContextExtractor(str(path), verification_config)
        context = extractor.extract()

        formula = context['actual_cell_contents']['profit_loss'][23]['quarter']['value']
        assert not formula.startswith('=-')

    def test_bs_quarter_should_be_empty(self, tmp_path, verification_config):
        """Balance Sheet quarter column should be empty (only LTM has formulas)."""
        path = tmp_path / "bs_quarter_test.xlsx"
        wb = Workbook()

        bdo = wb.active
        bdo.title = verification_config.bdo_sheet_name

        mc = wb.create_sheet(verification_config.summary_sheet_name)
        mc.cell(row=22, column=27, value="Q4 2025")
        mc.cell(row=22, column=28, value="LTM Q4 2025")
        mc.cell(row=3, column=27, value=12345)

        wb.save(str(path))
        wb.close()

        extractor = WorkbookContextExtractor(str(path), verification_config)
        context = extractor.extract()

        row_3_q = context['actual_cell_contents']['balance_sheet'][3]['quarter']
        assert row_3_q['value'] == 12345


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

class TestEdgeCases:

    def test_extractor_handles_missing_bdo_sheet(self, tmp_path, verification_config):
        """Gracefully handle workbook without BDO sheet."""
        path = tmp_path / "no_bdo.xlsx"
        wb = Workbook()
        mc = wb.active
        mc.title = verification_config.summary_sheet_name
        mc.cell(row=22, column=27, value="Q4 2025")
        mc.cell(row=22, column=28, value="LTM Q4 2025")
        wb.save(str(path))
        wb.close()

        extractor = WorkbookContextExtractor(str(path), verification_config)
        context = extractor.extract()

        assert context['bdo_account_to_row_map'] == {}

    def test_extractor_handles_missing_summary_sheet(self, tmp_path, verification_config):
        """Gracefully handle workbook without summary sheet."""
        path = tmp_path / "no_summary.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "SomeOtherSheet"
        wb.save(str(path))
        wb.close()

        extractor = WorkbookContextExtractor(str(path), verification_config)
        context = extractor.extract()

        assert context['actual_cell_contents'] == {}

    def test_patch_applier_handles_missing_sheet_gracefully(self, sample_workbook, verification_config):
        """PatchApplier resolves to existing MC sheet if exact name not found."""
        applier = PatchApplier(str(sample_workbook), verification_config.summary_sheet_name)

        patches = [{
            'sheet': 'Management Cijfers - Q99 2099',
            'row': 23,
            'col': 27,
            'kind': 'formula',
            'value': '=1+1',
        }]

        output = str(sample_workbook).replace('.xlsx', '_fallback.xlsx')
        applied, rejected = applier.apply(patches, output)

        assert applied == 1
