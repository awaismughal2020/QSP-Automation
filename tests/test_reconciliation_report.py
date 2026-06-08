"""
Tests for post-AI reconciliation reporting (formula evaluator warnings).
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from src.transformers.reconciliation_report import (
    build_reconciliation_report,
    format_reconciliation_messages,
    read_bdo_ground_truth_pair,
)


@pytest.fixture
def synthetic_ma_workbook(tmp_path):
    path = tmp_path / "ma.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Management Cijfers - Q1 2026"

    bdo = wb.create_sheet("BDO - Q1-26")
    bdo.cell(row=50, column=7, value=-100.0)
    bdo.cell(row=50, column=8, value=-1000.0)
    bdo.cell(row=51, column=7, value=-200.0)
    bdo.cell(row=51, column=8, value=-2000.0)
    bdo.cell(row=130, column=8, value=50.0)
    bdo.cell(row=125, column=2, value="Resultaat na belasting")
    bdo.cell(row=125, column=4, value=-100.0)
    bdo.cell(row=125, column=5, value=-100.0)
    bdo.cell(row=125, column=6, value=-100.0)
    bdo.cell(row=125, column=7, value=-100.0)
    bdo.cell(row=125, column=8, value=-400.0)

    q_col = 29
    ltm_col = 30
    ws.cell(row=22, column=q_col, value="Q1 2026")
    ws.cell(row=22, column=ltm_col, value="LTM Q1 2026")
    ws.cell(row=3, column=q_col, value="=-'BDO - Q1-26'!G50")
    ws.cell(row=3, column=ltm_col, value="=-'BDO - Q1-26'!H50")
    ws.cell(row=4, column=q_col, value="=-'BDO - Q1-26'!G51")
    ws.cell(row=4, column=ltm_col, value="=-'BDO - Q1-26'!H51")
    ws.cell(row=19, column=ltm_col, value="=SUM(AD3:AD18)-'BDO - Q1-26'!H130")
    ws.cell(row=66, column=ltm_col, value="=SUM(AD57:AD65)")
    ws.cell(row=67, column=ltm_col, value=42.0)
    ws.cell(row=68, column=ltm_col, value="=AD66+AD67")

    wb.save(str(path))
    wb.close()
    return path


class TestReadBdoGroundTruthPair:

    def test_reads_ltm_and_quarter(self, synthetic_ma_workbook):
        pair = read_bdo_ground_truth_pair(
            str(synthetic_ma_workbook),
            "BDO - Q1-26",
        )
        assert pair['ltm'] == pytest.approx(400.0)
        assert pair['quarter'] == pytest.approx(400.0)


class TestBuildReconciliationReport:

    def test_report_has_evaluated_totals(self, synthetic_ma_workbook):
        report = build_reconciliation_report(
            str(synthetic_ma_workbook),
            "Management Cijfers - Q1 2026",
            "BDO - Q1-26",
        )
        assert report['source'] == 'formula_evaluator_post_ai'
        assert report['ltm_col_letter'] == 'AD'
        assert isinstance(report.get('auth_row_value'), (int, float))
        assert isinstance(report.get('dep_row_value'), (int, float))
        assert report['auth_row_value'] == pytest.approx(2950.0)
        assert report['dep_row_value'] == pytest.approx(42.0)

    def test_messages_label_formula_evaluator(self, synthetic_ma_workbook):
        report = build_reconciliation_report(
            str(synthetic_ma_workbook),
            "Management Cijfers - Q1 2026",
            "BDO - Q1-26",
        )
        messages = format_reconciliation_messages(report)
        assert any(
            'Post-AI reconciliation (formula evaluator' in m
            for m in messages
        )
        assert any('AD19' in m or 'evaluated' in m.lower() for m in messages)
        assert not any('Template check' in m for m in messages)
