"""
Script to create test fixture Excel files.
Run this to generate the fixture files used in tests.
"""

from pathlib import Path
from openpyxl import Workbook

fixtures_dir = Path(__file__).parent


def create_bdo_8col():
    """Create BDO file with 8 columns (old format)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BalansenWinstverlies"
    
    # Header rows
    ws['A1'] = "BDO Quarterly Report"
    ws['A2'] = "Period: Q1 2019"
    ws['A3'] = "Date: 31-03-2019"
    
    # Column headers (row 4)
    headers = ["Grootboek", "Omschrijving", "Beginsaldo", "Mutatie 1", "Mutatie 2", 
               "Mutatie 3", "Mutatie 4", "Eindsaldo"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)
    
    # Sample account data
    accounts = [
        ("1600000", "Verkrijgingsprijs", 10000000.00, 0, 0, 0, 0, 10000000.00),
        ("1600200", "Cumulatieve afschrijving", -500000.00, -25000.00, 0, 0, 0, -525000.00),
        ("1760000", "Te vorderen DMRRP", 500000.00, 0, 0, 0, 0, 500000.00),
        ("1100000", "Kas", 100000.00, 50000.00, 0, 0, 0, 150000.00),
        ("1300000", "Debiteuren", 200000.00, 0, 0, 0, 0, 200000.00),
        ("0500000", "Aandelenkapitaal", 5000000.00, 0, 0, 0, 0, 5000000.00),
        ("0600000", "Ingehouden winst", 4500000.00, 25000.00, 0, 0, 0, 4525000.00),
        ("0900000", "Senior lening", 3000000.00, 0, 0, 0, 0, 3000000.00),
        ("8000000", "Huuropbrengsten", 0, 500000.00, 0, 0, 0, 500000.00),
        ("4000000", "Vastgoed kosten", 0, -100000.00, 0, 0, 0, -100000.00),
    ]
    
    for row_idx, account_data in enumerate(accounts, 5):
        for col_idx, value in enumerate(account_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(fixtures_dir / "bdo_q1_2019_sample.xlsx")
    print("Created bdo_q1_2019_sample.xlsx")


def create_bdo_9col():
    """Create BDO file with 9 columns (new format)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BalansenWinstverlies"
    
    # Header rows
    ws['A1'] = "BDO Quarterly Report"
    ws['A2'] = "Period: Q3 2025"
    ws['A3'] = "Date: 30-09-2025"
    
    # Column headers (row 4)
    headers = ["Grootboek", "Omschrijving", "Beginsaldo", "Mutatie 1", "Mutatie 2", 
               "Mutatie 3", "Mutatie 4", "Mutatie 5", "Eindsaldo"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)
    
    # Sample account data
    accounts = [
        ("1600000", "Verkrijgingsprijs", 12000000.00, 0, 0, 0, 0, 0, 12000000.00),
        ("1600200", "Cumulatieve afschrijving", -600000.00, -30000.00, 0, 0, 0, 0, -630000.00),
        ("1601000", "Verkoop verkrijgingsprijs", 0, -200000.00, 0, 0, 0, 0, -200000.00),
        ("1760000", "Te vorderen DMRRP", 600000.00, 0, 0, 0, 0, 0, 600000.00),
        ("1790002", "Latente belastingvordering", 100000.00, 0, 0, 0, 0, 0, 100000.00),
        ("1100000", "Kas", 150000.00, 50000.00, 0, 0, 0, 0, 200000.00),
        ("1110000", "Bank", 500000.00, 100000.00, 0, 0, 0, 0, 600000.00),
        ("1300000", "Debiteuren", 250000.00, 0, 0, 0, 0, 0, 250000.00),
        ("0500000", "Aandelenkapitaal", 5000000.00, 0, 0, 0, 0, 0, 5000000.00),
        ("0510000", "Agio", 1000000.00, 0, 0, 0, 0, 0, 1000000.00),
        ("0600000", "Ingehouden winst", 5000000.00, 120000.00, 0, 0, 0, 0, 5120000.00),
        ("0900000", "Senior lening", 3500000.00, 0, 0, 0, 0, 0, 3500000.00),
        ("0920000", "Opgelopen rente", 50000.00, 25000.00, 0, 0, 0, 0, 75000.00),
        ("8000000", "Huuropbrengsten", 0, 600000.00, 0, 0, 0, 0, 600000.00),
        ("8100000", "Servicekosten", 0, 50000.00, 0, 0, 0, 0, 50000.00),
        ("4000000", "Vastgoed kosten", 0, -120000.00, 0, 0, 0, 0, -120000.00),
        ("4200000", "Onderhoud", 0, -50000.00, 0, 0, 0, 0, -50000.00),
        ("4300000", "Afschrijving", 0, -30000.00, 0, 0, 0, 0, -30000.00),
        ("4400000", "Rentelasten", 0, -25000.00, 0, 0, 0, 0, -25000.00),
    ]
    
    for row_idx, account_data in enumerate(accounts, 5):
        for col_idx, value in enumerate(account_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(fixtures_dir / "bdo_q3_2025_sample.xlsx")
    print("Created bdo_q3_2025_sample.xlsx")


if __name__ == '__main__':
    create_bdo_8col()
    create_bdo_9col()
    print("All fixture files created successfully!")

