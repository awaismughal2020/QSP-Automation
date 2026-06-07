"""
Tests for the formula mirror and column evaluator used to keep the LTM
column in lock-step with the quarter column.

Covers:
- ``mirror_quarter_formula_to_ltm`` swapping G->H and AC->AD.
- ``evaluate_ma_column`` resolving BDO refs, in-sheet refs, ranges.
- ``ManagementAccountsBuilder.evaluate_ma_column_from_formulas`` end-to-end.
- Regression for the false-positive revalidation: ``_revalidate_detailed``
  now drives off the formula evaluator instead of trusting a pre-computed
  cache that was being overwritten by ``_enforce_bdo_alignment``.
"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook

from src.transformers.formula_mirror import (
    evaluate_ma_column,
    mirror_quarter_column_to_ltm,
    mirror_quarter_formula_to_ltm,
)
from src.transformers.management_accounts import ManagementAccountsBuilder


# ---------------------------------------------------------------------------
# Mirror utility — pure-string rewrite, no I/O.
# ---------------------------------------------------------------------------

class TestMirrorQuarterFormulaToLtm:

    def test_swaps_bdo_g_to_h(self):
        formula = "=-'BDO - Q4-25'!G50"
        result = mirror_quarter_formula_to_ltm(
            formula,
            bdo_sheet_name="BDO - Q1-26",
            quarter_col_letter="AC",
            ltm_col_letter="AD",
        )
        assert result == "=-'BDO - Q1-26'!H50"

    def test_swaps_quarter_letter_to_ltm_letter_in_calc(self):
        formula = "=SUM(AC23:AC24)"
        result = mirror_quarter_formula_to_ltm(
            formula,
            bdo_sheet_name="BDO",
            quarter_col_letter="AC",
            ltm_col_letter="AD",
        )
        assert result == "=SUM(AD23:AD24)"

    def test_combined_external_and_inhouse_refs(self):
        formula = "=AC23+AC24-'BDO'!G50-'BDO'!G51"
        result = mirror_quarter_formula_to_ltm(
            formula,
            bdo_sheet_name="BDO",
            quarter_col_letter="AC",
            ltm_col_letter="AD",
        )
        assert result == "=AD23+AD24-'BDO'!H50-'BDO'!H51"

    def test_leaves_non_bdo_external_refs_unchanged_except_sheet(self):
        # External refs always rewrite the sheet to current bdo, but only
        # swap the column when it matches the quarter BDO column.
        formula = "=-'OldBDO'!G42-'OldBDO'!Z99"
        result = mirror_quarter_formula_to_ltm(
            formula,
            bdo_sheet_name="BDO",
            quarter_col_letter="AC",
            ltm_col_letter="AD",
        )
        assert result == "=-'BDO'!H42-'BDO'!Z99"

    def test_non_formula_passes_through(self):
        assert mirror_quarter_formula_to_ltm(
            "plain text",
            bdo_sheet_name="BDO",
            quarter_col_letter="AC",
            ltm_col_letter="AD",
        ) == "plain text"


class TestMirrorQuarterColumnToLtm:

    def test_skip_rows_are_not_overwritten(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        # Row 23 in AC: BDO ref -> mirror to AD as H ref.
        ws.cell(row=23, column=29, value="=-'BDO'!G10")
        ws.cell(row=23, column=30, value="STALE")
        # Row 19 in AC: total — must NOT be mirrored (in skip set).
        ws.cell(row=19, column=29, value="=SUM(AC3:AC18)")
        ws.cell(row=19, column=30, value="=KEEP_ME")
        rewrites = mirror_quarter_column_to_ltm(
            ws,
            quarter_col=29,
            ltm_col=30,
            bdo_sheet_name="BDO",
            rows=[19, 23],
            skip_rows={19},
        )
        assert ws.cell(row=23, column=30).value == "=-'BDO'!H10"
        assert ws.cell(row=19, column=30).value == "=KEEP_ME"
        assert [r for r, _, _ in rewrites] == [23]


# ---------------------------------------------------------------------------
# Column evaluator — against a synthetic BDO + MA sheet.
# ---------------------------------------------------------------------------

@pytest.fixture
def synthetic_workbook(tmp_path):
    """A minimal MA workbook the evaluator can resolve end-to-end."""
    path = tmp_path / "ma.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Management Cijfers - Q1 2026"

    bdo = wb.create_sheet("BDO")
    bdo.cell(row=50, column=7, value=-100.0)   # G
    bdo.cell(row=51, column=7, value=-200.0)
    bdo.cell(row=50, column=8, value=-1000.0)  # H
    bdo.cell(row=51, column=8, value=-2000.0)
    bdo.cell(row=130, column=8, value=50.0)    # verschil

    q_col = 29   # AC
    ltm_col = 30  # AD
    ws.cell(row=22, column=q_col, value="Q1 2026")
    ws.cell(row=22, column=ltm_col, value="LTM Q1 2026")

    ws.cell(row=3, column=q_col, value="=-'BDO'!G50")
    ws.cell(row=3, column=ltm_col, value="=-'BDO'!H50")
    ws.cell(row=4, column=q_col, value="=-'BDO'!G51")
    ws.cell(row=4, column=ltm_col, value="=-'BDO'!H51")
    ws.cell(row=19, column=q_col, value="=SUM(AC3:AC18)")
    ws.cell(row=19, column=ltm_col, value="=SUM(AD3:AD18)-'BDO'!H130")
    ws.cell(row=66, column=ltm_col, value="=SUM(AD57:AD65)")
    ws.cell(row=67, column=ltm_col, value=42.0)
    ws.cell(row=68, column=ltm_col, value="=AD66+AD67")

    wb.save(str(path))
    wb.close()
    return path


class TestEvaluateMaColumn:

    def test_resolves_bdo_refs_and_internal_sums(self, synthetic_workbook):
        result = evaluate_ma_column(
            str(synthetic_workbook),
            summary_sheet_name="Management Cijfers - Q1 2026",
            bdo_sheet_name="BDO",
            target_col=30,
            rows_of_interest=[3, 4, 5, 19, 66, 67, 68],
        )
        values = result["values"]
        assert values[3] == pytest.approx(1000.0)
        assert values[4] == pytest.approx(2000.0)
        # Row 19 = 1000 + 2000 (other rows empty -> 0) - 50 verschil = 2950
        assert values[19] == pytest.approx(2950.0)
        # P&L chain
        assert values[68] == pytest.approx(42.0)

    def test_quarter_column_uses_g_independently(self, synthetic_workbook):
        result = evaluate_ma_column(
            str(synthetic_workbook),
            summary_sheet_name="Management Cijfers - Q1 2026",
            bdo_sheet_name="BDO",
            target_col=29,
            rows_of_interest=[3, 4, 19],
        )
        values = result["values"]
        assert values[3] == pytest.approx(100.0)
        assert values[4] == pytest.approx(200.0)
        assert values[19] == pytest.approx(300.0)

    def test_wrong_h_row_changes_evaluated_total(self, synthetic_workbook):
        """If an upstream LTM row points to a wrong BDO row, the evaluator
        returns the wrong total — this is the desired signal that the
        revalidation loop now uses to detect AD ground-truth misses."""
        wb = openpyxl.load_workbook(str(synthetic_workbook))
        ws = wb["Management Cijfers - Q1 2026"]
        # Replace the AD3 BDO ref with a row that exists but holds 0.
        ws.cell(row=3, column=30, value="=-'BDO'!H999")
        wb.save(str(synthetic_workbook))
        wb.close()

        result = evaluate_ma_column(
            str(synthetic_workbook),
            summary_sheet_name="Management Cijfers - Q1 2026",
            bdo_sheet_name="BDO",
            target_col=30,
            rows_of_interest=[3, 4, 19],
        )
        # Row 3 now resolves to 0 (BDO!H999 is empty), so row 19 is
        # 0 + 2000 - 50 = 1950, not 2950.
        assert result["values"][19] == pytest.approx(1950.0)


class TestEvaluateMaColumnFromFormulas:

    def test_static_method_resolves_ltm_total(self, synthetic_workbook):
        result = ManagementAccountsBuilder.evaluate_ma_column_from_formulas(
            str(synthetic_workbook),
            "Management Cijfers - Q1 2026",
            "BDO",
            column_scope='ltm',
        )
        assert result is not None
        assert result["scope"] == "ltm"
        assert result["values"][19] == pytest.approx(2950.0)


# ---------------------------------------------------------------------------
# Regression: cache is no longer overwritten by _enforce_bdo_alignment.
# ---------------------------------------------------------------------------

class TestEnforceBdoAlignmentDoesNotOverwriteCache:
    """
    Earlier versions of ``_enforce_bdo_alignment`` wrote the BDO ground
    truth straight into ``computed_values[(19,'ltm')]`` and
    ``[(68,'ltm')]``. That made ``_revalidate_detailed`` compare the
    cache against itself and pass even when the AD column formulas did
    not actually evaluate to the BDO H total. The fix stores the BDO
    ground truth on a separate ``bdo_ground_truth`` attribute so the
    cache stays formula-derived.
    """

    def test_ltm_totals_are_not_substituted_into_computed_values(
        self, tmp_path
    ):
        from src.transformers.management_accounts import (
            ManagementAccountsConfig,
        )

        wb = Workbook()
        bdo = wb.active
        bdo.title = "BDO - Q1-26"
        bdo.cell(row=1, column=1, value="Grootboek")
        bdo.cell(row=125, column=2, value="Resultaat na belasting")
        bdo.cell(row=125, column=4, value=-25.0)
        bdo.cell(row=125, column=5, value=-25.0)
        bdo.cell(row=125, column=6, value=-25.0)
        bdo.cell(row=125, column=7, value=-25.0)
        bdo.cell(row=125, column=8, value=-100.0)
        bdo.cell(row=130, column=1, value="Verschil balans en winst-en-verlies")

        mc = wb.create_sheet("Management Cijfers - Q1 2026")
        mc.cell(row=22, column=29, value="Q1 2026")
        mc.cell(row=22, column=30, value="LTM Q1 2026")

        path = tmp_path / "ma.xlsx"
        wb.save(str(path))
        wb.close()

        config = ManagementAccountsConfig(
            quarter="Q1 2026",
            period_end=datetime(2026, 3, 31),
            bdo_sheet_name="BDO - Q1-26",
            summary_sheet_name="Management Cijfers - Q1 2026",
        )
        builder = ManagementAccountsBuilder.__new__(ManagementAccountsBuilder)
        builder.config = config
        builder.workbook = openpyxl.load_workbook(str(path))
        builder._rules = builder._load_accounting_rules()
        builder._new_quarter_col = 29
        builder._new_ltm_col = 30
        builder.computed_values = {(19, 'ltm'): 12345.0, (68, 'ltm'): 67890.0}
        builder.bdo_ground_truth = {'ltm': None, 'quarter': None}
        builder._new_bdo_label_map = {}
        builder.bdo_source_path = None  # disable cached column-H lookup

        from src.parsers.bdo_parser import BDOParseResult
        bdo_result = BDOParseResult(
            accounts={}, period_end="2026-03-31", quarter="Q1 2026",
            column_mapping={}, schema_version="standard", warnings=[],
        )
        builder._enforce_bdo_alignment(bdo_result)

        # Cache must remain formula-derived.
        assert builder.computed_values[(19, 'ltm')] == 12345.0
        assert builder.computed_values[(68, 'ltm')] == 67890.0
        # Sign-adjusted BDO H = -(-100) = 100.0 stored separately.
        assert builder.bdo_ground_truth['ltm'] == pytest.approx(100.0)
        builder.workbook.close()
