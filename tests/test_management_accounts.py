"""
Tests for Management Accounts Builder
"""

import pytest
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook

from src.transformers.management_accounts import (
    ManagementAccountsBuilder,
    ManagementAccountsConfig,
    _shift_formula_column_letter,
    find_column_by_header_label,
    find_ltm_column_near_quarter,
    ltm_column_after_insert,
    resolve_built_ltm_column,
    resolve_built_quarter_column,
    resolve_quarter_column,
    scan_summary_column_layout,
)
from src.parsers.bdo_parser import BDOParseResult, AccountEntry


class TestManagementAccountsBuilder:
    """Test Management Accounts builder."""
    
    @pytest.fixture
    def previous_ma_file(self, tmp_path):
        """Create a previous quarter Management Accounts file."""
        file_path = tmp_path / "prev_ma.xlsx"
        
        wb = Workbook()
        
        # Create BDO sheet
        bdo_sheet = wb.active
        bdo_sheet.title = "BDO - Q2-25"
        bdo_sheet.cell(row=1, column=1, value="Account Code")
        bdo_sheet.cell(row=1, column=2, value="Account Name")
        bdo_sheet.cell(row=1, column=3, value="Opening Balance")
        bdo_sheet.cell(row=1, column=4, value="Mutations")
        bdo_sheet.cell(row=1, column=5, value="Closing Balance")
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Management Cijfers - Q2 2025")
        summary_sheet.cell(row=1, column=1, value="Management Accounts QSP ESS B.V. - Q2 2025")
        summary_sheet.cell(row=2, column=1, value="Item")
        summary_sheet.cell(row=2, column=2, value="2025-06-30")
        
        # Add line items
        line_items = [
            "Deferred Tax Asset",
            "Real estate",
            "Financial fixed assets",
            "Accounts receivable",
            "Service costs to be settled",
            "Prepaid expenses",
            "Cash",
            "Equity",
            "Senior Loan",
            "Accrued Interest",
        ]
        
        for row_idx, item in enumerate(line_items, 3):
            summary_sheet.cell(row=row_idx, column=1, value=item)
            summary_sheet.cell(row=row_idx, column=2, value=1000000.00)
        
        wb.save(file_path)
        return file_path
    
    @pytest.fixture
    def bdo_result(self):
        """Create sample BDO result."""
        accounts = {
            "1600000": AccountEntry(
                code="1600000",
                name="Verkrijgingsprijs",
                opening_balance=12000000.00,
                closing_balance=12000000.00,
                mutations={},
                raw_row=5
            ),
            "1600200": AccountEntry(
                code="1600200",
                name="Cumulatieve afschrijving",
                opening_balance=-600000.00,
                closing_balance=-630000.00,
                mutations={"mutation_4": -30000.00},
                raw_row=6
            ),
            "1760000": AccountEntry(
                code="1760000",
                name="Te vorderen DMRRP",
                opening_balance=600000.00,
                closing_balance=600000.00,
                mutations={},
                raw_row=7
            ),
            "1790002": AccountEntry(
                code="1790002",
                name="Latente belastingvordering",
                opening_balance=100000.00,
                closing_balance=100000.00,
                mutations={},
                raw_row=8
            ),
            "1100000": AccountEntry(
                code="1100000",
                name="Kas",
                opening_balance=150000.00,
                closing_balance=200000.00,
                mutations={"mutation_4": 50000.00},
                raw_row=9
            ),
            "1110000": AccountEntry(
                code="1110000",
                name="Bank",
                opening_balance=500000.00,
                closing_balance=600000.00,
                mutations={"mutation_4": 100000.00},
                raw_row=10
            ),
            "1300000": AccountEntry(
                code="1300000",
                name="Debiteuren",
                opening_balance=250000.00,
                closing_balance=250000.00,
                mutations={},
                raw_row=11
            ),
            "0500000": AccountEntry(
                code="0500000",
                name="Aandelenkapitaal",
                opening_balance=5000000.00,
                closing_balance=5000000.00,
                mutations={},
                raw_row=12
            ),
            "0600000": AccountEntry(
                code="0600000",
                name="Ingehouden winst",
                opening_balance=5000000.00,
                closing_balance=5120000.00,
                mutations={"mutation_4": 120000.00},
                raw_row=13
            ),
            "0900000": AccountEntry(
                code="0900000",
                name="Senior lening",
                opening_balance=3500000.00,
                closing_balance=3500000.00,
                mutations={},
                raw_row=14
            ),
            "0920000": AccountEntry(
                code="0920000",
                name="Opgelopen rente",
                opening_balance=50000.00,
                closing_balance=75000.00,
                mutations={"mutation_4": 25000.00},
                raw_row=15
            ),
        }
        
        return BDOParseResult(
            accounts=accounts,
            period_end="2025-09-30",
            quarter="Q3 2025",
            column_mapping={},
            schema_version="standard",
            warnings=[]
        )
    
    def test_add_bdo_sheet(self, previous_ma_file, bdo_result, tmp_path):
        """Test adding new BDO sheet."""
        output_path = tmp_path / "new_ma.xlsx"
        config = ManagementAccountsConfig(
            quarter="Q3 2025",
            period_end=datetime(2025, 9, 30),
            bdo_sheet_name="BDO - Q3-25",
            summary_sheet_name="Management Cijfers - Q3 2025"
        )
        
        builder = ManagementAccountsBuilder(
            str(previous_ma_file),
            str(output_path),
            config
        )
        
        builder._add_bdo_sheet(bdo_result)
        
        # Verify sheet was added
        wb = load_workbook(output_path)
        assert "BDO - Q3-25" in wb.sheetnames
        
        # Verify data in sheet
        bdo_sheet = wb["BDO - Q3-25"]
        assert bdo_sheet.cell(row=1, column=1).value == "Account Code"
        assert bdo_sheet.cell(row=2, column=1).value == "1600000"
    
    def test_update_summary_sheet(self, previous_ma_file, bdo_result, tmp_path):
        """Test updating summary sheet with new column."""
        output_path = tmp_path / "new_ma.xlsx"
        config = ManagementAccountsConfig(
            quarter="Q3 2025",
            period_end=datetime(2025, 9, 30),
            bdo_sheet_name="BDO - Q3-25",
            summary_sheet_name="Management Cijfers - Q3 2025"
        )
        
        builder = ManagementAccountsBuilder(
            str(previous_ma_file),
            str(output_path),
            config
        )
        
        builder._update_summary_sheet()
        
        # Verify new column was added
        wb = load_workbook(output_path)
        summary_sheet = wb["Management Cijfers - Q3 2025"]
        
        # Should have 3 columns now (Item, Q2, Q3)
        assert summary_sheet.cell(row=2, column=3).value == datetime(2025, 9, 30)
        
        # Verify values were populated
        # Real estate should be sum of 1600000, 1600200, etc.
        real_estate_row = None
        for row_idx in range(3, summary_sheet.max_row + 1):
            if summary_sheet.cell(row=row_idx, column=1).value == "Real estate":
                real_estate_row = row_idx
                break
        
        if real_estate_row:
            value = summary_sheet.cell(row=real_estate_row, column=3).value
            assert value is not None
            assert isinstance(value, (int, float))
    
    def test_build_complete(self, previous_ma_file, bdo_result, tmp_path):
        """Test complete build process."""
        output_path = tmp_path / "new_ma.xlsx"
        config = ManagementAccountsConfig(
            quarter="Q3 2025",
            period_end=datetime(2025, 9, 30),
            bdo_sheet_name="BDO - Q3-25",
            summary_sheet_name="Management Cijfers - Q3 2025"
        )
        
        builder = ManagementAccountsBuilder(
            str(previous_ma_file),
            str(output_path),
            config
        )
        
        result_path = builder.build(bdo_result)
        
        assert result_path.exists()
        assert result_path == output_path
        
        # Verify file structure
        wb = load_workbook(output_path)
        assert "BDO - Q3-25" in wb.sheetnames
        assert "Management Cijfers - Q3 2025" in wb.sheetnames
    
    def test_validate_calculations(self, previous_ma_file, bdo_result, tmp_path):
        """Test equity movement validation."""
        output_path = tmp_path / "new_ma.xlsx"
        config = ManagementAccountsConfig(
            quarter="Q3 2025",
            period_end=datetime(2025, 9, 30),
            bdo_sheet_name="BDO - Q3-25",
            summary_sheet_name="Management Cijfers - Q3 2025"
        )
        
        builder = ManagementAccountsBuilder(
            str(previous_ma_file),
            str(output_path),
            config
        )
        
        # This should not raise an exception
        builder._validate_calculations(bdo_result)

    def test_computed_values_populated(self, previous_ma_file, bdo_result, tmp_path):
        """After build(), computed_values dict must be populated for all key rows."""
        output_path = tmp_path / "new_ma.xlsx"
        config = ManagementAccountsConfig(
            quarter="Q3 2025",
            period_end=datetime(2025, 9, 30),
            bdo_sheet_name="BDO - Q3-25",
            summary_sheet_name="Management Cijfers - Q3 2025"
        )
        builder = ManagementAccountsBuilder(
            str(previous_ma_file), str(output_path), config
        )
        builder.build(bdo_result)

        cv = builder.computed_values
        assert cv, "computed_values should not be empty"

        assert (19, 'ltm') in cv, "BS row 19 LTM must be in computed_values"
        assert (68, 'quarter') in cv, "P&L row 68 quarter must be in computed_values"
        assert (68, 'ltm') in cv, "P&L row 68 LTM must be in computed_values"

    def test_computed_values_has_key_rows(self, previous_ma_file, bdo_result, tmp_path):
        """computed_values must contain BS and P&L totals after build()."""
        output_path = tmp_path / "new_ma.xlsx"
        config = ManagementAccountsConfig(
            quarter="Q3 2025",
            period_end=datetime(2025, 9, 30),
            bdo_sheet_name="BDO - Q3-25",
            summary_sheet_name="Management Cijfers - Q3 2025"
        )
        builder = ManagementAccountsBuilder(
            str(previous_ma_file), str(output_path), config
        )
        builder.build(bdo_result)

        cv = builder.computed_values
        # These keys must always be present even with minimal test data
        for key in [(19, 'ltm'), (68, 'quarter'), (68, 'ltm'), (25, 'quarter'), (66, 'ltm')]:
            assert key in cv, f"computed_values missing key {key}"


def _make_bdo_result(accounts=None):
    """Helper to create a minimal BDOParseResult for tests."""
    return BDOParseResult(
        accounts=accounts or {},
        period_end="2025-09-30",
        quarter="Q3 2025",
        column_mapping={},
        schema_version="v1",
    )


class TestDetectFirstQuarterColumn:
    def test_finds_leftmost_q_header(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=5, value="Q1 2025")
        ws.cell(row=22, column=10, value="Q2 2025")
        ws.cell(row=22, column=28, value="LTM Q3 2025")
        shell = ManagementAccountsBuilder.__new__(ManagementAccountsBuilder)
        shell._rules = ManagementAccountsBuilder._load_accounting_rules()
        col = shell._detect_first_quarter_column_from_header_row(ws, 22, 27)
        assert col == 5


class TestSummaryColumnLayout:
    """Column insert must follow Q headers, not FY or misplaced LTM anchors."""

    def test_standard_layout_inserts_before_ltm(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=27, value="Q3 2025")
        ws.cell(row=22, column=28, value="Q4 2025")
        ws.cell(row=22, column=29, value="LTM Q4 2025")
        layout = scan_summary_column_layout(ws, 22)
        assert layout.last_quarter_column == 28
        assert layout.insert_column == 29
        assert layout.previous_quarter_column == 28
        assert layout.ltm_column == 29
        assert ltm_column_after_insert(layout.ltm_column, 29, layout.fy_columns) == 30

    def test_fy_between_quarter_and_ltm_inserts_after_quarter_not_ltm(self):
        """Client template: Q4 | FY | LTM — new Q column belongs after Q4, before FY."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=28, value="Q3 2025")
        ws.cell(row=22, column=29, value="Q4 2025")
        ws.cell(row=22, column=30, value="FY 2025")
        ws.cell(row=22, column=31, value="LTM Q4 2025")
        layout = scan_summary_column_layout(ws, 22)
        assert layout.last_quarter_column == 29
        assert layout.insert_column == 30
        assert layout.previous_quarter_column == 29
        assert layout.ltm_column == 31
        assert ltm_column_after_insert(31, 30, layout.fy_columns) == 32

    def test_find_column_by_header_label(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=30, value="Q1 2026")
        assert find_column_by_header_label(ws, "Q1 2026", 22) == 30
        assert find_column_by_header_label(ws, "q1 2026", 22) == 30

    def test_update_summary_inserts_q1_before_fy(self, tmp_path):
        """Regression: Q1 2026 must appear before FY, not after hidden LTM."""
        from openpyxl.utils import get_column_letter

        prev_path = tmp_path / "prev_q4_fy.xlsx"
        wb = Workbook()
        bdo = wb.active
        bdo.title = "BDO - Q4-25"
        bdo.cell(row=6, column=1, value="8000003")
        bdo.cell(row=6, column=7, value=100)

        mc = wb.create_sheet("Management Cijfers - Q4 2025")
        mc.cell(row=1, column=1, value="Management Accounts QSP ESS B.V. - Q4 2025")
        mc.cell(row=22, column=28, value="Q3 2025")
        mc.cell(row=22, column=29, value="Q4 2025")
        mc.cell(row=22, column=30, value="FY 2025")
        mc.cell(row=22, column=31, value="LTM Q4 2025")
        mc.cell(row=2, column=31, value=datetime(2025, 12, 31))
        mc.cell(row=23, column=29, value=1000)
        mc.cell(row=23, column=31, value="='BDO - Q4-25'!H6")
        wb.save(prev_path)

        out_path = tmp_path / "ma_q1_2026.xlsx"
        config = ManagementAccountsConfig(
            quarter="Q1 2026",
            period_end=datetime(2026, 3, 31),
            bdo_sheet_name="BDO - Q1-26",
            summary_sheet_name="Management Cijfers - Q1 2026",
        )
        builder = ManagementAccountsBuilder(str(prev_path), str(out_path), config)
        builder.workbook = load_workbook(prev_path)
        builder.config = config
        builder._copy_bdo_data_to_new_sheet(_make_bdo_result())
        new_bdo = builder.workbook[config.bdo_sheet_name]
        builder._build_row_map(new_bdo, builder._new_bdo_row_map, builder._new_bdo_label_map)
        builder._update_summary_sheet()
        builder.workbook.save(out_path)

        result = load_workbook(out_path)
        sheet = result["Management Cijfers - Q1 2026"]
        assert sheet.cell(row=22, column=30).value == "Q1 2026"
        assert sheet.cell(row=22, column=31).value == "FY 2025"
        ltm_header = sheet.cell(row=22, column=32).value
        assert "LTM" in str(ltm_header)
        # Header is rewritten to the current LTM label, not the pre-existing one.
        assert ltm_header == "LTM Q1 2026"
        assert get_column_letter(builder._new_quarter_col) == "AD"

        # Quarter column (AD) references the current BDO sheet with column G.
        q_val = sheet.cell(row=23, column=builder._new_quarter_col).value
        assert isinstance(q_val, str) and "'BDO - Q1-26'!G" in q_val, q_val

        # LTM column (AF) references the current BDO sheet with column H —
        # the row-23 cell carried a stale 'BDO - Q4-25'!H6 before our fix.
        ltm_val = sheet.cell(row=23, column=builder._new_ltm_col).value
        assert (
            isinstance(ltm_val, str)
            and "'BDO - Q1-26'!H" in ltm_val
            and "Q4-25" not in ltm_val
        ), ltm_val
        result.close()

    def test_standard_layout_matches_legacy_ltm_insert(self):
        """Without FY, insert after last Q is the same as insert before LTM (Q3/Q4 2025 path)."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=27, value="Q3 2025")
        ws.cell(row=22, column=28, value="Q4 2025")
        ws.cell(row=22, column=29, value="LTM Q4 2025")
        layout = scan_summary_column_layout(ws, 22)
        assert layout.insert_column == layout.ltm_column
        assert layout.previous_quarter_column == 28

    @pytest.mark.parametrize(
        "headers,expected_insert,expected_prev,expected_ltm_after",
        [
            (
                [(28, "Q3 2025"), (29, "Q4 2025"), (30, "LTM Q4 2025")],
                30,
                29,
                31,
            ),
            (
                [(28, "Q3 2025"), (29, "Q4 2025"), (30, "FY 2025"), (31, "LTM Q4 2025")],
                30,
                29,
                32,
            ),
            (
                [(29, "Q4 2025"), (30, "Q1 2026"), (31, "FY 2025"), (32, "LTM Q1 2026")],
                31,
                30,
                33,
            ),
            (
                [(30, "Q1 2026"), (31, "Q2 2026"), (32, "FY 2026"), (33, "LTM Q2 2026")],
                32,
                31,
                34,
            ),
        ],
    )
    def test_future_quarter_insert_positions(
        self, headers, expected_insert, expected_prev, expected_ltm_after
    ):
        wb = Workbook()
        ws = wb.active
        for col, label in headers:
            ws.cell(row=22, column=col, value=label)
        layout = scan_summary_column_layout(ws, 22)
        assert layout.insert_column == expected_insert
        assert layout.previous_quarter_column == expected_prev
        assert ltm_column_after_insert(
            layout.ltm_column, layout.insert_column, layout.fy_columns
        ) == expected_ltm_after

    def test_resolve_quarter_column_skips_fy(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=29, value="Q4 2025")
        ws.cell(row=22, column=30, value="Q1 2026")
        ws.cell(row=22, column=31, value="FY 2025")
        ws.cell(row=22, column=32, value="LTM Q1 2026")
        assert resolve_quarter_column(ws, "Q1 2026", 22) == 30
        assert find_ltm_column_near_quarter(ws, 30, 22) == 32

    def test_historical_ltm_at_left_is_ignored(self):
        """Workbooks with years of history have an old 'LTM …' header far to the
        left of the current quarterly block (e.g. column E).  That column must NOT
        be treated as the active LTM anchor — the active LTM slot should be
        computed from the FY column position instead."""
        wb = Workbook()
        ws = wb.active
        # Historical LTM from 2019-era data, far left
        ws.cell(row=22, column=5, value="LTM Q4 2025")
        # Current quarterly block: Q4 at AB (28), FY at AC (29), with no adjacent LTM yet
        ws.cell(row=22, column=27, value="Q3 2025")
        ws.cell(row=22, column=28, value="Q4 2025")
        ws.cell(row=22, column=29, value="FY 2025")
        layout = scan_summary_column_layout(ws, 22)
        # The historical LTM (col 5) is left of all quarters — must be None
        assert layout.ltm_column is None, (
            f"Historical LTM at col 5 should be ignored; got {layout.ltm_column}"
        )
        assert layout.insert_column == 29   # insert before FY (= after Q4)
        assert layout.previous_quarter_column == 28

        # After insert at 29: FY shifts from 29→30 (AD); LTM = max([30])+1 = 31 (AE)
        new_ltm = ltm_column_after_insert(layout.ltm_column, 29, layout.fy_columns)
        assert new_ltm == 31, (
            f"Expected new LTM at col 31 (after shifted FY 29→30, LTM = FY+1), got {new_ltm}"
        )

    def test_historical_ltm_left_no_fy(self):
        """Historical LTM left of quarters, no FY column: new LTM = insert+1."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=5, value="LTM Q4 2025")
        ws.cell(row=22, column=27, value="Q3 2025")
        ws.cell(row=22, column=28, value="Q4 2025")
        layout = scan_summary_column_layout(ws, 22)
        assert layout.ltm_column is None
        assert layout.insert_column == 29
        new_ltm = ltm_column_after_insert(layout.ltm_column, 29, layout.fy_columns)
        assert new_ltm == 30  # insert+1

    def test_active_ltm_right_of_quarters_still_shifts(self):
        """An LTM column that IS to the right of the last quarter should still
        shift right after insert — the original behaviour must be preserved."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=5, value="LTM Q3 2024")   # historical — ignored
        ws.cell(row=22, column=28, value="Q3 2025")
        ws.cell(row=22, column=29, value="Q4 2025")
        ws.cell(row=22, column=30, value="LTM Q4 2025")  # active
        layout = scan_summary_column_layout(ws, 22)
        assert layout.ltm_column == 30                   # rightmost wins
        # Q4 at 29 → insert_column = 30 (after Q4)
        assert layout.insert_column == 30
        assert layout.previous_quarter_column == 29
        # LTM at 30 shifts to 31 because pre_ltm_col(30) >= insert_column(30)
        new_ltm = ltm_column_after_insert(layout.ltm_column, 30, layout.fy_columns)
        assert new_ltm == 31


class TestShiftFormulaColumnLetter:
    """Regression tests for interest row 60 formula extension."""

    def test_shifts_prev_column_refs_ab_to_ac(self):
        f = "=AB57+AB60-$AB$61+SUM(Y60:AB60)"
        out = _shift_formula_column_letter(f, "AB", "AC")
        assert "AB57" not in out and "AC57" in out
        assert "AC60" in out
        assert "$AC$61" in out
        assert "SUM(Y60:AC60)" in out

    def test_does_not_replace_a_inside_aa(self):
        f = "=AA10+A11"
        out = _shift_formula_column_letter(f, "A", "B")
        assert out == "=AA10+B11"


class TestFuzzyAccountLookup:
    """Test that shadow model uses fuzzy matching like _find_account_row."""

    def test_exact_match(self):
        entry = AccountEntry(
            code="1600000", name="Test", opening_balance=0, closing_balance=100,
            mutations={}, raw_row=5,
        )
        result = ManagementAccountsBuilder._fuzzy_account_lookup(
            "1600000", _make_bdo_result({"1600000": entry}),
        )
        assert result is entry

    def test_prefix_match(self):
        entry = AccountEntry(
            code="1600000.0", name="Test", opening_balance=0, closing_balance=100,
            mutations={}, raw_row=5,
        )
        result = ManagementAccountsBuilder._fuzzy_account_lookup(
            "1600000", _make_bdo_result({"1600000.0": entry}),
        )
        assert result is entry

    def test_no_match_returns_none(self):
        result = ManagementAccountsBuilder._fuzzy_account_lookup(
            "9999999", _make_bdo_result(),
        )
        assert result is None


class TestEvaluateCalcPattern:
    """Ensure _evaluate_calc_pattern handles every YAML pattern."""

    @pytest.fixture
    def builder(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "BDO - Q3-25"
        wb.create_sheet("Management Cijfers - Q3 2025")
        path = tmp_path / "dummy.xlsx"
        wb.save(path)
        config = ManagementAccountsConfig(
            quarter="Q3 2025", period_end=datetime(2025, 9, 30),
            bdo_sheet_name="BDO - Q3-25",
            summary_sheet_name="Management Cijfers - Q3 2025",
        )
        b = ManagementAccountsBuilder.__new__(ManagementAccountsBuilder)
        b.config = config
        b.formula_templates_path = Path("config/formula_templates.yaml")
        b.formula_templates = {}
        b.balance_sheet_templates = {}
        return b

    def test_sum_range(self, builder):
        shadow = {3: {'value': 10.0}, 4: {'value': 20.0}, 5: {'value': 30.0}}
        assert builder._evaluate_calc_pattern("=SUM({COL}3:{COL}5)", shadow) == 60.0

    def test_negated_subtraction(self, builder):
        shadow = {50: {'value': 100.0}, 53: {'value': 40.0}}
        assert builder._evaluate_calc_pattern("=-({COL}50-{COL}53)", shadow) == -60.0

    def test_division_minus_one(self, builder):
        shadow = {25: {'value': 200.0}, 23: {'value': 100.0}}
        assert builder._evaluate_calc_pattern("={COL}25/{COL}23-1", shadow) == pytest.approx(1.0)

    def test_addition_chain(self, builder):
        shadow = {25: {'value': 10.0}, 30: {'value': 20.0}, 44: {'value': 30.0}}
        assert builder._evaluate_calc_pattern("={COL}25+{COL}30+{COL}44", shadow) == 60.0

    def test_two_term_addition(self, builder):
        shadow = {67: {'value': 5.0}, 66: {'value': 3.0}}
        assert builder._evaluate_calc_pattern("={COL}67+{COL}66", shadow) == 8.0

    def test_two_term_with_plus(self, builder):
        shadow = {53: {'value': 10.0}, 48: {'value': 20.0}}
        assert builder._evaluate_calc_pattern("={COL}53+{COL}48", shadow) == 30.0


class TestCCDirectCopy:
    """Verify the CC builder uses computed_values for direct copy."""

    def test_copy_ma_data_from_cache(self, tmp_path):
        from src.transformers.compliance_builder import ComplianceBuilder, ComplianceConfig

        # Create a minimal previous CC file with a "Q3 Management Accounts" sheet
        prev_path = tmp_path / "prev_cc.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Q2 Management Accounts"
        for r in range(1, 80):
            ws.cell(row=r, column=1, value=f"Row {r}")
        wb.save(prev_path)

        output_path = tmp_path / "new_cc.xlsx"
        config = ComplianceConfig(year=2025, quarter=3)
        builder = ComplianceBuilder(str(prev_path), str(output_path), config)

        computed_values = {
            (3, 'ltm'): 100.0,
            (3, 'quarter'): 50.0,
            (19, 'ltm'): 555.0,
            (23, 'ltm'): 200.0,
            (23, 'quarter'): 150.0,
            (68, 'ltm'): 999.0,
            (68, 'quarter'): 888.0,
            (107, 'ltm'): 303.0,
            (114, 'ltm'): 303.0,
        }

        bdo_result = _make_bdo_result()
        builder.build(bdo_result, management_accounts_path=None, computed_values=computed_values)

        result_wb = load_workbook(str(output_path))
        ma_sheet = result_wb["Q3 Management Accounts"]

        # CC row 2 -> MA row 3 (BS): LTM in col C only, no quarter for BS
        assert ma_sheet.cell(row=2, column=3).value == 100.0
        assert ma_sheet.cell(row=2, column=2).value is None

        # CC row 18 -> MA row 19 (Total Equity Movement): LTM only
        assert ma_sheet.cell(row=18, column=3).value == 555.0

        # CC row 22 -> MA row 23 (P&L): both quarter and LTM
        assert ma_sheet.cell(row=22, column=3).value == 200.0
        assert ma_sheet.cell(row=22, column=2).value == 150.0

        result_wb.close()


# ---------------------------------------------------------------------------
# LTM column deterministic rebuild — regression guard for the Q1 2026 bugs
# (stale 'BDO - Q4-25' refs in AD and SUM(AC...) subtotals where SUM(AD...)
# is expected). Tests target both FY-present and no-FY layouts and verify
# that the LTM column is fully rebuilt by _finalize_ltm_column.
# ---------------------------------------------------------------------------


def _make_q1_2026_bdo_accounts():
    """Minimal BDO accounts for the rows referenced by formula_templates."""
    rows = {
        # P&L account codes -> rows
        "8000003": 80,
        "8000004": 81,
        "8400600": 83,
        "8400700": 84,
        "8400800": 85,
        "4100400": 90,
        "4101000": 91,
        "4310400": 92,
        # Balance sheet codes
        "1790002": 6,
        "1600000": 10,
        "1600200": 11,
        "1601000": 12,
        "1610803": 13,
        "1611003": 14,
        "1760000": 20,
        "1790000": 21,
        "2000000": 25,
        "2000100": 26,
        "2000300": 27,
        "2400100": 31,
        "2400200": 32,
        "2400201": 33,
        "2400202": 34,
        "2400203": 35,
        "2400204": 36,
        "2400205": 37,
        "2400206": 38,
        "1000000": 50,
        "1100000": 51,
        "1160000": 52,
        "1930001": 60,
        "1930101": 61,
    }
    accounts = {}
    for code, raw_row in rows.items():
        accounts[code] = AccountEntry(
            code=code,
            name=f"Account {code}",
            opening_balance=100.0,
            closing_balance=100.0,
            mutations={},
            raw_row=raw_row,
        )
    return accounts


def _build_prev_q4_workbook(prev_path: Path, *, with_fy: bool):
    """
    Build a minimal previous-quarter workbook that triggers the LTM rebuild
    paths. Includes either:
      - Q3 | Q4 | FY | LTM   (FY layout: insert at col 30, LTM at col 32)
      - Q3 | Q4 | LTM        (no-FY layout: insert at col 29, LTM at col 30)
    The LTM column is pre-populated with stale Q4-25 formulas in the exact
    pattern that the FreshAIs file exhibited — so we know the rebuild must
    overwrite them.
    """
    wb = Workbook()
    bdo = wb.active
    bdo.title = "BDO - Q4-25"
    # Populate BDO rows for the accounts referenced by templates so the
    # build doesn't skip rows with "account not found".
    for code, raw_row in {
        "8000003": 80, "8000004": 81, "8400600": 83, "8400700": 84, "8400800": 85,
        "4100400": 90, "4101000": 91, "4310400": 92,
        "1790002": 6, "1600000": 10, "1600200": 11, "1601000": 12, "1610803": 13, "1611003": 14,
        "1760000": 20, "1790000": 21,
        "2000000": 25, "2000100": 26, "2000300": 27,
        "2400100": 31, "2400200": 32, "2400201": 33, "2400202": 34,
        "2400203": 35, "2400204": 36, "2400205": 37, "2400206": 38,
        "1000000": 50, "1100000": 51, "1160000": 52,
        "1930001": 60, "1930101": 61,
    }.items():
        bdo.cell(row=raw_row, column=1, value=code)
        bdo.cell(row=raw_row, column=2, value=f"Account {code}")
        bdo.cell(row=raw_row, column=7, value=100)  # column G
        bdo.cell(row=raw_row, column=8, value=400)  # column H

    mc = wb.create_sheet("Management Cijfers - Q4 2025")
    mc.cell(row=1, column=1, value="Management Accounts QSP ESS B.V. - Q4 2025")

    if with_fy:
        mc.cell(row=22, column=28, value="Q3 2025")
        mc.cell(row=22, column=29, value="Q4 2025")
        mc.cell(row=22, column=30, value="FY 2025")
        mc.cell(row=22, column=31, value="LTM Q4 2025")
        ltm_col_pre = 31
    else:
        # Client draft layout: prev Q3=AA, Q4=AB, LTM=AC. After Q1 insert
        # we expect Q1=AC and LTM=AD, matching the bug report's AC/AD letters.
        mc.cell(row=22, column=27, value="Q3 2025")
        mc.cell(row=22, column=28, value="Q4 2025")
        mc.cell(row=22, column=29, value="LTM Q4 2025")
        ltm_col_pre = 29

    mc.cell(row=2, column=ltm_col_pre, value=datetime(2025, 12, 31))

    # Stale Q4 formulas in the OLD LTM column. After insert, these will shift
    # right and (without our fix) survive into the new LTM column.
    mc.cell(row=23, column=ltm_col_pre, value="=-'BDO - Q4-25'!H78")
    mc.cell(row=25, column=ltm_col_pre, value="=SUM(AC23:AC24)")
    mc.cell(row=30, column=ltm_col_pre, value="=SUM(AC27:AC29)")
    mc.cell(row=44, column=ltm_col_pre, value="=SUM(AC32:AC43)")
    mc.cell(row=45, column=ltm_col_pre, value="=AC25+AC30+AC44")
    mc.cell(row=68, column=ltm_col_pre, value="=AC67+AC66")
    mc.cell(row=114, column=ltm_col_pre, value="=SUM(AC107:AC113)")

    wb.save(prev_path)
    return wb


def _run_builder(prev_path: Path, out_path: Path) -> ManagementAccountsBuilder:
    """Run just the parts of build() needed for LTM verification."""
    config = ManagementAccountsConfig(
        quarter="Q1 2026",
        period_end=datetime(2026, 3, 31),
        bdo_sheet_name="BDO - Q1-26",
        summary_sheet_name="Management Cijfers - Q1 2026",
    )
    builder = ManagementAccountsBuilder(str(prev_path), str(out_path), config)
    builder.workbook = load_workbook(prev_path)
    builder.config = config
    builder._copy_bdo_data_to_new_sheet(
        _make_bdo_result(_make_q1_2026_bdo_accounts())
    )
    new_bdo = builder.workbook[config.bdo_sheet_name]
    builder._build_row_map(
        new_bdo, builder._new_bdo_row_map, builder._new_bdo_label_map
    )
    builder._update_summary_sheet()
    builder.workbook.save(out_path)
    return builder


class TestLtmColumnRebuild:
    """LTM column must be fully rebuilt after insert_cols, with current BDO
    sheet name and self-referencing subtotals."""

    def test_fy_layout_ltm_header_and_formulas(self, tmp_path):
        """Q3 | Q4 | FY | LTM → after Q1 insert: Q3 | Q4 | Q1 | FY | LTM."""
        from openpyxl.utils import get_column_letter

        prev_path = tmp_path / "prev_q4_fy.xlsx"
        out_path = tmp_path / "ma_q1_2026_fy.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)
        builder = _run_builder(prev_path, out_path)

        wb = load_workbook(out_path)
        sheet = wb[builder.config.summary_sheet_name]
        ltm_col = builder._new_ltm_col
        q_col = builder._new_quarter_col
        ltm_letter = get_column_letter(ltm_col)
        q_letter = get_column_letter(q_col)

        # Header sanity
        assert sheet.cell(row=22, column=q_col).value == "Q1 2026"
        assert sheet.cell(row=22, column=ltm_col).value == "LTM Q1 2026"

        # LTM P&L bdo_ref now points at Q1-26 with column H
        f23 = sheet.cell(row=23, column=ltm_col).value
        assert isinstance(f23, str) and "'BDO - Q1-26'!H" in f23, f23

        # LTM calc subtotals reference the LTM column, not the quarter column
        f25 = sheet.cell(row=25, column=ltm_col).value
        assert f25 == f"=SUM({ltm_letter}23:{ltm_letter}24)", f25
        f30 = sheet.cell(row=30, column=ltm_col).value
        assert f30 == f"=SUM({ltm_letter}27:{ltm_letter}29)", f30
        f44 = sheet.cell(row=44, column=ltm_col).value
        assert f44 == f"=SUM({ltm_letter}32:{ltm_letter}43)", f44
        f114 = sheet.cell(row=114, column=ltm_col).value
        assert f114 == f"=SUM({ltm_letter}107:{ltm_letter}113)", f114

        # Quarter column still uses column G against the current BDO sheet
        q23 = sheet.cell(row=23, column=q_col).value
        assert isinstance(q23, str) and "'BDO - Q1-26'!G" in q23, q23

        wb.close()

    def test_no_fy_layout_matches_client_draft(self, tmp_path):
        """Q3 | Q4 | LTM (client layout) → after insert: Q3 | Q4 | Q1 | LTM."""
        from openpyxl.utils import get_column_letter

        prev_path = tmp_path / "prev_q4_nofy.xlsx"
        out_path = tmp_path / "ma_q1_2026_nofy.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=False)
        builder = _run_builder(prev_path, out_path)

        wb = load_workbook(out_path)
        sheet = wb[builder.config.summary_sheet_name]
        ltm_col = builder._new_ltm_col
        q_col = builder._new_quarter_col
        ltm_letter = get_column_letter(ltm_col)

        # Matches the client's "AC = Q1, AD = LTM Q1 2026" layout
        assert get_column_letter(q_col) == "AC"
        assert ltm_letter == "AD"
        assert sheet.cell(row=22, column=ltm_col).value == "LTM Q1 2026"
        assert sheet.cell(row=22, column=q_col).value == "Q1 2026"

        # Bug 1 fix: LTM BDO references the current quarter's BDO sheet
        # at column H, not the previous quarter's.
        f23 = sheet.cell(row=23, column=ltm_col).value
        assert "'BDO - Q1-26'!H" in f23 and "Q4-25" not in f23, f23

        # Bug 2 fix: subtotal rows sum the LTM column, not the quarter column.
        assert sheet.cell(row=25, column=ltm_col).value == f"=SUM(AD23:AD24)"
        assert sheet.cell(row=30, column=ltm_col).value == f"=SUM(AD27:AD29)"
        assert sheet.cell(row=44, column=ltm_col).value == f"=SUM(AD32:AD43)"
        assert sheet.cell(row=114, column=ltm_col).value == f"=SUM(AD107:AD113)"

        wb.close()

    def test_no_stale_bdo_sheet_anywhere_in_ltm(self, tmp_path):
        """Scan every LTM cell — no formula may reference the previous BDO."""
        prev_path = tmp_path / "prev_q4_nofy.xlsx"
        out_path = tmp_path / "ma_q1_2026_scan.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=False)
        builder = _run_builder(prev_path, out_path)

        wb = load_workbook(out_path)
        sheet = wb[builder.config.summary_sheet_name]
        ltm_col = builder._new_ltm_col

        offending = []
        for row_idx in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=ltm_col).value
            if isinstance(val, str) and "'BDO -" in val:
                if builder.config.bdo_sheet_name not in val:
                    offending.append((row_idx, val))

        assert not offending, (
            f"LTM column still references stale BDO sheet(s): {offending}"
        )

        wb.close()

    def test_no_quarter_letter_self_ref_in_ltm_subtotals(self, tmp_path):
        """No LTM subtotal row may still reference the new quarter column."""
        from openpyxl.utils import get_column_letter

        prev_path = tmp_path / "prev_q4_fy.xlsx"
        out_path = tmp_path / "ma_q1_2026_calc.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)
        builder = _run_builder(prev_path, out_path)

        wb = load_workbook(out_path)
        sheet = wb[builder.config.summary_sheet_name]
        ltm_col = builder._new_ltm_col
        q_letter = get_column_letter(builder._new_quarter_col)

        # Subtotal rows defined in calc_formulas.
        for row in (25, 30, 44, 45, 48, 55, 57, 66, 68):
            val = sheet.cell(row=row, column=ltm_col).value
            assert isinstance(val, str), f"LTM row {row} is empty"
            # Strip out cross-sheet BDO refs before checking column letters
            import re
            stripped = re.sub(r"'[^']*'![A-Z]+\$?\d+", '', val)
            assert q_letter not in stripped, (
                f"LTM row {row} still references quarter letter {q_letter}: {val}"
            )

        wb.close()

    def test_validate_ltm_column_formulas_passes(self, tmp_path):
        """End-to-end: _validate_ltm_column_formulas reports no hard errors."""
        prev_path = tmp_path / "prev_q4_fy.xlsx"
        out_path = tmp_path / "ma_q1_2026_validate.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)
        builder = _run_builder(prev_path, out_path)

        sheet = builder.workbook[builder.config.summary_sheet_name]
        result = builder._validate_ltm_column_formulas(sheet)
        assert result['errors'] == [], (
            f"LTM validation should pass, got errors: {result['errors']}"
        )

    def test_resolve_built_ltm_column_uses_header_first(self):
        """``resolve_built_ltm_column`` should respect ``expected_col``
        whenever its header contains ``LTM``."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=30, value="Q1 2026")
        ws.cell(row=22, column=31, value="FY 2025")
        ws.cell(row=22, column=32, value="LTM Q1 2026")
        # expected_col matches the header → returned as-is
        assert resolve_built_ltm_column(
            ws, header_row=22, expected_col=32, quarter_col=30
        ) == 32

    def test_resolve_built_ltm_column_falls_back_to_near_quarter(self):
        """When ``expected_col`` is wrong, fall back to the LTM near quarter."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=30, value="Q1 2026")
        ws.cell(row=22, column=31, value="FY 2025")
        ws.cell(row=22, column=32, value="LTM Q1 2026")
        # expected_col points at FY — header check fails, fall back finds LTM
        assert resolve_built_ltm_column(
            ws, header_row=22, expected_col=31, quarter_col=30
        ) == 32

    def test_resolve_built_ltm_column_raises_when_no_ltm_anywhere(self):
        """No LTM header anywhere → raise with a layout dump."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=22, column=30, value="Q1 2026")
        ws.cell(row=22, column=31, value="FY 2025")
        with pytest.raises(ValueError, match="Could not resolve LTM column"):
            resolve_built_ltm_column(
                ws, header_row=22, expected_col=31, quarter_col=30
            )


# ---------------------------------------------------------------------------
# Fix Management Accounts Column Issues — new regression tests.
# ---------------------------------------------------------------------------


def _run_builder_with_quarter_num(
    prev_path: Path, out_path: Path, quarter_num: int
) -> ManagementAccountsBuilder:
    """Same as `_run_builder` but threads a quarter_num through the config so
    the FY-removal branch in `_update_summary_sheet` is exercised."""
    config = ManagementAccountsConfig(
        quarter="Q1 2026",
        period_end=datetime(2026, 3, 31),
        bdo_sheet_name="BDO - Q1-26",
        summary_sheet_name="Management Cijfers - Q1 2026",
        quarter_num=quarter_num,
    )
    builder = ManagementAccountsBuilder(str(prev_path), str(out_path), config)
    builder.workbook = load_workbook(prev_path)
    builder.config = config
    builder._copy_bdo_data_to_new_sheet(
        _make_bdo_result(_make_q1_2026_bdo_accounts())
    )
    new_bdo = builder.workbook[config.bdo_sheet_name]
    builder._build_row_map(
        new_bdo, builder._new_bdo_row_map, builder._new_bdo_label_map
    )
    builder._update_summary_sheet()
    builder.workbook.save(out_path)
    return builder


class TestFYColumnRemoval:
    """FY 2025 must disappear when the new quarter is Q1/Q2/Q3 and be kept
    when the new quarter is Q4. With FY gone, the new LTM lands directly
    after the new quarter column."""

    def test_fy_removed_for_q1(self, tmp_path):
        """Q4|FY|LTM input + quarter_num=1 → FY deleted, layout Q4|Q1|LTM."""
        from openpyxl.utils import get_column_letter
        prev_path = tmp_path / "prev_q4_fy_q1.xlsx"
        out_path = tmp_path / "out_q1.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)

        builder = _run_builder_with_quarter_num(
            prev_path, out_path, quarter_num=1
        )
        sheet = builder.workbook[builder.config.summary_sheet_name]

        # Input had Q3=AB(28), Q4=AC(29), FY=AD(30), LTM=AE(31).
        # After deleting FY: Q3=28, Q4=29, LTM=30.
        # After inserting Q1 at 30: Q3=28, Q4=29, Q1=30, LTM=31.
        assert builder._new_quarter_col == 30
        assert builder._new_ltm_col == 31
        assert get_column_letter(builder._new_quarter_col) == "AD"
        assert get_column_letter(builder._new_ltm_col) == "AE"

        # No FY column anywhere on row 22
        fy_columns = []
        for col_idx in range(1, sheet.max_column + 1):
            v = sheet.cell(row=22, column=col_idx).value
            if v and isinstance(v, str) and v.strip().upper().startswith("FY"):
                fy_columns.append((col_idx, v))
        assert fy_columns == [], (
            f"FY column should have been removed but found: {fy_columns}"
        )

        # Headers placed correctly
        assert sheet.cell(row=22, column=builder._new_quarter_col).value == "Q1 2026"
        assert sheet.cell(row=22, column=builder._new_ltm_col).value == "LTM Q1 2026"

    def test_fy_kept_for_q4(self, tmp_path):
        """Q3|FY|LTM input + quarter_num=4 → FY retained."""
        from openpyxl.utils import get_column_letter

        prev_path = tmp_path / "prev_q3_fy.xlsx"
        out_path = tmp_path / "out_q4.xlsx"

        # Build a Q3 2025 workbook with the FY-of-prior-year still present.
        wb = Workbook()
        bdo = wb.active
        bdo.title = "BDO - Q3-25"
        for code, raw_row in {
            "8000003": 80, "8400600": 83, "1790002": 6, "1600000": 10,
        }.items():
            bdo.cell(row=raw_row, column=1, value=code)
            bdo.cell(row=raw_row, column=2, value=f"Account {code}")
            bdo.cell(row=raw_row, column=7, value=100)
            bdo.cell(row=raw_row, column=8, value=400)
        mc = wb.create_sheet("Management Cijfers - Q3 2025")
        mc.cell(row=1, column=1, value="Management Accounts QSP ESS B.V. - Q3 2025")
        mc.cell(row=22, column=27, value="Q2 2025")
        mc.cell(row=22, column=28, value="Q3 2025")
        mc.cell(row=22, column=29, value="FY 2024")
        mc.cell(row=22, column=30, value="LTM Q3 2025")
        mc.cell(row=2, column=30, value=datetime(2025, 9, 30))
        wb.save(prev_path)

        # Reuse builder plumbing but with a Q4 config and quarter_num=4.
        config = ManagementAccountsConfig(
            quarter="Q4 2025",
            period_end=datetime(2025, 12, 31),
            bdo_sheet_name="BDO - Q4-25",
            summary_sheet_name="Management Cijfers - Q4 2025",
            quarter_num=4,
        )
        builder = ManagementAccountsBuilder(str(prev_path), str(out_path), config)
        builder.workbook = load_workbook(prev_path)
        builder.config = config
        builder._copy_bdo_data_to_new_sheet(
            _make_bdo_result(_make_q1_2026_bdo_accounts())
        )
        # Rename copied BDO sheet to Q4-25 for the test fixture
        if "BDO - Q1-26" in builder.workbook.sheetnames:
            builder.workbook["BDO - Q1-26"].title = "BDO - Q4-25"
        new_bdo = builder.workbook["BDO - Q4-25"]
        builder._build_row_map(
            new_bdo, builder._new_bdo_row_map, builder._new_bdo_label_map
        )
        builder._update_summary_sheet()

        sheet = builder.workbook[config.summary_sheet_name]
        # FY 2024 must still exist somewhere on row 22
        fy_seen = any(
            isinstance(sheet.cell(row=22, column=c).value, str)
            and sheet.cell(row=22, column=c).value.strip().upper().startswith("FY")
            for c in range(1, sheet.max_column + 1)
        )
        assert fy_seen, "FY column must be retained when quarter_num == 4"

    def test_fy_removal_is_skipped_when_quarter_num_none(self, tmp_path):
        """Backward compat: legacy callers without quarter_num keep FY."""
        prev_path = tmp_path / "prev_legacy.xlsx"
        out_path = tmp_path / "out_legacy.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)
        # _run_builder (no quarter_num) → quarter_num defaults to None → no removal
        builder = _run_builder(prev_path, out_path)
        sheet = builder.workbook[builder.config.summary_sheet_name]
        fy_seen = any(
            isinstance(sheet.cell(row=22, column=c).value, str)
            and sheet.cell(row=22, column=c).value.strip().upper().startswith("FY")
            for c in range(1, sheet.max_column + 1)
        )
        assert fy_seen, (
            "Legacy callers without quarter_num must not lose the FY column"
        )


class TestRow106DateLabel:
    """Row 106 'Per DD-MM-YYYY' label must be written unconditionally in the
    LTM column, even when the previous workbook left that cell empty."""

    def test_row_106_label_written_with_period_end(self, tmp_path):
        prev_path = tmp_path / "prev_q4_fy_dates.xlsx"
        out_path = tmp_path / "out_dates.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)
        # Wipe row 106 in the input LTM column so we know the fix is unconditional
        wb = load_workbook(prev_path)
        wb["Management Cijfers - Q4 2025"].cell(row=106, column=31).value = None
        wb.save(prev_path)

        builder = _run_builder_with_quarter_num(
            prev_path, out_path, quarter_num=1
        )
        sheet = builder.workbook[builder.config.summary_sheet_name]
        ltm_val = sheet.cell(row=106, column=builder._new_ltm_col).value
        assert ltm_val == "Per 31-3-2026", (
            f"Expected 'Per 31-3-2026' in LTM row 106, got {ltm_val!r}"
        )

    def test_row_2_ltm_date_is_period_end(self, tmp_path):
        prev_path = tmp_path / "prev_q4_fy_row2.xlsx"
        out_path = tmp_path / "out_row2.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)
        builder = _run_builder_with_quarter_num(
            prev_path, out_path, quarter_num=1
        )
        sheet = builder.workbook[builder.config.summary_sheet_name]
        row2 = sheet.cell(row=2, column=builder._new_ltm_col).value
        assert row2 == datetime(2026, 3, 31), (
            f"Expected period_end 2026-03-31 in row 2 of LTM, got {row2!r}"
        )


class TestQuarterColumnStylingPropagation:
    """Rows 2-21 in the new quarter column must inherit styling (font,
    number_format) from the previous quarter column, not Calibri 11/General."""

    def test_quarter_row_22_number_format_copied(self, tmp_path):
        from openpyxl.styles import Font
        prev_path = tmp_path / "prev_styling.xlsx"
        out_path = tmp_path / "out_styling.xlsx"
        # Build a workbook with explicit Avenir / accounting format on the
        # previous quarter header and row 2 so we can assert propagation.
        _build_prev_q4_workbook(prev_path, with_fy=True)
        wb = load_workbook(prev_path)
        mc = wb["Management Cijfers - Q4 2025"]
        target_num_fmt = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
        mc.cell(row=22, column=29).font = Font(name="Avenir Book", size=10, bold=True)
        mc.cell(row=22, column=29).number_format = target_num_fmt
        mc.cell(row=2, column=29).font = Font(name="Avenir Book", size=10)
        mc.cell(row=2, column=29).number_format = "mm-dd-yy"
        wb.save(prev_path)

        builder = _run_builder_with_quarter_num(
            prev_path, out_path, quarter_num=1
        )
        sheet = builder.workbook[builder.config.summary_sheet_name]
        # The new quarter header should carry the same number format as the
        # previous quarter header (was missing before — fell back to General).
        new_q22 = sheet.cell(row=22, column=builder._new_quarter_col)
        assert new_q22.number_format == target_num_fmt, (
            f"Quarter header number_format not copied: {new_q22.number_format!r}"
        )


class TestBankRowMap:
    """Bank rows 110 and 111 must reference BDO rows 33 and 32 respectively
    (DEP then GEN), matching the client reference workbook."""

    def test_accounting_rules_bank_map(self):
        import yaml
        with open("config/accounting_rules.yaml", encoding="utf-8") as fh:
            rules = yaml.safe_load(fh)
        bank = rules["bank_bdo_row_map"]
        assert bank[110] == 33, f"Row 110 must map to BDO H33 (DEP), got {bank[110]}"
        assert bank[111] == 32, f"Row 111 must map to BDO H32 (GEN), got {bank[111]}"

    def test_bank_formulas_written_to_ltm(self, tmp_path):
        prev_path = tmp_path / "prev_q4_bank.xlsx"
        out_path = tmp_path / "out_bank.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)
        # Add the bank rows in the fixture BDO so the bank section has data
        wb = load_workbook(prev_path)
        bdo = wb["BDO - Q4-25"]
        for raw_row in (31, 32, 33, 34, 35, 36, 37):
            bdo.cell(row=raw_row, column=1, value=f"bank_{raw_row}")
            bdo.cell(row=raw_row, column=8, value=raw_row * 1000)
        wb.save(prev_path)

        builder = _run_builder_with_quarter_num(
            prev_path, out_path, quarter_num=1
        )
        sheet = builder.workbook[builder.config.summary_sheet_name]
        ltm_col = builder._new_ltm_col

        f110 = sheet.cell(row=110, column=ltm_col).value
        f111 = sheet.cell(row=111, column=ltm_col).value
        assert isinstance(f110, str) and "H33" in f110, f"Row 110 should ref H33: {f110}"
        assert isinstance(f111, str) and "H32" in f111, f"Row 111 should ref H32: {f111}"


class TestRow19BdoCorrection:
    """
    Row 19 LTM must contain
    ``=SUM({ltm}3:{ltm}18)-'<bdo>'!H<verschil_row>`` where
    ``<verschil_row>`` is the dynamically-discovered row whose column A
    label is "Verschil balans en winst-en-verlies" in the new BDO sheet.

    The BDO position of that label varies between quarterly files, so
    the formula must NEVER hardcode H134 (which historically referenced
    "Afschrijving prepaid" in some quarterly BDO sheets and silently
    broke reconciliation).
    """

    def test_row19_uses_dynamically_resolved_verschil_row(self, tmp_path):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        prev_path = tmp_path / "prev_row19.xlsx"
        out_path = tmp_path / "out_row19.xlsx"
        _build_prev_q4_workbook(prev_path, with_fy=True)

        # Inject a "Verschil balans en winst-en-verlies" label at a
        # non-default row (135) in the prev-quarter BDO sheet AND in the
        # source data path used by _make_q1_2026_bdo_accounts. The Q1
        # BDO copy step writes account rows starting near row 5, so an
        # injected label at row 135 will be preserved and discoverable
        # by _find_verschil_row().
        wb = load_workbook(prev_path)
        bdo = wb["BDO - Q4-25"]
        verschil_row_in_prev = 135
        bdo.cell(row=verschil_row_in_prev, column=1, value="Verschil balans en winst-en-verlies")
        wb.save(prev_path)
        wb.close()

        builder = _run_builder_with_quarter_num(
            prev_path, out_path, quarter_num=1
        )
        sheet = builder.workbook[builder.config.summary_sheet_name]
        ltm_letter = get_column_letter(builder._new_ltm_col)

        # Discover the verschil row in the freshly-written BDO sheet —
        # the Q1 BDO copy step may place its data at a different offset
        # so we look up the label rather than assume the prev row index.
        new_bdo = builder.workbook[builder.config.bdo_sheet_name]
        resolved_verschil_row = None
        for r in range(1, new_bdo.max_row + 1):
            v = new_bdo.cell(row=r, column=1).value
            if isinstance(v, str) and 'verschil balans' in v.lower():
                resolved_verschil_row = r
                break
        assert resolved_verschil_row is not None, (
            "Test setup error: the new BDO sheet should contain a "
            "'Verschil balans en winst-en-verlies' label."
        )

        f19 = sheet.cell(row=19, column=builder._new_ltm_col).value
        expected = (
            f"=SUM({ltm_letter}3:{ltm_letter}18)"
            f"-'BDO - Q1-26'!H{resolved_verschil_row}"
        )
        assert f19 == expected, (
            f"Row 19 LTM should be {expected!r}, got {f19!r}"
        )
        # Defensive: must NOT be a direct reference to row 68 (a
        # tautology) and must NOT hardcode the legacy H134 row when the
        # Verschil label lives elsewhere.
        assert f19 != f"={ltm_letter}68", (
            "Row 19 must derive from SUM(3:18) - H<verschil>, never "
            "from a direct reference to row 68."
        )
        if resolved_verschil_row != 134:
            assert "H134" not in f19, (
                f"Row 19 must use the dynamically-resolved verschil "
                f"row {resolved_verschil_row}, not the legacy H134."
            )

