"""
Integration tests for Quarterly Report Orchestrator
"""

import pytest
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

from src.orchestrator import QuarterlyReportOrchestrator, QuarterlyReportConfig
from src.parsers.bdo_parser import BDOParseResult, AccountEntry


class TestQuarterlyReportOrchestrator:
    """Integration tests for orchestrator."""
    
    @pytest.fixture
    def sample_input_files(self, tmp_path):
        """Create sample input files for testing."""
        inputs = {}
        
        # BDO file
        bdo_file = tmp_path / "Cijfers_QSP_30-09-2025.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "BalansenWinstverlies"
        headers = ["Grootboek", "Omschrijving", "Beginsaldo", "Mutatie", "Eindsaldo"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=1, value="1600000")
        ws.cell(row=2, column=2, value="Verkrijgingsprijs")
        ws.cell(row=2, column=3, value=12000000.00)
        ws.cell(row=2, column=5, value=12000000.00)
        wb.save(bdo_file)
        inputs['bdo'] = bdo_file
        
        # Previous Management Accounts
        ma_file = tmp_path / "Management_Accounts_Q2_2025.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Management Cijfers - Q2 2025"
        ws.cell(row=1, column=1, value="Management Accounts")
        ws.cell(row=2, column=1, value="Item")
        ws.cell(row=2, column=2, value="2025-06-30")
        wb.save(ma_file)
        inputs['ma'] = ma_file
        
        # Rent roll
        rent_file = tmp_path / "QSP_huurlijst_1-10-2025.xlsx"
        wb = Workbook()
        ws = wb.active
        headers = ["Unit", "Address", "Annual Rent", "Status"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=1, value="A101")
        ws.cell(row=2, column=2, value="Street 1")
        ws.cell(row=2, column=3, value=14400.00)
        ws.cell(row=2, column=4, value="Occupied")
        wb.save(rent_file)
        inputs['rent_roll'] = rent_file
        
        # Sales tracker
        sales_file = tmp_path / "Unit_Sales_tracker_Q3.xlsx"
        wb = Workbook()
        ws = wb.active
        headers = ["Unit", "Address", "Sale Price", "Closing Date"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        ws.cell(row=2, column=1, value="B201")
        ws.cell(row=2, column=2, value="Richtersweg 344")
        ws.cell(row=2, column=3, value=203555.00)
        ws.cell(row=2, column=4, value="2025-09-20")
        wb.save(sales_file)
        inputs['sales'] = sales_file
        
        # Compliance calc
        compliance_file = tmp_path / "Compliance_Q2_2025.xlsx"
        wb = Workbook()
        wb.save(compliance_file)
        inputs['compliance'] = compliance_file
        
        # Word template
        word_file = tmp_path / "Quarterly_QSP_Q2_2025_Draft.docx"
        # Create minimal docx (would need python-docx for real file)
        word_file.touch()
        inputs['word'] = word_file
        
        return inputs
    
    @pytest.fixture
    def config(self, sample_input_files, tmp_path):
        """Create orchestrator config."""
        return QuarterlyReportConfig(
            year=2025,
            quarter=3,
            bdo_file=sample_input_files['bdo'],
            previous_management_accounts=sample_input_files['ma'],
            rent_roll_file=sample_input_files['rent_roll'],
            sales_tracker_file=sample_input_files['sales'],
            previous_compliance_calc=sample_input_files['compliance'],
            word_template=sample_input_files['word'],
            output_dir=tmp_path / "outputs"
        )
    
    def test_orchestrator_initialization(self, config):
        """Test orchestrator initialization."""
        orchestrator = QuarterlyReportOrchestrator(config)
        
        assert orchestrator.config == config
        assert orchestrator.bdo_parser is not None
        assert orchestrator.rent_roll_parser is not None
        assert orchestrator.sales_parser is not None
        assert orchestrator.compliance_calc is not None
        assert orchestrator.balance_validator is not None
    
    def test_dry_run_mode(self, config):
        """Test dry run mode (validation only)."""
        orchestrator = QuarterlyReportOrchestrator(config)
        results = orchestrator.run(dry_run=True)
        
        assert results['status'] == 'success'
        assert results['dry_run'] is True
        assert 'bdo_parse' in results['steps']
        assert 'rent_roll_parse' in results['steps']
        assert 'sales_parse' in results['steps']
        # Should not have file generation steps
        assert 'management_accounts' not in results['steps']
        assert 'word_update' not in results['steps']
        assert 'pdf_assembly' not in results['steps']
    
    def test_parse_steps(self, config):
        """Test parsing steps in workflow."""
        orchestrator = QuarterlyReportOrchestrator(config)
        results = orchestrator.run(dry_run=True)
        
        # Step 1: BDO parse
        assert results['steps']['bdo_parse']['status'] == 'success'
        assert 'accounts_parsed' in results['steps']['bdo_parse']
        
        # Step 2: Rent roll parse
        assert results['steps']['rent_roll_parse']['status'] == 'success'
        assert 'units_count' in results['steps']['rent_roll_parse']
        
        # Step 3: Sales parse
        assert results['steps']['sales_parse']['status'] == 'success'
        assert 'quarter_sales' in results['steps']['sales_parse']
    
    def test_results_structure(self, config):
        """Test results dictionary structure."""
        orchestrator = QuarterlyReportOrchestrator(config)
        results = orchestrator.run(dry_run=True)
        
        # Check required keys
        assert 'quarter' in results
        assert 'started_at' in results
        assert 'status' in results
        assert 'steps' in results
        assert 'warnings' in results
        assert 'errors' in results
        assert 'output_files' in results
        
        # Check quarter string
        assert results['quarter'] == "Q3 2025"
    
    def test_error_handling(self, tmp_path):
        """Test error handling for missing files."""
        config = QuarterlyReportConfig(
            year=2025,
            quarter=3,
            bdo_file=tmp_path / "nonexistent.xlsx",
            previous_management_accounts=tmp_path / "nonexistent.xlsx",
            rent_roll_file=tmp_path / "nonexistent.xlsx",
            sales_tracker_file=tmp_path / "nonexistent.xlsx",
            previous_compliance_calc=tmp_path / "nonexistent.xlsx",
            word_template=tmp_path / "nonexistent.docx",
            output_dir=tmp_path / "outputs"
        )
        
        orchestrator = QuarterlyReportOrchestrator(config)
        results = orchestrator.run(dry_run=True)
        
        # Should have failed status
        assert results['status'] == 'failed'
        assert len(results['errors']) > 0
    
    def test_config_properties(self, config):
        """Test QuarterlyReportConfig properties."""
        assert config.quarter_str == "Q3 2025"
        assert config.previous_quarter_str == "Q2 2025"
        assert config.period_end == datetime(2025, 9, 30)
        
        # Test Q1 previous quarter (should be Q4 of previous year)
        config_q1 = QuarterlyReportConfig(
            year=2025,
            quarter=1,
            bdo_file=Path("dummy.xlsx"),
            previous_management_accounts=Path("dummy.xlsx"),
            rent_roll_file=Path("dummy.xlsx"),
            sales_tracker_file=Path("dummy.xlsx"),
            previous_compliance_calc=Path("dummy.xlsx"),
            word_template=Path("dummy.docx"),
            output_dir=Path("dummy")
        )
        assert config_q1.previous_quarter_str == "Q4 2024"
        assert config_q1.period_end == datetime(2025, 3, 31)

